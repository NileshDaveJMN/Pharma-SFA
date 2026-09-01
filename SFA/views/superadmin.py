import openpyxl
from django.db import transaction
from django.contrib import messages
from django.contrib.auth.models import User, Group
from django.shortcuts import render, redirect
from SFA.models import Company, Territory, Employee, DARate, TARate, Product, Route, Stockist, SystemSetting

def onboard_company_view(request):
    if not request.user.is_superuser:
        messages.error(request, "Access Denied! Only Super Admins can onboard companies.")
        return redirect('/login/')

    if request.method == 'POST' and request.FILES.get('onboard_file'):
        excel_file = request.FILES['onboard_file']
        
        # 🌟 NAYA: Checkbox ki value POST request se nikali ('on' aata hai HTML se agar checked ho)
        is_digital_va = request.POST.get('is_digital_va_enabled') == 'on'
        
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True, read_only=True)

            with transaction.atomic():
                # 1. CREATE COMPANY & SETTINGS
                sheet_company = wb['Company']
                comp_name = str(sheet_company.cell(row=2, column=1).value).strip()
                comp_code = str(sheet_company.cell(row=2, column=2).value).strip().upper()
                
                # 🌟 NAYA: Company model mein is_digital_va_enabled save karwaya
                new_company = Company.objects.create(
                    name=comp_name, 
                    code=comp_code, 
                    slug=comp_code.lower(),
                    is_digital_va_enabled=is_digital_va
                )
                SystemSetting.objects.create(company=new_company)

                # 2. CREATE TERRITORIES
                sheet_terr = wb['Territories']
                territory_dict = {}
                for row in sheet_terr.iter_rows(min_row=2, values_only=True):
                    if row[0]:
                        t_name = str(row[0]).strip()
                        t = Territory.objects.create(company=new_company, name=t_name, city=t_name)
                        territory_dict[t_name.lower()] = t

                # 3. EXPENSE RATES (DA/TA)
                sheet_rates = wb['Expense Rates']
                for row in sheet_rates.iter_rows(min_row=2, values_only=True):
                    if row[0]:
                        desig = str(row[0]).strip()
                        DARate.objects.create(company=new_company, designation=desig, hq_da=row[1], exhq_da=row[2], outstation_da=row[3])
                        TARate.objects.create(company=new_company, designation=desig, slab1_upto_km=row[4], slab1_rate=row[5], slab2_upto_km=row[6], slab2_rate=row[7])

                # 4. EMPLOYEES & HIERARCHY
                sheet_emp = wb['Hierarchy']
                emp_objects = {}
                
                admin_group, _ = Group.objects.get_or_create(name='Company Admin')
                
                for row in sheet_emp.iter_rows(min_row=2, values_only=True):
                    if row[0]:
                        name, desig, phone, hq_name, manager_name = row
                        name = str(name).strip()
                        phone = str(phone).strip()
                        desig_str = str(desig).strip()
                        
                        django_username = f"{comp_code}_{phone}"
                        name_parts = name.split()
                        surname = name_parts[-1].capitalize() if len(name_parts) > 1 else name.capitalize()
                        default_password = f"{surname}@123"

                        user = User.objects.create_user(username=django_username, password=default_password)
                        
                        if desig_str.lower() == 'admin':
                            user.is_staff = True
                            user.save()
                            user.groups.add(admin_group)

                        hq_obj = territory_dict.get(str(hq_name).strip().lower()) if hq_name else None
                        emp_count = Employee.objects.filter(company=new_company).count() + 1
                        emp_code = f"{comp_code}-{emp_count:03d}"

                        emp = Employee.objects.create(
                            user=user, company=new_company, name=name, 
                            employee_code=emp_code, designation=desig_str, 
                            phone=phone, headquarter=hq_obj
                        )
                        emp_objects[name.lower()] = emp

                # PASS 2: Assign Managers
                for row in sheet_emp.iter_rows(min_row=2, values_only=True):
                    if row[0]:
                        name, _, _, _, manager_name = row
                        if manager_name:
                            emp_obj = emp_objects.get(str(name).strip().lower())
                            mgr_obj = emp_objects.get(str(manager_name).strip().lower())
                            if emp_obj and mgr_obj:
                                emp_obj.manager = mgr_obj
                                emp_obj.save()

                # 5. PRODUCTS (Updated with GST support)
                sheet_prod = wb['Products']
                for row in sheet_prod.iter_rows(min_row=2, values_only=True):
                    if row[0]:
                        raw_gst = str(row[5]).replace('%', '').strip() if len(row) > 5 and row[5] is not None else '12'
                        
                        try:
                            final_gst = int(float(raw_gst))
                            if final_gst not in [0, 5, 12, 18, 28]:
                                final_gst = 12
                        except ValueError:
                            final_gst = 12

                        Product.objects.create(
                            company=new_company, 
                            name=str(row[0]).strip(), 
                            pack_size=str(row[1]).strip(), 
                            mrp=row[2], 
                            ptr=row[3], 
                            pts=row[4],
                            gst_slab=final_gst
                        )

                # 6. ROUTES
                sheet_route = wb['Routes']
                for row in sheet_route.iter_rows(min_row=2, values_only=True):
                    if row[0]:
                        hq_name = str(row[1]).strip().lower() if row[1] else ''
                        hq_obj = territory_dict.get(hq_name)
                        if hq_obj:
                            Route.objects.create(company=new_company, name=str(row[0]).strip(), territory=hq_obj, status='Approved')

                # 7. STOCKISTS
                sheet_stockist = wb['Stockists']
                for row in sheet_stockist.iter_rows(min_row=2, values_only=True):
                    if row[0]:
                        hq_name = str(row[1]).strip().lower() if row[1] else ''
                        hq_obj = territory_dict.get(hq_name)
                        if hq_obj:
                            Stockist.objects.create(company=new_company, name=str(row[0]).strip(), territory=hq_obj, phone="9999999999", contact_person="Owner")

            messages.success(request, f"🎉 Successfully onboarded Company '{comp_name}' with {len(emp_objects)} employees!")
            return redirect('onboard_company')

        except Exception as e:
            messages.error(request, f"❌ Upload Failed! Error: {str(e)}")

    return render(request, 'company_onboard.html')
