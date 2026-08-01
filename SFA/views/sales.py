import io
import csv
import calendar
from collections import defaultdict
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from SFA.models import  SystemSetting
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.db.models import Sum
from django.http import HttpResponse

from SFA.models import (
    Employee, Stockist, Chemist, Doctor, Product, PrimarySale,
    StockistProductStatement, MonthlyTargetMaster, PartyWiseSaleReport, 
    TerritoryTarget, PartyWiseSaleLine, DoctorRxMapping, Territory, 
    PromoDispatch, MRInventory
)
from .auth import get_full_team_employees, get_dropdown_team, get_team_hq_territory_ids
from SFA.decorators import employee_required


# ==============================================================================
# 2. 📊 SMART SECONDARY STOCKIST STATEMENT (AUTO-GENERATED)
# ==============================================================================
@employee_required
def smart_secondary_report_view(request, employee):
    team_employees = get_dropdown_team(employee, ordered=False)
    default_emp_id = str(employee.id)
    if employee.designation != 'MR':
        first_sub = team_employees.exclude(id=employee.id).first()
        if first_sub:
            default_emp_id = str(first_sub.id)
            
    selected_emp_id = request.GET.get('employee_id', default_emp_id)
    selected_emp = get_object_or_404(Employee, id=int(selected_emp_id))
    
    today = timezone.now().date()
    selected_month = int(request.GET.get('month') or today.month)
    selected_year = int(request.GET.get('year') or today.year)

    selected_team_scope = get_full_team_employees(selected_emp)
    area_territory_ids = selected_team_scope.exclude(headquarter__isnull=True).values_list('headquarter_id', flat=True)
    
    stockists = Stockist.objects.filter(territory_id__in=area_territory_ids)
    stockist_dict = {st.id: st.name for st in stockists}
    products_dict = {p.id: {'name': p.name, 'price': float(p.price) if getattr(p, 'price', None) else 0.0} for p in Product.objects.filter(company=selected_emp.company).order_by('name')}

    curr_val = selected_year * 12 + selected_month
    statements = StockistProductStatement.objects.filter(stockist_id__in=stockist_dict.keys()).values('stockist_id', 'product_id', 'month', 'year').annotate(op=Sum('opening_qty'), rec=Sum('received_qty'))
    secondaries = PartyWiseSaleLine.objects.filter(report__stockist_id__in=stockist_dict.keys()).values('report__stockist_id', 'product_id', 'report__month', 'report__year').annotate(tb=Sum('billed_qty'), tf=Sum('free_qty'))

    data_map = defaultdict(lambda: defaultdict(lambda: {'op': 0, 'rec': 0, 'tb': 0, 'tf': 0}))

    for st in statements:
        st_val = st['year'] * 12 + st['month']
        if st_val < curr_val:
            data_map[st['stockist_id']][st['product_id']]['op'] += (st['op'] or 0) + (st['rec'] or 0)
        elif st_val == curr_val:
            data_map[st['stockist_id']][st['product_id']]['op'] += (st['op'] or 0)
            data_map[st['stockist_id']][st['product_id']]['rec'] += (st['rec'] or 0)

    for sec in secondaries:
        sec_val = sec['report__year'] * 12 + sec['report__month']
        if sec_val < curr_val:
            data_map[sec['report__stockist_id']][sec['product_id']]['op'] -= ((sec['tb'] or 0) + (sec['tf'] or 0))
        elif sec_val == curr_val:
            data_map[sec['report__stockist_id']][sec['product_id']]['tb'] += (sec['tb'] or 0)
            data_map[sec['report__stockist_id']][sec['product_id']]['tf'] += (sec['tf'] or 0)

    report_data, gt_op_qty, gt_op_val, gt_pr_qty, gt_pr_val, gt_sec_b, gt_sec_f, gt_sec_val, gt_cl_qty, gt_cl_val = [], 0, 0.0, 0, 0.0, 0, 0, 0.0, 0, 0.0

    for s_id, s_name in stockist_dict.items():
        if s_id not in data_map: continue
        st_row = {'name': s_name, 'products': []}
        
        for p_id, p_info in products_dict.items():
            if p_id not in data_map[s_id]: continue
            m_data = data_map[s_id][p_id]
            op, rec, tb, tf = m_data['op'], m_data['rec'], m_data['tb'], m_data['tf']
            cl = op + rec - (tb + tf)
            
            if op == 0 and rec == 0 and tb == 0 and tf == 0 and cl == 0: continue
                
            price = p_info['price']
            op_val, pr_val, sec_val, cl_val = op * price, rec * price, tb * price, cl * price
            
            st_row['products'].append({
                'name': p_info['name'], 'op': op, 'op_val': op_val, 'rec': rec, 'pr_val': pr_val,
                'sec_b': tb, 'sec_f': tf, 'sec_val': sec_val, 'cl': cl, 'cl_val': cl_val
            })
            gt_op_qty += op; gt_op_val += op_val; gt_pr_qty += rec; gt_pr_val += pr_val
            gt_sec_b += tb; gt_sec_f += tf; gt_sec_val += sec_val; gt_cl_qty += cl; gt_cl_val += cl_val
            
        if st_row['products']: report_data.append(st_row)
    report_data.sort(key=lambda x: x['name'])

    if request.GET.get('export') == 'excel':
        filename = f"Stockist_Statement_{selected_emp.name}_{selected_month}_{selected_year}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stockist Statement"
        
        period_str = f"{calendar.month_name[selected_month]} {selected_year}"
        ws.append(['SMART STOCKIST STATEMENT (MONTHLY)'])
        ws.append(['Employee:', selected_emp.name, 'Month:', period_str])
        ws.append([''])
        ws['A1'].font = Font(bold=True, size=14, color="107C41")
        ws['A2'].font = Font(bold=True); ws['C2'].font = Font(bold=True)
        
        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        
        headers = ['Stockist Name', 'Product Name', 'Opening Qty', 'Opening Val (₹)', 'Primary Qty (Recv)', 'Primary Val (₹)', 'Secondary Billed Qty', 'Secondary Free Qty', 'Secondary Val (₹)', 'Closing Qty', 'Closing Val (₹)']
        ws.append(headers)
        
        for col_num, cell in enumerate(ws[5], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 18 if col_num > 2 else 22
            
        for st_data in report_data:
            for prod in st_data['products']:
                ws.append([st_data['name'], prod['name'], prod['op'], round(prod['op_val'], 2), prod['rec'], round(prod['pr_val'], 2), prod['sec_b'], prod['sec_f'], round(prod['sec_val'], 2), prod['cl'], round(prod['cl_val'], 2)])
                
        ws.append([''])
        ws.append(['GRAND TOTAL', '', gt_op_qty, round(gt_op_val, 2), gt_pr_qty, round(gt_pr_val, 2), gt_sec_b, gt_sec_f, round(gt_sec_val, 2), gt_cl_qty, round(gt_cl_val, 2)])
        for cell in ws[ws.max_row]: cell.font = Font(bold=True)
            
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return render(request, 'smart_stockist_statement.html', {
        'team_employees': team_employees, 'selected_emp_id': int(selected_emp_id),
        'selected_month': selected_month, 'selected_year': selected_year,
        'months_choices': [(i, calendar.month_name[i]) for i in range(1, 13)],
        'report_data': report_data,
        'gt_op_qty': gt_op_qty, 'gt_op_val': gt_op_val, 'gt_pr_qty': gt_pr_qty, 'gt_pr_val': gt_pr_val,
        'gt_sec_b': gt_sec_b, 'gt_sec_f': gt_sec_f, 'gt_sec_val': gt_sec_val, 'gt_cl_qty': gt_cl_qty, 'gt_cl_val': gt_cl_val,
        'is_manager_view': employee.designation != 'MR'
    })
    
# ==============================================================================
# 3. 📦 PRIMARY SALES REPORT
# ==============================================================================
@employee_required
def primary_sales_report_view(request, employee):
    team_employees = get_dropdown_team(employee, ordered=False)
    default_emp_id = str(employee.id)
    if employee.designation != 'MR':
        first_sub = team_employees.exclude(id=employee.id).first()
        if first_sub:
            default_emp_id = str(first_sub.id)
            
    selected_emp_id = request.GET.get('employee_id', default_emp_id)
    selected_emp = get_object_or_404(Employee, id=int(selected_emp_id))
    
    today = timezone.now().date()
    from_month = int(request.GET.get('from_month') or today.month)
    to_month = int(request.GET.get('to_month') or today.month)
    selected_year = int(request.GET.get('year') or today.year)

    if from_month > to_month:
        from_month, to_month = to_month, from_month

    months_range = list(range(from_month, to_month + 1))
    months_headers = [(m, calendar.month_name[m][:3]) for m in months_range]

    selected_team_scope = get_full_team_employees(selected_emp)
    area_territory_ids = selected_team_scope.exclude(headquarter__isnull=True).values_list('headquarter_id', flat=True)
    
    primary_sales = PrimarySale.objects.filter(
        stockist__territory_id__in=area_territory_ids, date__month__gte=from_month, date__month__lte=to_month, date__year=selected_year
    ).select_related('stockist', 'product').order_by('-date')

    party_wise_data = defaultdict(lambda: {'party_name': '', 'products': defaultdict(lambda: {'monthly': {m: {'qty': 0, 'val': 0.0} for m in months_range}, 'total_qty': 0, 'total_val': 0.0}), 'total_qty': 0, 'total_val': 0.0})
    product_wise_data = defaultdict(lambda: {'product_name': '', 'monthly': {m: {'qty': 0, 'val': 0.0} for m in months_range}, 'total_qty': 0, 'total_val': 0.0})
    gt_monthly = {m: {'qty': 0, 'val': 0.0} for m in months_range}
    gt_qty, gt_val = 0, 0.0

    for sale in primary_sales:
        s_month, st_id, pr_id, pr_name, qty = sale.date.month, sale.stockist.id, sale.product.id, sale.product.name, sale.quantity
        price = float(sale.product.price) if getattr(sale.product, 'price', None) else 0.0
        val = qty * price

        party_wise_data[st_id]['party_name'] = sale.stockist.name
        party_wise_data[st_id]['products'][pr_name]['monthly'][s_month]['qty'] += qty
        party_wise_data[st_id]['products'][pr_name]['monthly'][s_month]['val'] += val
        party_wise_data[st_id]['products'][pr_name]['total_qty'] += qty
        party_wise_data[st_id]['products'][pr_name]['total_val'] += val
        party_wise_data[st_id]['total_qty'] += qty
        party_wise_data[st_id]['total_val'] += val

        product_wise_data[pr_id]['product_name'] = pr_name
        product_wise_data[pr_id]['monthly'][s_month]['qty'] += qty
        product_wise_data[pr_id]['monthly'][s_month]['val'] += val
        product_wise_data[pr_id]['total_qty'] += qty
        product_wise_data[pr_id]['total_val'] += val

        if s_month in gt_monthly:
            gt_monthly[s_month]['qty'] += qty
            gt_monthly[s_month]['val'] += val
        gt_qty += qty; gt_val += val

    party_list = []
    for pid, pdata in party_wise_data.items():
        prod_list = []
        for name, pd in pdata['products'].items():
            pd['name'] = name; pd['monthly_list'] = [pd['monthly'][m] for m in months_range]
            prod_list.append(pd)
        prod_list.sort(key=lambda x: x['name'])
        party_list.append({'party_name': pdata['party_name'], 'products': prod_list, 'total_qty': pdata['total_qty'], 'total_val': pdata['total_val']})
    party_list.sort(key=lambda x: x['party_name'])

    prod_list_final = []
    for pr_id, pdata in product_wise_data.items():
        pdata['monthly_list'] = [pdata['monthly'][m] for m in months_range]
        prod_list_final.append(pdata)
    prod_list_final.sort(key=lambda x: x['product_name'])
    gt_monthly_list = [gt_monthly[m] for m in months_range]

    if request.GET.get('export') == 'excel':
        filename = f"Primary_Sales_{selected_emp.name}_M{from_month}-M{to_month}_{selected_year}.xlsx"
        wb = openpyxl.Workbook()
        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        period_str = f"{calendar.month_name[from_month][:3]} to {calendar.month_name[to_month][:3]} {selected_year}"

        ws1 = wb.active
        ws1.title = "Stockist Wise"
        ws1.append(['PRIMARY SALES - STOCKIST WISE (MONTH TREND)'])
        ws1.append(['Employee:', selected_emp.name, 'Period:', period_str])
        ws1.append([''])
        ws1['A1'].font = Font(bold=True, size=12, color="107C41")
        
        headers1 = ['Stockist Name', 'Product Name']
        for m_num, m_name in months_headers:
            headers1.extend([f'{m_name} Qty', f'{m_name} Value (₹)'])
        headers1.extend(['Grand Total Qty', 'Grand Total Value (₹)'])
        ws1.append(headers1)
        for col_num, cell in enumerate(ws1[4], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws1.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 15 if col_num > 2 else 25

        for party in party_list:
            for prod in party['products']:
                row_data = [party['party_name'], prod['name']]
                for m_data in prod['monthly_list']: row_data.extend([m_data['qty'], round(m_data['val'], 2)])
                row_data.extend([prod['total_qty'], round(prod['total_val'], 2)])
                ws1.append(row_data)

        gt_row1 = ['GRAND TOTAL', '']
        for m_gt in gt_monthly_list: gt_row1.extend([m_gt['qty'], round(m_gt['val'], 2)])
        gt_row1.extend([gt_qty, round(gt_val, 2)])
        ws1.append(gt_row1)
        for cell in ws1[ws1.max_row]: cell.font = Font(bold=True)

        ws2 = wb.create_sheet(title="Product Wise")
        ws2.append(['PRIMARY SALES - PRODUCT WISE (MONTH TREND)'])
        ws2.append(['Employee:', selected_emp.name, 'Period:', period_str])
        ws2.append([''])
        ws2['A1'].font = Font(bold=True, size=12, color="107C41")
        
        headers2 = ['Product Name']
        for m_num, m_name in months_headers: headers2.extend([f'{m_name} Qty', f'{m_name} Value (₹)'])
        headers2.extend(['Grand Total Qty', 'Grand Total Value (₹)'])
        ws2.append(headers2)

        for col_num, cell in enumerate(ws2[4], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws2.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 15 if col_num > 1 else 25

        for prod in prod_list_final:
            row_data = [prod['product_name']]
            for m_data in prod['monthly_list']: row_data.extend([m_data['qty'], round(m_data['val'], 2)])
            row_data.extend([prod['total_qty'], round(prod['total_val'], 2)])
            ws2.append(row_data)

        gt_row2 = ['GRAND TOTAL']
        for m_gt in gt_monthly_list: gt_row2.extend([m_gt['qty'], round(m_gt['val'], 2)])
        gt_row2.extend([gt_qty, round(gt_val, 2)])
        ws2.append(gt_row2)
        for cell in ws2[ws2.max_row]: cell.font = Font(bold=True)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return render(request, 'primary_sales_report.html', {
        'team_employees': team_employees, 'selected_emp_id': int(selected_emp_id),
        'from_month': from_month, 'to_month': to_month, 'selected_year': selected_year,
        'months_choices': [(i, calendar.month_name[i]) for i in range(1, 13)], 'months_headers': months_headers,
        'primary_sales': primary_sales, 'party_list': party_list, 'product_list': prod_list_final,
        'gt_monthly_list': gt_monthly_list, 'grand_total_qty': gt_qty, 'grand_total_value': gt_val,
        'is_manager_view': employee.designation != 'MR'
    })

@employee_required
def party_wise_sale_entry_view(request, employee):
    team_employees = get_dropdown_team(employee).filter(designation='MR', is_active=True)
    is_manager_view = employee.designation != 'MR'
    
    selected_emp_id = request.GET.get('employee_id') or request.POST.get('employee_id')
    if not selected_emp_id:
        selected_emp_id = str(employee.id)
        
    selected_emp = get_object_or_404(Employee, id=selected_emp_id)
    
    # Hamesha Pichla Mahina (Last Month)
    today = timezone.now().date()
    current_month, current_year = (12, today.year - 1) if today.month == 1 else (today.month - 1, today.year)
    
    from SFA.models import SystemSetting
    setting = SystemSetting.objects.filter(company=employee.company).first()
    deadline = setting.sale_upload_deadline_day if setting and setting.sale_upload_deadline_day else 4
    is_locked = today.day > deadline
    
    my_terr_ids = [selected_emp.headquarter_id] if selected_emp.headquarter_id else []
    available_stockists = Stockist.objects.filter(territory_id__in=my_terr_ids).order_by('name')
    
    selected_stockist_id = request.GET.get('stockist_id') or request.POST.get('stockist_id')
    
    if selected_stockist_id and not available_stockists.filter(id=selected_stockist_id).exists():
        selected_stockist_id = None

    if not selected_stockist_id and available_stockists.exists():
        selected_stockist_id = str(available_stockists.first().id)
        
    selected_stockist = available_stockists.filter(id=selected_stockist_id).first()
    selected_stockist_name = selected_stockist.name if selected_stockist else "No Stockist"

    if request.method == 'POST':
        if is_locked and employee.designation not in ['Admin', 'System Administrator']:
            messages.error(request, f"⚠️ Edit Locked! Last month's entry is only allowed until the {deadline} of each month. Please contact Admin.")
            return redirect(f'/reports/party-wise-sale/?stockist_id={selected_stockist_id}&employee_id={selected_emp.id}')

        posted_stockist_id = request.POST.get('stockist_id') or selected_stockist_id
        if not posted_stockist_id:
            messages.error(request, "⚠️ Stockist missing! Please select a Stockist.")
            return redirect(f'/reports/party-wise-sale/?employee_id={selected_emp.id}')
            
        posted_stockist = get_object_or_404(Stockist, id=posted_stockist_id)
        chemist_id = request.POST.get('chemist_id')
        
        if not chemist_id:
            messages.error(request, "⚠️ Selecting a Chemist is required!")
            return redirect(f'/reports/party-wise-sale/?stockist_id={posted_stockist.id}&employee_id={selected_emp.id}')
            
        chemist = get_object_or_404(Chemist, id=chemist_id)
        report, created = PartyWiseSaleReport.objects.get_or_create(
            employee=selected_emp, stockist=posted_stockist, month=current_month, year=current_year
        )
        
        saved_any = False
        for prod in Product.objects.filter(company=selected_emp.company):
            b_qty = int(request.POST.get(f'billed_{prod.id}') or 0)
            f_qty = int(request.POST.get(f'free_{prod.id}') or 0)
            if b_qty > 0 or f_qty > 0:
                PartyWiseSaleLine.objects.create(report=report, chemist=chemist, product=prod, billed_qty=b_qty, free_qty=f_qty)
                saved_any = True

        if saved_any:
            messages.success(request, f"✅ Sale saved for {chemist.name} (By: {selected_emp.name}).")
        else:
            messages.warning(request, "⚠️ The form was blank or quantities were 0, so nothing was saved.")
            
        return redirect(f'/reports/party-wise-sale/?stockist_id={posted_stockist.id}&employee_id={selected_emp.id}')

    # ==========================================
    # 🌟 NEW STOCK BALANCE CALCULATION (Exact & Foolproof Logic)
    # ==========================================
    balances = []
    if selected_stockist:
        import calendar
        from datetime import date
        from django.db.models import Q, Sum
        
        # Target mahine (Last Month) ka aakhiri din nikal lo
        last_day_of_target_month = date(current_year, current_month, calendar.monthrange(current_year, current_month)[1])

        for prod in Product.objects.filter(company=selected_emp.company):
            # 1. Company se received Total Primary Sale (Sirf target mahine ke end tak)
            primary_agg = PrimarySale.objects.filter(
                stockist=selected_stockist, 
                product=prod,
                date__lte=last_day_of_target_month
            ).aggregate(tot_qty=Sum('quantity'), tot_free=Sum('free_quantity'))
            total_lifetime_primary = (primary_agg['tot_qty'] or 0) + (primary_agg['tot_free'] or 0)
            
            # 2. Pichle saare mahinon mein jo Secondary Sale ho chuki hai
            past_party_agg = PartyWiseSaleLine.objects.filter(
                report__stockist=selected_stockist, 
                product=prod
            ).filter(
                Q(report__year__lt=current_year) | Q(report__year=current_year, report__month__lt=current_month)
            ).aggregate(tb=Sum('billed_qty'), tf=Sum('free_qty'))
            total_past_secondary = (past_party_agg['tb'] or 0) + (past_party_agg['tf'] or 0)
            
            # 3. EXACT STOCK: Jo is mahine MR bill kar sakta hai (Total stock minus purani sale)
            stock_available_for_this_month = total_lifetime_primary - total_past_secondary
            
            # 4. Is current entry mein MR ne ab tak kitna bill kar diya hai
            current_party_agg = PartyWiseSaleLine.objects.filter(
                report__stockist=selected_stockist, 
                product=prod,
                report__month=current_month,
                report__year=current_year
            ).aggregate(tb=Sum('billed_qty'), tf=Sum('free_qty'))
            billed_this_month = (current_party_agg['tb'] or 0) + (current_party_agg['tf'] or 0)
            
            # 5. Remaining Balance
            current_balance = stock_available_for_this_month - billed_this_month
            
            # Agar stock zero hai toh line gayab kar do, taaki table clean rahe
            if stock_available_for_this_month > 0 or billed_this_month > 0:
                balances.append({
                    'product_id': prod.id, 
                    'product_name': prod.name, 
                    'total_sale': stock_available_for_this_month, 
                    'billed': billed_this_month, 
                    'balance': current_balance
                })

    chemists = Chemist.objects.filter(allocated_to=selected_emp, status='Approved').order_by('name')

    return render(request, 'party_wise_sale.html', {
        'team_employees': team_employees, 'is_manager_view': is_manager_view,
        'selected_emp_id': int(selected_emp_id), 'available_stockists': available_stockists,
        'selected_stockist_id': int(selected_stockist_id) if selected_stockist_id else '',
        'selected_stockist_name': selected_stockist_name, 'balances': balances,
        'chemists': chemists, 'products': Product.objects.filter(company=selected_emp.company),
        'current_month': current_month, 'current_year': current_year,
        'is_locked': is_locked, 'deadline': deadline
    })


# ==============================================================================
# CLASSIFY RX ENTRY VIEW UPDATE
# ==============================================================================
@employee_required
def classify_rx_entry_view(request, employee):
    team_employees = get_dropdown_team(employee).filter(designation='MR', is_active=True)
    is_manager_view = employee.designation != 'MR'
    
    selected_emp_id = request.GET.get('employee_id') or request.POST.get('employee_id')
    if not selected_emp_id: selected_emp_id = str(employee.id)
        
    selected_emp = get_object_or_404(Employee, id=selected_emp_id)
    
    # 🌟 NAYA RULE 1: Hamesha Pichla Mahina hi khulega
    today = timezone.now().date()
    month, year = (12, today.year - 1) if today.month == 1 else (today.month - 1, today.year)
    
    # 🌟 NAYA RULE 2: Deadline Lock Check
    from SFA.models import SystemSetting
    setting = SystemSetting.objects.filter(company=employee.company).first()
    deadline = setting.sale_upload_deadline_day if setting and setting.sale_upload_deadline_day else 4
    is_locked = today.day > deadline
    
    my_terr_ids = [selected_emp.headquarter_id] if selected_emp.headquarter_id else []
    stockists = Stockist.objects.filter(territory_id__in=my_terr_ids).order_by('name')
    
    stockist_id = request.GET.get('stockist_id') or request.POST.get('stockist_id')
    if stockist_id and not stockists.filter(id=stockist_id).exists(): stockist_id = None
    if not stockist_id and stockists.exists(): stockist_id = stockists.first().id

    stockist = None
    report = None
    lines_data = []

    if not stockist_id: 
        messages.error(request, f"No Stockist is mapped in {selected_emp.name}'s territory!")
    else:
        stockist = get_object_or_404(Stockist, id=stockist_id)
        report = PartyWiseSaleReport.objects.filter(stockist=stockist, month=month, year=year, employee=selected_emp).first()
        
        if report:
            for line in PartyWiseSaleLine.objects.filter(report=report).select_related('chemist', 'product'):
                mappings = line.dr_mappings.select_related('doctor').all()
                mapped_billed = sum(m.mapped_billed_qty for m in mappings)
                mapped_free = sum(m.mapped_free_qty for m in mappings)
                lines_data.append({
                    'line': line, 'mapped_billed': mapped_billed, 'mapped_free': mapped_free,
                    'bal_billed': line.billed_qty - mapped_billed, 'bal_free': line.free_qty - mapped_free,
                    'mappings': mappings
                })

    if request.method == "POST":
        # 🌟 STRICT BLOCKER: Deadline ke baad lock
        if is_locked and employee.designation not in ['Admin', 'System Administrator']:
            messages.error(request, f"⚠️ Edit Locked! Classify Rx entries are only allowed until the {deadline} of each month. Please contact Admin.")
            return redirect(f"{request.path}?stockist_id={stockist_id}&employee_id={selected_emp.id}")

        target_line = PartyWiseSaleLine.objects.get(id=request.POST.get('line_id'))
        m_billed = int(request.POST.get('mapped_billed', 0) or 0)
        m_free = int(request.POST.get('mapped_free', 0) or 0)
        
        current_mapped_billed = sum(m.mapped_billed_qty for m in target_line.dr_mappings.all())
        current_mapped_free = sum(m.mapped_free_qty for m in target_line.dr_mappings.all())
        
        if m_billed > (target_line.billed_qty - current_mapped_billed) or m_free > (target_line.free_qty - current_mapped_free): 
            messages.error(request, "Error! Allocation qty cannot be more than the Balance.")
        elif m_billed > 0 or m_free > 0: 
            DoctorRxMapping.objects.create(party_line=target_line, doctor_id=request.POST.get('doctor_id'), mapped_billed_qty=m_billed, mapped_free_qty=m_free)
            messages.success(request, f"Rx Classified for {selected_emp.name}!")
            
        return redirect(f"{request.path}?stockist_id={stockist_id}&month={month}&year={year}&employee_id={selected_emp.id}")

    return render(request, 'classify_rx.html', {
        'team_employees': team_employees, 'is_manager_view': is_manager_view, 'selected_emp_id': int(selected_emp_id),
        'stockists': stockists, 'selected_stockist_id': int(stockist_id) if stockist_id else '',
        'stockist': stockist, 'report': report, 'lines_data': lines_data, 
        'doctors': Doctor.objects.filter(allocated_to=selected_emp, status='Approved').order_by('name'), 
        'month': month, 'year': year,
        'is_locked': is_locked, 'deadline': deadline # Frontend ko status bheja
    })

# ==============================================================================
# 5. PARTY & RX SALE REPORT VIEW
# ==============================================================================
@employee_required
def party_rx_report_view(request, employee):
    team_employees = get_dropdown_team(employee, ordered=False)
    default_emp_id = str(employee.id)
    if employee.designation != 'MR':
        first_sub = team_employees.exclude(id=employee.id).first()
        if first_sub:
            default_emp_id = str(first_sub.id)
            
    selected_emp_id = request.GET.get('employee_id', default_emp_id)
    selected_emp = get_object_or_404(Employee, id=int(selected_emp_id))
    
    today = timezone.now().date()
    from_month = int(request.GET.get('from_month') or today.month)
    to_month = int(request.GET.get('to_month') or today.month)
    selected_year = int(request.GET.get('year') or today.year)
    selected_stockist_id = request.GET.get('stockist_id', '')

    if from_month > to_month: from_month, to_month = to_month, from_month
    months_range = list(range(from_month, to_month + 1))
    months_headers = [(m, calendar.month_name[m][:3]) for m in months_range]

    stockists = Stockist.objects.filter(territory=selected_emp.headquarter).order_by('name') if selected_emp.headquarter else Stockist.objects.none()

    report_filter = {'employee': selected_emp, 'month__gte': from_month, 'month__lte': to_month, 'year': selected_year}
    if selected_stockist_id: report_filter['stockist_id'] = selected_stockist_id

    reports = PartyWiseSaleReport.objects.filter(**report_filter)
    lines = PartyWiseSaleLine.objects.filter(report__in=reports).select_related('chemist', 'product', 'report__stockist')

    party_wise_data = defaultdict(lambda: {'chemist_name': '', 'stockist_name': '', 'monthly': {m: {'b_qty': 0, 'f_qty': 0, 'val': 0.0} for m in months_range}, 'tot_b': 0, 'tot_f': 0, 'tot_val': 0.0, 'products': defaultdict(lambda: {'monthly': {m: {'b_qty': 0, 'f_qty': 0, 'val': 0.0} for m in months_range}, 'tot_b': 0, 'tot_f': 0, 'tot_val': 0.0})})
    product_wise_data = defaultdict(lambda: {'product_name': '', 'monthly': {m: {'b_qty': 0, 'f_qty': 0, 'val': 0.0} for m in months_range}, 'tot_b': 0, 'tot_f': 0, 'tot_val': 0.0})
    # 🌟 NAYA: Product tab ke liye — har product ke andar chemist-wise breakdown (jaisa Chemist tab me product-wise breakdown hai)
    product_chemist_data = defaultdict(lambda: {'product_name': '', 'chemists': defaultdict(lambda: {'chemist_name': '', 'stockist_name': '', 'monthly': {m: {'b_qty': 0, 'f_qty': 0, 'val': 0.0} for m in months_range}, 'tot_b': 0, 'tot_f': 0, 'tot_val': 0.0})})
    gt_monthly = {m: {'b_qty': 0, 'f_qty': 0, 'val': 0.0} for m in months_range}
    gt_b_qty, gt_f_qty, gt_val = 0, 0, 0.0
    raw_logs = []

    for line in lines:
        r_month, ch_id, ch_name, st_name = line.report.month, line.chemist.id, line.chemist.name, line.report.stockist.name
        pr_id, pr_name, b_qty, f_qty = line.product.id, line.product.name, line.billed_qty, line.free_qty
        price = float(line.product.price) if getattr(line.product, 'price', None) else 0.0
        val = b_qty * price
            
        raw_logs.append({'date': f"{calendar.month_name[r_month][:3]}-{selected_year}", 'chemist': ch_name, 'stockist': st_name, 'product': pr_name, 'billed': b_qty, 'free': f_qty, 'value': val})

        party_wise_data[ch_id]['chemist_name'] = ch_name
        party_wise_data[ch_id]['stockist_name'] = st_name
        # 🌟 NAYA: chemist-level (product ke aar-paar) monthly subtotal
        party_wise_data[ch_id]['monthly'][r_month]['b_qty'] += b_qty; party_wise_data[ch_id]['monthly'][r_month]['f_qty'] += f_qty; party_wise_data[ch_id]['monthly'][r_month]['val'] += val
        party_wise_data[ch_id]['tot_b'] += b_qty; party_wise_data[ch_id]['tot_f'] += f_qty; party_wise_data[ch_id]['tot_val'] += val
        p_data = party_wise_data[ch_id]['products'][pr_name]
        p_data['monthly'][r_month]['b_qty'] += b_qty; p_data['monthly'][r_month]['f_qty'] += f_qty; p_data['monthly'][r_month]['val'] += val
        p_data['tot_b'] += b_qty; p_data['tot_f'] += f_qty; p_data['tot_val'] += val

        pr_data = product_wise_data[pr_id]
        pr_data['product_name'] = pr_name
        pr_data['monthly'][r_month]['b_qty'] += b_qty; pr_data['monthly'][r_month]['f_qty'] += f_qty; pr_data['monthly'][r_month]['val'] += val
        pr_data['tot_b'] += b_qty; pr_data['tot_f'] += f_qty; pr_data['tot_val'] += val

        # 🌟 NAYA: product_chemist_data bhi isi loop me fill karo
        pc_data = product_chemist_data[pr_id]
        pc_data['product_name'] = pr_name
        pcc_data = pc_data['chemists'][ch_id]
        pcc_data['chemist_name'] = ch_name
        pcc_data['stockist_name'] = st_name
        pcc_data['monthly'][r_month]['b_qty'] += b_qty; pcc_data['monthly'][r_month]['f_qty'] += f_qty; pcc_data['monthly'][r_month]['val'] += val
        pcc_data['tot_b'] += b_qty; pcc_data['tot_f'] += f_qty; pcc_data['tot_val'] += val

        gt_monthly[r_month]['b_qty'] += b_qty; gt_monthly[r_month]['f_qty'] += f_qty; gt_monthly[r_month]['val'] += val
        gt_b_qty += b_qty; gt_f_qty += f_qty; gt_val += val

    party_list = []
    for ch_id, ch_data in party_wise_data.items():
        prod_list = []
        for pr_name, pd in ch_data['products'].items():
            prod_list.append({'name': pr_name, 'monthly_list': [pd['monthly'][m] for m in months_range], 'tot_b': pd['tot_b'], 'tot_f': pd['tot_f'], 'tot_val': pd['tot_val']})
        prod_list.sort(key=lambda x: x['name'])
        party_list.append({
            'chemist_name': ch_data['chemist_name'], 'stockist_name': ch_data['stockist_name'], 'products': prod_list,
            'monthly_list': [ch_data['monthly'][m] for m in months_range],
            'tot_b': ch_data['tot_b'], 'tot_f': ch_data['tot_f'], 'tot_val': ch_data['tot_val']
        })
    party_list.sort(key=lambda x: x['chemist_name'])

    prod_list_final = []
    for pr_id, pr_data in product_wise_data.items():
        pr_data['monthly_list'] = [pr_data['monthly'][m] for m in months_range]
        prod_list_final.append(pr_data)
    prod_list_final.sort(key=lambda x: x['product_name'])
    gt_monthly_list = [gt_monthly[m] for m in months_range]

    # 🌟 NAYA: product_list_grouped — har product ke andar chemist-wise breakdown (Product tab ke liye)
    product_list_grouped = []
    for pr_id, pc_data in product_chemist_data.items():
        chem_list = []
        for ch_id, cd in pc_data['chemists'].items():
            chem_list.append({
                'chemist_name': cd['chemist_name'], 'stockist_name': cd['stockist_name'],
                'monthly_list': [cd['monthly'][m] for m in months_range],
                'tot_b': cd['tot_b'], 'tot_f': cd['tot_f'], 'tot_val': cd['tot_val']
            })
        chem_list.sort(key=lambda x: x['chemist_name'])
        pr_totals = product_wise_data[pr_id]
        product_list_grouped.append({
            'product_name': pc_data['product_name'], 'chemists': chem_list,
            'monthly_list': [pr_totals['monthly'][m] for m in months_range],
            'tot_b': pr_totals['tot_b'], 'tot_f': pr_totals['tot_f'], 'tot_val': pr_totals['tot_val']
        })
    product_list_grouped.sort(key=lambda x: x['product_name'])

    if request.GET.get('export') == 'excel':
        filename = f"Party_Sales_{selected_emp.name}_M{from_month}-M{to_month}_{selected_year}.xlsx"
        wb = openpyxl.Workbook()
        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        period_str = f"{calendar.month_name[from_month][:3]} to {calendar.month_name[to_month][:3]} {selected_year}"

        ws1 = wb.active
        ws1.title = "Party Wise Sale"
        ws1.append(['PARTY WISE SALE REPORT (MONTH TREND)'])
        ws1.append(['Employee:', selected_emp.name, 'Period:', period_str])
        ws1.append([''])
        ws1['A1'].font = Font(bold=True, size=12, color="107C41")
        
        headers1 = ['Product Name']
        for m_num, m_name in months_headers: headers1.extend([f'{m_name} Billed', f'{m_name} Free', f'{m_name} Value (₹)'])
        headers1.extend(['Total Billed', 'Total Free', 'Total Value (₹)'])
        ws1.append(headers1)
        total_cols1 = len(headers1)
        for col_num, cell in enumerate(ws1[4], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws1.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 15 if col_num > 1 else 25

        group_fill = PatternFill(start_color="0B5E31", end_color="0B5E31", fill_type="solid")
        group_font = Font(color="FFFFFF", bold=True)
        subtotal_fill = PatternFill(start_color="EEF6F1", end_color="EEF6F1", fill_type="solid")

        for party in party_list:
            # 🌟 NAYA: Chemist ke naam ka group header row (poori width merge karke)
            ws1.append([f"💊 {party['chemist_name']} — {party['stockist_name']}"])
            r = ws1.max_row
            ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=total_cols1)
            ws1.cell(row=r, column=1).fill = group_fill
            ws1.cell(row=r, column=1).font = group_font

            for prod in party['products']:
                row_data = [prod['name']]
                for m_data in prod['monthly_list']: row_data.extend([m_data['b_qty'], m_data['f_qty'], round(m_data['val'], 2)])
                row_data.extend([prod['tot_b'], prod['tot_f'], round(prod['tot_val'], 2)])
                ws1.append(row_data)

            # 🌟 NAYA: Chemist subtotal row
            sub_row = ['Subtotal']
            for m_data in party['monthly_list']: sub_row.extend([m_data['b_qty'], m_data['f_qty'], round(m_data['val'], 2)])
            sub_row.extend([party['tot_b'], party['tot_f'], round(party['tot_val'], 2)])
            ws1.append(sub_row)
            r = ws1.max_row
            for c in range(1, total_cols1 + 1):
                ws1.cell(row=r, column=c).fill = subtotal_fill
                ws1.cell(row=r, column=c).font = Font(bold=True)

        gt_row1 = ['GRAND TOTAL']
        for m_gt in gt_monthly_list: gt_row1.extend([m_gt['b_qty'], m_gt['f_qty'], round(m_gt['val'], 2)])
        gt_row1.extend([gt_b_qty, gt_f_qty, round(gt_val, 2)])
        ws1.append(gt_row1)
        for cell in ws1[ws1.max_row]: cell.font = Font(bold=True)

        ws2 = wb.create_sheet(title="Product Wise Sale")
        ws2.append(['PRODUCT WISE SALE REPORT (MONTH TREND)'])
        ws2.append(['Employee:', selected_emp.name, 'Period:', period_str])
        ws2.append([''])
        ws2['A1'].font = Font(bold=True, size=12, color="107C41")
        
        headers2 = ['Chemist Name', 'Stockist Name']
        for m_num, m_name in months_headers: headers2.extend([f'{m_name} Billed', f'{m_name} Free', f'{m_name} Value (₹)'])
        headers2.extend(['Total Billed', 'Total Free', 'Total Value (₹)'])
        ws2.append(headers2)
        total_cols2 = len(headers2)

        for col_num, cell in enumerate(ws2[4], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws2.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 15 if col_num > 2 else 25

        for prod in product_list_grouped:
            # 🌟 NAYA: Product ke naam ka group header row (poori width merge karke)
            ws2.append([f"📦 {prod['product_name']}"])
            r = ws2.max_row
            ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=total_cols2)
            ws2.cell(row=r, column=1).fill = group_fill
            ws2.cell(row=r, column=1).font = group_font

            for chem in prod['chemists']:
                row_data = [chem['chemist_name'], chem['stockist_name']]
                for m_data in chem['monthly_list']: row_data.extend([m_data['b_qty'], m_data['f_qty'], round(m_data['val'], 2)])
                row_data.extend([chem['tot_b'], chem['tot_f'], round(chem['tot_val'], 2)])
                ws2.append(row_data)

            # 🌟 NAYA: Product subtotal row
            sub_row = ['Subtotal', '']
            for m_data in prod['monthly_list']: sub_row.extend([m_data['b_qty'], m_data['f_qty'], round(m_data['val'], 2)])
            sub_row.extend([prod['tot_b'], prod['tot_f'], round(prod['tot_val'], 2)])
            ws2.append(sub_row)
            r = ws2.max_row
            for c in range(1, total_cols2 + 1):
                ws2.cell(row=r, column=c).fill = subtotal_fill
                ws2.cell(row=r, column=c).font = Font(bold=True)

        gt_row2 = ['GRAND TOTAL', '']
        for m_gt in gt_monthly_list: gt_row2.extend([m_gt['b_qty'], m_gt['f_qty'], round(m_gt['val'], 2)])
        gt_row2.extend([gt_b_qty, gt_f_qty, round(gt_val, 2)])
        ws2.append(gt_row2)
        for cell in ws2[ws2.max_row]: cell.font = Font(bold=True)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return render(request, 'party_rx_report.html', {
        'team_employees': team_employees, 'selected_emp_id': int(selected_emp_id), 
        'from_month': from_month, 'to_month': to_month, 'selected_year': selected_year, 
        'selected_stockist_id': int(selected_stockist_id) if selected_stockist_id else '', 
        'stockists': stockists, 'months_choices': [(i, calendar.month_name[i]) for i in range(1, 13)], 
        'months_headers': months_headers, 'party_list': party_list, 'product_list': prod_list_final,
        'product_list_grouped': product_list_grouped,
        'raw_logs': raw_logs, 'gt_monthly_list': gt_monthly_list, 'gt_b_qty': gt_b_qty, 'gt_f_qty': gt_f_qty, 'gt_val': gt_val, 
        'is_manager_view': employee.designation != 'MR'
    })

# ==============================================================================
# 6. DOCTOR WISE RX REPORT
# ==============================================================================
@employee_required
def dr_wise_sale_report_view(request, employee):
    team_employees = get_dropdown_team(employee, ordered=False)
    default_emp_id = str(employee.id)
    if employee.designation != 'MR':
        first_sub = team_employees.exclude(id=employee.id).first()
        if first_sub:
            default_emp_id = str(first_sub.id)
            
    selected_emp_id = request.GET.get('employee_id', default_emp_id)
    selected_emp = get_object_or_404(Employee, id=int(selected_emp_id))
    
    today = timezone.now().date()
    from_month = int(request.GET.get('from_month') or today.month)
    to_month = int(request.GET.get('to_month') or today.month)
    selected_year = int(request.GET.get('year') or today.year)

    if from_month > to_month: from_month, to_month = to_month, from_month
    months_range = list(range(from_month, to_month + 1))
    months_headers = [(m, calendar.month_name[m][:3]) for m in months_range]

    mappings = DoctorRxMapping.objects.filter(
        party_line__report__month__gte=from_month, party_line__report__month__lte=to_month,
        party_line__report__year=selected_year, party_line__report__employee=selected_emp
    ).select_related('doctor', 'party_line__chemist', 'party_line__product')

    agg_data = defaultdict(lambda: {'doctor_name': '', 'specialty': '', 'chemist_name': '', 'product_name': '', 'price': 0.0, 'monthly_data': {m: {'billed': 0, 'free': 0, 'val': 0.0} for m in months_range}, 'total_billed': 0, 'total_free': 0, 'total_value': 0.0})
    product_wise_data = defaultdict(lambda: {'product_name': '', 'doctors': defaultdict(lambda: {'doc_name': '', 'specialty': '', 'monthly': {m: {'billed': 0, 'free': 0, 'val': 0.0} for m in months_range}, 'tot_b': 0, 'tot_f': 0, 'tot_val': 0.0}), 'tot_b': 0, 'tot_f': 0, 'tot_val': 0.0})
    gt_monthly = {m: {'billed': 0, 'free': 0, 'val': 0.0} for m in months_range}
    gt_billed, gt_free, gt_value = 0, 0, 0.0

    for m in mappings:
        sale_month = m.party_line.report.month
        doc_id, doc_name, specialty = m.doctor.id, m.doctor.name, m.doctor.specialty or '-'
        chem_id, chem_name = m.party_line.chemist.id, m.party_line.chemist.name
        prod_id, prod_name = m.party_line.product.id, m.party_line.product.name
        b_qty, f_qty = m.mapped_billed_qty, m.mapped_free_qty
        price = float(m.party_line.product.price) if getattr(m.party_line.product, 'price', None) else 0.0
        val = b_qty * price
        
        key = f"{doc_id}_{chem_id}_{prod_id}"
        if agg_data[key]['doctor_name'] == '':
            agg_data[key].update({'doctor_name': doc_name, 'specialty': specialty, 'chemist_name': chem_name, 'product_name': prod_name, 'price': price})
            
        agg_data[key]['monthly_data'][sale_month]['billed'] += b_qty; agg_data[key]['monthly_data'][sale_month]['free'] += f_qty; agg_data[key]['monthly_data'][sale_month]['val'] += val
        agg_data[key]['total_billed'] += b_qty; agg_data[key]['total_free'] += f_qty; agg_data[key]['total_value'] += val
        
        pw = product_wise_data[prod_id]
        pw['product_name'] = prod_name
        dw = pw['doctors'][doc_id]
        dw['doc_name'] = doc_name; dw['specialty'] = specialty
        
        dw['monthly'][sale_month]['billed'] += b_qty; dw['monthly'][sale_month]['free'] += f_qty; dw['monthly'][sale_month]['val'] += val
        dw['tot_b'] += b_qty; dw['tot_f'] += f_qty; dw['tot_val'] += val
        pw['tot_b'] += b_qty; pw['tot_f'] += f_qty; pw['tot_val'] += val
        
        gt_monthly[sale_month]['billed'] += b_qty; gt_monthly[sale_month]['free'] += f_qty; gt_monthly[sale_month]['val'] += val
        gt_billed += b_qty; gt_free += f_qty; gt_value += val

    report_data = list(agg_data.values())
    report_data.sort(key=lambda x: (x['doctor_name'], x['chemist_name'], x['product_name']))
    for row in report_data: row['monthly_data_list'] = [row['monthly_data'][m[0]] for m in months_headers]
        
    # ==========================================
    # 🌟 NAYA LOGIC: FOR UI GROUPING (Sub-Headers)
    # ==========================================
    doc_grouped_data = []
    current_doc = None
    curr_group = None
    
    for row in report_data:
        if row['doctor_name'] != current_doc:
            if curr_group: doc_grouped_data.append(curr_group)
            current_doc = row['doctor_name']
            curr_group = {
                'doctor_name': current_doc,
                'specialty': row['specialty'],
                'doc_total_val': 0.0,
                'items': []
            }
        curr_group['items'].append(row)
        curr_group['doc_total_val'] += row['total_value']
        
    if curr_group: doc_grouped_data.append(curr_group)

    prod_tab_list = []
    for p_id, p_data in product_wise_data.items():
        doc_list = []
        for d_id, d_data in p_data['doctors'].items():
            d_data['monthly_list'] = [d_data['monthly'][m] for m in months_range]
            doc_list.append(d_data)
        doc_list.sort(key=lambda x: x['doc_name'])
        p_data['doctors_list'] = doc_list
        prod_tab_list.append(p_data)
    prod_tab_list.sort(key=lambda x: x['product_name'])
        
    gt_monthly_list = [gt_monthly[m[0]] for m in months_headers]

    # ==========================================
    # 📊 EXCEL EXPORT LOGIC
    # ==========================================
    if request.GET.get('export') == 'excel':
        filename = f"Dr_Rx_Report_{selected_emp.name}_M{from_month}-M{to_month}_{selected_year}.xlsx"
        wb = openpyxl.Workbook()
        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        period_str = f"{calendar.month_name[from_month][:3]} to {calendar.month_name[to_month][:3]} {selected_year}"

        ws1 = wb.active
        ws1.title = "Doctor Wise Rx"
        ws1.append(['DOCTOR WISE RX SALE (MONTH TREND)'])
        ws1.append(['Employee:', selected_emp.name, 'Period:', period_str])
        ws1.append([''])
        ws1['A1'].font = Font(bold=True, size=14, color="107C41")
        
        headers1 = ['Doctor Name', 'Specialty', 'Chemist Name', 'Product Name', 'Price (₹)']
        for m_num, m_name in months_headers: headers1.extend([f'{m_name} Billed', f'{m_name} Free', f'{m_name} Value (₹)'])
        headers1.extend(['Total Billed', 'Total Free', 'Total Value (₹)'])
        ws1.append(headers1)
        
        for col_num, cell in enumerate(ws1[4], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws1.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 15 if col_num > 4 else 20
            
        for row in report_data:
            row_data = [f"Dr. {row['doctor_name']}", row['specialty'], row['chemist_name'], row['product_name'], round(row['price'], 2)]
            for m_data in row['monthly_data_list']: row_data.extend([m_data['billed'], m_data['free'], round(m_data['val'], 2)])
            row_data.extend([row['total_billed'], row['total_free'], round(row['total_value'], 2)])
            ws1.append(row_data)

        ws2 = wb.create_sheet(title="Product Wise Rx")
        ws2.append(['PRODUCT WISE -> DOCTOR RX SALE (MONTH TREND)'])
        ws2.append(['Employee:', selected_emp.name, 'Period:', period_str])
        ws2.append([''])
        ws2['A1'].font = Font(bold=True, size=14, color="107C41")
        
        headers2 = ['Product Name', 'Doctor Name', 'Specialty']
        for m_num, m_name in months_headers: headers2.extend([f'{m_name} Billed', f'{m_name} Free', f'{m_name} Value (₹)'])
        headers2.extend(['Total Billed', 'Total Free', 'Total Value (₹)'])
        ws2.append(headers2)

        for col_num, cell in enumerate(ws2[4], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws2.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 15 if col_num > 3 else 25

        for prod in prod_tab_list:
            for doc in prod['doctors_list']:
                row_data = [prod['product_name'], f"Dr. {doc['doc_name']}", doc['specialty']]
                for m_data in doc['monthly_list']: row_data.extend([m_data['billed'], m_data['free'], round(m_data['val'], 2)])
                row_data.extend([doc['tot_b'], doc['tot_f'], round(doc['tot_val'], 2)])
                ws2.append(row_data)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return render(request, 'dr_wise_sale_report.html', {
        'team_employees': team_employees, 'selected_emp_id': int(selected_emp_id),
        'from_month': from_month, 'to_month': to_month, 'selected_year': selected_year,
        'months_choices': [(i, calendar.month_name[i]) for i in range(1, 13)], 'months_headers': months_headers,
        'report_data': report_data, 'doc_grouped_data': doc_grouped_data, 'prod_tab_list': prod_tab_list,
        'gt_monthly_list': gt_monthly_list, 'gt_billed': gt_billed, 'gt_free': gt_free, 'gt_value': gt_value,
        'is_manager_view': employee.designation != 'MR'
    })


# ==============================================================================
# 7. TARGET SETTING (Self Only)
# ==============================================================================
@employee_required
def target_setting_view(request, employee):
    selected_month = int(request.GET.get('month') or timezone.now().month)
    selected_year = int(request.GET.get('year') or timezone.now().year)
    
    # 🌟 FIX: 'selected_emp' yahan defined nahi tha, isliye 'employee' variable use kiya
    products = Product.objects.filter(company=employee.company)
    
    master = None
    is_readonly = True # Default True rakha hai taaki bina HQ walo ko error na aaye
    
    if employee.headquarter:
        master, _ = MonthlyTargetMaster.objects.get_or_create(
            territory=employee.headquarter, 
            month=selected_month, 
            year=selected_year
        )
        is_readonly = master.status not in ['Draft', 'Rejected']

    if request.method == "POST" and not is_readonly and employee.headquarter:
        action = request.POST.get('action')
        
        if action in ['Save_Draft', 'Submit_Manager']:
            for p in products:
                t_qty = int(request.POST.get(f'target_{p.id}', 0) or 0)
                if t_qty > 0: 
                    # 🌟 FIX: employee=employee ko territory=employee.headquarter kiya
                    TerritoryTarget.objects.update_or_create(
                        territory=employee.headquarter, 
                        product=p, 
                        month=selected_month, 
                        year=selected_year, 
                        defaults={'target_qty': t_qty}
                    )
                else: 
                    # 🌟 FIX: Yahan bhi territory=employee.headquarter kiya
                    TerritoryTarget.objects.filter(
                        territory=employee.headquarter, 
                        product=p, 
                        month=selected_month, 
                        year=selected_year
                    ).delete()
        
        if action == 'Save_Draft':
            master.status = 'Draft'
            messages.success(request, "Target Draft saved!")
        elif action == 'Submit_Manager':
            master.status = 'Pending_Manager'
            master.approved_by_managers = []
            messages.success(request, "Target submitted for Manager approval! 🚀")
            
        master.save()
        return redirect(f"{request.path}?month={selected_month}&year={selected_year}")

    # 🌟 FIX: Agar HQ hai tabhi target fetch karega, warna khali dictionary rahegi
    existing_targets = {}
    if employee.headquarter:
        existing_targets = {
            t.product_id: t.target_qty 
            for t in TerritoryTarget.objects.filter(territory=employee.headquarter, month=selected_month, year=selected_year)
        }
        
    targets = [{'product': p, 'target_qty': existing_targets.get(p.id, 0)} for p in products]

    return render(request, 'target_setting.html', {
        'selected_month': selected_month, 
        'selected_year': selected_year, 
        'targets': targets, 
        'master': master, 
        'is_readonly': is_readonly, 
        'months_choices': [(i, calendar.month_name[i]) for i in range(1, 13)], 
        'user_emp': employee
    })

# ==============================================================================
# 8. TARGET REVIEW PAGE (For Managers)
# ==============================================================================
@employee_required
def review_target_view(request, employee, target_id):
    master = get_object_or_404(MonthlyTargetMaster, id=target_id)
    
    # 🌟 FIX 1: master.employee ab nahi raha, isliye territory se current active MR nikalna padega
    target_emp = Employee.objects.filter(company=employee.company, headquarter=master.territory, is_active=True).first()
    
    # 🌟 FIX 2: 'selected_emp' undefined tha, usko hata kar 'employee.company' kar diya
    products = Product.objects.filter(company=employee.company)

    if request.method == "POST":
        action = request.POST.get('action')
        if action in ['Approve', 'Reject', 'Save_Only']:
            for p in products:
                t_qty = int(request.POST.get(f'target_{p.id}', 0) or 0)
                if t_qty > 0: 
                    # 🌟 FIX 2: employee=target_emp ki jagah territory=master.territory aayega
                    TerritoryTarget.objects.update_or_create(
                        territory=master.territory, 
                        product=p, 
                        month=master.month, 
                        year=master.year, 
                        defaults={'target_qty': t_qty}
                    )
                else: 
                    # 🌟 FIX 3: Yahan bhi territory=master.territory aayega
                    TerritoryTarget.objects.filter(
                        territory=master.territory, 
                        product=p, 
                        month=master.month, 
                        year=master.year
                    ).delete()

        if action == 'Reject':
            master.status = 'Rejected'
            master.manager_remark = request.POST.get('manager_remark', '')
            messages.error(request, "Target Rejected! ❌")
            master.save()
            return redirect('manager_approvals')

        elif action == 'Approve':
            master.manager_remark = request.POST.get('manager_remark', '')
            if employee.designation == 'Admin':
                master.status = 'Approved'
                messages.success(request, "Target Officially Approved by Admin! ✅")
            else:
                if employee.id not in master.approved_by_managers:
                    new_list = list(master.approved_by_managers)
                    new_list.append(employee.id)
                    master.approved_by_managers = new_list

                chain_managers = []
                # 🌟 FIX 4: Agar us territory me MR active hai tabhi Approval chain banegi
                creator_manager = target_emp.manager if target_emp else None
                while creator_manager is not None:
                    chain_managers.append(creator_manager.id)
                    creator_manager = creator_manager.manager

                if len(master.approved_by_managers) >= len(chain_managers):
                    master.status = 'Pending_Admin'
                    messages.success(request, "Approved by all managers! Pending Admin Approval.")
                else:
                    master.status = 'Pending_Manager'
                    messages.success(request, "Approved & Forwarded to next level.")
            master.save()
            return redirect('manager_approvals')

        elif action == 'Save_Only':
            messages.success(request, "Target modifications saved!")
            master.save()
            return redirect(request.path)

    # 🌟 FIX 5: Fetch karne ke liye bhi territory=master.territory use kiya
    existing_targets = {}
    if master.territory:
        existing_targets = {
            t.product_id: t.target_qty 
            for t in TerritoryTarget.objects.filter(territory=master.territory, month=master.month, year=master.year)
        }
    targets = [{'product': p, 'target_qty': existing_targets.get(p.id, 0)} for p in products]

    return render(request, 'review_target.html', {'targets': targets, 'master': master, 'target_emp': target_emp})


@employee_required
def sales_summary_report_view(request, employee):
    team_employees = get_dropdown_team(employee)
    default_emp_id = str(employee.id)
    if employee.designation != 'MR':
        first_sub = team_employees.exclude(id=employee.id).first()
        if first_sub:
            default_emp_id = str(first_sub.id)
            
    selected_emp_id = request.GET.get('employee_id', default_emp_id)
    selected_emp = get_object_or_404(Employee, id=int(selected_emp_id))
    
    today = timezone.now().date()
    from_month = int(request.GET.get('from_month') or today.month)
    to_month = int(request.GET.get('to_month') or today.month)
    selected_year = int(request.GET.get('year') or today.year)

    if from_month > to_month: from_month, to_month = to_month, from_month
    months_range = list(range(from_month, to_month + 1))
    months_headers = [(m, calendar.month_name[m][:3]) for m in months_range]

    target_team = get_full_team_employees(selected_emp)
    area_territory_ids = target_team.exclude(headquarter__isnull=True).values_list('headquarter_id', flat=True)

    span = to_month - from_month + 1
    curr_start_val = selected_year * 12 + from_month
    curr_end_val = selected_year * 12 + to_month

    prev_end_val = curr_start_val - 1
    prev_start_val = prev_end_val - span + 1
    
    ly_start_val = (selected_year - 1) * 12 + from_month
    ly_end_val = (selected_year - 1) * 12 + to_month

    years_needed = list(set([selected_year, selected_year - 1, (prev_start_val - 1) // 12]))

    # 🌟 FIX 1: MonthlyTargetMaster ab 'territory' check karega 'employee' ki jagah
    team_territories = [emp.headquarter for emp in target_team if emp.headquarter]
    approved_territory_ids = MonthlyTargetMaster.objects.filter(
        territory__in=team_territories, 
        status='Approved', 
        year__in=years_needed
    ).values_list('territory_id', flat=True)

    # 🌟 FIX 2: TerritoryTarget ab 'territory_id' se target uthayega 'employee_id' ki jagah
    all_targets = TerritoryTarget.objects.filter(territory_id__in=approved_territory_ids, year__in=years_needed).select_related('product')
    
    all_primary = StockistProductStatement.objects.filter(stockist__territory_id__in=area_territory_ids, year__in=years_needed).select_related('product')
    all_secondary = PartyWiseSaleLine.objects.filter(report__stockist__territory_id__in=area_territory_ids, report__year__in=years_needed).select_related('product')

    def calc_growth(curr, prev):
        if prev == 0: return 100.0 if curr > 0 else 0.0
        return ((curr - prev) / prev) * 100.0

    def calc_ach(val, tgt):
        if tgt == 0: return 100.0 if val > 0 else 0.0
        return (val / tgt) * 100.0

    products = Product.objects.filter(company=selected_emp.company).order_by('name')
    p_dict = {}
    
    gt = {
        'monthly': {m: {'t_qty':0, 't_val':0.0, 'p_qty':0, 'p_val':0.0, 's_qty':0, 's_val':0.0} for m in months_range},
        't_qty': 0, 't_val': 0.0, 'p_qty': 0, 'p_val': 0.0, 's_qty': 0, 's_val': 0.0,
        'prev_p_val': 0.0, 'prev_s_val': 0.0, 'ly_p_val': 0.0, 'ly_s_val': 0.0
    }

    for p in products:
        price = float(p.price) if getattr(p, 'price', None) else 0.0
        p_dict[p.id] = {
            'name': p.name, 'price': price,
            'monthly': {m: {'t_qty':0, 't_val':0.0, 'p_qty':0, 'p_val':0.0, 's_qty':0, 's_val':0.0} for m in months_range},
            'curr_t_qty': 0, 'curr_t_val': 0.0, 'curr_p_qty': 0, 'curr_p_val': 0.0, 'curr_s_qty': 0, 'curr_s_val': 0.0,
            'prev_p_val': 0.0, 'prev_s_val': 0.0, 'ly_p_val': 0.0, 'ly_s_val': 0.0
        }

    for t in all_targets:
        val = t.year * 12 + t.month
        if curr_start_val <= val <= curr_end_val and t.month in p_dict[t.product_id]['monthly']:
            t_val = t.target_qty * float(t.product.price)
            p_dict[t.product_id]['monthly'][t.month]['t_qty'] += t.target_qty; p_dict[t.product_id]['monthly'][t.month]['t_val'] += t_val
            p_dict[t.product_id]['curr_t_qty'] += t.target_qty; p_dict[t.product_id]['curr_t_val'] += t_val
            gt['monthly'][t.month]['t_qty'] += t.target_qty; gt['monthly'][t.month]['t_val'] += t_val
            gt['t_qty'] += t.target_qty; gt['t_val'] += t_val

    for p in all_primary:
        val = p.year * 12 + p.month
        p_val = p.received_qty * float(p.product.price)
        if curr_start_val <= val <= curr_end_val and p.month in p_dict[p.product_id]['monthly']:
            p_dict[p.product_id]['monthly'][p.month]['p_qty'] += p.received_qty; p_dict[p.product_id]['monthly'][p.month]['p_val'] += p_val
            p_dict[p.product_id]['curr_p_qty'] += p.received_qty; p_dict[p.product_id]['curr_p_val'] += p_val
            gt['monthly'][p.month]['p_qty'] += p.received_qty; gt['monthly'][p.month]['p_val'] += p_val
            gt['p_qty'] += p.received_qty; gt['p_val'] += p_val
        elif prev_start_val <= val <= prev_end_val:
            p_dict[p.product_id]['prev_p_val'] += p_val; gt['prev_p_val'] += p_val
        elif ly_start_val <= val <= ly_end_val:
            p_dict[p.product_id]['ly_p_val'] += p_val; gt['ly_p_val'] += p_val

    for s in all_secondary:
        val = s.report.year * 12 + s.report.month
        s_val = s.billed_qty * float(s.product.price)
        s_qty = s.billed_qty + s.free_qty
        if curr_start_val <= val <= curr_end_val and s.report.month in p_dict[s.product_id]['monthly']:
            p_dict[s.product_id]['monthly'][s.report.month]['s_qty'] += s_qty; p_dict[s.product_id]['monthly'][s.report.month]['s_val'] += s_val
            p_dict[s.product_id]['curr_s_qty'] += s_qty; p_dict[s.product_id]['curr_s_val'] += s_val
            gt['monthly'][s.report.month]['s_qty'] += s_qty; gt['monthly'][s.report.month]['s_val'] += s_val
            gt['s_qty'] += s_qty; gt['s_val'] += s_val
        elif prev_start_val <= val <= prev_end_val:
            p_dict[s.product_id]['prev_s_val'] += s_val; gt['prev_s_val'] += s_val
        elif ly_start_val <= val <= ly_end_val:
            p_dict[s.product_id]['ly_s_val'] += s_val; gt['ly_s_val'] += s_val

    report_data = []
    for pid, d in p_dict.items():
        if d['curr_t_val'] > 0 or d['curr_p_val'] > 0 or d['curr_s_val'] > 0:
            d['p_ach'] = calc_ach(d['curr_p_val'], d['curr_t_val'])
            d['s_ach'] = calc_ach(d['curr_s_val'], d['curr_t_val'])
            d['p_m2m'] = calc_growth(d['curr_p_val'], d['prev_p_val'])
            d['p_y2y'] = calc_growth(d['curr_p_val'], d['ly_p_val'])
            d['s_m2m'] = calc_growth(d['curr_s_val'], d['prev_s_val'])
            d['s_y2y'] = calc_growth(d['curr_s_val'], d['ly_s_val'])
            d['monthly_list'] = [d['monthly'][m] for m in months_range]
            report_data.append(d)

    gt['p_ach'] = calc_ach(gt['p_val'], gt['t_val']); gt['s_ach'] = calc_ach(gt['s_val'], gt['t_val'])
    gt['p_m2m'] = calc_growth(gt['p_val'], gt['prev_p_val']); gt['p_y2y'] = calc_growth(gt['p_val'], gt['ly_p_val'])
    gt['s_m2m'] = calc_growth(gt['s_val'], gt['prev_s_val']); gt['s_y2y'] = calc_growth(gt['s_val'], gt['ly_s_val'])
    gt['monthly_list'] = [gt['monthly'][m] for m in months_range]

    if request.GET.get('export') == 'excel':
        filename = f"Sales_Summary_{selected_emp.name}_M{from_month}-M{to_month}_{selected_year}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sales Summary Matrix"
        
        hq_name = selected_emp.headquarter.name if selected_emp.headquarter else "N/A"
        period_str = f"{calendar.month_name[from_month][:3]} to {calendar.month_name[to_month][:3]} {selected_year}"
        
        ws.append(['MASTER SALES SUMMARY REPORT (QTY & VAL MATRIX)'])
        ws.append(['Employee:', selected_emp.name, 'HQ:', hq_name])
        ws.append(['Period:', period_str])
        ws.append([''])
        
        ws['A1'].font = Font(bold=True, size=14, color="107C41")
        for cell in ws['A2:C3']: cell[0].font = Font(bold=True)
        
        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        
        headers1 = ['Product Name']
        headers2 = ['']
        for m in months_headers:
            headers1.extend([f"{m[1]} DATA", "", "", "", "", ""])
            headers2.extend(['Tgt Qty', 'Tgt (₹)', 'Pri Qty', 'Pri (₹)', 'Sec Qty', 'Sec (₹)'])
            
        headers1.extend(['OVERALL GRAND TOTAL', '', '', '', '', '', '', '', '', '', '', ''])
        headers2.extend(['Tot Tgt Qty', 'Tot Tgt (₹)', 'Tot Pri Qty', 'Tot Pri (₹)', 'Tot Sec Qty', 'Tot Sec (₹)', 'Pri Ach%', 'Pri M2M%', 'Pri Y2Y%', 'Sec Ach%', 'Sec M2M%', 'Sec Y2Y%'])
        
        ws.append(headers1)
        ws.append(headers2)
        
        for row_idx in [5, 6]:
            for col_num, cell in enumerate(ws[row_idx], 1):
                cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
                if row_idx == 6: ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 18 if col_num == 1 else 12
            
        for row in report_data:
            row_data = [row['name']]
            for m_data in row['monthly_list']:
                row_data.extend([m_data['t_qty'], round(m_data['t_val'], 2), m_data['p_qty'], round(m_data['p_val'], 2), m_data['s_qty'], round(m_data['s_val'], 2)])
            row_data.extend([
                row['curr_t_qty'], round(row['curr_t_val'], 2), row['curr_p_qty'], round(row['curr_p_val'], 2), row['curr_s_qty'], round(row['curr_s_val'], 2),
                f"{round(row['p_ach'], 1)}%", f"{round(row['p_m2m'], 1)}%", f"{round(row['p_y2y'], 1)}%", f"{round(row['s_ach'], 1)}%", f"{round(row['s_m2m'], 1)}%", f"{round(row['s_y2y'], 1)}%"
            ])
            ws.append(row_data)
            
        gt_row = ['OVERALL TOTAL']
        for m_data in gt['monthly_list']:
            gt_row.extend([m_data['t_qty'], round(m_data['t_val'], 2), m_data['p_qty'], round(m_data['p_val'], 2), m_data['s_qty'], round(m_data['s_val'], 2)])
            
        gt_row.extend([
            gt['t_qty'], round(gt['t_val'], 2), gt['p_qty'], round(gt['p_val'], 2), gt['s_qty'], round(gt['s_val'], 2),
            f"{round(gt['p_ach'], 1)}%", f"{round(gt['p_m2m'], 1)}%", f"{round(gt['p_y2y'], 1)}%", f"{round(gt['s_ach'], 1)}%", f"{round(gt['s_m2m'], 1)}%", f"{round(gt['s_y2y'], 1)}%"
        ])
        ws.append(gt_row)
        for cell in ws[ws.max_row]: cell.font = Font(bold=True)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return render(request, 'sales_summary_report.html', {
        'team_employees': team_employees, 'selected_emp_id': int(selected_emp_id), 'selected_employee_name': selected_emp.name,
        'from_month': from_month, 'to_month': to_month, 'selected_year': selected_year,
        'months_choices': [(i, calendar.month_name[i]) for i in range(1, 13)], 'months_headers': months_headers,
        'report_data': report_data, 'gt': gt, 'is_manager_view': employee.designation != 'MR'
    })