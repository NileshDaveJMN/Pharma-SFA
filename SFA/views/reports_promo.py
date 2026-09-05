import io
import calendar
from collections import defaultdict
from datetime import date, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.db.models import Sum

from SFA.models import (
    Employee, Stockist, Product, PartyWiseSaleLine, SystemSetting,
    FreeQtyClaimMaster, FreeQtyClaimLine, PromoDispatch, MRInventory,
    GiftCampaignPlan, DoctorROILedger, Doctor, DoctorRxMapping
)
from .auth import get_full_team_employees, get_dropdown_team
from SFA.decorators import employee_required

# ==============================================================================
# 1. 🎁 FREE CLAIM VIEW (GENERATE & SUBMIT)
# ==============================================================================
@employee_required
def free_claim_view(request, employee):
    team_employees = get_dropdown_team(employee)
    is_manager_view = employee.designation != 'MR'

    default_emp_id = str(employee.id)
    if is_manager_view:
        first_sub = team_employees.exclude(id=employee.id).first()
        if first_sub:
            default_emp_id = str(first_sub.id)

    selected_emp_id = request.GET.get('employee_id') or request.POST.get('employee_id') or default_emp_id
    selected_emp = get_object_or_404(Employee, id=int(selected_emp_id), company=employee.company)

    selected_month = int(request.GET.get('month') or timezone.localdate().month)
    selected_year = int(request.GET.get('year') or timezone.localdate().year)

    today = timezone.localdate()
    setting = SystemSetting.objects.filter(company=employee.company).first()
    deadline = setting.free_claim_deadline_day if setting and setting.free_claim_deadline_day else 4
    prev_month, prev_year = (12, today.year - 1) if today.month == 1 else (today.month - 1, today.year)
    is_immediate_prev_month = (selected_month == prev_month and selected_year == prev_year)
    is_locked = (today.day > deadline) if is_immediate_prev_month else True

    my_terr_ids = [selected_emp.headquarter_id] if selected_emp.headquarter_id else []
    available_stockists = Stockist.objects.filter(territory_id__in=my_terr_ids).order_by('name')

    selected_stockist_id = request.GET.get('stockist_id') or request.POST.get('stockist_id')
    if selected_stockist_id and not available_stockists.filter(id=selected_stockist_id).exists():
        selected_stockist_id = None
    if not selected_stockist_id and available_stockists.exists():
        selected_stockist_id = str(available_stockists.first().id)

    selected_stockist = available_stockists.filter(id=selected_stockist_id).first()

    master = None
    if selected_stockist:
        master = FreeQtyClaimMaster.objects.filter(employee=selected_emp, stockist=selected_stockist, month=selected_month, year=selected_year).first()

    if request.method == "POST":
        action = request.POST.get('action')
        
        if selected_year > today.year or (selected_year == today.year and selected_month >= today.month):
            messages.error(request, "⚠️ The Free Claim for the current month cannot be generated yet. Please wait until the month ends.")
            return redirect(f"{request.path}?employee_id={selected_emp.id}&stockist_id={selected_stockist_id}&month={selected_month}&year={selected_year}")      
        
        if is_locked and employee.designation not in ['Admin', 'System Administrator']:
            messages.error(request, f"⚠️ Action Locked! Free Claim entries are only allowed until the {deadline} of each month. Please contact Admin.")
            return redirect(f"{request.path}?employee_id={selected_emp.id}&stockist_id={selected_stockist_id}&month={selected_month}&year={selected_year}")
        
        if not selected_stockist:
            messages.error(request, "⚠️ Stockist is missing!")
            return redirect(request.path)

        if action == 'generate':
            if master and master.status not in ['Draft', 'Rejected']:
                messages.error(request, f"⚠️ This report is currently marked as '{master.status}'. It cannot be regenerated.")
            else:
                sales = PartyWiseSaleLine.objects.filter(
                    report__employee=selected_emp,
                    report__stockist=selected_stockist,
                    report__month=selected_month,
                    report__year=selected_year,
                    free_qty__gt=0
                ).values('product_id').annotate(
                    tot_billed=Sum('billed_qty'),
                    tot_free=Sum('free_qty')
                )

                if not sales:
                    if master and master.status in ['Draft', 'Rejected']:
                        master.delete()
                        master = None
                    messages.warning(request, f"No free scheme (secondary sale) entries found for {selected_stockist.name} this month. Any previous drafts have been cleared.")
                else:
                    if not master:
                        master = FreeQtyClaimMaster.objects.create(
                            employee=selected_emp, stockist=selected_stockist, month=selected_month, year=selected_year, status='Draft'
                        )
                    else:
                        master.claim_lines.all().delete()
                        master.status = 'Draft'
                        master.save()

                    # 🚀 N+1 FIXED: products EK query + bulk_create (API twin)
                    prod_ids = [s['product_id'] for s in sales]
                    prod_map = {p.id: p for p in Product.objects.filter(id__in=prod_ids, company=employee.company)}

                    line_objs = []
                    for s in sales:
                        prod = prod_map.get(s['product_id'])
                        if not prod:
                            continue
                        price = float(prod.price) if getattr(prod, 'price', None) else 0.0
                        val = s['tot_free'] * price
                        line_objs.append(FreeQtyClaimLine(
                            master=master, stockist=selected_stockist, product=prod,
                            total_billed_qty=s['tot_billed'], total_free_qty=s['tot_free'], claim_value=val
                        ))
                    if line_objs:
                        FreeQtyClaimLine.objects.bulk_create(line_objs)
                    messages.success(request, f"🎉 Free Claim Report for {selected_stockist.name} has been synced successfully!")

        elif action == 'submit':
            if master and master.status in ['Draft', 'Rejected']:
                master.status = 'Pending_Manager'
                master.save()
                messages.success(request, f"✅ Claim for {selected_stockist.name} has been submitted for approval.")

        return redirect(f"{request.path}?employee_id={selected_emp.id}&stockist_id={selected_stockist_id}&month={selected_month}&year={selected_year}")

    lines = []
    grand_total = 0
    if master:
        lines = master.claim_lines.select_related('product').order_by('product__name')
        grand_total = sum(l.claim_value for l in lines)

    return render(request, 'free_claim_report.html', {
        'team_employees': team_employees,
        'is_manager_view': is_manager_view,
        'selected_emp_id': int(selected_emp_id),
        'selected_emp': selected_emp,
        'selected_month': selected_month,
        'selected_year': selected_year,
        'available_stockists': available_stockists,
        'selected_stockist_id': int(selected_stockist_id) if selected_stockist_id else '',
        'months_choices': [(i, calendar.month_name[i]) for i in range(1, 13)],
        'master': master,
        'lines': lines,
        'grand_total': grand_total,
        'is_locked': is_locked, 'deadline': deadline
    })

# ==============================================================================
# 2. 👁️ FREE CLAIM VIEW (READONLY)
# ==============================================================================
@employee_required
def free_claim_view_readonly(request, employee):
    team_employees = get_dropdown_team(employee)
    is_manager_view = employee.designation != 'MR'

    default_emp_id = str(employee.id)
    if is_manager_view:
        first_sub = team_employees.exclude(id=employee.id).first()
        if first_sub:
            default_emp_id = str(first_sub.id)

    selected_emp_id = request.GET.get('employee_id') or default_emp_id
    selected_emp = get_object_or_404(Employee, id=int(selected_emp_id), company=employee.company)

    selected_month = int(request.GET.get('month') or timezone.localdate().month)
    selected_year = int(request.GET.get('year') or timezone.localdate().year)

    my_terr_ids = [selected_emp.headquarter_id] if selected_emp.headquarter_id else []
    available_stockists = Stockist.objects.filter(territory_id__in=my_terr_ids).order_by('name')

    selected_stockist_id = request.GET.get('stockist_id')
    if selected_stockist_id and not available_stockists.filter(id=selected_stockist_id).exists():
        selected_stockist_id = None
    if not selected_stockist_id and available_stockists.exists():
        selected_stockist_id = str(available_stockists.first().id)

    selected_stockist = available_stockists.filter(id=selected_stockist_id).first()

    master = None
    if selected_stockist:
        master = FreeQtyClaimMaster.objects.filter(
            employee=selected_emp, stockist=selected_stockist, month=selected_month, year=selected_year
        ).first()

    lines = []
    grand_total = 0
    if master:
        lines = master.claim_lines.select_related('product').order_by('product__name')
        grand_total = sum(l.claim_value for l in lines)

    return render(request, 'free_claim_view_readonly.html', {
        'team_employees': team_employees,
        'is_manager_view': is_manager_view,
        'selected_emp_id': int(selected_emp_id),
        'selected_emp': selected_emp,
        'selected_month': selected_month,
        'selected_year': selected_year,
        'available_stockists': available_stockists,
        'selected_stockist_id': int(selected_stockist_id) if selected_stockist_id else '',
        'months_choices': [(i, calendar.month_name[i]) for i in range(1, 13)],
        'master': master,
        'lines': lines,
        'grand_total': grand_total,
    })

# ==============================================================================
# 3. ✔️ APPROVE FREE CLAIMS
# ==============================================================================
@employee_required
def approve_free_claims_view(request, employee):
    if employee.designation == 'MR':
        messages.error(request, "You do not have permission to access this page.")
        return redirect('request_hub')

    if request.method == "POST":
        claim_id = request.POST.get('claim_id')
        action = request.POST.get('action')
        remark = request.POST.get('remark', '')
        claim = get_object_or_404(FreeQtyClaimMaster, id=claim_id, employee__company=employee.company)   # 🛡️

        if action == 'approve':
            if employee.designation == 'Admin':
                claim.status = 'Approved'; claim.admin_remark = remark
            else:
                mgr_list = list(claim.approved_by_managers)
                if employee.id not in mgr_list: mgr_list.append(employee.id)
                claim.approved_by_managers = mgr_list
                claim.manager_remark = remark
                claim.status = 'Pending_Manager' if employee.manager_id else 'Pending_Admin'
            claim.save()
            messages.success(request, f"✅ Free Claim Report for {claim.employee.name} has been approved.")
        elif action == 'reject':
            claim.status = 'Rejected'
            if employee.designation == 'Admin': claim.admin_remark = remark
            else: claim.manager_remark = remark
            claim.save()
            messages.error(request, f"❌ Claim has been rejected.")
        return redirect('approve_free_claims')

    if employee.designation == 'Admin':
        pending_claims = FreeQtyClaimMaster.objects.filter(employee__company=employee.company, status='Pending_Admin').order_by('-created_at')
        history_claims = FreeQtyClaimMaster.objects.filter(employee__company=employee.company, status__in=['Approved', 'Rejected']).order_by('-updated_at')[:50]
    else:
        team = get_full_team_employees(employee).exclude(id=employee.id)
        pending_claims = FreeQtyClaimMaster.objects.filter(status='Pending_Manager', employee__in=team).order_by('-created_at')
        history_claims = FreeQtyClaimMaster.objects.filter(status__in=['Approved', 'Rejected', 'Pending_Admin'], employee__in=team).order_by('-updated_at')[:50]

    return render(request, 'approve_free_claims.html', {'pending_claims': pending_claims, 'history_claims': history_claims, 'is_admin': employee.designation == 'Admin'})

# ==============================================================================
# 4. 📦 MR INVENTORY (SAMPLES & GIFTS)
# ==============================================================================
@employee_required
def mr_inventory_view(request, employee):
    team_employees = get_dropdown_team(employee)
    is_manager_view = employee.designation != 'MR'

    if request.method == "POST" and "receive_dispatch" in request.POST:
        dispatch_id = request.POST.get('dispatch_id')
        if dispatch_id:
            try:
                dispatch = PromoDispatch.objects.get(id=dispatch_id, employee=employee, status='In-Transit')
                dispatch.status = 'Received'
                dispatch.save()
                
                inv_obj, created = MRInventory.objects.get_or_create(
                    employee=employee, 
                    item=dispatch.item,
                    defaults={'stock_qty': 0}
                )
                inv_obj.stock_qty += dispatch.quantity
                inv_obj.save()
                
                messages.success(request, f"✅ Stock of {dispatch.item.name} (Qty: {dispatch.quantity}) has been added to your inventory.")
            except PromoDispatch.DoesNotExist:
                messages.error(request, "⚠️ This dispatch has either already been received, or the ID is invalid.")
                
        return redirect(f"{request.path}?employee_id={employee.id}")

    default_emp_id = str(employee.id)
    if is_manager_view:
        first_sub = team_employees.exclude(id=employee.id).first()
        if first_sub: default_emp_id = str(first_sub.id)

    selected_emp_id = request.GET.get('employee_id', default_emp_id)
    selected_emp = get_object_or_404(Employee, id=int(selected_emp_id), company=employee.company)

    in_transit_items = PromoDispatch.objects.filter(employee=selected_emp, status='In-Transit').order_by('-dispatch_date')
    my_stock_qs = MRInventory.objects.filter(employee=selected_emp).select_related('item').order_by('item__item_type', 'item__name')

    sample_gift_data = []
    hv_stock_data = []
    # 🚀 N+1 KILLED: dispatch totals + ledger EK-2 queries mein (API twin)
    recv_map = {}
    for item_id, qty in PromoDispatch.objects.filter(employee=selected_emp, status='Received').values_list('item_id', 'quantity'):
        recv_map[item_id] = recv_map.get(item_id, 0) + qty

    ledger_map = {}   # (doctor_id, item_id) → latest ledger entry
    for led in DoctorROILedger.objects.filter(employee=selected_emp).select_related('doctor', 'item').order_by('date_given'):
        ledger_map[(led.doctor_id, led.item_id)] = led   # last wali jeetegi (order ascending)

    sample_gift_data = []
    hv_stock_data = []
    for stock in my_stock_qs:
        total_rec = recv_map.get(stock.item_id, 0)      # 🚀 dict lookup
        distributed = total_rec - stock.stock_qty
        if distributed < 0: distributed = 0
        if total_rec > 0 or stock.stock_qty > 0 or distributed > 0:
            row = {'item_name': stock.item.name, 'category': stock.item.item_type, 'received': total_rec, 'distributed': distributed, 'balance': stock.stock_qty}
            if stock.item.item_type == 'HighValue':
                hv_stock_data.append(row)
            else:
                sample_gift_data.append(row)

    hv_plans = GiftCampaignPlan.objects.filter(employee=selected_emp, item__item_type='HighValue').select_related('doctor', 'item').order_by('-year', '-month')
    hv_tracker = []
    for plan in hv_plans:
        ledger_entry = ledger_map.get((plan.doctor_id, plan.item_id))   # 🚀 dict lookup
        dist_date = ledger_entry.date_given if ledger_entry else None
        dist_qty = ledger_entry.quantity if ledger_entry else 0
        
        
        if dist_qty >= 1: status_badge = 'Delivered'
        elif plan.status == 'Approved': status_badge = 'Pending Delivery'
        else: status_badge = plan.status

        hv_tracker.append({'item_name': plan.item.name, 'doctor_name': plan.doctor.name, 'specialty': plan.doctor.specialty or '-', 'month_year': f"{calendar.month_name[plan.month][:3]} {plan.year}", 'plan_status': plan.status, 'dist_qty': dist_qty, 'dist_date': dist_date, 'final_status': status_badge})

    recent_logs = DoctorROILedger.objects.filter(employee=selected_emp).select_related('doctor', 'item').order_by('-date_given')[:100]

    if request.GET.get('export') == 'excel':
        filename = f"Inventory_Report_{selected_emp.name}.xlsx"
        wb = openpyxl.Workbook()
        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")

        ws1 = wb.active
        ws1.title = "Stock Summary"
        ws1.append(['INVENTORY SUMMARY REPORT'])
        ws1.append(['Employee:', selected_emp.name])
        ws1.append([''])
        ws1['A1'].font = Font(bold=True, size=14, color="107C41")
        ws1.append(['Item Name', 'Category', 'Total Received', 'Total Distributed', 'Current Balance (Pending)'])
        for col_num, cell in enumerate(ws1[4], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws1.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20
        for row in sample_gift_data: ws1.append([row['item_name'], row['category'], row['received'], row['distributed'], row['balance']])

        ws2 = wb.create_sheet(title="High Value Tracker")
        ws2.append(['HIGH VALUE GIFT TRACKER (PROPOSED vs ACTUAL)'])
        ws2.append(['Employee:', selected_emp.name])
        ws2.append([''])
        ws2['A1'].font = Font(bold=True, size=14, color="107C41")
        ws2.append(['Doctor Name', 'Specialty', 'Item Name', 'Plan Month', 'Approval Status', 'Distributed Qty', 'Distributed Date', 'Execution Status'])
        for col_num, cell in enumerate(ws2[4], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws2.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 18
        for row in hv_tracker: ws2.append([f"Dr. {row['doctor_name']}", row['specialty'], row['item_name'], row['month_year'], row['plan_status'], row['dist_qty'], row['dist_date'] or '-', row['final_status']])

        ws3 = wb.create_sheet(title="Distribution Log")
        ws3.append(['RECENT DISTRIBUTION LOG'])
        ws3.append(['Employee:', selected_emp.name])
        ws3.append([''])
        ws3['A1'].font = Font(bold=True, size=14, color="107C41")
        ws3.append(['Date Given', 'Doctor Name', 'Item Name', 'Category', 'Quantity Distributed', 'Value (₹)'])
        for col_num, cell in enumerate(ws3[4], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws3.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 18
        for log in recent_logs: ws3.append([log.date_given, f"Dr. {log.doctor.name}", log.item.name, log.item.item_type, log.quantity, float(log.total_value)])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return render(request, 'mr_inventory.html', {
        'team_employees': team_employees, 'selected_emp_id': int(selected_emp_id),
        'sample_gift_data': sample_gift_data, 'hv_stock_data': hv_stock_data,
        'hv_tracker': hv_tracker,
        'recent_logs': recent_logs, 'in_transit_items': in_transit_items, 'is_manager_view': is_manager_view
    })

# ==============================================================================
# 5. 📉 GIFT DISTRIBUTION REPORT
# ==============================================================================
@employee_required
def gift_distribution_report(request, employee):
    team_employees = get_dropdown_team(employee)
    is_manager_view = employee.designation != 'MR'

    today = timezone.localdate()
    
    try:
        raw_from = request.GET.get('from_month')
        from_month = int(raw_from) if raw_from else today.month
    except (TypeError, ValueError):
        from_month = today.month

    try:
        raw_to = request.GET.get('to_month')
        to_month = int(raw_to) if raw_to else today.month
    except (TypeError, ValueError):
        to_month = today.month

    try:
        raw_year = request.GET.get('year')
        selected_year = int(raw_year) if raw_year else today.year
    except (TypeError, ValueError):
        selected_year = today.year
        
    if from_month > to_month:
        from_month, to_month = to_month, from_month

    selected_emp_id = request.GET.get('employee_id', '')
    inv_type = request.GET.get('inv_type', 'any_gift')

    if is_manager_view and selected_emp_id:
        # 🛡️ IDOR FIX: selected employee apni team ka ho
        sel = Employee.objects.filter(id=selected_emp_id, company=employee.company).first()
        if sel and not team_employees.filter(id=sel.id).exists():
            qs = DoctorROILedger.objects.none(); plan_qs = GiftCampaignPlan.objects.none()
        else:
            qs = DoctorROILedger.objects.filter(employee_id=selected_emp_id)
            plan_qs = GiftCampaignPlan.objects.filter(employee_id=selected_emp_id)
        plan_qs = GiftCampaignPlan.objects.filter(employee_id=selected_emp_id)
    elif is_manager_view: 
        qs = DoctorROILedger.objects.filter(employee__in=team_employees)
        plan_qs = GiftCampaignPlan.objects.filter(employee__in=team_employees)
    else: 
        qs = DoctorROILedger.objects.filter(employee=employee)
        plan_qs = GiftCampaignPlan.objects.filter(employee=employee)

    qs = qs.filter(
        date_given__month__gte=from_month,
        date_given__month__lte=to_month,
        date_given__year=selected_year
    ).select_related('doctor', 'employee', 'item').order_by('-date_given')

    if inv_type == 'high_value':
        qs = qs.filter(item__item_type='HighValue')
    elif inv_type == 'normal':
        qs = qs.filter(item__item_type='Routine')

    total_qty = qs.aggregate(t_qty=Sum('quantity'))['t_qty'] or 0
    total_val = qs.aggregate(t_val=Sum('total_value'))['t_val'] or 0

    plan_qs = plan_qs.filter(
        month__gte=from_month,
        month__lte=to_month,
        year=selected_year,
        status='Approved'
    ).select_related('doctor', 'item', 'employee')
    
    delivered_combinations = set(
        qs.filter(item__item_type='HighValue').values_list('employee_id', 'doctor_id', 'item_id')
    )
    
    pending_hv_list = []
    for plan in plan_qs:
        if (plan.employee_id, plan.doctor_id, plan.item_id) not in delivered_combinations:
            pending_hv_list.append(plan)

    if request.GET.get('export') == 'excel':
        filename = f"Gift_Distribution_{calendar.month_abbr[from_month]}_to_{calendar.month_abbr[to_month]}_{selected_year}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Gift Distribution"
        
        ws.append(['GIFT DISTRIBUTION REPORT'])
        ws.append(['Period:', f"{calendar.month_name[from_month]} to {calendar.month_name[to_month]} {selected_year}"])
        ws.append(['Filter:', inv_type.replace('_', ' ').title()])
        ws.append([''])
        
        ws['A1'].font = Font(bold=True, size=14, color="107C41")
        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        
        headers = ['Date Given', 'Employee Name', 'Doctor Name', 'Specialty', 'Item Name', 'Category', 'Quantity', 'Value (₹)']
        ws.append(headers)
        
        for col_num, cell in enumerate(ws[5], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20 if col_num in [2,3,5] else 15

        for log in qs:
            ws.append([log.date_given.strftime('%d-%b-%Y'), log.employee.name, f"Dr. {log.doctor.name}", log.doctor.specialty or '-', log.item.name, log.item.item_type, log.quantity, float(log.total_value)])
            
        ws.append(['GRAND TOTAL', '', '', '', '', '', total_qty, round(float(total_val), 2)])
        for cell in ws[ws.max_row]: cell.font = Font(bold=True)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return render(request, 'gift_distribution_report.html', {
        'reports': qs, 
        'pending_hv_list': pending_hv_list,
        'team_employees': team_employees, 
        'is_manager_view': is_manager_view,
        'from_month': from_month,
        'to_month': to_month,
        'selected_year': selected_year,
        'selected_emp_id': int(selected_emp_id) if selected_emp_id else '',
        'total_qty': total_qty, 
        'total_val': total_val, 
        'months_choices': [(i, calendar.month_name[i]) for i in range(1, 13)],
        'inv_type': inv_type,
    })

# ==============================================================================
# 6. 📊 DOCTOR ROI REPORT
# ==============================================================================
@login_required
def doctor_roi_report(request):
    employee = request.user.employee 
    
    team_employees = get_dropdown_team(employee)
    is_manager_view = employee.designation != 'MR'

    today = timezone.localdate()
    
    try:
        raw_from = request.GET.get('from_month')
        from_month = int(raw_from) if raw_from else today.month
    except (TypeError, ValueError):
        from_month = today.month

    try:
        raw_to = request.GET.get('to_month')
        to_month = int(raw_to) if raw_to else today.month
    except (TypeError, ValueError):
        to_month = today.month

    try:
        raw_year = request.GET.get('year')
        selected_year = int(raw_year) if raw_year else today.year
    except (TypeError, ValueError):
        selected_year = today.year
        
    if from_month > to_month:
        from_month, to_month = to_month, from_month
        
    months_range = list(range(from_month, to_month + 1))
    months_headers = [(m, calendar.month_name[m][:3]) for m in months_range]

    selected_emp_id = request.GET.get('employee_id', '')
    inv_type = request.GET.get('inv_type', 'any_gift')

    if is_manager_view and selected_emp_id:
        # 🛡️ IDOR FIX: selected employee apni team ka ho
        sel = Employee.objects.filter(id=selected_emp_id, company=employee.company).first()
        if sel and not team_employees.filter(id=sel.id).exists():
            qs = DoctorROILedger.objects.none(); plan_qs = GiftCampaignPlan.objects.none()
        else:
            qs = DoctorROILedger.objects.filter(employee_id=selected_emp_id)
            plan_qs = GiftCampaignPlan.objects.filter(employee_id=selected_emp_id)
        base_docs = Doctor.objects.filter(allocated_to_id=selected_emp_id)
    elif is_manager_view: 
        qs = DoctorROILedger.objects.filter(employee__in=team_employees)
        base_docs = Doctor.objects.filter(allocated_to__in=team_employees)
    else: 
        qs = DoctorROILedger.objects.filter(employee=employee)
        base_docs = Doctor.objects.filter(allocated_to=employee)

    qs = qs.filter(
        date_given__month__gte=from_month,
        date_given__month__lte=to_month,
        date_given__year=selected_year
    ).select_related('doctor', 'employee', 'item')

    if inv_type == 'high_value':
        qs = qs.filter(item__item_type='HighValue')
    elif inv_type == 'normal':
        qs = qs.filter(item__item_type='Routine')

    agg_data = defaultdict(lambda: {
        'doc_name': '', 'items': set(), 'emp_name': '', 'doc_id': 0,
        'monthly': {m: {'inv_qty': 0, 'inv_val': 0.0} for m in months_range},
        'tot_inv_qty': 0, 'tot_inv_val': 0.0
    })
    
    for row in qs:
        m = row.date_given.month
        key = row.doctor_id
        
        if not agg_data[key]['doc_name']:
            agg_data[key]['doc_name'] = row.doctor.name
            agg_data[key]['emp_name'] = row.employee.name if row.employee else 'Unknown'
            agg_data[key]['doc_id'] = row.doctor_id
            
        agg_data[key]['items'].add((row.item.name, row.item.item_type))
        agg_data[key]['monthly'][m]['inv_qty'] += row.quantity
        agg_data[key]['monthly'][m]['inv_val'] += float(row.total_value)
        agg_data[key]['tot_inv_qty'] += row.quantity
        agg_data[key]['tot_inv_val'] += float(row.total_value)

    if inv_type == 'all_doctors':
        for doc in base_docs.select_related('allocated_to'):
            if doc.id not in agg_data:
                agg_data[doc.id]['doc_name'] = doc.name
                agg_data[doc.id]['emp_name'] = doc.allocated_to.name if doc.allocated_to else 'Unknown'
                agg_data[doc.id]['doc_id'] = doc.id
                agg_data[doc.id]['items'].add(('No Gifts', 'Routine'))

    doc_ids = set(agg_data.keys())
    rx_qs = DoctorRxMapping.objects.filter(
        doctor_id__in=doc_ids,
        party_line__report__month__gte=from_month,
        party_line__report__month__lte=to_month,
        party_line__report__year=selected_year
    ).select_related('party_line__product')
    rx_dict = defaultdict(lambda: defaultdict(float))
    for rx in rx_qs:
        m = rx.party_line.report.month
        price = float(rx.party_line.product.price) if getattr(rx.party_line.product, 'price', None) else 0.0
        rx_dict[rx.doctor_id][m] += (rx.mapped_billed_qty * price)

    report_data = []
    gt_monthly = {m: {'inv_qty': 0, 'inv_val': 0.0, 'rx_val': 0.0} for m in months_range}
    gt_inv_qty, gt_inv_val, gt_rx_val = 0, 0.0, 0.0
    
    for doc_id, data in agg_data.items():
        monthly_list = []
        doc_tot_rx = 0.0
        
        for m in months_range:
            i_q = data['monthly'][m]['inv_qty']
            i_v = data['monthly'][m]['inv_val']
            r_v = rx_dict[doc_id][m]
            doc_tot_rx += r_v
            
            roi = (r_v / i_v * 100) if i_v > 0 else 0.0
            monthly_list.append({
                'inv_qty': i_q, 'inv_val': i_v, 'rx_val': r_v, 'roi': roi
            })
            
            gt_monthly[m]['inv_qty'] += i_q
            gt_monthly[m]['inv_val'] += i_v
            gt_monthly[m]['rx_val'] += r_v 
                
        gt_inv_qty += data['tot_inv_qty']
        gt_inv_val += data['tot_inv_val']
        gt_rx_val += doc_tot_rx
        
        tot_roi = (doc_tot_rx / data['tot_inv_val'] * 100) if data['tot_inv_val'] > 0 else 0.0
        
        item_names = " + ".join(sorted([i[0] for i in data['items']]))
        item_types = set([i[1] for i in data['items']])
        final_item_type = 'HighValue' if 'HighValue' in item_types else 'Routine'

        report_data.append({
            'doc_name': data['doc_name'],
            'item_name': item_names,
            'item_type': final_item_type,
            'emp_name': data['emp_name'],
            'monthly_list': monthly_list,
            'tot_inv_qty': data['tot_inv_qty'],
            'tot_inv_val': data['tot_inv_val'],
            'tot_rx_val': doc_tot_rx,
            'tot_roi': tot_roi
        })
        
    report_data.sort(key=lambda x: x['doc_name'])
    gt_monthly_list = [gt_monthly[m] for m in months_range]
    overall_roi = (gt_rx_val / gt_inv_val * 100) if gt_inv_val > 0 else 0.0

    if request.GET.get('export') == 'excel':
        filename = f"Doctor_ROI_Matrix_{calendar.month_abbr[from_month]}_to_{calendar.month_abbr[to_month]}_{selected_year}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ROI Matrix"

        ws.append(['DOCTOR ROI MATRIX (MONTH-WISE)'])
        ws.append(['Period:', f"{calendar.month_name[from_month]} to {calendar.month_name[to_month]} {selected_year}"])
        ws.append(['Filter:', inv_type.replace('_', ' ').title()])
        ws.append([''])
        ws['A1'].font = Font(bold=True, size=14, color="107C41")

        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")

        headers1 = ['Doctor Name', 'Item Details', 'Given By']
        headers2 = ['', '', '']
        
        for m_num, m_name in months_headers:
            headers1.extend([f"{m_name}", "", "", ""])
            headers2.extend(['Inv Qty', 'Inv Val', 'Rx Val', 'ROI %'])
            
        headers1.extend(["GRAND TOTAL", "", "", ""])
        headers2.extend(['Inv Qty', 'Inv Val', 'Rx Val', 'ROI %'])

        ws.append(headers1)
        ws.append(headers2)

        for row in report_data:
            row_data = [f"Dr. {row['doc_name']}", row['item_name'], row['emp_name']]
            for m in row['monthly_list']:
                row_data.extend([m['inv_qty'], m['inv_val'], m['rx_val'], round(m['roi'], 1)])
            row_data.extend([row['tot_inv_qty'], row['tot_inv_val'], row['tot_rx_val'], round(row['tot_roi'], 1)])
            ws.append(row_data)

        gt_row = ['GRAND TOTAL', '', '']
        for m_gt in gt_monthly_list:
            m_roi = (m_gt['rx_val'] / m_gt['inv_val'] * 100) if m_gt['inv_val'] > 0 else 0.0
            gt_row.extend([m_gt['inv_qty'], m_gt['inv_val'], m_gt['rx_val'], round(m_roi, 1)])
        gt_row.extend([gt_inv_qty, gt_inv_val, gt_rx_val, round(overall_roi, 1)])
        ws.append(gt_row)
        
        for cell in ws[ws.max_row]: cell.font = Font(bold=True)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return render(request, 'doctor_roi_report.html', {
        'team_employees': team_employees,
        'selected_emp_id': int(selected_emp_id) if selected_emp_id else '',
        'inv_type': inv_type,
        'from_month': from_month,
        'to_month': to_month,
        'selected_year': selected_year,
        'months_choices': [(i, calendar.month_name[i]) for i in range(1, 13)],
        'months_headers': months_headers,
        'report_data': report_data,
        'gt_monthly_list': gt_monthly_list,
        'gt_inv_qty': gt_inv_qty,
        'gt_inv_val': gt_inv_val,
        'gt_rx_val': gt_rx_val,
        'overall_roi': overall_roi,
        'is_manager_view': is_manager_view,
    })