import io
import csv
import json
import calendar
from datetime import date, datetime
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404
from django.utils import timezone
from django.db.models import Sum

from SFA.models import (
    Employee, Doctor, Chemist, DailyDCR, DCRProductDetail,
    MonthlyTourProgram, Territory, MonthlyExpenseReport, DailyTourPlan, Route, Holiday,
    DayStart, DayEnd, SystemSetting
)
from .auth import get_dropdown_team, get_team_territory_ids, get_team_requested_routes
from SFA.decorators import employee_required

# ==============================================================================
# 1. 📝 DCR REPORT VIEW (Speed Optimized)
# ==============================================================================
@employee_required
def dcr_report_view(request, employee):
    team_employees = get_dropdown_team(employee)
    default_emp_id = str(employee.id)
    if employee.designation != 'MR':
        first_sub = team_employees.exclude(id=employee.id).first()
        if first_sub:
            default_emp_id = str(first_sub.id)
            
    selected_emp_id = request.GET.get('employee_id', default_emp_id)
    selected_emp = get_object_or_404(Employee, id=selected_emp_id)
    
    today = timezone.localdate()
    selected_month = int(request.GET.get('month') or today.month)
    selected_year = int(request.GET.get('year') or today.year)

    # 🚀 SPEED HACK: select_related aur prefetch_related laga diya gaya hai
    day_starts = DayStart.objects.filter(
        employee=selected_emp, date__month=selected_month, date__year=selected_year
    ).select_related('employee', 'territory').prefetch_related('routes__territory').order_by('-date')

    actual_dcrs = DailyDCR.objects.filter(
        employee=selected_emp, date__month=selected_month, date__year=selected_year
    ).prefetch_related('visits__doctor', 'visits__chemist', 'visits__product_details__product')
    dcr_dict = {d.date: d for d in actual_dcrs}

    report_list, total_dr_visits, total_chem_visits = [], 0, 0

    for ds in day_starts:
        dcr_obj = dcr_dict.get(ds.date)
        visits = dcr_obj.visits.all() if dcr_obj else []

        if dcr_obj:
            dr_v = len([v for v in visits if v.doctor_id])
            chem_v = len([v for v in visits if v.chemist_id])
            total_dr_visits += dr_v
            total_chem_visits += chem_v
            
        route_list = ds.routes.all() if hasattr(ds, 'routes') else []
        territory_names = []
        for r in route_list:
            if r.territory and r.territory.name not in territory_names:
                territory_names.append(r.territory.name)
        final_route = {'name': ", ".join([r.name for r in route_list]), 'territory': ", ".join(territory_names)} if route_list else None

        report_list.append({'id': ds.id, 'date': ds.date, 'work_type': ds.work_type, 'visits': visits, 'employee': ds.employee, 'route': final_route})

    months = [(i, calendar.month_name[i]) for i in range(1, 13)]
    years = [today.year - 1, today.year, today.year + 1]
    total_days_worked = day_starts.count()
    dr_avg   = round(total_dr_visits   / total_days_worked, 1) if total_days_worked > 0 else 0
    chem_avg = round(total_chem_visits / total_days_worked, 1) if total_days_worked > 0 else 0

    total_samples = DCRProductDetail.objects.filter(visit__daily_dcr__employee=selected_emp, visit__daily_dcr__date__month=selected_month, visit__daily_dcr__date__year=selected_year).aggregate(s=Sum('sample_qty'))['s'] or 0
    total_orders = DCRProductDetail.objects.filter(visit__daily_dcr__employee=selected_emp, visit__daily_dcr__date__month=selected_month, visit__daily_dcr__date__year=selected_year).aggregate(o=Sum('order_qty'))['o'] or 0

    return render(request, 'dcr_report.html', {
        'dcrs': report_list, 'team_employees': team_employees, 'selected_emp_id': int(selected_emp_id),
        'selected_employee_name': selected_emp.name, 'is_manager_view': employee.designation != 'MR',
        'selected_month': selected_month, 'selected_year': selected_year, 'months': months, 'years': years,
        'total_days_worked': total_days_worked, 'total_dr_visits': total_dr_visits, 'total_chem_visits': total_chem_visits,
        'dr_avg': dr_avg, 'chem_avg': chem_avg, 'total_samples': total_samples, 'total_orders': total_orders,
    })

# ==============================================================================
# 2. 💸 EXPENSE REPORT VIEW (Speed Optimized)
# ==============================================================================
@employee_required
def expense_report_view(request, employee):
    team_employees = get_dropdown_team(employee)
    default_emp_id = str(employee.id)
    if employee.designation != 'MR':
        first_sub = team_employees.exclude(id=employee.id).first()
        if first_sub: default_emp_id = str(first_sub.id)
            
    selected_emp_id = request.GET.get('employee_id', default_emp_id)
    selected_emp = get_object_or_404(Employee, id=selected_emp_id)

    selected_month  = int(request.GET.get('month') or timezone.localdate().month)
    selected_year   = int(request.GET.get('year') or timezone.localdate().year)

    # 🚀 SPEED HACK: prefetch_related('daily_lines') added here
    expenses = MonthlyExpenseReport.objects.filter(
        employee=selected_emp, month=selected_month, year=selected_year
    ).prefetch_related('daily_lines').order_by('-year', '-month')
    
    day_starts = DayStart.objects.filter(
        employee=selected_emp, date__month=selected_month, date__year=selected_year
    ).prefetch_related('routes', 'territory')
    ds_dict = {ds.date: ds for ds in day_starts}

    for exp in expenses:
        total = 0
        processed_lines = []
        # No extra queries will be fired inside this loop now
        for line in exp.daily_lines.all():
            app_ta   = line.approved_ta   if line.approved_ta   is not None else line.ta_amount
            app_da   = line.approved_da   if line.approved_da   is not None else line.da_amount
            app_misc = line.approved_misc if line.approved_misc is not None else line.misc_amount
            total += float(app_ta) + float(app_da) + float(app_misc)

            line_date = line.date if hasattr(line.date, 'year') else line.date.date()
            ds = ds_dict.get(line_date)
            location_display = ""
            if ds:
                route_str = ", ".join([r.name for r in ds.routes.all()])
                hq_str = ds.territory.name if ds.territory else ""
                if route_str:
                    location_display = f"{route_str} ({hq_str})" if hq_str else route_str
                elif hq_str:
                    location_display = hq_str

            line.working_area  = location_display or '-'
            line.display_cat   = line.get_territory_category_display() if line.territory_category else 'HQ'
            line.display_km    = float(line.distance_km or 0)
            line.display_ta    = app_ta
            line.display_da    = app_da
            line.display_misc  = app_misc
            line.display_total = round(float(app_ta) + float(app_da) + float(app_misc), 2)
            processed_lines.append(line)
            
        exp.grand_total = round(total, 2)
        # Sort processed lines by date for UI
        exp.processed_lines = sorted(processed_lines, key=lambda x: x.date)

    if request.GET.get('export') in ['csv', 'excel']:
        filename = f"Expense_Report_{selected_emp.name}_{selected_month}_{selected_year}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Expense Report"

        ws.append(['MONTHLY EXPENSE REPORT'])
        ws.append(['Employee:', selected_emp.name, 'Period:', f"{calendar.month_name[selected_month]} {selected_year}"])
        ws.append([''])
        
        ws['A1'].font = Font(bold=True, size=14, color="107C41")
        ws['A2'].font = Font(bold=True); ws['C2'].font = Font(bold=True)

        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")

        headers = ['Date', 'Category', 'Route / HQ', 'Distance (KM)', 'TA (Rs)', 'DA (Rs)', 'Misc (Rs)', 'Total (Rs)', 'Status']
        ws.append(headers)

        for col_num, cell in enumerate(ws[5], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 15
        ws.column_dimensions['C'].width = 30 

        grand_total_all = 0.0
        for exp in expenses:
            for line in exp.processed_lines:
                grand_total_all += line.display_total
                date_str = line.date.strftime('%d-%b-%Y') if line.date else 'N/A'
                ws.append([
                    date_str, line.display_cat, line.working_area, line.display_km,
                    float(line.display_ta), float(line.display_da), float(line.display_misc), line.display_total, exp.status
                ])
                
        ws.append([''])
        gt_row = ['GRAND TOTAL', '', '', '', '', '', '', round(grand_total_all, 2), '']
        ws.append(gt_row)
        for cell in ws[ws.max_row]: cell.font = Font(bold=True)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return render(request, 'expense_report.html', {
        'team_employees': team_employees, 'selected_emp_id': int(selected_emp_id), 'selected_employee_name': selected_emp.name,
        'selected_month': selected_month, 'selected_year': selected_year,
        'months': [(i, calendar.month_name[i]) for i in range(1, 13)],
        'is_manager_view': employee.designation != 'MR', 'expenses': expenses,
    })


# ==============================================================================
# 3. 🌐 NETWORK REPORT VIEW
# ==============================================================================
@employee_required
def network_report_view(request, employee):
    team_employees = get_dropdown_team(employee)
    default_emp_id = str(employee.id)
    if employee.designation != 'MR':
        first_sub = team_employees.exclude(id=employee.id).first()
        if first_sub: default_emp_id = str(first_sub.id)
            
    selected_emp_id = request.GET.get('employee_id', default_emp_id)
    selected_emp = get_object_or_404(Employee, id=selected_emp_id)
    
    sub_team = get_dropdown_team(selected_emp, ordered=False)
    my_terr_ids = get_team_territory_ids(sub_team)
    routes = get_team_requested_routes(sub_team, my_terr_ids)

    active_tab = request.GET.get('tab', 'doctor')
    route_id = request.GET.get('route', '')
    specialty = request.GET.get('specialty', '')
    category = request.GET.get('category', '')
    
    doctors = Doctor.objects.filter(allocated_to__in=sub_team, status='Approved')
    chemists = Chemist.objects.filter(allocated_to__in=sub_team, status='Approved')
    
    if active_tab == 'doctor':
        if route_id: doctors = doctors.filter(route_id=route_id)
        if specialty: doctors = doctors.filter(specialty=specialty)
        if category: doctors = doctors.filter(category=category)
    elif active_tab == 'chemist':
        if route_id: chemists = chemists.filter(route_id=route_id)
        
    setting, _ = SystemSetting.objects.get_or_create(company=employee.company)

    if request.GET.get('export') == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")

        if active_tab == 'doctor':
            filename = f"Doctor_Network_{selected_emp.name}.xlsx"
            ws.title = "Doctor List"
            ws.append(['MASTER DOCTOR DIRECTORY'])
            ws.append(['Employee:', selected_emp.name])
            ws.append([''])
            ws['A1'].font = Font(bold=True, size=14, color="107C41")
            
            headers = ['Doctor Name', 'Specialty', 'Category', 'Route / Patch', 'Territory / HQ', 'Allocated MR', 'Status']
            ws.append(headers)
            for col_num, cell in enumerate(ws[4], 1):
                cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20

            for doc in doctors:
                terr_name = getattr(doc.territory, 'name', None)
                if not terr_name and doc.route and getattr(doc.route, 'territory', None):
                    terr_name = doc.route.territory.name
                ws.append([f"Dr. {doc.name}", doc.specialty or '-', doc.category or '-', doc.route.name if doc.route else '-', terr_name or '-', doc.allocated_to.name if doc.allocated_to else '-', doc.status or 'Approved'])

        elif active_tab == 'chemist':
            filename = f"Chemist_Network_{selected_emp.name}.xlsx"
            ws.title = "Chemist List"
            ws.append(['MASTER CHEMIST / PHARMACY DIRECTORY'])
            ws.append(['Employee:', selected_emp.name])
            ws.append([''])
            ws['A1'].font = Font(bold=True, size=14, color="107C41")
            
            headers = ['Pharmacy / Chemist Name', 'Route / Patch', 'Territory / HQ', 'Allocated MR', 'Status']
            ws.append(headers)
            for col_num, cell in enumerate(ws[4], 1):
                cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 25

            for chem in chemists:
                terr_name = getattr(chem.territory, 'name', None) if hasattr(chem, 'territory') else None
                if not terr_name and chem.route and getattr(chem.route, 'territory', None): terr_name = chem.route.territory.name
                ws.append([chem.name, chem.route.name if chem.route else '-', terr_name or '-', chem.allocated_to.name if chem.allocated_to else '-', getattr(chem, 'status', 'Approved')])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return render(request, 'network_report.html', {
        'doctors': doctors, 'chemists': chemists, 'routes': routes,
        'selected_emp_id': int(selected_emp_id), 'team_employees': team_employees,
        'is_manager_view': employee.designation != 'MR',
        'active_tab': active_tab, 'selected_route': int(route_id) if route_id else '',
        'selected_specialty': specialty, 'selected_category': category,
        'specialty_choices': Doctor.SPECIALTY_CHOICES, 'category_choices': Doctor.CATEGORY_CHOICES,
        'allow_location_capture': setting.allow_location_capture,
    })


# ==============================================================================
# 4. 🗺️ ROUTE REPORT VIEW
# ==============================================================================
@employee_required
def route_report_view(request, employee):
    team_employees = get_dropdown_team(employee)
    default_emp_id = str(employee.id)
    if employee.designation != 'MR':
        first_sub = team_employees.exclude(id=employee.id).first()
        if first_sub: default_emp_id = str(first_sub.id)
            
    selected_emp_id = request.GET.get('employee_id', default_emp_id)
    selected_emp = get_object_or_404(Employee, id=int(selected_emp_id))
    
    sub_team = get_dropdown_team(selected_emp, ordered=False)
    my_terr_ids = get_team_territory_ids(sub_team)
    routes = get_team_requested_routes(sub_team, my_terr_ids)

    report_data = []
    gt_docs = 0; gt_chems = 0

    for r in routes:
        doc_count = Doctor.objects.filter(route=r, status='Approved').count()
        chem_count = Chemist.objects.filter(route=r, status='Approved').count()
        report_data.append({
            'route_name': r.name, 'territory': r.territory.name if r.territory else 'N/A',
            'category': r.get_category_display() if r.category else 'HQ', 'distance': float(r.distance_from_hq or 0),
            'doc_count': doc_count, 'chem_count': chem_count, 'total_customers': doc_count + chem_count
        })
        gt_docs += doc_count; gt_chems += chem_count

    report_data = sorted(report_data, key=lambda x: (x['territory'], x['route_name']))
    
    if request.GET.get('export') == 'excel':
        filename = f"Route_Report_{selected_emp.name}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Route Coverage"
        ws.append(['ROUTE & CUSTOMER COVERAGE REPORT'])
        ws.append(['Employee:', selected_emp.name])
        ws.append([''])
        ws['A1'].font = Font(bold=True, size=14, color="107C41")
        ws['A2'].font = Font(bold=True)
        
        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        
        headers = ['Route Name', 'Territory / HQ', 'Category', 'Distance (KM)', 'Total Doctors', 'Total Chemists', 'Total Customers']
        ws.append(headers)
        
        for col_num, cell in enumerate(ws[4], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20 if col_num <= 2 else 15
            
        for row in report_data:
            ws.append([row['route_name'], row['territory'], row['category'], row['distance'], row['doc_count'], row['chem_count'], row['total_customers']])
            
        ws.append(['GRAND TOTAL', '', '', '', gt_docs, gt_chems, gt_docs + gt_chems])
        for cell in ws[ws.max_row]: cell.font = Font(bold=True)
            
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return render(request, 'route_report.html', {
        'team_employees': team_employees, 'selected_emp_id': int(selected_emp_id),
        'selected_employee_name': selected_emp.name, 'is_manager_view': employee.designation != 'MR',
        'report_data': report_data, 'gt_docs': gt_docs, 'gt_chems': gt_chems, 'gt_total': gt_docs + gt_chems
    })


# ==============================================================================
# 5. 📅 TOUR PLAN (MTP) REPORT
# ==============================================================================
@employee_required
def tour_plan_report_view(request, employee):
    team_employees = get_dropdown_team(employee)
    default_emp_id = str(employee.id)
    if employee.designation != 'MR':
        first_sub = team_employees.exclude(id=employee.id).first()
        if first_sub: default_emp_id = str(first_sub.id)
            
    selected_emp_id = request.GET.get('employee_id', default_emp_id)
    selected_emp = get_object_or_404(Employee, id=selected_emp_id)
    
    selected_month = int(request.GET.get('month') or timezone.localdate().month)
    selected_year = int(request.GET.get('year') or timezone.localdate().year)
    
    daily_plans = DailyTourPlan.objects.filter(mtp__employee=selected_emp, date__month=selected_month, date__year=selected_year).select_related('route')
    plan_map = {p.date.day: p.route.name for p in daily_plans if p.route}
    
    mtp = MonthlyTourProgram.objects.filter(employee=selected_emp, month=selected_month, year=selected_year).first()
    mtp_status = mtp.status if mtp else 'Not Created'
    
    days_in_month = calendar.monthrange(selected_year, selected_month)[1]
    report_data = [{'day': d, 'date_str': date(selected_year, selected_month, d).strftime("%d %b, %A"), 'route_name': plan_map.get(d, 'No Plan / Weekly Off')} for d in range(1, days_in_month + 1)]
        
    months = [(i, calendar.month_name[i]) for i in range(1, 13)]
    if request.GET.get('export') == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="TourPlan_{selected_emp.name}_{selected_month}_{selected_year}.csv"'
        writer = csv.writer(response)
        writer.writerow(['Employee Name', 'Month', 'Year', 'MTP Status'])
        writer.writerow([selected_emp.name, selected_month, selected_year, mtp_status])
        writer.writerow([]) 
        writer.writerow(['Date & Day', 'Planned Target Route'])
        for row in report_data: writer.writerow([row['date_str'], row['route_name']])
        return response

    return render(request, 'tour_plan_report.html', {
        'team_employees': team_employees, 'selected_emp_id': int(selected_emp_id),
        'selected_month': selected_month, 'selected_year': selected_year,
        'months': months, 'report_data': report_data, 'selected_employee_name': selected_emp.name, 'mtp_status': mtp_status
    })

# ==============================================================================
# 6. 🏖️ HOLIDAY LIST VIEW
# ==============================================================================
@employee_required
def holiday_list_view(request, employee):
    selected_emp_id = request.GET.get('employee_id', str(employee.id))
    selected_emp = get_object_or_404(Employee, id=selected_emp_id)
    
    rbm_emp = None
    curr = selected_emp
    while curr:
        if curr.designation == 'RBM':
            rbm_emp = curr
            break
        curr = curr.manager
        
    admin_ids = list(Employee.objects.filter(company=selected_emp.company, designation='Admin').values_list('id', flat=True))
    holiday_creators = admin_ids.copy()
    if rbm_emp: holiday_creators.append(rbm_emp.id)
        
    holidays = Holiday.objects.filter(proposed_by_id__in=holiday_creators, status='Approved').order_by('-date')

    if request.GET.get('export') == 'excel':
        region_name = rbm_emp.name if rbm_emp else "Consolidated"
        filename = f"Approved_Holidays_{region_name}.xlsx"
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Holidays"
        
        ws.append(['APPROVED HOLIDAYS LIST'])
        ws.append(['Region:', region_name])
        ws.append([''])
        
        ws['A1'].font = Font(bold=True, size=12, color="107C41")
        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        ws.append(['Date', 'Day', 'Holiday Name'])
        for col_num, cell in enumerate(ws[4], 1):
            cell.fill = header_fill
            cell.font = header_font
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20
        
        for h in holidays:
            ws.append([h.date.strftime('%d-%b-%Y'), h.date.strftime('%A'), h.name])
            
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return render(request, 'holiday_list.html', {
        'holidays': holidays, 
        'rbm_name': rbm_emp.name if rbm_emp else "Consolidated (All States)", 
        'selected_emp_id': selected_emp.id
    })

# ==============================================================================
# 7. 📍 ROUTE PLAYBACK VIEW
# ==============================================================================
@employee_required
def route_playback_view(request, employee, employee_id, date_str):
    emp = get_object_or_404(Employee, id=employee_id)
    target_date = datetime.strptime(date_str, '%Y-%m-%d').date()

    day_start = DayStart.objects.filter(employee=emp, date=target_date).first()
    day_end = DayEnd.objects.filter(employee=emp, date=target_date).first()
    daily_dcr = DailyDCR.objects.filter(employee=emp, date=target_date).first()

    waypoints = []
    if day_start and day_start.latitude and day_start.longitude:
        waypoints.append({'lat': float(day_start.latitude), 'lng': float(day_start.longitude), 'title': 'Day Start', 'time': day_start.started_at.strftime('%I:%M %p'), 'type': 'start'})
    
    if daily_dcr:
        visits = daily_dcr.visits.all().select_related('doctor', 'chemist').order_by('created_at')
        for v in visits:
            if v.latitude and v.longitude:
                name = f"Dr. {v.doctor.name}" if v.doctor else v.chemist.name
                waypoints.append({'lat': float(v.latitude), 'lng': float(v.longitude), 'title': name, 'time': v.created_at.strftime('%I:%M %p'), 'type': 'visit'})
                
    if day_end and day_end.latitude and day_end.longitude:
        waypoints.append({'lat': float(day_end.latitude), 'lng': float(day_end.longitude), 'title': 'Day End', 'time': day_end.closed_at.strftime('%I:%M %p'), 'type': 'end'})

    return render(request, 'route_playback.html', {'employee': emp, 'target_date': target_date, 'waypoints_json': json.dumps(waypoints), 'has_data': len(waypoints) > 0})

# ==============================================================================
# 8. 👁️ VIEW SINGLE DCR REPORT
# ==============================================================================
@employee_required
def view_dcr_report(request, employee, dcr_id):
    daily_dcr = get_object_or_404(DailyDCR.objects.prefetch_related('visits__product_details__product', 'visits__doctor', 'visits__chemist'), id=dcr_id, employee=employee)
    return render(request, 'view_dcr_report.html', {'daily_dcr': daily_dcr, 'visits': daily_dcr.visits.all()})