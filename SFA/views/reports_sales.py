import io
import csv
import calendar
from collections import defaultdict
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.db.models import Sum, Q

from SFA.models import (
    Employee, Stockist, Chemist, Doctor, Product, PrimarySale,
    StockistProductStatement, PartyWiseSaleReport, PartyWiseSaleLine, 
    DoctorRxMapping, DCRProductDetail, SystemSetting
)
from .auth import get_full_team_employees, get_dropdown_team
from SFA.decorators import employee_required

# ==============================================================================
# 1. 📊 SMART SECONDARY STOCKIST STATEMENT
# ==============================================================================
@employee_required
def smart_secondary_report_view(request, employee):
    team_employees = get_dropdown_team(employee, ordered=False)
    default_emp_id = str(employee.id)
    if employee.designation != 'MR':
        first_sub = team_employees.exclude(id=employee.id).first()
        if first_sub:
            default_emp_id = str(first_sub.id)
            
    selected_emp_id = request.GET.get('employee_id', default_emp_id)
    selected_emp = get_object_or_404(Employee, id=int(selected_emp_id))
    
    today = timezone.now().date()
    selected_month = int(request.GET.get('month') or today.month)
    selected_year = int(request.GET.get('year') or today.year)

    selected_team_scope = get_full_team_employees(selected_emp)
    area_territory_ids = selected_team_scope.exclude(headquarter__isnull=True).values_list('headquarter_id', flat=True)
    
    stockists = Stockist.objects.filter(territory_id__in=area_territory_ids)
    stockist_dict = {st.id: st.name for st in stockists}
    products_dict = {p.id: {'name': p.name, 'price': float(p.price) if getattr(p, 'price', None) else 0.0} for p in Product.objects.filter(company=selected_emp.company).order_by('name')}

    curr_val = selected_year * 12 + selected_month
    statements = StockistProductStatement.objects.filter(stockist_id__in=stockist_dict.keys()).values('stockist_id', 'product_id', 'month', 'year').annotate(op=Sum('opening_qty'), rec=Sum('received_qty'))
    secondaries = PartyWiseSaleLine.objects.filter(report__stockist_id__in=stockist_dict.keys()).values('report__stockist_id', 'product_id', 'report__month', 'report__year').annotate(tb=Sum('billed_qty'), tf=Sum('free_qty'))

    data_map = defaultdict(lambda: defaultdict(lambda: {'op': 0, 'rec': 0, 'tb': 0, 'tf': 0}))

    for st in statements:
        st_val = st['year'] * 12 + st['month']
        if st_val < curr_val:
            data_map[st['stockist_id']][st['product_id']]['op'] += (st['op'] or 0) + (st['rec'] or 0)
        elif st_val == curr_val:
            data_map[st['stockist_id']][st['product_id']]['op'] += (st['op'] or 0)
            data_map[st['stockist_id']][st['product_id']]['rec'] += (st['rec'] or 0)

    for sec in secondaries:
        sec_val = sec['report__year'] * 12 + sec['report__month']
        if sec_val < curr_val:
            data_map[sec['report__stockist_id']][sec['product_id']]['op'] -= ((sec['tb'] or 0) + (sec['tf'] or 0))
        elif sec_val == curr_val:
            data_map[sec['report__stockist_id']][sec['product_id']]['tb'] += (sec['tb'] or 0)
            data_map[sec['report__stockist_id']][sec['product_id']]['tf'] += (sec['tf'] or 0)

    report_data, gt_op_qty, gt_op_val, gt_pr_qty, gt_pr_val, gt_sec_b, gt_sec_f, gt_sec_val, gt_cl_qty, gt_cl_val = [], 0, 0.0, 0, 0.0, 0, 0, 0.0, 0, 0.0

    for s_id, s_name in stockist_dict.items():
        if s_id not in data_map: continue
        st_row = {'name': s_name, 'products': []}
        
        for p_id, p_info in products_dict.items():
            if p_id not in data_map[s_id]: continue
            m_data = data_map[s_id][p_id]
            op, rec, tb, tf = m_data['op'], m_data['rec'], m_data['tb'], m_data['tf']
            cl = op + rec - (tb + tf)
            
            if op == 0 and rec == 0 and tb == 0 and tf == 0 and cl == 0: continue
                
            price = p_info['price']
            op_val, pr_val, sec_val, cl_val = op * price, rec * price, tb * price, cl * price
            
            st_row['products'].append({
                'name': p_info['name'], 'op': op, 'op_val': op_val, 'rec': rec, 'pr_val': pr_val,
                'sec_b': tb, 'sec_f': tf, 'sec_val': sec_val, 'cl': cl, 'cl_val': cl_val
            })
            gt_op_qty += op; gt_op_val += op_val; gt_pr_qty += rec; gt_pr_val += pr_val
            gt_sec_b += tb; gt_sec_f += tf; gt_sec_val += sec_val; gt_cl_qty += cl; gt_cl_val += cl_val
            
        if st_row['products']: report_data.append(st_row)
    report_data.sort(key=lambda x: x['name'])

    if request.GET.get('export') == 'excel':
        filename = f"Stockist_Statement_{selected_emp.name}_{selected_month}_{selected_year}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stockist Statement"
        
        period_str = f"{calendar.month_name[selected_month]} {selected_year}"
        ws.append(['SMART STOCKIST STATEMENT (MONTHLY)'])
        ws.append(['Employee:', selected_emp.name, 'Month:', period_str])
        ws.append([''])
        ws['A1'].font = Font(bold=True, size=14, color="107C41")
        ws['A2'].font = Font(bold=True); ws['C2'].font = Font(bold=True)
        
        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        
        headers = ['Stockist Name', 'Product Name', 'Opening Qty', 'Opening Val (₹)', 'Primary Qty (Recv)', 'Primary Val (₹)', 'Secondary Billed Qty', 'Secondary Free Qty', 'Secondary Val (₹)', 'Closing Qty', 'Closing Val (₹)']
        ws.append(headers)
        
        for col_num, cell in enumerate(ws[5], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 18 if col_num > 2 else 22
            
        for st_data in report_data:
            for prod in st_data['products']:
                ws.append([st_data['name'], prod['name'], prod['op'], round(prod['op_val'], 2), prod['rec'], round(prod['pr_val'], 2), prod['sec_b'], prod['sec_f'], round(prod['sec_val'], 2), prod['cl'], round(prod['cl_val'], 2)])
                
        ws.append([''])
        ws.append(['GRAND TOTAL', '', gt_op_qty, round(gt_op_val, 2), gt_pr_qty, round(gt_pr_val, 2), gt_sec_b, gt_sec_f, round(gt_sec_val, 2), gt_cl_qty, round(gt_cl_val, 2)])
        for cell in ws[ws.max_row]: cell.font = Font(bold=True)
            
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return render(request, 'smart_stockist_statement.html', {
        'team_employees': team_employees, 'selected_emp_id': int(selected_emp_id),
        'selected_month': selected_month, 'selected_year': selected_year,
        'months_choices': [(i, calendar.month_name[i]) for i in range(1, 13)],
        'report_data': report_data,
        'gt_op_qty': gt_op_qty, 'gt_op_val': gt_op_val, 'gt_pr_qty': gt_pr_qty, 'gt_pr_val': gt_pr_val,
        'gt_sec_b': gt_sec_b, 'gt_sec_f': gt_sec_f, 'gt_sec_val': gt_sec_val, 'gt_cl_qty': gt_cl_qty, 'gt_cl_val': gt_cl_val,
        'is_manager_view': employee.designation != 'MR'
    })
    
# ==============================================================================
# 2. 📦 PRIMARY SALES REPORT
# ==============================================================================
@employee_required
def primary_sales_report_view(request, employee):
    team_employees = get_dropdown_team(employee, ordered=False)
    default_emp_id = str(employee.id)
    if employee.designation != 'MR':
        first_sub = team_employees.exclude(id=employee.id).first()
        if first_sub:
            default_emp_id = str(first_sub.id)
            
    selected_emp_id = request.GET.get('employee_id', default_emp_id)
    selected_emp = get_object_or_404(Employee, id=int(selected_emp_id))
    
    today = timezone.now().date()
    from_month = int(request.GET.get('from_month') or today.month)
    to_month = int(request.GET.get('to_month') or today.month)
    selected_year = int(request.GET.get('year') or today.year)

    if from_month > to_month:
        from_month, to_month = to_month, from_month

    months_range = list(range(from_month, to_month + 1))
    months_headers = [(m, calendar.month_name[m][:3]) for m in months_range]

    selected_team_scope = get_full_team_employees(selected_emp)
    area_territory_ids = selected_team_scope.exclude(headquarter__isnull=True).values_list('headquarter_id', flat=True)
    
    primary_sales = PrimarySale.objects.filter(
        stockist__territory_id__in=area_territory_ids, date__month__gte=from_month, date__month__lte=to_month, date__year=selected_year
    ).select_related('stockist', 'product').order_by('-date')

    party_wise_data = defaultdict(lambda: {'party_name': '', 'products': defaultdict(lambda: {'monthly': {m: {'qty': 0, 'val': 0.0} for m in months_range}, 'total_qty': 0, 'total_val': 0.0}), 'total_qty': 0, 'total_val': 0.0})
    product_wise_data = defaultdict(lambda: {'product_name': '', 'monthly': {m: {'qty': 0, 'val': 0.0} for m in months_range}, 'total_qty': 0, 'total_val': 0.0})
    gt_monthly = {m: {'qty': 0, 'val': 0.0} for m in months_range}
    gt_qty, gt_val = 0, 0.0

    for sale in primary_sales:
        s_month, st_id, pr_id, pr_name, qty = sale.date.month, sale.stockist.id, sale.product.id, sale.product.name, sale.quantity
        price = float(sale.product.price) if getattr(sale.product, 'price', None) else 0.0
        val = qty * price

        party_wise_data[st_id]['party_name'] = sale.stockist.name
        party_wise_data[st_id]['products'][pr_name]['monthly'][s_month]['qty'] += qty
        party_wise_data[st_id]['products'][pr_name]['monthly'][s_month]['val'] += val
        party_wise_data[st_id]['products'][pr_name]['total_qty'] += qty
        party_wise_data[st_id]['products'][pr_name]['total_val'] += val
        party_wise_data[st_id]['total_qty'] += qty
        party_wise_data[st_id]['total_val'] += val

        product_wise_data[pr_id]['product_name'] = pr_name
        product_wise_data[pr_id]['monthly'][s_month]['qty'] += qty
        product_wise_data[pr_id]['monthly'][s_month]['val'] += val
        product_wise_data[pr_id]['total_qty'] += qty
        product_wise_data[pr_id]['total_val'] += val

        if s_month in gt_monthly:
            gt_monthly[s_month]['qty'] += qty
            gt_monthly[s_month]['val'] += val
        gt_qty += qty; gt_val += val

    party_list = []
    for pid, pdata in party_wise_data.items():
        prod_list = []
        for name, pd in pdata['products'].items():
            pd['name'] = name; pd['monthly_list'] = [pd['monthly'][m] for m in months_range]
            prod_list.append(pd)
        prod_list.sort(key=lambda x: x['name'])
        party_list.append({'party_name': pdata['party_name'], 'products': prod_list, 'total_qty': pdata['total_qty'], 'total_val': pdata['total_val']})
    party_list.sort(key=lambda x: x['party_name'])

    prod_list_final = []
    for pr_id, pdata in product_wise_data.items():
        pdata['monthly_list'] = [pdata['monthly'][m] for m in months_range]
        prod_list_final.append(pdata)
    prod_list_final.sort(key=lambda x: x['product_name'])
    gt_monthly_list = [gt_monthly[m] for m in months_range]

    if request.GET.get('export') == 'excel':
        filename = f"Primary_Sales_{selected_emp.name}_M{from_month}-M{to_month}_{selected_year}.xlsx"
        wb = openpyxl.Workbook()
        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        period_str = f"{calendar.month_name[from_month][:3]} to {calendar.month_name[to_month][:3]} {selected_year}"

        ws1 = wb.active
        ws1.title = "Stockist Wise"
        ws1.append(['PRIMARY SALES - STOCKIST WISE (MONTH TREND)'])
        ws1.append(['Employee:', selected_emp.name, 'Period:', period_str])
        ws1.append([''])
        ws1['A1'].font = Font(bold=True, size=12, color="107C41")
        
        headers1 = ['Stockist Name', 'Product Name']
        for m_num, m_name in months_headers:
            headers1.extend([f'{m_name} Qty', f'{m_name} Value (₹)'])
        headers1.extend(['Grand Total Qty', 'Grand Total Value (₹)'])
        ws1.append(headers1)
        for col_num, cell in enumerate(ws1[4], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws1.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 15 if col_num > 2 else 25

        for party in party_list:
            for prod in party['products']:
                row_data = [party['party_name'], prod['name']]
                for m_data in prod['monthly_list']: row_data.extend([m_data['qty'], round(m_data['val'], 2)])
                row_data.extend([prod['total_qty'], round(prod['total_val'], 2)])
                ws1.append(row_data)

        gt_row1 = ['GRAND TOTAL', '']
        for m_gt in gt_monthly_list: gt_row1.extend([m_gt['qty'], round(m_gt['val'], 2)])
        gt_row1.extend([gt_qty, round(gt_val, 2)])
        ws1.append(gt_row1)
        for cell in ws1[ws1.max_row]: cell.font = Font(bold=True)

        ws2 = wb.create_sheet(title="Product Wise")
        ws2.append(['PRIMARY SALES - PRODUCT WISE (MONTH TREND)'])
        ws2.append(['Employee:', selected_emp.name, 'Period:', period_str])
        ws2.append([''])
        ws2['A1'].font = Font(bold=True, size=12, color="107C41")
        
        headers2 = ['Product Name']
        for m_num, m_name in months_headers: headers2.extend([f'{m_name} Qty', f'{m_name} Value (₹)'])
        headers2.extend(['Grand Total Qty', 'Grand Total Value (₹)'])
        ws2.append(headers2)

        for col_num, cell in enumerate(ws2[4], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws2.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 15 if col_num > 1 else 25

        for prod in prod_list_final:
            row_data = [prod['product_name']]
            for m_data in prod['monthly_list']: row_data.extend([m_data['qty'], round(m_data['val'], 2)])
            row_data.extend([prod['total_qty'], round(prod['total_val'], 2)])
            ws2.append(row_data)

        gt_row2 = ['GRAND TOTAL']
        for m_gt in gt_monthly_list: gt_row2.extend([m_gt['qty'], round(m_gt['val'], 2)])
        gt_row2.extend([gt_qty, round(gt_val, 2)])
        ws2.append(gt_row2)
        for cell in ws2[ws2.max_row]: cell.font = Font(bold=True)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return render(request, 'primary_sales_report.html', {
        'team_employees': team_employees, 'selected_emp_id': int(selected_emp_id),
        'from_month': from_month, 'to_month': to_month, 'selected_year': selected_year,
        'months_choices': [(i, calendar.month_name[i]) for i in range(1, 13)], 'months_headers': months_headers,
        'primary_sales': primary_sales, 'party_list': party_list, 'product_list': prod_list_final,
        'gt_monthly_list': gt_monthly_list, 'grand_total_qty': gt_qty, 'grand_total_value': gt_val,
        'is_manager_view': employee.designation != 'MR'
    })

# ==============================================================================
# 3. ✍️ PARTY WISE SALE ENTRY
# ==============================================================================
@employee_required
def party_wise_sale_entry_view(request, employee):
    team_employees = get_dropdown_team(employee).filter(designation='MR', is_active=True)
    is_manager_view = employee.designation != 'MR'
    
    selected_emp_id = request.GET.get('employee_id') or request.POST.get('employee_id')
    if not selected_emp_id:
        selected_emp_id = str(employee.id)
        
    selected_emp = get_object_or_404(Employee, id=selected_emp_id)
    
    today = timezone.now().date()
    current_month, current_year = (12, today.year - 1) if today.month == 1 else (today.month - 1, today.year)
    
    setting = SystemSetting.objects.filter(company=employee.company).first()
    deadline = setting.sale_upload_deadline_day if setting and setting.sale_upload_deadline_day else 4
    is_locked = today.day > deadline
    
    my_terr_ids = [selected_emp.headquarter_id] if selected_emp.headquarter_id else []
    available_stockists = Stockist.objects.filter(territory_id__in=my_terr_ids).order_by('name')
    
    selected_stockist_id = request.GET.get('stockist_id') or request.POST.get('stockist_id')
    
    if selected_stockist_id and not available_stockists.filter(id=selected_stockist_id).exists():
        selected_stockist_id = None

    if not selected_stockist_id and available_stockists.exists():
        selected_stockist_id = str(available_stockists.first().id)
        
    selected_stockist = available_stockists.filter(id=selected_stockist_id).first()
    selected_stockist_name = selected_stockist.name if selected_stockist else "No Stockist"

    if request.method == 'POST':
        if is_locked and employee.designation not in ['Admin', 'System Administrator']:
            messages.error(request, f"⚠️ Edit Locked! Last month's entry is only allowed until the {deadline} of each month. Please contact Admin.")
            return redirect(f'/reports/party-wise-sale/?stockist_id={selected_stockist_id}&employee_id={selected_emp.id}')

        posted_stockist_id = request.POST.get('stockist_id') or selected_stockist_id
        if not posted_stockist_id:
            messages.error(request, "⚠️ Stockist missing! Please select a Stockist.")
            return redirect(f'/reports/party-wise-sale/?employee_id={selected_emp.id}')
            
        posted_stockist = get_object_or_404(Stockist, id=posted_stockist_id)
        chemist_id = request.POST.get('chemist_id')
        
        if not chemist_id:
            messages.error(request, "⚠️ Selecting a Chemist is required!")
            return redirect(f'/reports/party-wise-sale/?stockist_id={posted_stockist.id}&employee_id={selected_emp.id}')
            
        chemist = get_object_or_404(Chemist, id=chemist_id)
        report, created = PartyWiseSaleReport.objects.get_or_create(
            employee=selected_emp, stockist=posted_stockist, month=current_month, year=current_year
        )
        
        saved_any = False
        for prod in Product.objects.filter(company=selected_emp.company):
            b_qty = int(request.POST.get(f'billed_{prod.id}') or 0)
            f_qty = int(request.POST.get(f'free_{prod.id}') or 0)
            if b_qty > 0 or f_qty > 0:
                PartyWiseSaleLine.objects.create(report=report, chemist=chemist, product=prod, billed_qty=b_qty, free_qty=f_qty)
                saved_any = True

        if saved_any:
            messages.success(request, f"✅ Sale saved for {chemist.name} (By: {selected_emp.name}).")
        else:
            messages.warning(request, "⚠️ The form was blank or quantities were 0, so nothing was saved.")
            
        return redirect(f'/reports/party-wise-sale/?stockist_id={posted_stockist.id}&employee_id={selected_emp.id}')

    balances = []
    if selected_stockist:
        last_day_of_target_month = date(current_year, current_month, calendar.monthrange(current_year, current_month)[1])

        for prod in Product.objects.filter(company=selected_emp.company):
            primary_agg = PrimarySale.objects.filter(
                stockist=selected_stockist, product=prod, date__lte=last_day_of_target_month
            ).aggregate(tot_qty=Sum('quantity'), tot_free=Sum('free_quantity'))
            total_lifetime_primary = (primary_agg['tot_qty'] or 0) + (primary_agg['tot_free'] or 0)
            
            past_party_agg = PartyWiseSaleLine.objects.filter(
                report__stockist=selected_stockist, product=prod
            ).filter(
                Q(report__year__lt=current_year) | Q(report__year=current_year, report__month__lt=current_month)
            ).aggregate(tb=Sum('billed_qty'), tf=Sum('free_qty'))
            total_past_secondary = (past_party_agg['tb'] or 0) + (past_party_agg['tf'] or 0)
            
            stock_available_for_this_month = total_lifetime_primary - total_past_secondary
            
            current_party_agg = PartyWiseSaleLine.objects.filter(
                report__stockist=selected_stockist, product=prod, report__month=current_month, report__year=current_year
            ).aggregate(tb=Sum('billed_qty'), tf=Sum('free_qty'))
            billed_this_month = (current_party_agg['tb'] or 0) + (current_party_agg['tf'] or 0)
            
            current_balance = stock_available_for_this_month - billed_this_month
            
            if stock_available_for_this_month > 0 or billed_this_month > 0:
                balances.append({
                    'product_id': prod.id, 'product_name': prod.name, 
                    'total_sale': stock_available_for_this_month, 
                    'billed': billed_this_month, 'balance': current_balance
                })

    chemists = Chemist.objects.filter(allocated_to=selected_emp, status='Approved').order_by('name')

    return render(request, 'party_wise_sale.html', {
        'team_employees': team_employees, 'is_manager_view': is_manager_view,
        'selected_emp_id': int(selected_emp_id), 'available_stockists': available_stockists,
        'selected_stockist_id': int(selected_stockist_id) if selected_stockist_id else '',
        'selected_stockist_name': selected_stockist_name, 'balances': balances,
        'chemists': chemists, 'products': Product.objects.filter(company=selected_emp.company),
        'current_month': current_month, 'current_year': current_year,
        'is_locked': is_locked, 'deadline': deadline
    })

# ==============================================================================
# 4. 👨‍⚕️ CLASSIFY RX ENTRY
# ==============================================================================
@employee_required
def classify_rx_entry_view(request, employee):
    team_employees = get_dropdown_team(employee).filter(designation='MR', is_active=True)
    is_manager_view = employee.designation != 'MR'
    
    selected_emp_id = request.GET.get('employee_id') or request.POST.get('employee_id')
    if not selected_emp_id: selected_emp_id = str(employee.id)
        
    selected_emp = get_object_or_404(Employee, id=selected_emp_id)
    
    today = timezone.now().date()
    month, year = (12, today.year - 1) if today.month == 1 else (today.month - 1, today.year)
    
    setting = SystemSetting.objects.filter(company=employee.company).first()
    deadline = setting.sale_upload_deadline_day if setting and setting.sale_upload_deadline_day else 4
    is_locked = today.day > deadline
    
    my_terr_ids = [selected_emp.headquarter_id] if selected_emp.headquarter_id else []
    stockists = Stockist.objects.filter(territory_id__in=my_terr_ids).order_by('name')
    
    stockist_id = request.GET.get('stockist_id') or request.POST.get('stockist_id')
    if stockist_id and not stockists.filter(id=stockist_id).exists(): stockist_id = None
    if not stockist_id and stockists.exists(): stockist_id = stockists.first().id

    stockist = None
    report = None
    lines_data = []

    if not stockist_id: 
        messages.error(request, f"No Stockist is mapped in {selected_emp.name}'s territory!")
    else:
        stockist = get_object_or_404(Stockist, id=stockist_id)
        report = PartyWiseSaleReport.objects.filter(stockist=stockist, month=month, year=year, employee=selected_emp).first()
        
        if report:
            for line in PartyWiseSaleLine.objects.filter(report=report).select_related('chemist', 'product'):
                mappings = line.dr_mappings.select_related('doctor').all()
                mapped_billed = sum(m.mapped_billed_qty for m in mappings)
                mapped_free = sum(m.mapped_free_qty for m in mappings)
                lines_data.append({
                    'line': line, 'mapped_billed': mapped_billed, 'mapped_free': mapped_free,
                    'bal_billed': line.billed_qty - mapped_billed, 'bal_free': line.free_qty - mapped_free,
                    'mappings': mappings
                })

    if request.method == "POST":
        if is_locked and employee.designation not in ['Admin', 'System Administrator']:
            messages.error(request, f"⚠️ Edit Locked! Classify Rx entries are only allowed until the {deadline} of each month. Please contact Admin.")
            return redirect(f"{request.path}?stockist_id={stockist_id}&employee_id={selected_emp.id}")

        target_line = PartyWiseSaleLine.objects.get(id=request.POST.get('line_id'))
        m_billed = int(request.POST.get('mapped_billed', 0) or 0)
        m_free = int(request.POST.get('mapped_free', 0) or 0)
        
        current_mapped_billed = sum(m.mapped_billed_qty for m in target_line.dr_mappings.all())
        current_mapped_free = sum(m.mapped_free_qty for m in target_line.dr_mappings.all())
        
        if m_billed > (target_line.billed_qty - current_mapped_billed) or m_free > (target_line.free_qty - current_mapped_free): 
            messages.error(request, "Error! Allocation qty cannot be more than the Balance.")
        elif m_billed > 0 or m_free > 0: 
            DoctorRxMapping.objects.create(party_line=target_line, doctor_id=request.POST.get('doctor_id'), mapped_billed_qty=m_billed, mapped_free_qty=m_free)
            messages.success(request, f"Rx Classified for {selected_emp.name}!")
            
        return redirect(f"{request.path}?stockist_id={stockist_id}&month={month}&year={year}&employee_id={selected_emp.id}")

    return render(request, 'classify_rx.html', {
        'team_employees': team_employees, 'is_manager_view': is_manager_view, 'selected_emp_id': int(selected_emp_id),
        'stockists': stockists, 'selected_stockist_id': int(stockist_id) if stockist_id else '',
        'stockist': stockist, 'report': report, 'lines_data': lines_data, 
        'doctors': Doctor.objects.filter(allocated_to=selected_emp, status='Approved').order_by('name'), 
        'month': month, 'year': year,
        'is_locked': is_locked, 'deadline': deadline
    })

# ==============================================================================
# 5. 💊 PARTY & RX SALE REPORT 
# ==============================================================================
@employee_required
def party_rx_report_view(request, employee):
    team_employees = get_dropdown_team(employee, ordered=False)
    default_emp_id = str(employee.id)
    if employee.designation != 'MR':
        first_sub = team_employees.exclude(id=employee.id).first()
        if first_sub:
            default_emp_id = str(first_sub.id)
            
    selected_emp_id = request.GET.get('employee_id', default_emp_id)
    selected_emp = get_object_or_404(Employee, id=int(selected_emp_id))
    
    today = timezone.now().date()
    from_month = int(request.GET.get('from_month') or today.month)
    to_month = int(request.GET.get('to_month') or today.month)
    selected_year = int(request.GET.get('year') or today.year)
    selected_stockist_id = request.GET.get('stockist_id', '')

    if from_month > to_month: from_month, to_month = to_month, from_month
    months_range = list(range(from_month, to_month + 1))
    months_headers = [(m, calendar.month_name[m][:3]) for m in months_range]

    stockists = Stockist.objects.filter(territory=selected_emp.headquarter).order_by('name') if selected_emp.headquarter else Stockist.objects.none()

    report_filter = {'employee': selected_emp, 'month__gte': from_month, 'month__lte': to_month, 'year': selected_year}
    if selected_stockist_id: report_filter['stockist_id'] = selected_stockist_id

    reports = PartyWiseSaleReport.objects.filter(**report_filter)
    lines = PartyWiseSaleLine.objects.filter(report__in=reports).select_related('chemist', 'product', 'report__stockist')

    party_wise_data = defaultdict(lambda: {'chemist_name': '', 'stockist_name': '', 'monthly': {m: {'b_qty': 0, 'f_qty': 0, 'val': 0.0} for m in months_range}, 'tot_b': 0, 'tot_f': 0, 'tot_val': 0.0, 'products': defaultdict(lambda: {'monthly': {m: {'b_qty': 0, 'f_qty': 0, 'val': 0.0} for m in months_range}, 'tot_b': 0, 'tot_f': 0, 'tot_val': 0.0})})
    product_wise_data = defaultdict(lambda: {'product_name': '', 'monthly': {m: {'b_qty': 0, 'f_qty': 0, 'val': 0.0} for m in months_range}, 'tot_b': 0, 'tot_f': 0, 'tot_val': 0.0})
    product_chemist_data = defaultdict(lambda: {'product_name': '', 'chemists': defaultdict(lambda: {'chemist_name': '', 'stockist_name': '', 'monthly': {m: {'b_qty': 0, 'f_qty': 0, 'val': 0.0} for m in months_range}, 'tot_b': 0, 'tot_f': 0, 'tot_val': 0.0})})
    gt_monthly = {m: {'b_qty': 0, 'f_qty': 0, 'val': 0.0} for m in months_range}
    gt_b_qty, gt_f_qty, gt_val = 0, 0, 0.0
    raw_logs = []

    for line in lines:
        r_month, ch_id, ch_name, st_name = line.report.month, line.chemist.id, line.chemist.name, line.report.stockist.name
        pr_id, pr_name, b_qty, f_qty = line.product.id, line.product.name, line.billed_qty, line.free_qty
        price = float(line.product.price) if getattr(line.product, 'price', None) else 0.0
        val = b_qty * price
            
        raw_logs.append({'date': f"{calendar.month_name[r_month][:3]}-{selected_year}", 'chemist': ch_name, 'stockist': st_name, 'product': pr_name, 'billed': b_qty, 'free': f_qty, 'value': val})

        party_wise_data[ch_id]['chemist_name'] = ch_name
        party_wise_data[ch_id]['stockist_name'] = st_name
        party_wise_data[ch_id]['monthly'][r_month]['b_qty'] += b_qty; party_wise_data[ch_id]['monthly'][r_month]['f_qty'] += f_qty; party_wise_data[ch_id]['monthly'][r_month]['val'] += val
        party_wise_data[ch_id]['tot_b'] += b_qty; party_wise_data[ch_id]['tot_f'] += f_qty; party_wise_data[ch_id]['tot_val'] += val
        p_data = party_wise_data[ch_id]['products'][pr_name]
        p_data['monthly'][r_month]['b_qty'] += b_qty; p_data['monthly'][r_month]['f_qty'] += f_qty; p_data['monthly'][r_month]['val'] += val
        p_data['tot_b'] += b_qty; p_data['tot_f'] += f_qty; p_data['tot_val'] += val

        pr_data = product_wise_data[pr_id]
        pr_data['product_name'] = pr_name
        pr_data['monthly'][r_month]['b_qty'] += b_qty; pr_data['monthly'][r_month]['f_qty'] += f_qty; pr_data['monthly'][r_month]['val'] += val
        pr_data['tot_b'] += b_qty; pr_data['tot_f'] += f_qty; pr_data['tot_val'] += val

        pc_data = product_chemist_data[pr_id]
        pc_data['product_name'] = pr_name
        pcc_data = pc_data['chemists'][ch_id]
        pcc_data['chemist_name'] = ch_name
        pcc_data['stockist_name'] = st_name
        pcc_data['monthly'][r_month]['b_qty'] += b_qty; pcc_data['monthly'][r_month]['f_qty'] += f_qty; pcc_data['monthly'][r_month]['val'] += val
        pcc_data['tot_b'] += b_qty; pcc_data['tot_f'] += f_qty; pcc_data['tot_val'] += val

        gt_monthly[r_month]['b_qty'] += b_qty; gt_monthly[r_month]['f_qty'] += f_qty; gt_monthly[r_month]['val'] += val
        gt_b_qty += b_qty; gt_f_qty += f_qty; gt_val += val

    party_list = []
    for ch_id, ch_data in party_wise_data.items():
        prod_list = []
        for pr_name, pd in ch_data['products'].items():
            prod_list.append({'name': pr_name, 'monthly_list': [pd['monthly'][m] for m in months_range], 'tot_b': pd['tot_b'], 'tot_f': pd['tot_f'], 'tot_val': pd['tot_val']})
        prod_list.sort(key=lambda x: x['name'])
        party_list.append({
            'chemist_name': ch_data['chemist_name'], 'stockist_name': ch_data['stockist_name'], 'products': prod_list,
            'monthly_list': [ch_data['monthly'][m] for m in months_range],
            'tot_b': ch_data['tot_b'], 'tot_f': ch_data['tot_f'], 'tot_val': ch_data['tot_val']
        })
    party_list.sort(key=lambda x: x['chemist_name'])

    prod_list_final = []
    for pr_id, pr_data in product_wise_data.items():
        pr_data['monthly_list'] = [pr_data['monthly'][m] for m in months_range]
        prod_list_final.append(pr_data)
    prod_list_final.sort(key=lambda x: x['product_name'])
    gt_monthly_list = [gt_monthly[m] for m in months_range]
    product_list_grouped = []
    for pr_id, pc_data in product_chemist_data.items():
        chem_list = []
        for ch_id, cd in pc_data['chemists'].items():
            chem_list.append({
                'chemist_name': cd['chemist_name'], 'stockist_name': cd['stockist_name'],
                'monthly_list': [cd['monthly'][m] for m in months_range],
                'tot_b': cd['tot_b'], 'tot_f': cd['tot_f'], 'tot_val': cd['tot_val']
            })
        chem_list.sort(key=lambda x: x['chemist_name'])
        pr_totals = product_wise_data[pr_id]
        product_list_grouped.append({
            'product_name': pc_data['product_name'], 'chemists': chem_list,
            'monthly_list': [pr_totals['monthly'][m] for m in months_range],
            'tot_b': pr_totals['tot_b'], 'tot_f': pr_totals['tot_f'], 'tot_val': pr_totals['tot_val']
        })
    product_list_grouped.sort(key=lambda x: x['product_name'])

    if request.GET.get('export') == 'excel':
        filename = f"Party_Sales_{selected_emp.name}_M{from_month}-M{to_month}_{selected_year}.xlsx"
        wb = openpyxl.Workbook()
        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        period_str = f"{calendar.month_name[from_month][:3]} to {calendar.month_name[to_month][:3]} {selected_year}"

        ws1 = wb.active
        ws1.title = "Party Wise Sale"
        ws1.append(['PARTY WISE SALE REPORT (MONTH TREND)'])
        ws1.append(['Employee:', selected_emp.name, 'Period:', period_str])
        ws1.append([''])
        ws1['A1'].font = Font(bold=True, size=12, color="107C41")
        
        headers1 = ['Product Name']
        for m_num, m_name in months_headers: headers1.extend([f'{m_name} Billed', f'{m_name} Free', f'{m_name} Value (₹)'])
        headers1.extend(['Total Billed', 'Total Free', 'Total Value (₹)'])
        ws1.append(headers1)
        total_cols1 = len(headers1)
        for col_num, cell in enumerate(ws1[4], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws1.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 15 if col_num > 1 else 25

        group_fill = PatternFill(start_color="0B5E31", end_color="0B5E31", fill_type="solid")
        group_font = Font(color="FFFFFF", bold=True)
        subtotal_fill = PatternFill(start_color="EEF6F1", end_color="EEF6F1", fill_type="solid")

        for party in party_list:
            ws1.append([f"💊 {party['chemist_name']} — {party['stockist_name']}"])
            r = ws1.max_row
            ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=total_cols1)
            ws1.cell(row=r, column=1).fill = group_fill
            ws1.cell(row=r, column=1).font = group_font

            for prod in party['products']:
                row_data = [prod['name']]
                for m_data in prod['monthly_list']: row_data.extend([m_data['b_qty'], m_data['f_qty'], round(m_data['val'], 2)])
                row_data.extend([prod['tot_b'], prod['tot_f'], round(prod['tot_val'], 2)])
                ws1.append(row_data)

            sub_row = ['Subtotal']
            for m_data in party['monthly_list']: sub_row.extend([m_data['b_qty'], m_data['f_qty'], round(m_data['val'], 2)])
            sub_row.extend([party['tot_b'], party['tot_f'], round(party['tot_val'], 2)])
            ws1.append(sub_row)
            r = ws1.max_row
            for c in range(1, total_cols1 + 1):
                ws1.cell(row=r, column=c).fill = subtotal_fill
                ws1.cell(row=r, column=c).font = Font(bold=True)

        gt_row1 = ['GRAND TOTAL']
        for m_gt in gt_monthly_list: gt_row1.extend([m_gt['b_qty'], m_gt['f_qty'], round(m_gt['val'], 2)])
        gt_row1.extend([gt_b_qty, gt_f_qty, round(gt_val, 2)])
        ws1.append(gt_row1)
        for cell in ws1[ws1.max_row]: cell.font = Font(bold=True)

        ws2 = wb.create_sheet(title="Product Wise Sale")
        ws2.append(['PRODUCT WISE SALE REPORT (MONTH TREND)'])
        ws2.append(['Employee:', selected_emp.name, 'Period:', period_str])
        ws2.append([''])
        ws2['A1'].font = Font(bold=True, size=12, color="107C41")
        
        headers2 = ['Chemist Name', 'Stockist Name']
        for m_num, m_name in months_headers: headers2.extend([f'{m_name} Billed', f'{m_name} Free', f'{m_name} Value (₹)'])
        headers2.extend(['Total Billed', 'Total Free', 'Total Value (₹)'])
        ws2.append(headers2)
        total_cols2 = len(headers2)

        for col_num, cell in enumerate(ws2[4], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws2.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 15 if col_num > 2 else 25

        for prod in product_list_grouped:
            ws2.append([f"📦 {prod['product_name']}"])
            r = ws2.max_row
            ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=total_cols2)
            ws2.cell(row=r, column=1).fill = group_fill
            ws2.cell(row=r, column=1).font = group_font

            for chem in prod['chemists']:
                row_data = [chem['chemist_name'], chem['stockist_name']]
                for m_data in chem['monthly_list']: row_data.extend([m_data['b_qty'], m_data['f_qty'], round(m_data['val'], 2)])
                row_data.extend([chem['tot_b'], chem['tot_f'], round(chem['tot_val'], 2)])
                ws2.append(row_data)

            sub_row = ['Subtotal', '']
            for m_data in prod['monthly_list']: sub_row.extend([m_data['b_qty'], m_data['f_qty'], round(m_data['val'], 2)])
            sub_row.extend([prod['tot_b'], prod['tot_f'], round(prod['tot_val'], 2)])
            ws2.append(sub_row)
            r = ws2.max_row
            for c in range(1, total_cols2 + 1):
                ws2.cell(row=r, column=c).fill = subtotal_fill
                ws2.cell(row=r, column=c).font = Font(bold=True)

        gt_row2 = ['GRAND TOTAL', '']
        for m_gt in gt_monthly_list: gt_row2.extend([m_gt['b_qty'], m_gt['f_qty'], round(m_gt['val'], 2)])
        gt_row2.extend([gt_b_qty, gt_f_qty, round(gt_val, 2)])
        ws2.append(gt_row2)
        for cell in ws2[ws2.max_row]: cell.font = Font(bold=True)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return render(request, 'party_rx_report.html', {
        'team_employees': team_employees, 'selected_emp_id': int(selected_emp_id), 
        'from_month': from_month, 'to_month': to_month, 'selected_year': selected_year, 
        'selected_stockist_id': int(selected_stockist_id) if selected_stockist_id else '', 
        'stockists': stockists, 'months_choices': [(i, calendar.month_name[i]) for i in range(1, 13)], 
        'months_headers': months_headers, 'party_list': party_list, 'product_list': prod_list_final,
        'product_list_grouped': product_list_grouped,
        'raw_logs': raw_logs, 'gt_monthly_list': gt_monthly_list, 'gt_b_qty': gt_b_qty, 'gt_f_qty': gt_f_qty, 'gt_val': gt_val, 
        'is_manager_view': employee.designation != 'MR'
    })

# ==============================================================================
# 6. 🩺 DOCTOR WISE RX REPORT
# ==============================================================================
@employee_required
def dr_wise_sale_report_view(request, employee):
    team_employees = get_dropdown_team(employee, ordered=False)
    default_emp_id = str(employee.id)
    if employee.designation != 'MR':
        first_sub = team_employees.exclude(id=employee.id).first()
        if first_sub:
            default_emp_id = str(first_sub.id)
            
    selected_emp_id = request.GET.get('employee_id', default_emp_id)
    selected_emp = get_object_or_404(Employee, id=int(selected_emp_id))
    
    today = timezone.now().date()
    from_month = int(request.GET.get('from_month') or today.month)
    to_month = int(request.GET.get('to_month') or today.month)
    selected_year = int(request.GET.get('year') or today.year)

    if from_month > to_month: from_month, to_month = to_month, from_month
    months_range = list(range(from_month, to_month + 1))
    months_headers = [(m, calendar.month_name[m][:3]) for m in months_range]

    mappings = DoctorRxMapping.objects.filter(
        party_line__report__month__gte=from_month, party_line__report__month__lte=to_month,
        party_line__report__year=selected_year, party_line__report__employee=selected_emp
    ).select_related('doctor', 'party_line__chemist', 'party_line__product')

    agg_data = defaultdict(lambda: {'doctor_name': '', 'specialty': '', 'chemist_name': '', 'product_name': '', 'price': 0.0, 'monthly_data': {m: {'billed': 0, 'free': 0, 'val': 0.0} for m in months_range}, 'total_billed': 0, 'total_free': 0, 'total_value': 0.0})
    product_wise_data = defaultdict(lambda: {'product_name': '', 'doctors': defaultdict(lambda: {'doc_name': '', 'specialty': '', 'monthly': {m: {'billed': 0, 'free': 0, 'val': 0.0} for m in months_range}, 'tot_b': 0, 'tot_f': 0, 'tot_val': 0.0}), 'tot_b': 0, 'tot_f': 0, 'tot_val': 0.0})
    gt_monthly = {m: {'billed': 0, 'free': 0, 'val': 0.0} for m in months_range}
    gt_billed, gt_free, gt_value = 0, 0, 0.0

    for m in mappings:
        sale_month = m.party_line.report.month
        doc_id, doc_name, specialty = m.doctor.id, m.doctor.name, m.doctor.specialty or '-'
        chem_id, chem_name = m.party_line.chemist.id, m.party_line.chemist.name
        prod_id, prod_name = m.party_line.product.id, m.party_line.product.name
        b_qty, f_qty = m.mapped_billed_qty, m.mapped_free_qty
        price = float(m.party_line.product.price) if getattr(m.party_line.product, 'price', None) else 0.0
        val = b_qty * price
        
        key = f"{doc_id}_{chem_id}_{prod_id}"
        if agg_data[key]['doctor_name'] == '':
            agg_data[key].update({'doctor_name': doc_name, 'specialty': specialty, 'chemist_name': chem_name, 'product_name': prod_name, 'price': price})
            
        agg_data[key]['monthly_data'][sale_month]['billed'] += b_qty; agg_data[key]['monthly_data'][sale_month]['free'] += f_qty; agg_data[key]['monthly_data'][sale_month]['val'] += val
        agg_data[key]['total_billed'] += b_qty; agg_data[key]['total_free'] += f_qty; agg_data[key]['total_value'] += val
        
        pw = product_wise_data[prod_id]
        pw['product_name'] = prod_name
        dw = pw['doctors'][doc_id]
        dw['doc_name'] = doc_name; dw['specialty'] = specialty
        
        dw['monthly'][sale_month]['billed'] += b_qty; dw['monthly'][sale_month]['free'] += f_qty; dw['monthly'][sale_month]['val'] += val
        dw['tot_b'] += b_qty; dw['tot_f'] += f_qty; dw['tot_val'] += val
        pw['tot_b'] += b_qty; pw['tot_f'] += f_qty; pw['tot_val'] += val
        
        gt_monthly[sale_month]['billed'] += b_qty; gt_monthly[sale_month]['free'] += f_qty; gt_monthly[sale_month]['val'] += val
        gt_billed += b_qty; gt_free += f_qty; gt_value += val

    report_data = list(agg_data.values())
    report_data.sort(key=lambda x: (x['doctor_name'], x['chemist_name'], x['product_name']))
    for row in report_data: row['monthly_data_list'] = [row['monthly_data'][m[0]] for m in months_headers]
        
    doc_grouped_data = []
    current_doc = None
    curr_group = None
    
    for row in report_data:
        if row['doctor_name'] != current_doc:
            if curr_group: doc_grouped_data.append(curr_group)
            current_doc = row['doctor_name']
            curr_group = {
                'doctor_name': current_doc,
                'specialty': row['specialty'],
                'doc_total_val': 0.0,
                'items': []
            }
        curr_group['items'].append(row)
        curr_group['doc_total_val'] += row['total_value']
        
    if curr_group: doc_grouped_data.append(curr_group)

    prod_tab_list = []
    for p_id, p_data in product_wise_data.items():
        doc_list = []
        for d_id, d_data in p_data['doctors'].items():
            d_data['monthly_list'] = [d_data['monthly'][m] for m in months_range]
            doc_list.append(d_data)
        doc_list.sort(key=lambda x: x['doc_name'])
        p_data['doctors_list'] = doc_list
        prod_tab_list.append(p_data)
    prod_tab_list.sort(key=lambda x: x['product_name'])
        
    gt_monthly_list = [gt_monthly[m[0]] for m in months_headers]

    if request.GET.get('export') == 'excel':
        filename = f"Dr_Rx_Report_{selected_emp.name}_M{from_month}-M{to_month}_{selected_year}.xlsx"
        wb = openpyxl.Workbook()
        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        period_str = f"{calendar.month_name[from_month][:3]} to {calendar.month_name[to_month][:3]} {selected_year}"

        ws1 = wb.active
        ws1.title = "Doctor Wise Rx"
        ws1.append(['DOCTOR WISE RX SALE (MONTH TREND)'])
        ws1.append(['Employee:', selected_emp.name, 'Period:', period_str])
        ws1.append([''])
        ws1['A1'].font = Font(bold=True, size=14, color="107C41")
        
        headers1 = ['Doctor Name', 'Specialty', 'Chemist Name', 'Product Name', 'Price (₹)']
        for m_num, m_name in months_headers: headers1.extend([f'{m_name} Billed', f'{m_name} Free', f'{m_name} Value (₹)'])
        headers1.extend(['Total Billed', 'Total Free', 'Total Value (₹)'])
        ws1.append(headers1)
        
        for col_num, cell in enumerate(ws1[4], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws1.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 15 if col_num > 4 else 20
            
        for row in report_data:
            row_data = [f"Dr. {row['doctor_name']}", row['specialty'], row['chemist_name'], row['product_name'], round(row['price'], 2)]
            for m_data in row['monthly_data_list']: row_data.extend([m_data['billed'], m_data['free'], round(m_data['val'], 2)])
            row_data.extend([row['total_billed'], row['total_free'], round(row['total_value'], 2)])
            ws1.append(row_data)

        ws2 = wb.create_sheet(title="Product Wise Rx")
        ws2.append(['PRODUCT WISE -> DOCTOR RX SALE (MONTH TREND)'])
        ws2.append(['Employee:', selected_emp.name, 'Period:', period_str])
        ws2.append([''])
        ws2['A1'].font = Font(bold=True, size=14, color="107C41")
        
        headers2 = ['Product Name', 'Doctor Name', 'Specialty']
        for m_num, m_name in months_headers: headers2.extend([f'{m_name} Billed', f'{m_name} Free', f'{m_name} Value (₹)'])
        headers2.extend(['Total Billed', 'Total Free', 'Total Value (₹)'])
        ws2.append(headers2)

        for col_num, cell in enumerate(ws2[4], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws2.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 15 if col_num > 3 else 25

        for prod in prod_tab_list:
            for doc in prod['doctors_list']:
                row_data = [prod['product_name'], f"Dr. {doc['doc_name']}", doc['specialty']]
                for m_data in doc['monthly_list']: row_data.extend([m_data['billed'], m_data['free'], round(m_data['val'], 2)])
                row_data.extend([doc['tot_b'], doc['tot_f'], round(doc['tot_val'], 2)])
                ws2.append(row_data)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return render(request, 'dr_wise_sale_report.html', {
        'team_employees': team_employees, 'selected_emp_id': int(selected_emp_id),
        'from_month': from_month, 'to_month': to_month, 'selected_year': selected_year,
        'months_choices': [(i, calendar.month_name[i]) for i in range(1, 13)], 'months_headers': months_headers,
        'report_data': report_data, 'doc_grouped_data': doc_grouped_data, 'prod_tab_list': prod_tab_list,
        'gt_monthly_list': gt_monthly_list, 'gt_billed': gt_billed, 'gt_free': gt_free, 'gt_value': gt_value,
        'is_manager_view': employee.designation != 'MR'
    })

# ==============================================================================
# 7. 💊 PRODUCT POB & SAMPLE REPORT
# ==============================================================================
@employee_required
def product_sales_report_view(request, employee):
    team_employees = get_dropdown_team(employee, ordered=False)
    default_emp_id = str(employee.id)
    if employee.designation != 'MR':
        first_sub = team_employees.exclude(id=employee.id).first()
        if first_sub:
            default_emp_id = str(first_sub.id)
            
    selected_emp_id = request.GET.get('employee_id', default_emp_id)
    selected_emp = get_object_or_404(Employee, id=int(selected_emp_id))
    
    today = timezone.localdate()
    from_month = int(request.GET.get('from_month') or today.month)
    to_month = int(request.GET.get('to_month') or today.month)
    selected_year = int(request.GET.get('year') or today.year)

    if from_month > to_month:
        from_month, to_month = to_month, from_month

    months_range = list(range(from_month, to_month + 1))
    months_headers = [(m, calendar.month_name[m][:3]) for m in months_range]

    product_entries = DCRProductDetail.objects.filter(
        visit__daily_dcr__employee=selected_emp, visit__daily_dcr__date__month__gte=from_month,
        visit__daily_dcr__date__month__lte=to_month, visit__daily_dcr__date__year=selected_year
    ).select_related('product', 'visit__daily_dcr')

    products_dict = {p.id: {'name': p.name, 'price': float(p.price) if getattr(p, 'price', None) else 0.0} for p in Product.objects.filter(company=selected_emp.company)}
    
    agg_data = defaultdict(lambda: {'monthly': {m: {'samples': 0, 'orders': 0, 'val': 0.0} for m in months_range}, 'tot_samples': 0, 'tot_orders': 0, 'tot_val': 0.0})
    gt_monthly = {m: {'samples': 0, 'orders': 0, 'val': 0.0} for m in months_range}
    gt_samples, gt_orders, gt_val = 0, 0, 0.0

    for entry in product_entries:
        p_id, m = entry.product_id, entry.visit.daily_dcr.date.month
        sq, oq = entry.sample_qty or 0, entry.order_qty or 0
        if p_id not in products_dict: continue
        price = products_dict[p_id]['price']
        val = oq * price
        
        agg_data[p_id]['monthly'][m]['samples'] += sq; agg_data[p_id]['monthly'][m]['orders'] += oq; agg_data[p_id]['monthly'][m]['val'] += val
        agg_data[p_id]['tot_samples'] += sq; agg_data[p_id]['tot_orders'] += oq; agg_data[p_id]['tot_val'] += val
        
        gt_monthly[m]['samples'] += sq; gt_monthly[m]['orders'] += oq; gt_monthly[m]['val'] += val
        gt_samples += sq; gt_orders += oq; gt_val += val

    report_data = []
    for p_id, p_info in products_dict.items():
        if p_id in agg_data and (agg_data[p_id]['tot_samples'] > 0 or agg_data[p_id]['tot_orders'] > 0):
            p_data = agg_data[p_id]
            report_data.append({'product_name': p_info['name'], 'price': p_info['price'], 'monthly_list': [p_data['monthly'][m] for m in months_range], 'tot_samples': p_data['tot_samples'], 'tot_orders': p_data['tot_orders'], 'tot_val': p_data['tot_val']})
            
    report_data.sort(key=lambda x: x['product_name'])
    gt_monthly_list = [gt_monthly[m] for m in months_range]

    if request.GET.get('export') == 'excel':
        filename = f"POB_Report_{selected_emp.name}_M{from_month}-M{to_month}_{selected_year}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "POB Report"
        
        period_str = f"{calendar.month_name[from_month][:3]} to {calendar.month_name[to_month][:3]} {selected_year}"
        ws.append(['PRODUCT POB & SAMPLES REPORT (MONTH TREND)'])
        ws.append(['Employee:', selected_emp.name, 'Period:', period_str])
        ws.append([''])
        ws['A1'].font = Font(bold=True, size=14, color="107C41")
        ws['A2'].font = Font(bold=True); ws['C2'].font = Font(bold=True)
        
        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        
        headers = ['Product Name', 'Price (₹)']
        for m_num, m_name in months_headers: headers.extend([f'{m_name} Samples', f'{m_name} POB', f'{m_name} Value (₹)'])
        headers.extend(['Total Samples', 'Total POB', 'Total Value (₹)'])
        ws.append(headers)
        
        for col_num, cell in enumerate(ws[5], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20 if col_num == 1 else 15
            
        for row in report_data:
            row_data = [row['product_name'], round(row['price'], 2)]
            for m_data in row['monthly_list']: row_data.extend([m_data['samples'], m_data['orders'], round(m_data['val'], 2)])
            row_data.extend([row['tot_samples'], row['tot_orders'], round(row['tot_val'], 2)])
            ws.append(row_data)
            
        ws.append([''])
        gt_row = ['GRAND TOTAL', '']
        for m_gt in gt_monthly_list: gt_row.extend([m_gt['samples'], m_gt['orders'], round(m_gt['val'], 2)])
        gt_row.extend([gt_samples, gt_orders, round(gt_val, 2)])
        
        ws.append(gt_row)
        for cell in ws[ws.max_row]: cell.font = Font(bold=True)
            
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return render(request, 'pob_report.html', {
        'team_employees': team_employees, 'selected_emp_id': int(selected_emp_id),
        'from_month': from_month, 'to_month': to_month, 'selected_year': selected_year,
        'months_choices': [(i, calendar.month_name[i]) for i in range(1, 13)], 'months_headers': months_headers,
        'report_data': report_data, 'gt_monthly_list': gt_monthly_list, 'gt_samples': gt_samples, 'gt_orders': gt_orders, 'gt_val': gt_val,
        'is_manager_view': employee.designation != 'MR'
    })

# ==============================================================================
# 8. 📦 PRODUCT MASTER VIEW
# ==============================================================================
@employee_required
def product_master_view(request, employee):
    products = Product.objects.filter(company=employee.company).order_by('name')

    if request.GET.get('export') == 'excel':
        filename = "Product_Master_List.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Product List"
        ws.append(['PRODUCT MASTER REPORT'])
        ws.append(['Generated By:', employee.name])
        ws.append([''])
        ws['A1'].font = Font(bold=True, size=14, color="107C41")
        ws['A2'].font = Font(bold=True)

        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")

        headers = ['Sr No.', 'Product Name', 'Pack Size', 'MRP (₹)', 'PTR (₹)', 'PTS (₹)', 'GST Slab (%)']
        ws.append(headers)
        for col_num, cell in enumerate(ws[4], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 25 if col_num == 2 else 15

        for idx, p in enumerate(products, 1):
            mrp_val = float(p.mrp) if getattr(p, 'mrp', None) else 0.0
            ptr_val = float(p.ptr) if getattr(p, 'ptr', None) else 0.0
            pts_val = float(p.pts) if getattr(p, 'pts', None) else 0.0
            gst_val = p.gst_slab if getattr(p, 'gst_slab', None) is not None else 0
            ws.append([idx, p.name, p.pack_size, mrp_val, ptr_val, pts_val, gst_val])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return render(request, 'product_master.html', {'products': products})