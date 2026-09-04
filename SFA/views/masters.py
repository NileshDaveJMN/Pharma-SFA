import calendar
import csv
import io
from io import TextIOWrapper
from datetime import date, datetime, timedelta
from SFA.models import ChemistEditRequest, DoctorEditRequest
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from SFA.models import (
    Employee, Doctor, Chemist, Route, Territory,
    MonthlyTourProgram, DailyTourPlan, PharmaActivity,
    DoctorChemistProductMapping, Holiday,
    LeaveBalance, LeaveApplication,
    Product, Stockist, PrimarySale, StockistProductStatement,
    MRInventory, PromoItem, PromoDispatch,
    GiftCampaignPlan, SystemNotification, SystemSetting
)
from .auth import (
    get_full_team_employees, get_team_territory_ids, get_team_route_ids,
    get_team_requested_routes, get_own_territories_and_routes,
    get_dropdown_team, get_data_scope
)
from SFA.decorators import employee_required
from .core import compress_photo

@employee_required
def add_doctor_view(request, employee):
    # 1. Team ke hisaab se Territories aur Routes laao
    team_employees = get_dropdown_team(employee)
    my_terr_ids = get_team_territory_ids(team_employees)
    territories = Territory.objects.filter(id__in=my_terr_ids).order_by('name')
    routes = get_team_requested_routes(team_employees, my_terr_ids)

    if request.method == "POST":
        # 2. Check karo kis MR ko allocate karna hai (Manager ne select kiya ya khud MR hai)
        allocated_id = request.POST.get('allocated_to')
        if allocated_id and employee.designation != 'MR':
            allocated_emp = get_object_or_404(Employee, id=allocated_id, company=employee.company)
        else:
            allocated_emp = employee

        doc_photo = compress_photo(request.FILES.get('photo'))
        doc_vcard = compress_photo(request.FILES.get('vcard_photo'))

        Doctor.objects.create(
            company=employee.company,
            name=request.POST.get('name', '').strip(), 
            specialty=request.POST.get('specialty', '').strip(),
            territory_id=request.POST.get('territory'), 
            route_id=request.POST.get('route') or None, 
            allocated_to=allocated_emp,  # 🌟 MAGIC: Yahan pe sahi employee aayega
            address=request.POST.get('address', '').strip() or None, 
            mobile=request.POST.get('mobile', '').strip() or None,
            email=request.POST.get('email', '').strip() or None, 
            degree=request.POST.get('degree', '').strip() or None,
            category=request.POST.get('category', '').strip() or None, 
            dob=request.POST.get('dob') or None,
            dom=request.POST.get('dom') or None, 
            spouse_dob=request.POST.get('spouse_dob') or None,
            latitude=request.POST.get('latitude', '').strip() or None, 
            longitude=request.POST.get('longitude', '').strip() or None,
            photo=doc_photo,
            vcard_photo=doc_vcard
        )
        messages.success(request, "Doctor successfully added with photos!")
        return redirect('request_hub')
        
    return render(request, 'add_doctor.html', {
        'territories': territories, 
        'routes': routes,
        'team_employees': team_employees,
        'is_manager_view': employee.designation != 'MR'
    })

@employee_required
def add_chemist_view(request, employee):
    team_employees = get_dropdown_team(employee)
    my_terr_ids = get_team_territory_ids(team_employees)
    territories = Territory.objects.filter(id__in=my_terr_ids).order_by('name')
    routes = get_team_requested_routes(team_employees, my_terr_ids)

    if request.method == "POST":
        allocated_id = request.POST.get('allocated_to')
        if allocated_id and employee.designation != 'MR':
            allocated_emp = get_object_or_404(Employee, id=allocated_id, company=employee.company)
        else:
            allocated_emp = employee

        Chemist.objects.create(
            company=employee.company,
            name=request.POST.get('name'), 
            phone=request.POST.get('phone'), 
            address=request.POST.get('address', '').strip() or None,
            territory_id=request.POST.get('territory'), 
            route_id=request.POST.get('route') or None, 
            allocated_to=allocated_emp 
        )
        messages.success(request, "Chemist successfully added!")
        return redirect('request_hub')
        
    return render(request, 'add_chemist.html', {
        'territories': territories, 
        'routes': routes,
        'team_employees': team_employees,
        'is_manager_view': employee.designation != 'MR'
    })

@employee_required
def add_route_view(request, employee):
    team_employees = get_full_team_employees(employee)
    my_territory_ids = get_team_territory_ids(team_employees)
    territories = Territory.objects.filter(id__in=my_territory_ids).order_by('name') if my_territory_ids else Territory.objects.filter(company=employee.company).order_by('name')
    
    if request.method == "POST":
        name, territory_id = request.POST.get('name', '').strip(), request.POST.get('territory')
        # 🌟 NAYA: Category aur Distance fetch kar rahe hain
        category = request.POST.get('category', 'HQ')
        distance_from_hq = float(request.POST.get('distance_from_hq', 0) or 0)
        
        if name and territory_id:
            if not Territory.objects.filter(id=territory_id, company=employee.company).exists():
                messages.error(request, "Invalid territory selected.")
                return redirect('add_route') 
            Route.objects.create(
                company=employee.company,
                name=name, 
                territory_id=territory_id, 
                category=category,
                distance_from_hq=distance_from_hq,
                requested_by=employee, 
                status='Pending'
            )
            messages.success(request, "Route request submitted!")
            return redirect('add_route')
            
    return render(request, 'add_route.html', {'territories': territories, 'my_routes': Route.objects.filter(requested_by__in=team_employees).order_by('-id'), 'is_manager_view': employee.designation != 'MR'})

@employee_required
def add_tour_program_view(request, employee):
    today = timezone.localdate()
    current_year = today.year
    curr_month = today.month
    
    # 🌟 NAYA RULE: Agla mahina aur saal nikalo
    next_month, next_year = (1, current_year + 1) if curr_month == 12 else (curr_month + 1, current_year)

    # 🌟 NAYA RULE: Dropdown list logic (New joiner ko current aur next, purane ko sirf next)
    is_new_joiner = employee.joining_date and (today - employee.joining_date).days <= 7
    # 🌟 FIX: get_or_create crash karta hai agar duplicate rows already ban chuki hain,
    # isliye filter().first() use karo — jo bhi pehli row mile use lo, warna nayi banao.
    setting = SystemSetting.objects.filter(company=employee.company).order_by('id').first()
    if not setting:
        setting = SystemSetting.objects.create(company=employee.company)

    allowed_months = [
        {'month': next_month, 'year': next_year, 'name': calendar.month_name[next_month], 'label': f"{calendar.month_name[next_month]} {next_year}"}
    ]

    if is_new_joiner or setting.allow_current_month_mtp:
        allowed_months.insert(0, {'month': curr_month, 'year': current_year, 'name': calendar.month_name[curr_month], 'label': f"{calendar.month_name[curr_month]} {current_year}"})

    # 1. HANDLE POST REQUESTS
    if request.method == "POST":
        action = request.POST.get('action')

        if action == 'create_mtp':
            m = int(request.POST.get('month'))
            y = int(request.POST.get('year'))
            
            # 🌟 STRICT BLOCKER: Koi URL manipulation (Inspect Element) karke past mahine na bhej de
            is_allowed = any(am['month'] == m and am['year'] == y for am in allowed_months)
            if not is_allowed:
                messages.error(request, "⚠️ Please create a Tour Plan only for the allowed month.")
                return redirect('add_tour_program')
            
            # 🌟 SMART DOJ BLOCKER: Joining se pehle ka MTP block karna (Extra Security)
            if employee.joining_date:
                mtp_val = y * 12 + m
                join_val = employee.joining_date.year * 12 + employee.joining_date.month
                if mtp_val < join_val:
                    messages.error(request, f"❌ You cannot create a Tour Plan for a date before your joining date ({employee.joining_date.strftime('%d %b %Y')})!")
                    return redirect('add_tour_program')

            if not MonthlyTourProgram.objects.filter(employee=employee, month=m, year=y).exists():
                MonthlyTourProgram.objects.create(employee=employee, month=m, year=y, status='Draft')
                messages.success(request, f"Draft created for {calendar.month_name[m]} {y}. Now select routes.")
            else:
                messages.warning(request, "A plan for this month already exists!")
            return redirect('add_tour_program')

        elif action in ['bulk_save_plan', 'submit_plan']:
            mtp_id = request.POST.get('mtp_id')
            mtp = get_object_or_404(MonthlyTourProgram, id=mtp_id, employee=employee)
            DailyTourPlan.objects.filter(mtp=mtp).delete()
            
            # 🌟 DOJ PATCH (POST LOGIC): Save karte waqt DOJ se pehle ki dates skip karo
            start_day = 1
            if employee.joining_date:
                mtp_val = mtp.year * 12 + mtp.month
                join_val = employee.joining_date.year * 12 + employee.joining_date.month
                if mtp_val == join_val:
                    start_day = employee.joining_date.day
                elif mtp_val < join_val:
                    start_day = 32 # Skip saving anything

            num_days = calendar.monthrange(mtp.year, mtp.month)[1]
            plan_objs = []
            for day in range(start_day, num_days + 1):
                route_id = request.POST.get(f'route_{day}')
                if route_id:
                    curr_date = date(mtp.year, mtp.month, day)
                    if curr_date.weekday() == 6:
                        continue
                    plan_objs.append(DailyTourPlan(mtp=mtp, date=curr_date, route_id=route_id))
            DailyTourPlan.objects.bulk_create(plan_objs)   # 🚀 EK INSERT            
                    
            if action == 'submit_plan':
                mtp.status = 'Pending'
                mtp.save()
                messages.success(request, f"🚀 MTP for {calendar.month_name[mtp.month]} submitted to Manager!")
                return redirect('request_hub')
            else:
                messages.success(request, "💾 Draft saved successfully! You can submit it later.")
                return redirect('add_tour_program')

    # 2. GET REQUEST
    draft_mtps = MonthlyTourProgram.objects.filter(
        employee=employee, 
        status__in=['Draft', 'Rejected']
    ).order_by('-year', '-month')

    # Sare routes fetch karo
    team_employees = get_full_team_employees(employee)
    all_terr_ids = get_team_territory_ids(team_employees)
    all_route_ids = get_team_route_ids(team_employees, all_terr_ids, approved_only=True)

    routes = Route.objects.filter(id__in=all_route_ids).select_related('territory').order_by('category', 'name')

    # 🚀 N+1 FIXED: Leaves, Holidays, Plans — sab loop se PEHLE (API twin pattern)
    all_leaves, all_holidays, plans_by_mtp = [], [], {}
    if draft_mtps:
        # RBM chain — EK baar (per-MTP walk khatam)
        rbm_emp = None
        for m in employee.get_my_managers():
            if m.designation == 'RBM':
                rbm_emp = m
                break
        holiday_creators = list(Employee.objects.filter(company=employee.company, designation='Admin').values_list('id', flat=True))
        if rbm_emp:
            holiday_creators.append(rbm_emp.id)

        span_start = min(date(m.year, m.month, 1) for m in draft_mtps)
        span_end = max(date(m.year, m.month, calendar.monthrange(m.year, m.month)[1]) for m in draft_mtps)

        all_leaves = list(LeaveApplication.objects.filter(
            employee=employee, status='Approved', start_date__lte=span_end, end_date__gte=span_start
        ))
        all_holidays = list(Holiday.objects.filter(
            proposed_by_id__in=holiday_creators, status='Approved', date__gte=span_start, date__lte=span_end
        ))
        # Saare drafts ki daily plans — 1 query
        for p in DailyTourPlan.objects.filter(mtp__in=draft_mtps).values('mtp_id', 'date', 'route_id'):
            plans_by_mtp.setdefault(p['mtp_id'], {})[p['date'].day] = p['route_id']

    mtp_data = []
    for mtp in draft_mtps:
        # 🚀 Leaves — prefetched list se Python filter (0 queries)
        leave_dates = {}
        month_first = date(mtp.year, mtp.month, 1)
        month_last = date(mtp.year, mtp.month, calendar.monthrange(mtp.year, mtp.month)[1])
        for l in all_leaves:
            if l.start_date <= month_last and l.end_date >= month_first:
                d = l.start_date
                while d <= l.end_date:
                    if d.month == mtp.month and d.year == mtp.year:
                        leave_dates[d.day] = l.leave_type
                    d += timedelta(days=1)

        # 🚀 Holidays — prefetched se Python filter (0 queries)
        holiday_dates = {h.date.day: h.name for h in all_holidays if h.date.month == mtp.month and h.date.year == mtp.year}

        # 🚀 Plans — prefetched dict se (0 queries)
        existing_plans = plans_by_mtp.get(mtp.id, {})

        # 🌟 DOJ PATCH (GET LOGIC): Form generate karte waqt loop start point set karein
        start_day = 1
        if employee.joining_date:
            mtp_val = mtp.year * 12 + mtp.month
            join_val = employee.joining_date.year * 12 + employee.joining_date.month
            if mtp_val == join_val:
                start_day = employee.joining_date.day
            elif mtp_val < join_val:
                start_day = 32 # Skip UI generation entirely

        days_list = []
        num_days = calendar.monthrange(mtp.year, mtp.month)[1]

        # Loop ab Joining Date ke din se shuru hoga
        for day in range(start_day, num_days + 1):
            curr_date = date(mtp.year, mtp.month, day)
            is_sunday = curr_date.weekday() == 6
            is_leave = day in leave_dates
            is_holiday = day in holiday_dates

            status_text = ""
            if is_leave:
                status_text = f"🏖️ On Leave ({leave_dates[day]})"
            elif is_holiday:
                status_text = f"⛱️ {holiday_dates[day]}"
            elif is_sunday:
                status_text = "🔴 Sunday (Weekly Off)"

            days_list.append({
                'day_number': day,
                'date_str': curr_date.strftime("%d %b, %a"),
                'is_sunday': is_sunday,
                'is_leave': is_leave,
                'is_holiday': is_holiday,
                'is_locked': is_leave or is_holiday or is_sunday,
                'status_text': status_text,
                'selected_route_id': existing_plans.get(day)
            })

        # Agar UI ke liye valid days hain, tabhi append karo
        if days_list:
            mtp_data.append({
                'mtp': mtp,
                'days_list': days_list
            })

    return render(request, 'add_tour_program.html', {
        'mtp_data': mtp_data,
        'routes': routes,
        'current_year': current_year,
        'allowed_months': allowed_months
    })

# 🌟 Helper function (Agar pehle add nahi kiya tha to ise upar rakhein)
def send_auto_alert(employee, title, message):
    if employee:
        SystemNotification.objects.create(employee=employee, title=title, message=message)

@employee_required
def review_mtp_view(request, employee, mtp_id):
    manager = employee
    mtp = get_object_or_404(MonthlyTourProgram, id=mtp_id)
    
    # Permission Check
    if mtp.employee not in get_full_team_employees(manager) and manager.designation != 'NSM': 
        return redirect('manager_approvals')
        
    daily_plans = mtp.daily_plans.all().order_by('date')
    
    if request.method == "POST":
        action = request.POST.get('action')
        remark = request.POST.get('manager_remark', '').strip()
        
        # ==================================
        # ❌ REJECT LOGIC
        # ==================================
        if action == 'Reject': 
            mtp.status = 'Rejected'
            mtp.manager_remark = remark
            mtp.save()
            
            # 🔔 AUTO-ALERT (To MR and Approver's Boss)
            
            send_auto_alert(mtp.employee, "Tour Plan Rejected ❌", f"Your Tour Plan for {mtp.month}/{mtp.year} has been rejected by {manager.name}.")
            if manager.manager:
                send_auto_alert(manager.manager, "Team MTP Action 📊", f"{manager.name} has rejected the Tour Plan of {mtp.employee.name} for {mtp.month}/{mtp.year}.")
                
            return redirect('manager_approvals')
            
        # ==================================
        # ✅ APPROVE LOGIC
        # ==================================
        elif action == 'Approve':
            is_changed = False
            for dp in daily_plans:
                new_route_id = request.POST.get(f'route_{dp.id}')
                if new_route_id and int(new_route_id) != dp.route_id: 
                    dp.route_id = new_route_id
                    dp.save()
                    is_changed = True
                    
            mtp.status = 'Approved'
            mtp.manager_remark = remark
            mtp.is_modified = is_changed
            mtp.save()
            
            # 🔔 AUTO-ALERT (To MR and Approver's Boss)
            send_auto_alert(mtp.employee, "Tour Plan Approved ✅", f"Your Tour Plan for {mtp.month}/{mtp.year} has been approved by {manager.name}.")
            if manager.manager:
                send_auto_alert(manager.manager, "Team MTP Action 📊", f"{manager.name} has approved the Tour Plan of {mtp.employee.name} for {mtp.month}/{mtp.year}.")
                
            return redirect('manager_approvals')
            
    return render(request, 'review_mtp.html', {'mtp': mtp, 'daily_plans': daily_plans, 'all_routes': Route.objects.filter(company=employee.company)})

@employee_required
def approve_activity_view(request, employee, activity_id):
    activity = get_object_or_404(PharmaActivity, id=activity_id)
    if request.method == "POST":
        action, remark = request.POST.get('action'), request.POST.get('remark', '')
        if action == 'Reject': activity.status = 'Rejected'; activity.manager_remark = remark; activity.save(); messages.error(request, "Activity rejected."); return redirect('manager_approvals')
        elif action == 'Approve':
            chain_managers, creator_manager = [], activity.employee.manager
            while creator_manager is not None: chain_managers.append(creator_manager.id); creator_manager = creator_manager.manager
            if employee.id in chain_managers:
                if employee.id not in activity.approved_by_managers: activity.approved_by_managers.append(employee.id)
                if len(activity.approved_by_managers) >= len(chain_managers): activity.status = 'Pending_Admin'; messages.success(request, "Approved by all managers! Pending Admin Approval.")
                else: messages.info(request, "Your approval recorded.")
                activity.save()
    return redirect('manager_approvals')

def get_chemists_for_doctor(request, doctor_id):
    chemists = [{'id': m.chemist.id, 'name': m.chemist.name} for m in DoctorChemistProductMapping.objects.filter(doctor_id=doctor_id).select_related('chemist')]
    seen = set()
    return JsonResponse({'chemists': [seen.add(c['id']) or c for c in chemists if c['id'] not in seen]})

def get_products_for_dr_chemist(request, doctor_id, chemist_id):
    return JsonResponse({'products': [{'id': m.product.id, 'name': m.product.name} for m in DoctorChemistProductMapping.objects.filter(doctor_id=doctor_id, chemist_id=chemist_id).select_related('product')]})      

# ==============================================================================
# ⛱️ REQUEST HOLIDAY VIEW (For RBM/ZBM/NSM/Admin)
# ==============================================================================
@employee_required
def request_holiday_view(request, employee):
    if employee.designation not in ['RBM', 'ZBM', 'NSM', 'Admin', 'System Administrator']:
        messages.error(request, "Only RBM/ZBM/NSM can propose state holidays.")
        return redirect('request_hub')
        
    # 🌟 NAYA LOGIC: RBM ke upar wale managers ko unke RBMs ki list dikhegi
    subordinate_rbms = []
    is_rbm = (employee.designation == 'RBM')
    if not is_rbm:
        team = get_full_team_employees(employee)
        subordinate_rbms = team.filter(designation='RBM', is_active=True).order_by('name')

    if request.method == "POST":
        name = request.POST.get('name')
        h_date = request.POST.get('date')
        is_national = request.POST.get('is_national') == 'on'
        selected_rbms = request.POST.getlist('rbm_ids')

        if name and h_date:
            status_val = 'Approved' if employee.designation in ['Admin', 'System Administrator'] else 'Pending'

            if is_rbm:
                # Agar RBM khud propose kar raha hai
                Holiday.objects.get_or_create(date=h_date, proposed_by=employee, defaults={'name': name, 'status': status_val})
                messages.success(request, "State Holiday proposal submitted to Admin!")
            else:
                # Agar RBM ke upar ka manager (ZBM/NSM/Admin) propose kar raha hai
                if is_national:
                    # National Holiday -> Applicable to all
                    Holiday.objects.get_or_create(date=h_date, proposed_by=employee, defaults={'name': name, 'status': status_val})
                    messages.success(request, "National Holiday submitted successfully!")
                else:
                    # State Holiday -> Sirf selected RBMs ke liye
                    if not selected_rbms:
                        messages.error(request, "Please select at least one RBM or mark as 'National Holiday'.")
                        return redirect('request_holiday')
                        
                    for r_id in selected_rbms:
                        rbm_emp = get_object_or_404(Employee, id=r_id, company=employee.company)
                        # 🌟 PROXY CREATION: Proposed_by me RBM ka naam dalega
                        Holiday.objects.get_or_create(date=h_date, proposed_by=rbm_emp, defaults={'name': name, 'status': status_val})
                    messages.success(request, f"Holiday successfully applied to {len(selected_rbms)} RBM(s) states!")
                    
            return redirect('request_holiday')
            
    holidays = Holiday.objects.filter(proposed_by=employee).order_by('-date')
    return render(request, 'request_holiday.html', {
        'holidays': holidays, 
        'subordinate_rbms': subordinate_rbms,
        'is_rbm': is_rbm
    })

# ==============================================================================
# 🏖️ APPLY LEAVE VIEW (For All Employees)
# ==============================================================================
def _get_leave_balance_data(employee):
    """
    🌟 SHARED HELPER: apply_leave_view (form+history) aur leave_status_view
    (sirf history, View Hub ke liye) — dono yahi se balance nikalte hain,
    taaki calculation duplicate na ho aur dono jagah hamesha sync rahein.
    """
    current_year = timezone.localdate().year
    balance, _ = LeaveBalance.objects.get_or_create(employee=employee, year=current_year)
    rem_cl = balance.cl_total - balance.cl_used
    rem_sl = balance.sl_total - balance.sl_used
    rem_pl = balance.pl_total - balance.pl_used
    applications = LeaveApplication.objects.filter(employee=employee).order_by('-applied_on')
    return balance, rem_cl, rem_sl, rem_pl, applications

@employee_required
def apply_leave_view(request, employee):
    balance, rem_cl, rem_sl, rem_pl, applications = _get_leave_balance_data(employee)

    if request.method == 'POST':
        l_type = request.POST.get('leave_type')
        s_date = datetime.strptime(request.POST.get('start_date'), '%Y-%m-%d').date()
        e_date = datetime.strptime(request.POST.get('end_date'), '%Y-%m-%d').date()
        reason = request.POST.get('reason')
        
        days = (e_date - s_date).days + 1
        
        if days <= 0:
            messages.error(request, "Error! The end date must be after the start date.")
        else:
            # Check if sufficient balance is available
            is_valid = True
            if l_type == 'CL' and rem_cl < days: is_valid = False
            elif l_type == 'SL' and rem_sl < days: is_valid = False
            elif l_type == 'PL' and rem_pl < days: is_valid = False
            
            if is_valid or l_type == 'LWP':
                LeaveApplication.objects.create(
                    employee=employee, leave_type=l_type, start_date=s_date, end_date=e_date, reason=reason
                )
                messages.success(request, f"🚀 {l_type} for {days} day(s) applied successfully! Sent to manager.")
            else:
                # 🛡️ SECURITY: eval hatao — dict lookup
                balance_map = {'CL': rem_cl, 'SL': rem_sl, 'PL': rem_pl, 'LWP': 0}
                messages.error(request, f"❌ Insufficient {l_type} balance! You only have {balance_map.get(l_type, 0)} left.")              
        return redirect('apply_leave')

    return render(request, 'apply_leave.html', {
        'balance': balance, 'rem_cl': rem_cl, 'rem_sl': rem_sl, 'rem_pl': rem_pl,
        'applications': applications
    })

@employee_required
def leave_status_view(request, employee):
    """
    🌟 NAYA: View Hub ke 'Leave Status' card ke liye — apply_leave_view jaisa
    hi balance + history dikhata hai, par 'New Application' form ke bina.
    Read-only summary, taaki MR/Manager apna leave-status check kar sake
    bina kahi aur jana.
    """
    balance, rem_cl, rem_sl, rem_pl, applications = _get_leave_balance_data(employee)
    return render(request, 'leave_status.html', {
        'balance': balance, 'rem_cl': rem_cl, 'rem_sl': rem_sl, 'rem_pl': rem_pl,
        'applications': applications
    })

import openpyxl
import csv
import io
import calendar
from datetime import datetime, date
from collections import defaultdict
from django.shortcuts import render, redirect
from django.contrib import messages
from django.utils import timezone
from SFA.models import Employee, Stockist, Product, PrimarySale, StockistProductStatement
from .auth import get_full_team_employees
from SFA.decorators import employee_required

@employee_required
def upload_primary_sales_view(request, employee):
    if employee.designation not in ['Admin', 'System Administrator']:
        messages.error(request, "🚫 Access Denied.")
        return redirect('view_hub')

    uploaded_file = request.FILES.get('upload_file') or request.FILES.get('excel_file') or request.FILES.get('csv_file')

    if request.method == 'POST' and uploaded_file:
        selected_month = int(request.POST.get('month', timezone.localdate().month))
        selected_year = int(request.POST.get('year', timezone.localdate().year))
        selected_rbm_id = request.POST.get('rbm_id')
        
        if not selected_rbm_id:
            messages.error(request, "❌ Please select an RBM (State) first!")
            return redirect('upload_primary_sales')

        try:
            rbm_emp = Employee.objects.get(id=selected_rbm_id, company=employee.company)
            state_team = get_full_team_employees(rbm_emp)
            state_terr_ids = state_team.exclude(headquarter__isnull=True).values_list('headquarter_id', flat=True)
            state_stockists = Stockist.objects.filter(territory_id__in=state_terr_ids)
            
            file_name = uploaded_file.name.lower()
            rows = []
            headers = []

            # ==========================================
            # 1. READ FILE (MEMORY SAFE & OOM FIXED)
            # ==========================================
            if file_name.endswith('.csv'):
                file_data = uploaded_file.read().decode('utf-8-sig')
                io_string = io.StringIO(file_data)
                reader = csv.reader(io_string)
                all_rows = list(reader)
                if len(all_rows) < 2:
                    messages.error(request, "❌ The CSV file is empty or missing headers.")
                    return redirect('upload_primary_sales')
                headers = [str(h).strip().lower().replace(' ', '_') for h in all_rows[0] if h]
                rows = all_rows[1:]
            
            elif file_name.endswith(('.xlsx', '.xls')):
                # read_only=True ensures no RAM leak for large files
                wb = openpyxl.load_workbook(uploaded_file, data_only=True, read_only=True)
                sheet = wb.active
                is_first_row = True
                for row in sheet.iter_rows(values_only=True):
                    if not any(row): continue  # Skip blank ghost rows
                    if is_first_row:
                        if not headers: headers = [str(h).strip().lower().replace(' ', '_') for h in row if h]
                        is_first_row = False
                    else:
                        rows.append(row)
                wb.close()
                if not rows:
                    messages.error(request, "❌ The Excel file is empty.")
                    return redirect('upload_primary_sales')
            else:
                messages.error(request, "❌ Please upload only a .csv or .xlsx file.")
                return redirect('upload_primary_sales')

            # ==========================================
            # 2. DELETE OLD DATA (FOR THE SELECTED MONTH)
            # ==========================================
            PrimarySale.objects.filter(date__month=selected_month, date__year=selected_year, stockist__in=state_stockists).delete()
            StockistProductStatement.objects.filter(month=selected_month, year=selected_year, stockist__in=state_stockists).delete()
            
            count = 0
            errors = []
            
            # ==========================================
            # 3. FAST CACHING (To stop server freeze)
            # ==========================================
            stockist_cache = {s.name.strip().lower(): s for s in state_stockists}
            product_cache = {p.name.strip().lower(): p for p in Product.objects.filter(company=employee.company)}
            emp_cache = {e.headquarter_id: e for e in state_team if e.headquarter_id}

            sales_to_create = []
            statement_agg = defaultdict(int) # This will sum up quantities automatically in python memory

            # ==========================================
            # 4. DATA PROCESSING LOOP
            # ==========================================
            for row_num, row_data in enumerate(rows, start=2):
                row_data = list(row_data) + [''] * (len(headers) - len(row_data))
                clean_row = dict(zip(headers, row_data))
                
                # Cleanup spaces and nan
                for k, v in clean_row.items():
                    val = str(v).strip() if v is not None else ''
                    clean_row[k] = '' if val.lower() == 'nan' else val
                
                raw_date = clean_row.get('date')
                st_name = clean_row.get('stockist_name', '').strip().lower()
                pr_name = clean_row.get('product_name', '').strip().lower()
                qty_str = clean_row.get('qty', '0').strip()
                batch = clean_row.get('batch_no', 'NA').strip()
                free_qty_str = clean_row.get('free_qty', '0').strip()
                
                if raw_date and st_name and pr_name and qty_str != '0':
                    # DATE PARSER
                    if isinstance(raw_date, datetime): sale_date = raw_date.date()
                    elif isinstance(raw_date, date): sale_date = raw_date
                    else:
                        raw_date_str = str(raw_date).strip().split()[0]
                        try:
                            if '/' in raw_date_str: sale_date = datetime.strptime(raw_date_str, '%d/%m/%Y').date()
                            elif '-' in raw_date_str:
                                if len(raw_date_str.split('-')[0]) == 4: sale_date = datetime.strptime(raw_date_str, '%Y-%m-%d').date()
                                else: sale_date = datetime.strptime(raw_date_str, '%d-%m-%Y').date()
                            else: raise ValueError
                        except ValueError:
                            errors.append(f"Row {row_num}: Invalid date format '{raw_date}'.")
                            continue
                    
                    if sale_date.month != selected_month or sale_date.year != selected_year:
                        errors.append(f"Row {row_num}: Date {sale_date.strftime('%d-%b-%Y')} does not belong to the selected month ({selected_month}/{selected_year}).")
                        continue

                    # Instant O(1) dictionary lookups instead of Database hits
                    stockist = stockist_cache.get(st_name)
                    prod = product_cache.get(pr_name)
                    
                    if not stockist:
                        errors.append(f"Row {row_num}: Stockist '{clean_row.get('stockist_name')}' not found in {rbm_emp.name}'s territory.")
                        continue
                    if not prod:
                        errors.append(f"Row {row_num}: Product '{clean_row.get('product_name')}' not found.")
                        continue
                        
                    emp = emp_cache.get(stockist.territory_id)
                    if emp:
                        q = int(qty_str) if qty_str.replace('-','').isdigit() else 0
                        fq = int(free_qty_str) if free_qty_str.replace('-','').isdigit() else 0
                        
                        sales_to_create.append(PrimarySale(
                            date=sale_date, stockist=stockist, product=prod,
                            quantity=q, free_quantity=fq, batch_number=batch
                        ))
                        
                        # In-memory aggregation for StockistProductStatement
                        agg_key = (emp.id, stockist.id, prod.id, sale_date.month, sale_date.year)
                        statement_agg[agg_key] += (q + fq)
                        count += 1
                    else:
                        errors.append(f"Row {row_num}: No MR is mapped to Stockist '{clean_row.get('stockist_name')}'.")
                else:
                    errors.append(f"Row {row_num}: Data Missing (Date, Stockist, Product or Qty=0).")

            # ==========================================
            # 5. BULK CREATE DATABASE INSERTS (SUPER FAST)
            # ==========================================
            if sales_to_create:
                PrimarySale.objects.bulk_create(sales_to_create, batch_size=1000)
            
            statements_to_create = []
            for (emp_id, stockist_id, prod_id, m, y), total_qty in statement_agg.items():
                statements_to_create.append(StockistProductStatement(
                    employee_id=emp_id, stockist_id=stockist_id, product_id=prod_id,
                    month=m, year=y, received_qty=total_qty, opening_qty=0
                ))
            
            if statements_to_create:
                StockistProductStatement.objects.bulk_create(statements_to_create, batch_size=1000)
                        
            messages.success(request, f"🚀 {count} Primary Sales records uploaded instantly for {rbm_emp.name}!")
            if errors:
                for err in errors[:5]: messages.warning(request, err)
                if len(errors) > 5: messages.warning(request, f"...and {len(errors) - 5} more similar errors.")
                    
        except Exception as e:
            if "list index out of range" in str(e):
                messages.error(request, "❌ File Formatting Error! Corrupt Excel file. Please use CSV.")
            else:
                messages.error(request, f"❌ File processing error: {str(e)}")
            
        return redirect('upload_primary_sales')
        
    today = timezone.localdate()
    rbms = Employee.objects.filter(company=employee.company, designation__in=['RBM', 'ZBM', 'NSM'], is_active=True).order_by('name')
    
    return render(request, 'upload_primary_sales.html', {
        'months': [(i, calendar.month_name[i]) for i in range(1, 13)], 
        'years': [today.year - 1, today.year, today.year + 1], 
        'today': today,
        'rbms': rbms
    })

# ==============================================================================
# ✏️ EDIT HUB & EDIT DOCTOR LOGIC
# ==============================================================================
@employee_required
def edit_hub_view(request, employee):
    return render(request, 'edit_hub.html')
@employee_required
def edit_doctor_list_view(request, employee):
    team_employees = get_dropdown_team(employee)
    
    default_emp_id = str(employee.id)
    if employee.designation != 'MR':
        first_subordinate = team_employees.exclude(id=employee.id).first()
        if first_subordinate:
            default_emp_id = str(first_subordinate.id)
            
    selected_emp_id = request.GET.get('employee_id', default_emp_id)
    selected_emp = get_object_or_404(Employee, id=selected_emp_id, company=employee.company)

    # 🌟 FIX: Sirf Approved doctors dikhao jinko edit kiya ja sake
    doctors = Doctor.objects.filter(allocated_to=selected_emp, status='Approved').order_by('name')
    
    return render(request, 'edit_doctor_list.html', {
        'doctors': doctors,
        'team_employees': team_employees,
        'selected_emp_id': int(selected_emp_id),
        'is_manager_view': employee.designation != 'MR'
    })

@employee_required
def edit_chemist_view(request, employee, chem_id):
    team_employees = get_dropdown_team(employee, ordered=False)
    chem = get_object_or_404(Chemist, id=chem_id, allocated_to__in=team_employees)
    
    if ChemistEditRequest.objects.filter(chemist=chem, status='Pending').exists():
        messages.warning(request, "⚠️ An Edit Request for this Chemist is already pending with the Manager!")
        return redirect('edit_chemist_list')
        
    sub_team = Employee.objects.filter(id=chem.allocated_to.id)
    my_terr_ids = get_team_territory_ids(sub_team)
    territories = Territory.objects.filter(id__in=my_terr_ids).order_by('name')
    routes = get_team_requested_routes(sub_team, my_terr_ids)

    if request.method == "POST":
        ChemistEditRequest.objects.create(
            chemist=chem, employee=employee,
            req_name=request.POST.get('name'),
            req_phone=request.POST.get('phone'),
            req_address=request.POST.get('address', '').strip() or None, # 🌟 NAYA FIELD MAP KIYA
            req_territory_id=request.POST.get('territory') or None,
            req_route_id=request.POST.get('route') or None,
            status='Pending'
        )
        messages.success(request, f"🚀 The Update Request for Chemist '{chem.name}' has been sent to the Manager for approval!")
        return redirect('edit_chemist_list')

    return render(request, 'edit_chemist.html', {'chemist': chem, 'territories': territories, 'routes': routes})

@employee_required
def edit_doctor_view(request, employee, doc_id):
    team_employees = get_dropdown_team(employee, ordered=False)
    doc = get_object_or_404(Doctor, id=doc_id, allocated_to__in=team_employees)
    
    if DoctorEditRequest.objects.filter(doctor=doc, status='Pending').exists():
        messages.warning(request, "⚠️ An Edit Request for this Doctor is already pending with the Manager!")
        return redirect('edit_doctor_list')
        
    sub_team = Employee.objects.filter(id=doc.allocated_to.id)
    my_terr_ids = get_team_territory_ids(sub_team)
    territories = Territory.objects.filter(id__in=my_terr_ids).order_by('name')
    routes = get_team_requested_routes(sub_team, my_terr_ids)

    if request.method == "POST":
        # 🌟 Photos capture aur compress karna (Main Photo & V-Card)
        photo_file = request.FILES.get('photo')
        vcard_file = request.FILES.get('vcard_photo')
        
        req_photo_compressed = compress_photo(photo_file) if photo_file else None
        req_vcard_compressed = compress_photo(vcard_file) if vcard_file else None

        DoctorEditRequest.objects.create(
            doctor=doc, employee=employee,
            req_name=request.POST.get('name'),
            req_degree=request.POST.get('degree') or None,
            req_specialty=request.POST.get('specialty') or None,
            req_category=request.POST.get('category') or None,
            req_territory_id=request.POST.get('territory') or None,
            req_route_id=request.POST.get('route') or None,
            req_mobile=request.POST.get('mobile') or None,
            req_email=request.POST.get('email') or None,
            req_dob=request.POST.get('dob') or None,
            req_address=request.POST.get('address') or None,          # 📍 Address Field
            req_dom=request.POST.get('dom') or None,                  # 💑 DOM Field
            req_spouse_dob=request.POST.get('spouse_dob') or None,    # 🍰 Spouse DOB Field
            req_photo=req_photo_compressed,                           # 📸 Compressed Main Photo
            req_vcard_photo=req_vcard_compressed,                     # 🪪 Compressed V-Card Photo
            status='Pending'
        )
        messages.success(request, f"🚀 The Update Request for Doctor '{doc.name}' has been sent to the Manager!")
        return redirect('edit_doctor_list')

    return render(request, 'edit_doctor.html', {
        'doctor': doc, 
        'territories': territories,
        'routes': routes, 
        'specialty_choices': Doctor.SPECIALTY_CHOICES,
        'degree_choices': Doctor.DEGREE_CHOICES,
        'category_choices': Doctor.CATEGORY_CHOICES
    })

# ==============================================================================
# 💊 EDIT CHEMIST LOGIC
# ==============================================================================

@employee_required
def edit_chemist_list_view(request, employee):
    # 1. Team fetch karo
    team_employees = get_dropdown_team(employee)
    
    # 2. Smart Default: Agar Manager hai, toh default pehle MR ko select karo
    default_emp_id = str(employee.id)
    if employee.designation != 'MR':
        first_subordinate = team_employees.exclude(id=employee.id).first()
        if first_subordinate:
            default_emp_id = str(first_subordinate.id)
            
    selected_emp_id = request.GET.get('employee_id', default_emp_id)
    selected_emp = get_object_or_404(Employee, id=selected_emp_id, company=employee.company)

    # 3. Sirf Approved chemists dikhao
    chemists = Chemist.objects.filter(allocated_to=selected_emp, status='Approved').order_by('name')
    
    # 4. Context bhejo taaki template mein dropdown ban sake
    return render(request, 'edit_chemist_list.html', {
        'chemists': chemists,
        'team_employees': team_employees,
        'selected_emp_id': int(selected_emp_id),
        'is_manager_view': employee.designation != 'MR'
    })

import openpyxl
import csv
import io
from django.contrib import messages
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from SFA.models import Employee, Territory, Route, Doctor, Chemist

@login_required
def bulk_network_upload_view(request):
    allowed_roles = ['Admin', 'System Administrator', 'Manager', 'NSM']
    if request.user.employee.designation not in allowed_roles:
        messages.error(request, "You do not have permission to access this page.")
        return redirect('view_hub')

    employees = Employee.objects.filter(company=request.user.employee.company, is_active=True).order_by('name')

    if request.method == 'POST':
        upload_type = request.POST.get('upload_type')
        selected_emp_id = request.POST.get('employee_id')
        
        uploaded_file = request.FILES.get('upload_file') or request.FILES.get('csv_file')

        if not uploaded_file or not selected_emp_id:
            messages.error(request, "❌ File or Employee missing!")
            return redirect('bulk_network_upload')

        emp_obj = Employee.objects.filter(company=request.user.employee.company, id=selected_emp_id).first()
        file_name = uploaded_file.name.lower()
        rows = []
        headers = []

        try:
            # ==========================================
            # 1. FILE READING (ACTIVE SHEET + MEMORY SAFE)
            # ==========================================
            if file_name.endswith('.csv'):
                file_data = uploaded_file.read().decode('utf-8-sig')
                io_string = io.StringIO(file_data)
                reader = csv.reader(io_string)
                all_rows = list(reader)
                if len(all_rows) < 2:
                    messages.error(request, "❌ The CSV file is empty or headers are missing.")
                    return redirect('bulk_network_upload')
                headers = [str(h).strip().lower().replace(' ', '_') for h in all_rows[0] if h]
                rows = all_rows[1:]
            
            elif file_name.endswith(('.xlsx', '.xls')):
                # 🚀 SPEED & MEMORY HACK with ACTIVE SHEET ONLY
                wb = openpyxl.load_workbook(uploaded_file, data_only=True, read_only=True)
                sheet = wb.active  # 🌟 Sirf pehli sheet load karega
                
                is_first_row = True
                for row in sheet.iter_rows(values_only=True):
                    if not any(row): continue # 🌟 Blank Ghost Rows Ignore
                    
                    if is_first_row:
                        if not headers:
                            headers = [str(h).strip().lower().replace(' ', '_') for h in row if h]
                        is_first_row = False
                    else:
                        rows.append(row)
                
                wb.close() # 🧹 Memory clean
                        
                if not rows:
                    messages.error(request, "❌ The Excel file is empty.")
                    return redirect('bulk_network_upload')
            else:
                messages.error(request, "❌ Please upload only a .csv or .xlsx file.")
                return redirect('bulk_network_upload')

            error_count = 0
            error_messages = []
            
            # ==========================================
            # 2. FAST CACHING (0 DB Queries Inside Loop)
            # ==========================================
            terr_cache = {t.name.strip().lower(): t for t in Territory.objects.filter(company=emp_obj.company)}
            route_cache = {f"{r.territory_id}_{r.name.strip().lower()}": r for r in Route.objects.filter(territory__company=emp_obj.company)}

            objects_to_create = []

            # ==========================================
            # 3. DATA CLEANING & MATCHING
            # ==========================================
            for row_number, row_data in enumerate(rows, start=2):
                row_data = list(row_data) + [''] * (len(headers) - len(row_data))
                clean_row = dict(zip(headers, row_data))
                
                # Extreme Cleanup (Spaces & nan)
                for k, v in clean_row.items():
                    val = str(v).strip() if v is not None else ''
                    clean_row[k] = '' if val.lower() == 'nan' else val

                name = clean_row.get('name')
                if not name:
                    continue 

                terr_name = clean_row.get('territory_name', '').strip().lower()
                if not terr_name:
                    error_count += 1
                    error_messages.append(f"Row {row_number}: 'territory_name' is blank.")
                    continue

                terr_obj = terr_cache.get(terr_name)
                if not terr_obj:
                    error_count += 1
                    error_messages.append(f"Row {row_number}: Territory '{clean_row.get('territory_name')}' not found.")
                    continue

                route_name = clean_row.get('route_name', '').strip().lower()
                route_obj = None
                if route_name:
                    cache_key = f"{terr_obj.id}_{route_name}"
                    route_obj = route_cache.get(cache_key)
                    if not route_obj:
                        error_count += 1
                        error_messages.append(f"Row {row_number}: Route '{clean_row.get('route_name')}' not found under Territory '{terr_obj.name}'.")
                        continue

                try:
                    if upload_type == 'doctor':
                        doc = Doctor(
                            company=emp_obj.company,
                            name=name,
                            specialty=(clean_row.get('specialty') or 'General')[:50],
                            mobile=clean_row.get('mobile', '').split('.')[0][:15], # Decimal removal
                            category=(clean_row.get('category') or 'C')[:1].upper(), # A+ to A
                            degree=clean_row.get('degree', '')[:50],
                            address=clean_row.get('address', ''),
                            territory=terr_obj,
                            route=route_obj,
                            allocated_to=emp_obj,
                            status='Approved'
                        )
                        objects_to_create.append(doc)
                    elif upload_type == 'chemist':
                        chem = Chemist(
                            company=emp_obj.company,
                            name=name,
                            phone=clean_row.get('phone', '').split('.')[0][:15],
                            territory=terr_obj,
                            route=route_obj,
                            allocated_to=emp_obj,
                            status='Approved'
                        )
                        objects_to_create.append(chem)
                except Exception as e:
                    error_count += 1
                    error_messages.append(f"Row {row_number}: Data Format Error - {str(e)}")

            # ==========================================
            # 4. BULK CREATE (Fastest Insert)
            # ==========================================
            if objects_to_create:
                if upload_type == 'doctor':
                    Doctor.objects.bulk_create(objects_to_create, batch_size=500)
                elif upload_type == 'chemist':
                    Chemist.objects.bulk_create(objects_to_create, batch_size=500)
            
            success_count = len(objects_to_create)

            if success_count > 0:
                messages.success(request, f"✅ Upload Complete! {success_count} records saved instantly for {emp_obj.name}.")
            
            if error_count > 0:
                messages.error(request, f"❌ {error_count} records skipped due to errors.")
                for err in error_messages[:5]:
                    messages.warning(request, err)
                if len(error_messages) > 5:
                    messages.warning(request, f"...and {len(error_messages) - 5} more similar errors.")

            return redirect('bulk_network_upload')

        except Exception as e:
            messages.error(request, f"❌ File processing error: {e}")
            return redirect('bulk_network_upload')

    return render(request, 'bulk_upload_network.html', {'employees': employees})

@employee_required
def promo_dispatch_view(request, employee):
    if employee.designation not in ['Admin', 'System Administrator']:
        messages.error(request, "🚫 Access Denied.")
        return redirect('view_hub')

    if request.method == 'POST':
        item_id = request.POST.get('item_id')
        emp_ids = request.POST.getlist('employee_ids') # Ek sath multiple MRs ko bhej sakte hain
        qty = request.POST.get('quantity')

        try:
            item = get_object_or_404(PromoItem, id=item_id, company=employee.company)
            count = 0
            for emp_id in emp_ids:
                emp = get_object_or_404(Employee, id=emp_id, company=employee.company)
                PromoDispatch.objects.create(
                    employee=emp, 
                    item=item, 
                    quantity=int(qty),
                    status='In-Transit'
                )
                count += 1
            messages.success(request, f"🚀 Successfully dispatched {item.name} to {count} MR(s)!")
        except Exception as e:
            messages.error(request, f"❌ Error: {str(e)}")
            
        return redirect('promo_dispatch')

    # Dropdowns ke liye data
    items = PromoItem.objects.filter(company=employee.company, is_active=True).order_by('name')
    # FIX: Company specific fetch
    mrs = Employee.objects.filter(company=employee.company, designation='MR', is_active=True).order_by('name')
    # FIX: Company specific fetch
    recent_dispatches = PromoDispatch.objects.filter(employee__company=employee.company).order_by('-dispatch_date', '-id')[:50]

    return render(request, 'promo_dispatch.html', {
        'items': items, 
        'mrs': mrs, 
        'recent_dispatches': recent_dispatches
    })

@employee_required
def gift_campaign_view(request, employee):
    # Sirf MR access kar sakta hai
    current_month = timezone.localdate().month
    current_year = timezone.localdate().year

    selected_month = int(request.GET.get('month', current_month))
    selected_year = int(request.GET.get('year', current_year))

    if request.method == 'POST':
        item_id = request.POST.get('item_id')
        doctor_ids = request.POST.getlist('doctor_ids')

        if not item_id or not doctor_ids:
            messages.error(request, "❌ Please select an Item and at least one Doctor.")
            return redirect('gift_campaign')

        # 🛡️ IDOR FIX: item company-scoped
        item = get_object_or_404(PromoItem, id=item_id, item_type='HighValue', company=employee.company)

        # Check karo MR ke paas yeh item inventory mein hai
        inventory = MRInventory.objects.filter(employee=employee, item=item, stock_qty__gt=0).first()
        if not inventory:
            messages.error(request, f"❌ You do not have stock of {item.name}.")
            return redirect('gift_campaign')

        # 🚀 N+1 FIXED: existing assignments EK query mein (per-doctor exists() hatao)
        existing_doc_ids = set(GiftCampaignPlan.objects.filter(
            employee=employee, item=item, month=selected_month, year=selected_year,
            status__in=['Pending', 'Approved']
        ).values_list('doctor_id', flat=True))

        # Rejected wale dobara allow honge (Purani logic jaisi hi)
        new_docs_to_add = [d for d in doctor_ids if d not in existing_doc_ids]

        # 🚀 Stock limit check (purane 'already_allocated' count ki jagah set ka len — same value)
        already_allocated = len(existing_doc_ids)

        if already_allocated + len(new_docs_to_add) > inventory.stock_qty:
            available_to_assign = inventory.stock_qty - already_allocated
            messages.error(
                request,
                f"❌ Stock Limit Exceeded! Total stock of {item.name} is {inventory.stock_qty} "
                f"(of which {already_allocated} are already assigned). You can select only {available_to_assign} more."
            )
            return redirect('gift_campaign')

        if not new_docs_to_add:
            messages.warning(request, "⚠️ All doctors are already in this month's plan.")
            return redirect('gift_campaign')

        # 🚀 N+1 FIXED: doctors EK query (🛡️ company-scoped) + bulk_create
        valid_doctors = list(Doctor.objects.filter(id__in=new_docs_to_add, company=employee.company))
        if len(set(new_docs_to_add)) != len(valid_doctors):
            messages.error(request, "❌ Invalid doctor in selection.")
            return redirect('gift_campaign')

        # 🛡️ TRANSACTION: partial save se bachav
        from django.db import transaction
        with transaction.atomic():
            GiftCampaignPlan.objects.bulk_create([
                GiftCampaignPlan(
                    employee=employee, doctor=d, item=item,
                    month=selected_month, year=selected_year,
                    status='Pending'
                ) for d in valid_doctors
            ])

        created = len(valid_doctors)
        messages.success(request, f"✅ Gift Campaign sent to Manager for {created} Doctor(s)!")
        return redirect('gift_campaign')

    # GET — MR ki HighValue inventory
    high_value_stock = MRInventory.objects.filter(
        employee=employee, item__item_type='HighValue', stock_qty__gt=0
    ).select_related('item')

    # MR ke doctors
    my_doctors = Doctor.objects.filter(
        route__territory=employee.headquarter, status='Approved'
    ).order_by('name')

    # Is month ke plans
    my_plans = GiftCampaignPlan.objects.filter(
        employee=employee, month=selected_month, year=selected_year
    ).select_related('doctor', 'item').order_by('-id')

    import calendar
    return render(request, 'gift_campaign.html', {
        'high_value_stock': high_value_stock,
        'my_doctors': my_doctors,
        'my_plans': my_plans,
        'selected_month': selected_month,
        'selected_year': selected_year,
        'months_choices': [(i, calendar.month_name[i]) for i in range(1, 13)],
        'current_year': current_year,
    })
# ==============================================================================
# 👨‍⚕️ VIEW DOCTOR PROFILE
# ==============================================================================
@employee_required
def view_doctor_profile(request, employee, doc_id):
    # Manager ki team fetch karo taaki security bani rahe
    team_employees = get_dropdown_team(employee, ordered=False)
    doctor = get_object_or_404(Doctor, id=doc_id, allocated_to__in=team_employees)
    
    return render(request, 'doctor_profile.html', {'doctor': doctor})

from django.shortcuts import render, get_object_or_404
from SFA.models import Chemist
from SFA.decorators import employee_required

@employee_required
def view_chemist_profile(request, employee, chem_id):
    team_employees = get_dropdown_team(employee, ordered=False)
    chemist = get_object_or_404(Chemist, id=chem_id, allocated_to__in=team_employees)
    return render(request, 'chemist_profile.html', {'chemist': chemist})