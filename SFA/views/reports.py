import io
import csv
import json
import calendar
from datetime import date, datetime
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from SFA.models import DCRVisit
from django.utils import timezone
from datetime import datetime
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.db.models import Sum, Count, Q

from SFA.models import (
    Employee, Doctor, Chemist, Product, DailyDCR, DCRVisit, DCRProductDetail,
    MonthlyTourProgram, Territory, MonthlyExpenseReport, DailyTourPlan, Route, Holiday,
    MonthlyTargetMaster, LeaveApplication, LeaveBalance, DayStart, DayEnd, TerritoryTarget,
    StockistProductStatement, Stockist, PartyWiseSaleReport, PartyWiseSaleLine, DoctorRxMapping,
    FreeQtyClaimMaster, FreeQtyClaimLine, GiftCampaignPlan, SystemSetting, DoctorROILedger,
    MRInventory, PromoDispatch,    DoctorRxMapping, ChemistEditRequest, DoctorEditRequest
)
from .auth import get_full_team_employees, get_team_territory_ids, get_team_requested_routes, get_dropdown_team
from SFA.decorators import employee_required

@employee_required
def product_sales_report_view(request, employee):
    team_employees = get_dropdown_team(employee, ordered=False)
    default_emp_id = str(employee.id)
    if employee.designation != 'MR':
        first_sub = team_employees.exclude(id=employee.id).first()
        if first_sub:
            default_emp_id = str(first_sub.id)
            
    selected_emp_id = request.GET.get('employee_id', default_emp_id)
    selected_emp = get_object_or_404(Employee, id=int(selected_emp_id))
    
    today = timezone.localdate()
    from_month = int(request.GET.get('from_month') or today.month)
    to_month = int(request.GET.get('to_month') or today.month)
    selected_year = int(request.GET.get('year') or today.year)

    if from_month > to_month:
        from_month, to_month = to_month, from_month

    months_range = list(range(from_month, to_month + 1))
    months_headers = [(m, calendar.month_name[m][:3]) for m in months_range]

    product_entries = DCRProductDetail.objects.filter(
        visit__daily_dcr__employee=selected_emp, visit__daily_dcr__date__month__gte=from_month,
        visit__daily_dcr__date__month__lte=to_month, visit__daily_dcr__date__year=selected_year
    ).select_related('product', 'visit__daily_dcr')

    products_dict = {p.id: {'name': p.name, 'price': float(p.price) if getattr(p, 'price', None) else 0.0} for p in Product.objects.filter(company=selected_emp.company)}
    
    agg_data = defaultdict(lambda: {'monthly': {m: {'samples': 0, 'orders': 0, 'val': 0.0} for m in months_range}, 'tot_samples': 0, 'tot_orders': 0, 'tot_val': 0.0})
    gt_monthly = {m: {'samples': 0, 'orders': 0, 'val': 0.0} for m in months_range}
    gt_samples, gt_orders, gt_val = 0, 0, 0.0

    for entry in product_entries:
        p_id, m = entry.product_id, entry.visit.daily_dcr.date.month
        sq, oq = entry.sample_qty or 0, entry.order_qty or 0
        if p_id not in products_dict: continue
        price = products_dict[p_id]['price']
        val = oq * price
        
        agg_data[p_id]['monthly'][m]['samples'] += sq; agg_data[p_id]['monthly'][m]['orders'] += oq; agg_data[p_id]['monthly'][m]['val'] += val
        agg_data[p_id]['tot_samples'] += sq; agg_data[p_id]['tot_orders'] += oq; agg_data[p_id]['tot_val'] += val
        
        gt_monthly[m]['samples'] += sq; gt_monthly[m]['orders'] += oq; gt_monthly[m]['val'] += val
        gt_samples += sq; gt_orders += oq; gt_val += val

    report_data = []
    for p_id, p_info in products_dict.items():
        if p_id in agg_data and (agg_data[p_id]['tot_samples'] > 0 or agg_data[p_id]['tot_orders'] > 0):
            p_data = agg_data[p_id]
            report_data.append({'product_name': p_info['name'], 'price': p_info['price'], 'monthly_list': [p_data['monthly'][m] for m in months_range], 'tot_samples': p_data['tot_samples'], 'tot_orders': p_data['tot_orders'], 'tot_val': p_data['tot_val']})
            
    report_data.sort(key=lambda x: x['product_name'])
    gt_monthly_list = [gt_monthly[m] for m in months_range]

    if request.GET.get('export') == 'excel':
        filename = f"POB_Report_{selected_emp.name}_M{from_month}-M{to_month}_{selected_year}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "POB Report"
        
        period_str = f"{calendar.month_name[from_month][:3]} to {calendar.month_name[to_month][:3]} {selected_year}"
        ws.append(['PRODUCT POB & SAMPLES REPORT (MONTH TREND)'])
        ws.append(['Employee:', selected_emp.name, 'Period:', period_str])
        ws.append([''])
        ws['A1'].font = Font(bold=True, size=14, color="107C41")
        ws['A2'].font = Font(bold=True); ws['C2'].font = Font(bold=True)
        
        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        
        headers = ['Product Name', 'Price (₹)']
        for m_num, m_name in months_headers: headers.extend([f'{m_name} Samples', f'{m_name} POB', f'{m_name} Value (₹)'])
        headers.extend(['Total Samples', 'Total POB', 'Total Value (₹)'])
        ws.append(headers)
        
        for col_num, cell in enumerate(ws[5], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20 if col_num == 1 else 15
            
        for row in report_data:
            row_data = [row['product_name'], round(row['price'], 2)]
            for m_data in row['monthly_list']: row_data.extend([m_data['samples'], m_data['orders'], round(m_data['val'], 2)])
            row_data.extend([row['tot_samples'], row['tot_orders'], round(row['tot_val'], 2)])
            ws.append(row_data)
            
        ws.append([''])
        gt_row = ['GRAND TOTAL', '']
        for m_gt in gt_monthly_list: gt_row.extend([m_gt['samples'], m_gt['orders'], round(m_gt['val'], 2)])
        gt_row.extend([gt_samples, gt_orders, round(gt_val, 2)])
        
        ws.append(gt_row)
        for cell in ws[ws.max_row]: cell.font = Font(bold=True)
            
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return render(request, 'pob_report.html', {
        'team_employees': team_employees, 'selected_emp_id': int(selected_emp_id),
        'from_month': from_month, 'to_month': to_month, 'selected_year': selected_year,
        'months_choices': [(i, calendar.month_name[i]) for i in range(1, 13)], 'months_headers': months_headers,
        'report_data': report_data, 'gt_monthly_list': gt_monthly_list, 'gt_samples': gt_samples, 'gt_orders': gt_orders, 'gt_val': gt_val,
        'is_manager_view': employee.designation != 'MR'
    })


@employee_required
def holiday_list_view(request, employee):
    selected_emp_id = request.GET.get('employee_id', str(employee.id))
    selected_emp = get_object_or_404(Employee, id=selected_emp_id)
    
    rbm_emp = None
    curr = selected_emp
    while curr:
        if curr.designation == 'RBM':
            rbm_emp = curr
            break
        curr = curr.manager
        
    admin_ids = list(Employee.objects.filter(company=selected_emp.company, designation='Admin').values_list('id', flat=True))
    holiday_creators = admin_ids.copy()
    if rbm_emp: holiday_creators.append(rbm_emp.id)
        
    holidays = Holiday.objects.filter(proposed_by_id__in=holiday_creators, status='Approved').order_by('-date')

    # 🌟 NAYA EXCEL EXPORT LOGIC YAHAN SE SHURU
    if request.GET.get('export') == 'excel':
        region_name = rbm_emp.name if rbm_emp else "Consolidated"
        filename = f"Approved_Holidays_{region_name}.xlsx"
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Holidays"
        
        ws.append(['APPROVED HOLIDAYS LIST'])
        ws.append(['Region:', region_name])
        ws.append([''])
        
        ws['A1'].font = Font(bold=True, size=12, color="107C41")
        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        ws.append(['Date', 'Day', 'Holiday Name'])
        for col_num, cell in enumerate(ws[4], 1):
            cell.fill = header_fill
            cell.font = header_font
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20
        
        for h in holidays:
            ws.append([h.date.strftime('%d-%b-%Y'), h.date.strftime('%A'), h.name])
            
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    # 🌟 EXCEL EXPORT LOGIC KHATAM

    return render(request, 'holiday_list.html', {
        'holidays': holidays, 
        'rbm_name': rbm_emp.name if rbm_emp else "Consolidated (All States)", 
        'selected_emp_id': selected_emp.id
    })

@employee_required
def view_dcr_report(request, employee, dcr_id):
    daily_dcr = get_object_or_404(DailyDCR.objects.prefetch_related('visits__product_details__product', 'visits__doctor', 'visits__chemist'), id=dcr_id, employee=employee)
    return render(request, 'view_dcr_report.html', {'daily_dcr': daily_dcr, 'visits': daily_dcr.visits.all()})

@employee_required
def manager_report_view(request, employee):
    today = timezone.localdate()
    mr_reports = DailyDCR.objects.filter(employee__company=employee.company, date=today).select_related('employee').annotate(
        total_visits=Count('visits', distinct=True), total_samples=Sum('visits__product_details__sample_qty'), total_orders=Sum('visits__product_details__order_qty')
    ).values('employee__name', 'total_visits', 'total_samples', 'total_orders')
    grand_total_orders = DCRProductDetail.objects.filter(visit__daily_dcr__employee__company=employee.company, visit__daily_dcr__date=today).aggregate(total=Sum('order_qty'))['total'] or 0
    return render(request, 'manager_report.html', {'today': today, 'mr_reports': mr_reports, 'grand_total_orders': grand_total_orders})

@employee_required
def dcr_report_view(request, employee):
    team_employees = get_dropdown_team(employee)
    default_emp_id = str(employee.id)
    if employee.designation != 'MR':
        first_sub = team_employees.exclude(id=employee.id).first()
        if first_sub:
            default_emp_id = str(first_sub.id)
            
    selected_emp_id = request.GET.get('employee_id', default_emp_id)
    selected_emp = get_object_or_404(Employee, id=selected_emp_id)
    
    today = timezone.localdate()
    selected_month = int(request.GET.get('month') or today.month)
    selected_year = int(request.GET.get('year') or today.year)

    day_starts = DayStart.objects.filter(employee=selected_emp, date__month=selected_month, date__year=selected_year).order_by('-date')
    actual_dcrs = DailyDCR.objects.filter(employee=selected_emp, date__month=selected_month, date__year=selected_year).prefetch_related('visits__doctor', 'visits__chemist', 'visits__product_details__product')
    dcr_dict = {d.date: d for d in actual_dcrs}

    report_list, total_dr_visits, total_chem_visits = [], 0, 0

    for ds in day_starts:
        dcr_obj = dcr_dict.get(ds.date)
        visits = dcr_obj.visits.all() if dcr_obj else []

        if dcr_obj:
            dr_v = dcr_obj.visits.filter(doctor__isnull=False).count()
            chem_v = dcr_obj.visits.filter(chemist__isnull=False).count()
            total_dr_visits += dr_v; total_chem_visits += chem_v
            
        route_list = ds.routes.select_related('territory').all() if hasattr(ds, 'routes') else []
        # 🌟 NAYA: Route ke saath uski territory (duplicate-free) bhi nikal rahe hain
        territory_names = []
        for r in route_list:
            if r.territory and r.territory.name not in territory_names:
                territory_names.append(r.territory.name)
        final_route = {'name': ", ".join([r.name for r in route_list]), 'territory': ", ".join(territory_names)} if route_list else None

        report_list.append({'id': ds.id, 'date': ds.date, 'work_type': ds.work_type, 'visits': visits, 'employee': ds.employee, 'route': final_route})

    months = [(i, calendar.month_name[i]) for i in range(1, 13)]
    years = [today.year - 1, today.year, today.year + 1]
    total_days_worked = day_starts.count()
    dr_avg   = round(total_dr_visits   / total_days_worked, 1) if total_days_worked > 0 else 0
    chem_avg = round(total_chem_visits / total_days_worked, 1) if total_days_worked > 0 else 0

    total_samples = DCRProductDetail.objects.filter(visit__daily_dcr__employee=selected_emp, visit__daily_dcr__date__month=selected_month, visit__daily_dcr__date__year=selected_year).aggregate(s=Sum('sample_qty'))['s'] or 0
    total_orders = DCRProductDetail.objects.filter(visit__daily_dcr__employee=selected_emp, visit__daily_dcr__date__month=selected_month, visit__daily_dcr__date__year=selected_year).aggregate(o=Sum('order_qty'))['o'] or 0

    return render(request, 'dcr_report.html', {
        'dcrs': report_list, 'team_employees': team_employees, 'selected_emp_id': int(selected_emp_id),
        'selected_employee_name': selected_emp.name, 'is_manager_view': employee.designation != 'MR',
        'selected_month': selected_month, 'selected_year': selected_year, 'months': months, 'years': years,
        'total_days_worked': total_days_worked, 'total_dr_visits': total_dr_visits, 'total_chem_visits': total_chem_visits,
        'dr_avg': dr_avg, 'chem_avg': chem_avg, 'total_samples': total_samples, 'total_orders': total_orders,
    })


def _get_target_chain_starter(target):
    """
    🌟 FIX: MonthlyTargetMaster ka 'employee' attribute nahi hota — ye model
    Employee ki jagah Territory se linked hai. Lekin approval-chain
    (manager -> upar ke manager -> ... -> Admin) follow karne ke liye humein
    ek "starting employee" chahiye jiske manager-chain upar traverse kar
    sakein.
    Isliye target.territory (jo Employee.headquarter hai) se us territory
    ka current active employee dhoondte hain — agar territory abhi 'Vacant_'
    placeholder ke paas hai, to uska manager bhi wahi hota hai jo asli
    (resigned) employee ka tha, isliye chain sahi traverse hoti hai.
    """
    if not target.territory_id:
        return None
    return Employee.objects.filter(headquarter_id=target.territory_id, is_active=True).first()


@employee_required
def manager_approval_hub(request, employee):

    manager = employee
    team_members = get_full_team_employees(manager).exclude(id=manager.id)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        itype = request.POST.get('item_type')
        iid = request.POST.get('item_id')
        
        obj = None
        if itype == 'mtp': obj = get_object_or_404(MonthlyTourProgram, id=iid)
        elif itype == 'doctor': obj = get_object_or_404(Doctor, id=iid)
        elif itype == 'chemist': obj = get_object_or_404(Chemist, id=iid)
        elif itype == 'monthly_expense': obj = get_object_or_404(MonthlyExpenseReport, id=iid)
        elif itype == 'route': obj = get_object_or_404(Route, id=iid)
        elif itype == 'holiday': obj = get_object_or_404(Holiday, id=iid)
        elif itype == 'leave': obj = get_object_or_404(LeaveApplication, id=iid)
        elif itype == 'target': obj = get_object_or_404(MonthlyTargetMaster, id=iid)
        elif itype == 'free_claim': obj = get_object_or_404(FreeQtyClaimMaster, id=iid)
        elif itype == 'gift_campaign': obj = get_object_or_404(GiftCampaignPlan, id=iid)
        elif itype == 'chemist_edit': obj = get_object_or_404(ChemistEditRequest, id=iid)
        elif itype == 'doctor_edit': obj = get_object_or_404(DoctorEditRequest, id=iid)
        
        if obj and action in ['approve', 'reject']:
            
            # ✅ APPROVE ACTION
            if action == 'approve':
                if itype == 'free_claim':
                    if manager.designation in ['Admin', 'System Administrator']:
                        obj.status = 'Approved'; obj.admin_remark = request.POST.get('remark', '')
                    else:
                        m_list = list(obj.approved_by_managers)
                        if manager.id not in m_list: m_list.append(manager.id)
                        obj.approved_by_managers = m_list
                        obj.manager_remark = request.POST.get('remark', '')
                        chain_complete = True
                        curr = obj.employee.manager
                        while curr and curr.designation not in ['Admin', 'System Administrator']:
                            if curr.id not in obj.approved_by_managers:
                                chain_complete = False; break
                            curr = curr.manager
                        obj.status = 'Pending_Admin' if chain_complete else 'Pending_Manager'
                            
                elif itype == 'target':
                    if manager.designation in ['Admin', 'System Administrator']:
                        obj.status = 'Approved'
                    else:
                        m_list = list(obj.approved_by_managers)
                        if manager.id not in m_list: m_list.append(manager.id)
                        obj.approved_by_managers = m_list
                        chain_complete = True
                        chain_emp = _get_target_chain_starter(obj)  # 🌟 FIX
                        curr = chain_emp.manager if chain_emp else None
                        while curr and curr.designation not in ['Admin', 'System Administrator']:
                            if curr.id not in obj.approved_by_managers:
                                chain_complete = False; break
                            curr = curr.manager
                        obj.status = 'Pending_Admin' if chain_complete else 'Pending_Manager'

                elif itype == 'gift_campaign':
                    obj.status = 'Approved'; obj.manager_remark = request.POST.get('remark', '')
                    
                # 🌟 NAYA & FIXED: Doctor/Chemist Edit Logic ab APPROVE block ke ANDAR aayega
                elif itype == 'chemist_edit':
                    obj.chemist.name = obj.req_name
                    obj.chemist.phone = obj.req_phone
                    obj.chemist.address = obj.req_address  # 🌟 FIX: address copy missing thi, isliye approve karne par bhi update nahi hota tha
                    if obj.req_territory: obj.chemist.territory = obj.req_territory
                    if obj.req_route: obj.chemist.route = obj.req_route
                    obj.chemist.save() # Main DB update hua!
                    obj.status = 'Approved'
                    
                elif itype == 'doctor_edit':
                    obj.doctor.name = obj.req_name
                    if obj.req_degree: obj.doctor.degree = obj.req_degree
                    if obj.req_specialty: obj.doctor.specialty = obj.req_specialty
                    if obj.req_category: obj.doctor.category = obj.req_category
                    if obj.req_territory: obj.doctor.territory = obj.req_territory
                    if obj.req_route: obj.doctor.route = obj.req_route
                    obj.doctor.mobile = obj.req_mobile
                    obj.doctor.email = obj.req_email
                    obj.doctor.dob = obj.req_dob
                    # 🌟 FIX: Ye 4 fields pehle se DoctorEditRequest me save ho
                    # rahe the (edit_doctor_view se), par yahan Doctor model
                    # me kabhi copy nahi ho rahe the — Manager approve karta
                    # tha par Address/Anniversary/Spouse-DOB/V-Card actually
                    # update hi nahi hota tha. Ab fix kiya.
                    obj.doctor.address = obj.req_address
                    obj.doctor.dom = obj.req_dom
                    obj.doctor.spouse_dob = obj.req_spouse_dob
                    if obj.req_vcard_photo: obj.doctor.vcard_photo = obj.req_vcard_photo
                    # 🌟 NAYE FIELDS: Residential Address + Children DOB
                    obj.doctor.residential_address = obj.req_residential_address
                    obj.doctor.child_1_dob = obj.req_child_1_dob
                    obj.doctor.child_2_dob = obj.req_child_2_dob
                    if obj.req_photo: obj.doctor.photo = obj.req_photo
                    obj.doctor.save() # Main DB update hua!
                    obj.status = 'Approved'

                else:
                    obj.status = 'Approved'
                    if itype == 'leave' and obj.leave_type != 'LWP':
                        bal = LeaveBalance.objects.get(employee=obj.employee, year=obj.start_date.year)
                        if obj.leave_type == 'CL': bal.cl_used += obj.no_of_days
                        elif obj.leave_type == 'SL': bal.sl_used += obj.no_of_days
                        elif obj.leave_type == 'PL': bal.pl_used += obj.no_of_days
                        bal.save()
                        
            # ❌ REJECT ACTION (Ab yahan sirf rejection hoga!)
            else:
                obj.status = 'Rejected'
                if itype == 'free_claim':
                    if manager.designation in ['Admin', 'System Administrator']: obj.admin_remark = request.POST.get('remark', '')
                    else: obj.manager_remark = request.POST.get('remark', '')
                elif itype == 'gift_campaign':
                    obj.manager_remark = request.POST.get('remark', '')
            
            # Request/Action finally save kardi
            obj.save()

            # 🔔 NOTIFICATION SYSTEM (Dynamically shows Approved or Rejected)
            try:
                from SFA.models import SystemNotification
                if itype == 'target':
                    applicant = _get_target_chain_starter(obj)  # 🌟 FIX: Territory se employee
                else:
                    applicant = getattr(obj, 'employee', getattr(obj, 'requested_by', getattr(obj, 'proposed_by', getattr(obj, 'allocated_to', None))))
                if applicant:
                    formatted_type = itype.replace('_', ' ').title()
                    # 🌟 FIX: Message ab status ke hisaab se badlega
                    msg = f"Your {formatted_type} has been {obj.status} by {manager.name}."
                    SystemNotification.objects.create(
                        employee=applicant, 
                        title=f"{formatted_type} {obj.status}!", 
                        message=msg
                    )
            except Exception: pass 
            
            return redirect('manager_approvals')


    pending_targets, pending_free_claims, pending_gift_campaigns = [], [], []

    if manager.designation in ['Admin', 'System Administrator']:
        pt_admin = MonthlyTargetMaster.objects.filter(territory__company=employee.company).exclude(status__in=['Draft', 'Approved', 'Rejected']).order_by('-year', '-month')
        for t in pt_admin:
            if t.status == 'Pending_Admin': pending_targets.append(t)
            else:
                chain_approved = True
                chain_emp = _get_target_chain_starter(t)  # 🌟 FIX
                curr = chain_emp.manager if chain_emp else None
                while curr and curr.designation not in ['Admin', 'System Administrator']:
                    if curr.id not in t.approved_by_managers: chain_approved = False; break
                    curr = curr.manager
                if chain_approved and t not in pending_targets: pending_targets.append(t)

        fc_admin = FreeQtyClaimMaster.objects.filter(employee__company=employee.company).exclude(status__in=['Draft', 'Approved', 'Rejected']).order_by('created_at')
        for c in fc_admin:
            if c.status == 'Pending_Admin': pending_free_claims.append(c)
            else:
                chain_approved = True
                curr = c.employee.manager
                while curr and curr.designation not in ['Admin', 'System Administrator']:
                    if curr.id not in c.approved_by_managers: chain_approved = False; break
                    curr = curr.manager
                if chain_approved and c not in pending_free_claims: pending_free_claims.append(c)

        pending_gift_campaigns = GiftCampaignPlan.objects.filter(employee__in=team_members, status='Pending').select_related('employee', 'doctor', 'item').order_by('-id')

    else:
        team_territories = Employee.objects.filter(id__in=team_members).exclude(headquarter__isnull=True).values_list('headquarter_id', flat=True)
        pt = MonthlyTargetMaster.objects.filter(territory_id__in=team_territories, status='Pending_Manager').order_by('-year', '-month')

        for t in pt:
            if manager.id in t.approved_by_managers: continue
            chain_approved = True
            chain_emp = _get_target_chain_starter(t)  # 🌟 FIX
            curr = chain_emp.manager if chain_emp else None
            while curr and curr.id != manager.id:
                if curr.id not in t.approved_by_managers: chain_approved = False; break
                curr = curr.manager
            if chain_approved: pending_targets.append(t)
                
        pfc = FreeQtyClaimMaster.objects.filter(employee__in=team_members, status='Pending_Manager').order_by('created_at')
        for c in pfc:
            if manager.id in c.approved_by_managers: continue
            chain_approved = True
            curr = c.employee.manager
            while curr and curr.id != manager.id:
                if curr.id not in c.approved_by_managers: chain_approved = False; break
                curr = curr.manager
            if chain_approved: pending_free_claims.append(c)

        pending_gift_campaigns = list(GiftCampaignPlan.objects.filter(employee__in=team_members, status='Pending').select_related('employee', 'doctor', 'item'))

    return render(request, 'manager_approvals.html', {
        'pending_mtps': MonthlyTourProgram.objects.filter(employee__in=team_members, status='Pending').order_by('-year', '-month'),
        'pending_doctors': Doctor.objects.filter(allocated_to__in=team_members, status='Pending'),
        'pending_chemists': Chemist.objects.filter(allocated_to__in=team_members, status='Pending'),
        'pending_expenses': MonthlyExpenseReport.objects.filter(employee__in=team_members, status='Pending'),
        'pending_routes': Route.objects.filter(requested_by__in=team_members, status='Pending').select_related('territory', 'requested_by'),
        'pending_holidays': Holiday.objects.filter(company=employee.company, status='Pending') if manager.designation in ['Admin', 'System Administrator'] else [], 
        'pending_targets': pending_targets,
        'pending_free_claims': pending_free_claims,
        'pending_chemist_edits': ChemistEditRequest.objects.filter(employee__in=team_members, status='Pending'),
        'pending_doctor_edits': DoctorEditRequest.objects.filter(employee__in=team_members, status='Pending'),
        'pending_gift_campaigns': pending_gift_campaigns,
        'manager_name': manager.name, 'designation': manager.designation,
        'pending_leaves': LeaveApplication.objects.filter(employee__in=team_members, status='Pending').order_by('start_date'),
    })

@employee_required
def tour_plan_report_view(request, employee):
    team_employees = get_dropdown_team(employee)
    default_emp_id = str(employee.id)
    if employee.designation != 'MR':
        first_sub = team_employees.exclude(id=employee.id).first()
        if first_sub: default_emp_id = str(first_sub.id)
            
    selected_emp_id = request.GET.get('employee_id', default_emp_id)
    selected_emp = get_object_or_404(Employee, id=selected_emp_id)
    
    selected_month = int(request.GET.get('month') or timezone.localdate().month)
    selected_year = int(request.GET.get('year') or timezone.localdate().year)
    
    daily_plans = DailyTourPlan.objects.filter(mtp__employee=selected_emp, date__month=selected_month, date__year=selected_year).select_related('route')
    plan_map = {p.date.day: p.route.name for p in daily_plans if p.route}
    
    mtp = MonthlyTourProgram.objects.filter(employee=selected_emp, month=selected_month, year=selected_year).first()
    mtp_status = mtp.status if mtp else 'Not Created'
    
    days_in_month = calendar.monthrange(selected_year, selected_month)[1]
    report_data = [{'day': d, 'date_str': date(selected_year, selected_month, d).strftime("%d %b, %A"), 'route_name': plan_map.get(d, 'No Plan / Weekly Off')} for d in range(1, days_in_month + 1)]
        
    months = [(i, calendar.month_name[i]) for i in range(1, 13)]
    if request.GET.get('export') == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="TourPlan_{selected_emp.name}_{selected_month}_{selected_year}.csv"'
        writer = csv.writer(response)
        writer.writerow(['Employee Name', 'Month', 'Year', 'MTP Status'])
        writer.writerow([selected_emp.name, selected_month, selected_year, mtp_status])
        writer.writerow([]) 
        writer.writerow(['Date & Day', 'Planned Target Route'])
        for row in report_data: writer.writerow([row['date_str'], row['route_name']])
        return response

    return render(request, 'tour_plan_report.html', {
        'team_employees': team_employees, 'selected_emp_id': int(selected_emp_id),
        'selected_month': selected_month, 'selected_year': selected_year,
        'months': months, 'report_data': report_data, 'selected_employee_name': selected_emp.name, 'mtp_status': mtp_status
    })

@employee_required
def network_report_view(request, employee):
    team_employees = get_dropdown_team(employee)
    default_emp_id = str(employee.id)
    if employee.designation != 'MR':
        first_sub = team_employees.exclude(id=employee.id).first()
        if first_sub: default_emp_id = str(first_sub.id)
            
    selected_emp_id = request.GET.get('employee_id', default_emp_id)
    selected_emp = get_object_or_404(Employee, id=selected_emp_id)
    
    sub_team = get_dropdown_team(selected_emp, ordered=False)
    my_terr_ids = get_team_territory_ids(sub_team)
    routes = get_team_requested_routes(sub_team, my_terr_ids)

    active_tab = request.GET.get('tab', 'doctor')
    route_id = request.GET.get('route', '')
    specialty = request.GET.get('specialty', '')
    category = request.GET.get('category', '')
    
    # 🌟 FIX: Sirf Approved Doctors/Chemists ko Network Report me dikhao
    doctors = Doctor.objects.filter(allocated_to__in=sub_team, status='Approved')
    chemists = Chemist.objects.filter(allocated_to__in=sub_team, status='Approved')
    
    if active_tab == 'doctor':
        if route_id: doctors = doctors.filter(route_id=route_id)
        if specialty: doctors = doctors.filter(specialty=specialty)
        if category: doctors = doctors.filter(category=category)
    elif active_tab == 'chemist':
        if route_id: chemists = chemists.filter(route_id=route_id)
        
    setting, _ = SystemSetting.objects.get_or_create(company=employee.company)

    if request.GET.get('export') == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")

        if active_tab == 'doctor':
            filename = f"Doctor_Network_{selected_emp.name}.xlsx"
            ws.title = "Doctor List"
            ws.append(['MASTER DOCTOR DIRECTORY'])
            ws.append(['Employee:', selected_emp.name])
            ws.append([''])
            ws['A1'].font = Font(bold=True, size=14, color="107C41")
            
            headers = ['Doctor Name', 'Specialty', 'Category', 'Route / Patch', 'Territory / HQ', 'Allocated MR', 'Status']
            ws.append(headers)
            for col_num, cell in enumerate(ws[4], 1):
                cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20

            for doc in doctors:
                terr_name = getattr(doc.territory, 'name', None)
                if not terr_name and doc.route and getattr(doc.route, 'territory', None):
                    terr_name = doc.route.territory.name
                ws.append([f"Dr. {doc.name}", doc.specialty or '-', doc.category or '-', doc.route.name if doc.route else '-', terr_name or '-', doc.allocated_to.name if doc.allocated_to else '-', doc.status or 'Approved'])

        elif active_tab == 'chemist':
            filename = f"Chemist_Network_{selected_emp.name}.xlsx"
            ws.title = "Chemist List"
            ws.append(['MASTER CHEMIST / PHARMACY DIRECTORY'])
            ws.append(['Employee:', selected_emp.name])
            ws.append([''])
            ws['A1'].font = Font(bold=True, size=14, color="107C41")
            
            headers = ['Pharmacy / Chemist Name', 'Route / Patch', 'Territory / HQ', 'Allocated MR', 'Status']
            ws.append(headers)
            for col_num, cell in enumerate(ws[4], 1):
                cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 25

            for chem in chemists:
                terr_name = getattr(chem.territory, 'name', None) if hasattr(chem, 'territory') else None
                if not terr_name and chem.route and getattr(chem.route, 'territory', None): terr_name = chem.route.territory.name
                ws.append([chem.name, chem.route.name if chem.route else '-', terr_name or '-', chem.allocated_to.name if chem.allocated_to else '-', getattr(chem, 'status', 'Approved')])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return render(request, 'network_report.html', {
        'doctors': doctors, 'chemists': chemists, 'routes': routes,
        'selected_emp_id': int(selected_emp_id), 'team_employees': team_employees,
        'is_manager_view': employee.designation != 'MR',
        'active_tab': active_tab, 'selected_route': int(route_id) if route_id else '',
        'selected_specialty': specialty, 'selected_category': category,
        'specialty_choices': Doctor.SPECIALTY_CHOICES, 'category_choices': Doctor.CATEGORY_CHOICES,
        'allow_location_capture': setting.allow_location_capture,
    })
@employee_required
def route_report_view(request, employee):
    team_employees = get_dropdown_team(employee)
    default_emp_id = str(employee.id)
    if employee.designation != 'MR':
        first_sub = team_employees.exclude(id=employee.id).first()
        if first_sub: default_emp_id = str(first_sub.id)
            
    selected_emp_id = request.GET.get('employee_id', default_emp_id)
    selected_emp = get_object_or_404(Employee, id=int(selected_emp_id))
    
    sub_team = get_dropdown_team(selected_emp, ordered=False)
    my_terr_ids = get_team_territory_ids(sub_team)
    routes = get_team_requested_routes(sub_team, my_terr_ids)

    report_data = []
    gt_docs = 0; gt_chems = 0

    for r in routes:
        doc_count = Doctor.objects.filter(route=r, status='Approved').count()
        chem_count = Chemist.objects.filter(route=r, status='Approved').count()
        report_data.append({
            'route_name': r.name, 'territory': r.territory.name if r.territory else 'N/A',
            'category': r.get_category_display() if r.category else 'HQ', 'distance': float(r.distance_from_hq or 0),
            'doc_count': doc_count, 'chem_count': chem_count, 'total_customers': doc_count + chem_count
        })
        gt_docs += doc_count; gt_chems += chem_count

    report_data = sorted(report_data, key=lambda x: (x['territory'], x['route_name']))
    
    if request.GET.get('export') == 'excel':
        filename = f"Route_Report_{selected_emp.name}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Route Coverage"
        ws.append(['ROUTE & CUSTOMER COVERAGE REPORT'])
        ws.append(['Employee:', selected_emp.name])
        ws.append([''])
        ws['A1'].font = Font(bold=True, size=14, color="107C41")
        ws['A2'].font = Font(bold=True)
        
        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        
        headers = ['Route Name', 'Territory / HQ', 'Category', 'Distance (KM)', 'Total Doctors', 'Total Chemists', 'Total Customers']
        ws.append(headers)
        
        for col_num, cell in enumerate(ws[4], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20 if col_num <= 2 else 15
            
        for row in report_data:
            ws.append([row['route_name'], row['territory'], row['category'], row['distance'], row['doc_count'], row['chem_count'], row['total_customers']])
            
        ws.append(['GRAND TOTAL', '', '', '', gt_docs, gt_chems, gt_docs + gt_chems])
        for cell in ws[ws.max_row]: cell.font = Font(bold=True)
            
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return render(request, 'route_report.html', {
        'team_employees': team_employees, 'selected_emp_id': int(selected_emp_id),
        'selected_employee_name': selected_emp.name, 'is_manager_view': employee.designation != 'MR',
        'report_data': report_data, 'gt_docs': gt_docs, 'gt_chems': gt_chems, 'gt_total': gt_docs + gt_chems
    })

@employee_required
def expense_report_view(request, employee):
    team_employees = get_dropdown_team(employee)
    default_emp_id = str(employee.id)
    if employee.designation != 'MR':
        first_sub = team_employees.exclude(id=employee.id).first()
        if first_sub: default_emp_id = str(first_sub.id)
            
    selected_emp_id = request.GET.get('employee_id', default_emp_id)
    selected_emp = get_object_or_404(Employee, id=selected_emp_id)

    selected_month  = int(request.GET.get('month') or timezone.localdate().month)
    selected_year   = int(request.GET.get('year') or timezone.localdate().year)

    expenses = MonthlyExpenseReport.objects.filter(employee=selected_emp, month=selected_month, year=selected_year).order_by('-year', '-month')
    
    day_starts = DayStart.objects.filter(
        employee=selected_emp, date__month=selected_month, date__year=selected_year
    ).prefetch_related('routes', 'territory')
    ds_dict = {ds.date: ds for ds in day_starts}

    for exp in expenses:
        total = 0
        processed_lines = []
        for line in exp.daily_lines.all().order_by('date'):
            app_ta   = line.approved_ta   if line.approved_ta   is not None else line.ta_amount
            app_da   = line.approved_da   if line.approved_da   is not None else line.da_amount
            app_misc = line.approved_misc if line.approved_misc is not None else line.misc_amount
            total += float(app_ta) + float(app_da) + float(app_misc)

            line_date = line.date if hasattr(line.date, 'year') else line.date.date()
            ds = ds_dict.get(line_date)
            location_display = ""
            if ds:
                route_str = ", ".join([r.name for r in ds.routes.all()])
                hq_str = ds.territory.name if ds.territory else ""
                if route_str:
                    location_display = f"{route_str} ({hq_str})" if hq_str else route_str
                elif hq_str:
                    location_display = hq_str

            line.working_area  = location_display or '-'
            line.display_cat   = line.get_territory_category_display() if line.territory_category else 'HQ'
            line.display_km    = float(line.distance_km or 0)
            line.display_ta    = app_ta
            line.display_da    = app_da
            line.display_misc  = app_misc
            line.display_total = round(float(app_ta) + float(app_da) + float(app_misc), 2)
            processed_lines.append(line)
            
        exp.grand_total = round(total, 2)
        exp.processed_lines = processed_lines

    # 🌟 NAYA: PREMIUM EXCEL (.XLSX) EXPORT FOR EXPENSES
    if request.GET.get('export') in ['csv', 'excel']:
        filename = f"Expense_Report_{selected_emp.name}_{selected_month}_{selected_year}.xlsx"
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Expense Report"

        ws.append(['MONTHLY EXPENSE REPORT'])
        ws.append(['Employee:', selected_emp.name, 'Period:', f"{calendar.month_name[selected_month]} {selected_year}"])
        ws.append([''])
        
        ws['A1'].font = Font(bold=True, size=14, color="107C41")
        ws['A2'].font = Font(bold=True); ws['C2'].font = Font(bold=True)

        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")

        headers = ['Date', 'Category', 'Route / HQ', 'Distance (KM)', 'TA (Rs)', 'DA (Rs)', 'Misc (Rs)', 'Total (Rs)', 'Status']
        ws.append(headers)

        for col_num, cell in enumerate(ws[5], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 15
        ws.column_dimensions['C'].width = 30 

        grand_total_all = 0.0
        for exp in expenses:
            for line in exp.processed_lines:
                grand_total_all += line.display_total
                date_str = line.date.strftime('%d-%b-%Y') if line.date else 'N/A'
                ws.append([
                    date_str, line.display_cat, line.working_area, line.display_km,
                    float(line.display_ta), float(line.display_da), float(line.display_misc), line.display_total, exp.status
                ])
                
        ws.append([''])
        gt_row = ['GRAND TOTAL', '', '', '', '', '', '', round(grand_total_all, 2), '']
        ws.append(gt_row)
        for cell in ws[ws.max_row]: cell.font = Font(bold=True)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return render(request, 'expense_report.html', {
        'team_employees': team_employees, 'selected_emp_id': int(selected_emp_id), 'selected_employee_name': selected_emp.name,
        'selected_month': selected_month, 'selected_year': selected_year,
        'months': [(i, calendar.month_name[i]) for i in range(1, 13)],
        'is_manager_view': employee.designation != 'MR', 'expenses': expenses,
    })

@employee_required
def product_master_view(request, employee):
    products = Product.objects.filter(company=employee.company).order_by('name')

    if request.GET.get('export') == 'excel':
        filename = "Product_Master_List.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Product List"
        ws.append(['PRODUCT MASTER REPORT'])
        ws.append(['Generated By:', employee.name])
        ws.append([''])
        ws['A1'].font = Font(bold=True, size=14, color="107C41")
        ws['A2'].font = Font(bold=True)

        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")

        # 🌟 NAYE COLUMNS: MRP, PTR, PTS, GST Slab
        headers = ['Sr No.', 'Product Name', 'Pack Size', 'MRP (₹)', 'PTR (₹)', 'PTS (₹)', 'GST Slab (%)']
        ws.append(headers)
        for col_num, cell in enumerate(ws[4], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 25 if col_num == 2 else 15

        for idx, p in enumerate(products, 1):
            mrp_val = float(p.mrp) if getattr(p, 'mrp', None) else 0.0
            ptr_val = float(p.ptr) if getattr(p, 'ptr', None) else 0.0
            pts_val = float(p.pts) if getattr(p, 'pts', None) else 0.0
            gst_val = p.gst_slab if getattr(p, 'gst_slab', None) is not None else 0
            ws.append([idx, p.name, p.pack_size, mrp_val, ptr_val, pts_val, gst_val])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return render(request, 'product_master.html', {'products': products})

    return render(request, 'product_master.html', {'products': products})

@employee_required
def route_playback_view(request, employee, employee_id, date_str):
    emp = get_object_or_404(Employee, id=employee_id)
    target_date = datetime.strptime(date_str, '%Y-%m-%d').date()

    day_start = DayStart.objects.filter(employee=emp, date=target_date).first()
    day_end = DayEnd.objects.filter(employee=emp, date=target_date).first()
    daily_dcr = DailyDCR.objects.filter(employee=emp, date=target_date).first()

    waypoints = []
    if day_start and day_start.latitude and day_start.longitude:
        waypoints.append({'lat': float(day_start.latitude), 'lng': float(day_start.longitude), 'title': 'Day Start', 'time': day_start.started_at.strftime('%I:%M %p'), 'type': 'start'})
    if daily_dcr:
        visits = daily_dcr.visits.all().order_by('created_at')
        for v in visits:
            if v.latitude and v.longitude:
                name = f"Dr. {v.doctor.name}" if v.doctor else v.chemist.name
                waypoints.append({'lat': float(v.latitude), 'lng': float(v.longitude), 'title': name, 'time': v.created_at.strftime('%I:%M %p'), 'type': 'visit'})
    if day_end and day_end.latitude and day_end.longitude:
        waypoints.append({'lat': float(day_end.latitude), 'lng': float(day_end.longitude), 'title': 'Day End', 'time': day_end.closed_at.strftime('%I:%M %p'), 'type': 'end'})

    return render(request, 'route_playback.html', {'employee': emp, 'target_date': target_date, 'waypoints_json': json.dumps(waypoints), 'has_data': len(waypoints) > 0})


@employee_required
def analysis_hub_view(request, employee):
    team_employees = get_dropdown_team(employee, ordered=False)
    is_manager_view = employee.designation != 'MR'

    default_emp_id = str(employee.id)
    if is_manager_view:
        first_sub = team_employees.exclude(id=employee.id).first()
        if first_sub: default_emp_id = str(first_sub.id)

    selected_emp_id = request.GET.get('employee_id', default_emp_id)
    selected_emp = get_object_or_404(Employee, id=int(selected_emp_id))

    today = timezone.localdate()
    from_month = int(request.GET.get('from_month') or today.month)
    to_month = int(request.GET.get('to_month') or today.month)
    selected_year = int(request.GET.get('year') or today.year)

    if from_month > to_month: from_month, to_month = to_month, from_month

    months_range = list(range(from_month, to_month + 1))
    months_headers = [(m, calendar.month_name[m][:3]) for m in months_range]

    is_mr = selected_emp.designation == 'MR'
    total_doctors = Doctor.objects.filter(allocated_to=selected_emp, status='Approved').count() if is_mr else 0
    total_chemists = Chemist.objects.filter(allocated_to=selected_emp, status='Approved').count() if is_mr else 0

    metrics = {
        'field_days': {'label': '🏃‍♂️ Field Days', 'data': {m: 0 for m in months_range}, 'tot': 0},
        'joint_days': {'label': '🤝 Joint Days', 'data': {m: 0 for m in months_range}, 'tot': 0},
        'meeting_transit': {'label': '🏢 Meet/Transit', 'data': {m: 0 for m in months_range}, 'tot': 0},
        'leaves': {'label': '🏖️ Leaves', 'data': {m: 0 for m in months_range}, 'tot': 0},
        'holidays': {'label': '⛱️ Holidays', 'data': {m: 0 for m in months_range}, 'tot': 0},
        'pob': {'label': '💰 Tot POB', 'data': {m: 0 for m in months_range}, 'tot': 0.0, 'is_currency': True},
        'dr_avg': {'label': '👨‍⚕️ Dr Avg', 'data': {m: 0 for m in months_range}, 'tot': 0.0, 'is_avg': True},
        'chem_avg': {'label': '💊 Chem Avg', 'data': {m: 0 for m in months_range}, 'tot': 0.0, 'is_avg': True},
    }

    if is_mr:
        metrics['dr_in_list'] = {'label': '📋 Tot Docs in List', 'data': {m: 0 for m in months_range}, 'tot': total_doctors}
        metrics['dr_visited'] = {'label': '✅ Visited Docs', 'data': {m: 0 for m in months_range}, 'tot': 0}
        metrics['dr_cov'] = {'label': '🎯 Dr Cov %', 'data': {m: 0 for m in months_range}, 'tot': 0.0, 'is_pct': True}
        
        metrics['chem_in_list'] = {'label': '📋 Tot Chems in List', 'data': {m: 0 for m in months_range}, 'tot': total_chemists}
        metrics['chem_visited'] = {'label': '✅ Visited Chems', 'data': {m: 0 for m in months_range}, 'tot': 0}
        metrics['chem_cov'] = {'label': '🎯 Chem Cov %', 'data': {m: 0 for m in months_range}, 'tot': 0.0, 'is_pct': True}

    gt_dr_visits, gt_chem_visits, gt_field_days = 0, 0, 0
    gt_visited_docs, gt_visited_chems = set(), set()
    joint_workers = {}

    for m in months_range:
        num_days = calendar.monthrange(selected_year, m)[1]
        
        # Day Start with Joint Worker details
        day_starts = DayStart.objects.filter(employee=selected_emp, date__month=m, date__year=selected_year).select_related('joint_worked_with')
        
        f_days = 0; j_days = 0; m_t_days = 0; h_days = 0
        for ds in day_starts:
            if ds.work_type == 'Field Work':
                f_days += 1
                if ds.joint_worked_with:
                    j_days += 1
                    jw_name = ds.joint_worked_with.name
                    # Dynamic Bifurcation of Joint Work
                    if jw_name not in joint_workers:
                        joint_workers[jw_name] = {'label': f'↳ Joint: {jw_name}', 'data': {mm: 0 for mm in months_range}, 'tot': 0, 'is_sub_row': True}
                    joint_workers[jw_name]['data'][m] += 1
                    joint_workers[jw_name]['tot'] += 1
            elif ds.work_type in ['Meeting', 'Transit']:
                m_t_days += 1
            elif ds.work_type == 'Holiday':
                h_days += 1
        
        l_days = 0
        leaves = LeaveApplication.objects.filter(
            employee=selected_emp, status='Approved', start_date__lte=date(selected_year, m, num_days), end_date__gte=date(selected_year, m, 1)
        )
        for l in leaves:
            start = max(l.start_date, date(selected_year, m, 1))
            end = min(l.end_date, date(selected_year, m, num_days))
            l_days += (end - start).days + 1
            
        dcrs = DailyDCR.objects.filter(employee=selected_emp, date__month=m, date__year=selected_year).prefetch_related('visits__doctor', 'visits__chemist', 'visits__product_details__product')
        m_dr_visits, m_chem_visits, m_pob = 0, 0, 0.0
        m_visited_docs, m_visited_chems = set(), set()
        
        for dcr in dcrs:
            for v in dcr.visits.all():
                if v.doctor:
                    m_dr_visits += 1; m_visited_docs.add(v.doctor_id); gt_visited_docs.add(v.doctor_id)
                if v.chemist:
                    m_chem_visits += 1; m_visited_chems.add(v.chemist_id); gt_visited_chems.add(v.chemist_id)
                for pd in v.product_details.all():
                    price = float(pd.product.price) if getattr(pd.product, 'price', None) else 0.0
                    m_pob += pd.order_qty * price
                    
        gt_dr_visits += m_dr_visits; gt_chem_visits += m_chem_visits; gt_field_days += f_days
        
        metrics['field_days']['data'][m] = f_days; metrics['field_days']['tot'] += f_days
        metrics['joint_days']['data'][m] = j_days; metrics['joint_days']['tot'] += j_days
        metrics['meeting_transit']['data'][m] = m_t_days; metrics['meeting_transit']['tot'] += m_t_days
        metrics['leaves']['data'][m] = l_days; metrics['leaves']['tot'] += l_days
        metrics['holidays']['data'][m] = h_days; metrics['holidays']['tot'] += h_days
        metrics['pob']['data'][m] = m_pob; metrics['pob']['tot'] += m_pob
        metrics['dr_avg']['data'][m] = round(m_dr_visits / f_days, 1) if f_days > 0 else 0
        metrics['chem_avg']['data'][m] = round(m_chem_visits / f_days, 1) if f_days > 0 else 0
        
        if is_mr:
            metrics['dr_in_list']['data'][m] = total_doctors
            metrics['dr_visited']['data'][m] = len(m_visited_docs)
            metrics['dr_cov']['data'][m] = round((len(m_visited_docs) / total_doctors * 100), 1) if total_doctors > 0 else 0
            
            metrics['chem_in_list']['data'][m] = total_chemists
            metrics['chem_visited']['data'][m] = len(m_visited_chems)
            metrics['chem_cov']['data'][m] = round((len(m_visited_chems) / total_chemists * 100), 1) if total_chemists > 0 else 0

    metrics['dr_avg']['tot'] = round(gt_dr_visits / gt_field_days, 1) if gt_field_days > 0 else 0
    metrics['chem_avg']['tot'] = round(gt_chem_visits / gt_field_days, 1) if gt_field_days > 0 else 0
    
    if is_mr:
        metrics['dr_in_list']['tot'] = total_doctors
        metrics['dr_visited']['tot'] = len(gt_visited_docs)
        metrics['dr_cov']['tot'] = round((len(gt_visited_docs) / total_doctors * 100), 1) if total_doctors > 0 else 0
        
        metrics['chem_in_list']['tot'] = total_chemists
        metrics['chem_visited']['tot'] = len(gt_visited_chems)
        metrics['chem_cov']['tot'] = round((len(gt_visited_chems) / total_chemists * 100), 1) if total_chemists > 0 else 0

    report_data = []
    def add_metric(key):
        if key in metrics:
            item = metrics[key]
            item['monthly_list'] = [item['data'][m] for m in months_range]
            report_data.append(item)

    # Maintain strict professional order
    add_metric('field_days')
    add_metric('joint_days')
    
    # 🌟 NAYA: Joint workers details will slide exactly below Joint Days
    for jw_name in sorted(joint_workers.keys()):
        jw_item = joint_workers[jw_name]
        jw_item['monthly_list'] = [jw_item['data'][m] for m in months_range]
        report_data.append(jw_item)
        
    add_metric('meeting_transit')
    add_metric('leaves')
    add_metric('holidays')
    add_metric('pob')
    
    if is_mr:
        add_metric('dr_in_list')
        add_metric('dr_visited')
        add_metric('dr_avg')
        add_metric('dr_cov')
        add_metric('chem_in_list')
        add_metric('chem_visited')
        add_metric('chem_avg')
        add_metric('chem_cov')
    else:
        add_metric('dr_avg')
        add_metric('chem_avg')

    if request.GET.get('export') == 'excel':
        filename = f"Analytical_Matrix_{selected_emp.name}_M{from_month}-M{to_month}_{selected_year}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Analytics Matrix"
        
        period_str = f"{calendar.month_name[from_month][:3]} to {calendar.month_name[to_month][:3]} {selected_year}"
        ws.append(['ANALYTICAL HUB - PERFORMANCE TREND MATRIX'])
        ws.append(['Employee:', selected_emp.name, 'Period:', period_str])
        ws.append([''])
        ws['A1'].font = Font(bold=True, size=14, color="107C41")
        ws['A2'].font = Font(bold=True); ws['C2'].font = Font(bold=True)
        
        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        
        headers = ['Key Performance Indicator (KPI)']
        for m_num, m_name in months_headers: headers.append(m_name)
        headers.append('Grand Total / Average')
        ws.append(headers)
        
        for col_num, cell in enumerate(ws[5], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 35 if col_num == 1 else 18
            
        for row in report_data:
            clean_label = row['label'].split(' ', 1)[1] if ' ' in row['label'] else row['label']
            row_data = [clean_label]
            for val in row['monthly_list']:
                if row.get('is_currency'): val = f"₹{round(val, 2)}"
                elif row.get('is_pct'): val = f"{val}%"
                row_data.append(val)
                
            tot = row['tot']
            if row.get('is_currency'): tot = f"₹{round(tot, 2)}"
            elif row.get('is_pct'): tot = f"{tot}%"
            row_data.append(tot)
            ws.append(row_data)
            
        for cell in ws[ws.max_row]: cell.font = Font(bold=True)
            
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return render(request, 'analysis_hub.html', {
        'team_employees': team_employees, 'selected_emp_id': int(selected_emp_id), 'selected_emp': selected_emp,
        'from_month': from_month, 'to_month': to_month, 'selected_year': selected_year,
        'months_choices': [(i, calendar.month_name[i]) for i in range(1, 13)], 'months_headers': months_headers,
        'report_data': report_data, 'is_manager_view': is_manager_view
    })


@employee_required
def free_claim_view(request, employee):
    # 🌟 MANAGER OVERRIDE LOGIC
    team_employees = get_dropdown_team(employee)
    is_manager_view = employee.designation != 'MR'

    default_emp_id = str(employee.id)
    if is_manager_view:
        first_sub = team_employees.exclude(id=employee.id).first()
        if first_sub:
            default_emp_id = str(first_sub.id)

    selected_emp_id = request.GET.get('employee_id') or request.POST.get('employee_id') or default_emp_id
    selected_emp = get_object_or_404(Employee, id=int(selected_emp_id))

    selected_month = int(request.GET.get('month') or timezone.localdate().month)
    selected_year = int(request.GET.get('year') or timezone.localdate().year)

    # 🌟 STRICT BLOCKER: Sirf turant-pichle mahine ke liye deadline-day tak, baaki purani months hamesha lock (Admin override sirf submit ke waqt)
    today = timezone.localdate()
    setting = SystemSetting.objects.filter(company=employee.company).first()
    deadline = setting.free_claim_deadline_day if setting and setting.free_claim_deadline_day else 4
    prev_month, prev_year = (12, today.year - 1) if today.month == 1 else (today.month - 1, today.year)
    is_immediate_prev_month = (selected_month == prev_month and selected_year == prev_year)
    is_locked = (today.day > deadline) if is_immediate_prev_month else True

    # 🌟 NAYA: Fetch Stockists for the selected MR
    my_terr_ids = [selected_emp.headquarter_id] if selected_emp.headquarter_id else []
    available_stockists = Stockist.objects.filter(territory_id__in=my_terr_ids).order_by('name')

    selected_stockist_id = request.GET.get('stockist_id') or request.POST.get('stockist_id')
    if selected_stockist_id and not available_stockists.filter(id=selected_stockist_id).exists():
        selected_stockist_id = None
    if not selected_stockist_id and available_stockists.exists():
        selected_stockist_id = str(available_stockists.first().id)

    selected_stockist = available_stockists.filter(id=selected_stockist_id).first()

    master = None
    if selected_stockist:
        # Ab master report Stockist-wise dhoondhega
        master = FreeQtyClaimMaster.objects.filter(employee=selected_emp, stockist=selected_stockist, month=selected_month, year=selected_year).first()

    # 🌟 INDENT FIX: Ye POST block ab 'if selected_stockist:' ke bahar same level par aa gaya hai
    if request.method == "POST":
        action = request.POST.get('action')
        
        # 🌟 NAYA RULE: Current month ka claim generate/submit hone se rokein
        if selected_year > today.year or (selected_year == today.year and selected_month >= today.month):
            messages.error(request, "⚠️ The Free Claim for the current month cannot be generated yet. Please wait for the secondary sale to be uploaded after the month ends.")
            return redirect(f"{request.path}?employee_id={selected_emp.id}&stockist_id={selected_stockist_id}&month={selected_month}&year={selected_year}")      
        
        # 🌟 STRICT BLOCKER: Deadline ke baad lock (is_locked upar pehle se compute ho chuka hai)
        if is_locked and employee.designation not in ['Admin', 'System Administrator']:
            messages.error(request, f"⚠️ Edit Locked! Free Claim entries are only allowed until the {deadline} of each month. Please contact Admin.")
            return redirect(f"{request.path}?employee_id={selected_emp.id}&stockist_id={selected_stockist_id}&month={selected_month}&year={selected_year}")
        
        if not selected_stockist:
            messages.error(request, "⚠️ Stockist is missing!")
            return redirect(request.path)

        # 🚀 ACTION 1: AUTO-GENERATE / RE-GENERATE REPORT
        if action == 'generate':
            if master and master.status not in ['Draft', 'Rejected']:
                messages.error(request, f"⚠️ This report is in '{master.status}' state. It can no longer be regenerated.")
            else:
                sales = PartyWiseSaleLine.objects.filter(
                    report__employee=selected_emp,
                    report__stockist=selected_stockist, # 🌟 Sirf is stockist ka data
                    report__month=selected_month,
                    report__year=selected_year,
                    free_qty__gt=0
                ).values('product_id').annotate(
                    tot_billed=Sum('billed_qty'),
                    tot_free=Sum('free_qty')
                )

                if not sales:
                    # 🌟 ORPHAN CLEANUP PATCH: Agar sale udayi gayi hai, toh purana claim bhi uda do!
                    if master and master.status in ['Draft', 'Rejected']:
                        master.delete()
                        master = None
                    messages.warning(request, f"There is no free scheme (secondary sale) entry for {selected_stockist.name} this month. If there was an old Draft, it has been cleared as well.")
                else:
                    if not master:
                        master = FreeQtyClaimMaster.objects.create(
                            employee=selected_emp, stockist=selected_stockist, month=selected_month, year=selected_year, status='Draft'
                        )
                    else:
                        master.claim_lines.all().delete()
                        master.status = 'Draft'
                        master.save()

                    for s in sales:
                        prod = Product.objects.get(id=s['product_id'])
                        price = float(prod.price) if getattr(prod, 'price', None) else 0.0
                        val = s['tot_free'] * price

                        FreeQtyClaimLine.objects.create(
                            master=master, stockist=selected_stockist, product=prod,
                            total_billed_qty=s['tot_billed'], total_free_qty=s['tot_free'], claim_value=val
                        )
                    messages.success(request, f"🎉 {selected_stockist.name}'s Free Claim Report has been synced successfully!")

        # 🚀 ACTION 2: SUBMIT TO MANAGER
        elif action == 'submit':
            if master and master.status in ['Draft', 'Rejected']:
                master.status = 'Pending_Manager'
                master.save()
                messages.success(request, f"✅ Claim for {selected_stockist.name} submitted for approval!")

        return redirect(f"{request.path}?employee_id={selected_emp.id}&stockist_id={selected_stockist_id}&month={selected_month}&year={selected_year}")

    lines = []
    grand_total = 0
    if master:
        lines = master.claim_lines.select_related('product').order_by('product__name')
        grand_total = sum(l.claim_value for l in lines)

    return render(request, 'free_claim_report.html', {
        'team_employees': team_employees,
        'is_manager_view': is_manager_view,
        'selected_emp_id': int(selected_emp_id),
        'selected_emp': selected_emp,
        'selected_month': selected_month,
        'selected_year': selected_year,
        'available_stockists': available_stockists,
        'selected_stockist_id': int(selected_stockist_id) if selected_stockist_id else '',
        'months_choices': [(i, calendar.month_name[i]) for i in range(1, 13)],
        'master': master,
        'lines': lines,
        'grand_total': grand_total,
        'is_locked': is_locked, 'deadline': deadline  # Frontend ko status bheja
    })


@employee_required
def free_claim_view_readonly(request, employee):
    """
    🌟 NAYA: View Hub ke 'Free Claims' card ke liye — bilkul same
    employee/stockist/month/year select karne wala UI jo Request Hub ke
    'free_claim_view' (Generate/Sync/Submit) me hai, par SIRF DEKHNE ke
    liye — koi POST/Generate/Sync/Submit action yahan nahi hota.
    Jab tak claim Approved na ho, isi page se uska latest status/data
    dekha ja sakta hai (Request Hub me jaake generate/sync karna padega).
    """
    team_employees = get_dropdown_team(employee)
    is_manager_view = employee.designation != 'MR'

    default_emp_id = str(employee.id)
    if is_manager_view:
        first_sub = team_employees.exclude(id=employee.id).first()
        if first_sub:
            default_emp_id = str(first_sub.id)

    selected_emp_id = request.GET.get('employee_id') or default_emp_id
    selected_emp = get_object_or_404(Employee, id=int(selected_emp_id))

    selected_month = int(request.GET.get('month') or timezone.localdate().month)
    selected_year = int(request.GET.get('year') or timezone.localdate().year)

    my_terr_ids = [selected_emp.headquarter_id] if selected_emp.headquarter_id else []
    available_stockists = Stockist.objects.filter(territory_id__in=my_terr_ids).order_by('name')

    selected_stockist_id = request.GET.get('stockist_id')
    if selected_stockist_id and not available_stockists.filter(id=selected_stockist_id).exists():
        selected_stockist_id = None
    if not selected_stockist_id and available_stockists.exists():
        selected_stockist_id = str(available_stockists.first().id)

    selected_stockist = available_stockists.filter(id=selected_stockist_id).first()

    master = None
    if selected_stockist:
        master = FreeQtyClaimMaster.objects.filter(
            employee=selected_emp, stockist=selected_stockist, month=selected_month, year=selected_year
        ).first()

    lines = []
    grand_total = 0
    if master:
        lines = master.claim_lines.select_related('product').order_by('product__name')
        grand_total = sum(l.claim_value for l in lines)

    return render(request, 'free_claim_view_readonly.html', {
        'team_employees': team_employees,
        'is_manager_view': is_manager_view,
        'selected_emp_id': int(selected_emp_id),
        'selected_emp': selected_emp,
        'selected_month': selected_month,
        'selected_year': selected_year,
        'available_stockists': available_stockists,
        'selected_stockist_id': int(selected_stockist_id) if selected_stockist_id else '',
        'months_choices': [(i, calendar.month_name[i]) for i in range(1, 13)],
        'master': master,
        'lines': lines,
        'grand_total': grand_total,
    })


@employee_required
def approve_free_claims_view(request, employee):
    if employee.designation == 'MR':
        messages.error(request, "You do not have permission to access this page.")
        return redirect('request_hub')

    if request.method == "POST":
        claim_id = request.POST.get('claim_id')
        action = request.POST.get('action')
        remark = request.POST.get('remark', '')
        claim = get_object_or_404(FreeQtyClaimMaster, id=claim_id)

        if action == 'approve':
            if employee.designation == 'Admin':
                claim.status = 'Approved'; claim.admin_remark = remark
            else:
                mgr_list = list(claim.approved_by_managers)
                if employee.id not in mgr_list: mgr_list.append(employee.id)
                claim.approved_by_managers = mgr_list
                claim.manager_remark = remark
                claim.status = 'Pending_Manager' if employee.manager_id else 'Pending_Admin'
            claim.save()
            messages.success(request, f"✅ {claim.employee.name}'s Free Claim Report has been approved!")
        elif action == 'reject':
            claim.status = 'Rejected'
            if employee.designation == 'Admin': claim.admin_remark = remark
            else: claim.manager_remark = remark
            claim.save()
            messages.error(request, f"❌ Claim has been rejected.")
        return redirect('approve_free_claims')

    if employee.designation == 'Admin':
        pending_claims = FreeQtyClaimMaster.objects.filter(employee__company=employee.company, status='Pending_Admin').order_by('-created_at')
        history_claims = FreeQtyClaimMaster.objects.filter(employee__company=employee.company, status__in=['Approved', 'Rejected']).order_by('-updated_at')[:50]

    else:
        team = get_full_team_employees(employee).exclude(id=employee.id)
        pending_claims = FreeQtyClaimMaster.objects.filter(status='Pending_Manager', employee__in=team).order_by('-created_at')
        history_claims = FreeQtyClaimMaster.objects.filter(status__in=['Approved', 'Rejected', 'Pending_Admin'], employee__in=team).order_by('-updated_at')[:50]

    return render(request, 'approve_free_claims.html', {'pending_claims': pending_claims, 'history_claims': history_claims, 'is_admin': employee.designation == 'Admin'})

@employee_required
def mr_inventory_view(request, employee):
    team_employees = get_dropdown_team(employee)
    is_manager_view = employee.designation != 'MR'

    # =======================================================
    # 🌟 NAYA FIX: IN-TRANSIT STOCK RECEIVE KARNE KA LOGIC
    # =======================================================
    if request.method == "POST" and "receive_dispatch" in request.POST:
        dispatch_id = request.POST.get('dispatch_id')
        if dispatch_id:
            try:
                # 1. Dispatch record dhoondho aur 'Received' mark karo
                dispatch = PromoDispatch.objects.get(id=dispatch_id, employee=employee, status='In-Transit')
                dispatch.status = 'Received'
                dispatch.save()
                
                # 2. MR Inventory mein quantity plus (+) karo
                inv_obj, created = MRInventory.objects.get_or_create(
                    employee=employee, 
                    item=dispatch.item,
                    defaults={'stock_qty': 0}
                )
                inv_obj.stock_qty += dispatch.quantity
                inv_obj.save()
                
                messages.success(request, f"✅ Stock of {dispatch.item.name} (Qty: {dispatch.quantity}) has been added to your inventory!")
            except PromoDispatch.DoesNotExist:
                messages.error(request, "⚠️ This dispatch has either already been received, or the ID is invalid.")
                
        # Redirect taaki form resubmit na ho
        return redirect(f"{request.path}?employee_id={employee.id}")

    # =======================================================
    # BAAKI KA ORIGINAL LOGIC (Yahan se neeche sab same hai)
    # =======================================================
    default_emp_id = str(employee.id)
    if is_manager_view:
        first_sub = team_employees.exclude(id=employee.id).first()
        if first_sub: default_emp_id = str(first_sub.id)

    selected_emp_id = request.GET.get('employee_id', default_emp_id)
    selected_emp = get_object_or_404(Employee, id=int(selected_emp_id))

    in_transit_items = PromoDispatch.objects.filter(employee=selected_emp, status='In-Transit').order_by('-dispatch_date')
    my_stock_qs = MRInventory.objects.filter(employee=selected_emp).select_related('item').order_by('item__item_type', 'item__name')

    # 🌟 FIX: pehle sab kuch (Sample/Routine/HighValue) ek hi 'inventory_data'
    # list me mix ho jata tha. Ab 2 alag lists:
    #   - sample_gift_data -> "Sample and Gift" card (HighValue EXCLUDE)
    #   - hv_stock_data    -> "HV Gifts" card (sirf HighValue, stock-table
    #     format me — doctor-wise nahi, jaisa Stock card dikhata tha)
    sample_gift_data = []
    hv_stock_data = []
    for stock in my_stock_qs:
        total_rec = PromoDispatch.objects.filter(employee=selected_emp, item=stock.item, status='Received').aggregate(total=Sum('quantity'))['total'] or 0
        distributed = total_rec - stock.stock_qty
        if distributed < 0: distributed = 0
        if total_rec > 0 or stock.stock_qty > 0 or distributed > 0:
            row = {'item_name': stock.item.name, 'category': stock.item.item_type, 'received': total_rec, 'distributed': distributed, 'balance': stock.stock_qty}
            if stock.item.item_type == 'HighValue':
                hv_stock_data.append(row)
            else:
                sample_gift_data.append(row)

    hv_plans = GiftCampaignPlan.objects.filter(employee=selected_emp, item__item_type='HighValue').select_related('doctor', 'item').order_by('-year', '-month')
    hv_tracker = []
    for plan in hv_plans:
        ledger_entry = DoctorROILedger.objects.filter(employee=selected_emp, doctor=plan.doctor, item=plan.item).order_by('-date_given').first()
        dist_date = ledger_entry.date_given if ledger_entry else None
        dist_qty = ledger_entry.quantity if ledger_entry else 0
        
        if dist_qty >= 1: status_badge = 'Delivered'
        elif plan.status == 'Approved': status_badge = 'Pending Delivery'
        else: status_badge = plan.status

        hv_tracker.append({'item_name': plan.item.name, 'doctor_name': plan.doctor.name, 'specialty': plan.doctor.specialty or '-', 'month_year': f"{calendar.month_name[plan.month][:3]} {plan.year}", 'plan_status': plan.status, 'dist_qty': dist_qty, 'dist_date': dist_date, 'final_status': status_badge})

    recent_logs = DoctorROILedger.objects.filter(employee=selected_emp).select_related('doctor', 'item').order_by('-date_given')[:100]

    if request.GET.get('export') == 'excel':
        filename = f"Inventory_Report_{selected_emp.name}.xlsx"
        wb = openpyxl.Workbook()
        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")

        ws1 = wb.active
        ws1.title = "Stock Summary"
        ws1.append(['INVENTORY SUMMARY REPORT'])
        ws1.append(['Employee:', selected_emp.name])
        ws1.append([''])
        ws1['A1'].font = Font(bold=True, size=14, color="107C41")
        ws1.append(['Item Name', 'Category', 'Total Received', 'Total Distributed', 'Current Balance (Pending)'])
        for col_num, cell in enumerate(ws1[4], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws1.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20
        for row in sample_gift_data: ws1.append([row['item_name'], row['category'], row['received'], row['distributed'], row['balance']])

        ws2 = wb.create_sheet(title="High Value Tracker")
        ws2.append(['HIGH VALUE GIFT TRACKER (PROPOSED vs ACTUAL)'])
        ws2.append(['Employee:', selected_emp.name])
        ws2.append([''])
        ws2['A1'].font = Font(bold=True, size=14, color="107C41")
        ws2.append(['Doctor Name', 'Specialty', 'Item Name', 'Plan Month', 'Approval Status', 'Distributed Qty', 'Distributed Date', 'Execution Status'])
        for col_num, cell in enumerate(ws2[4], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws2.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 18
        for row in hv_tracker: ws2.append([f"Dr. {row['doctor_name']}", row['specialty'], row['item_name'], row['month_year'], row['plan_status'], row['dist_qty'], row['dist_date'] or '-', row['final_status']])

        ws3 = wb.create_sheet(title="Distribution Log")
        ws3.append(['RECENT DISTRIBUTION LOG'])
        ws3.append(['Employee:', selected_emp.name])
        ws3.append([''])
        ws3['A1'].font = Font(bold=True, size=14, color="107C41")
        ws3.append(['Date Given', 'Doctor Name', 'Item Name', 'Category', 'Quantity Distributed', 'Value (₹)'])
        for col_num, cell in enumerate(ws3[4], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws3.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 18
        for log in recent_logs: ws3.append([log.date_given, f"Dr. {log.doctor.name}", log.item.name, log.item.item_type, log.quantity, float(log.total_value)])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return render(request, 'mr_inventory.html', {
        'team_employees': team_employees, 'selected_emp_id': int(selected_emp_id),
        'sample_gift_data': sample_gift_data, 'hv_stock_data': hv_stock_data,  # 🌟 FIX: naye split fields
        'hv_tracker': hv_tracker,
        'recent_logs': recent_logs, 'in_transit_items': in_transit_items, 'is_manager_view': is_manager_view
    })

@employee_required
def gift_distribution_report(request, employee):
    team_employees = get_dropdown_team(employee)
    is_manager_view = employee.designation != 'MR'

    today = timezone.localdate()
    
    # --- STRONG DATE PARSING ---
    try:
        raw_from = request.GET.get('from_month')
        from_month = int(raw_from) if raw_from else today.month
    except (TypeError, ValueError):
        from_month = today.month

    try:
        raw_to = request.GET.get('to_month')
        to_month = int(raw_to) if raw_to else today.month
    except (TypeError, ValueError):
        to_month = today.month

    try:
        raw_year = request.GET.get('year')
        selected_year = int(raw_year) if raw_year else today.year
    except (TypeError, ValueError):
        selected_year = today.year
        
    if from_month > to_month:
        from_month, to_month = to_month, from_month

    selected_emp_id = request.GET.get('employee_id', '')
    inv_type = request.GET.get('inv_type', 'any_gift')

    # ==========================================
    # 1. DISTRIBUTED GIFTS LOGIC
    # ==========================================
    if is_manager_view and selected_emp_id: 
        qs = DoctorROILedger.objects.filter(employee_id=selected_emp_id)
        plan_qs = GiftCampaignPlan.objects.filter(employee_id=selected_emp_id)
    elif is_manager_view: 
        qs = DoctorROILedger.objects.filter(employee__in=team_employees)
        plan_qs = GiftCampaignPlan.objects.filter(employee__in=team_employees)
    else: 
        qs = DoctorROILedger.objects.filter(employee=employee)
        plan_qs = GiftCampaignPlan.objects.filter(employee=employee)

    qs = qs.filter(
        date_given__month__gte=from_month,
        date_given__month__lte=to_month,
        date_given__year=selected_year
    ).select_related('doctor', 'employee', 'item').order_by('-date_given')

    if inv_type == 'high_value':
        qs = qs.filter(item__item_type='HighValue')
    elif inv_type == 'normal':
        qs = qs.filter(item__item_type='Routine')

    total_qty = qs.aggregate(t_qty=Sum('quantity'))['t_qty'] or 0
    total_val = qs.aggregate(t_val=Sum('total_value'))['t_val'] or 0

    # ==========================================
    # 2. 🌟 PENDING HV GIFTS LOGIC
    # ==========================================
    # Jo approved hain par unka ledger entry (distribution) nahi hua
    plan_qs = plan_qs.filter(
        month__gte=from_month,
        month__lte=to_month,
        year=selected_year,
        status='Approved'
    ).select_related('doctor', 'item', 'employee')
    
    # Delivered entries nikal lo taaki unko filter out kar sakein
    delivered_combinations = set(
        qs.filter(item__item_type='HighValue').values_list('employee_id', 'doctor_id', 'item_id')
    )
    
    pending_hv_list = []
    for plan in plan_qs:
        if (plan.employee_id, plan.doctor_id, plan.item_id) not in delivered_combinations:
            pending_hv_list.append(plan)

    # ==========================================
    # EXCEL EXPORT (Same as before)
    # ==========================================
    if request.GET.get('export') == 'excel':
        filename = f"Gift_Distribution_{calendar.month_abbr[from_month]}_to_{calendar.month_abbr[to_month]}_{selected_year}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Gift Distribution"
        # ... (Excel Logic as it is)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return render(request, 'gift_distribution_report.html', {
        'reports': qs, 
        'pending_hv_list': pending_hv_list,  # 🌟 UI KE LIYE NAYA VARIABLE BHEJA HAI
        'team_employees': team_employees, 
        'is_manager_view': is_manager_view,
        'from_month': from_month,
        'to_month': to_month,
        'selected_year': selected_year,
        'selected_emp_id': int(selected_emp_id) if selected_emp_id else '',
        'total_qty': total_qty, 
        'total_val': total_val, 
        'months_choices': [(i, calendar.month_name[i]) for i in range(1, 13)],
        'inv_type': inv_type,
    })

@login_required
def doctor_roi_report(request):
    # 🌟 FIX: Employee ko request se fetch kar liya
    employee = request.user.employee 
    
    team_employees = get_dropdown_team(employee)
    is_manager_view = employee.designation != 'MR'

    today = timezone.localdate()
    
    
    # --- STRONG DATE PARSING ---
    try:
        raw_from = request.GET.get('from_month')
        from_month = int(raw_from) if raw_from else today.month
    except (TypeError, ValueError):
        from_month = today.month

    try:
        raw_to = request.GET.get('to_month')
        to_month = int(raw_to) if raw_to else today.month
    except (TypeError, ValueError):
        to_month = today.month

    try:
        raw_year = request.GET.get('year')
        selected_year = int(raw_year) if raw_year else today.year
    except (TypeError, ValueError):
        selected_year = today.year
        
    if from_month > to_month:
        from_month, to_month = to_month, from_month
        
    months_range = list(range(from_month, to_month + 1))
    months_headers = [(m, calendar.month_name[m][:3]) for m in months_range]

    selected_emp_id = request.GET.get('employee_id', '')
    inv_type = request.GET.get('inv_type', 'any_gift')

    # 1. Base Query for Investment (Gifts)
    if is_manager_view and selected_emp_id: 
        qs = DoctorROILedger.objects.filter(employee_id=selected_emp_id)
        base_docs = Doctor.objects.filter(allocated_to_id=selected_emp_id)
    elif is_manager_view: 
        qs = DoctorROILedger.objects.filter(employee__in=team_employees)
        base_docs = Doctor.objects.filter(allocated_to__in=team_employees)
    else: 
        qs = DoctorROILedger.objects.filter(employee=employee)
        base_docs = Doctor.objects.filter(allocated_to=employee)

    qs = qs.filter(
        date_given__month__gte=from_month,
        date_given__month__lte=to_month,
        date_given__year=selected_year
    ).select_related('doctor', 'employee', 'item')

    if inv_type == 'high_value':
        qs = qs.filter(item__item_type='HighValue')
    elif inv_type == 'normal':
        qs = qs.filter(item__item_type='Routine')

    # 🌟 NAYA LOGIC: Group by DOCTOR only (Not Doctor+Item)
    agg_data = defaultdict(lambda: {
        'doc_name': '', 'items': set(), 'emp_name': '', 'doc_id': 0,
        'monthly': {m: {'inv_qty': 0, 'inv_val': 0.0} for m in months_range},
        'tot_inv_qty': 0, 'tot_inv_val': 0.0
    })
    
    for row in qs:
        m = row.date_given.month
        key = row.doctor_id # Bug fix: Strictly grouping by Doctor ID
        
        if not agg_data[key]['doc_name']:
            agg_data[key]['doc_name'] = row.doctor.name
            agg_data[key]['emp_name'] = row.employee.name if row.employee else 'Unknown'
            agg_data[key]['doc_id'] = row.doctor_id
            
        agg_data[key]['items'].add((row.item.name, row.item.item_type))
        agg_data[key]['monthly'][m]['inv_qty'] += row.quantity
        agg_data[key]['monthly'][m]['inv_val'] += float(row.total_value)
        agg_data[key]['tot_inv_qty'] += row.quantity
        agg_data[key]['tot_inv_val'] += float(row.total_value)

    # 🌟 NAYA LOGIC: "All Doctors" filter ab sach me sabko dikhayega (0 investment wale bhi)
    if inv_type == 'all_doctors':
        for doc in base_docs.select_related('allocated_to'):
            if doc.id not in agg_data:
                agg_data[doc.id]['doc_name'] = doc.name
                agg_data[doc.id]['emp_name'] = doc.allocated_to.name if doc.allocated_to else 'Unknown'
                agg_data[doc.id]['doc_id'] = doc.id
                agg_data[doc.id]['items'].add(('No Gifts', 'Routine')) # Taki table UI render ho sake

    # 🌟 RX CALCULATION
    doc_ids = set(agg_data.keys())
    rx_qs = DoctorRxMapping.objects.filter(
        doctor_id__in=doc_ids,
        party_line__report__month__gte=from_month,
        party_line__report__month__lte=to_month,
        party_line__report__year=selected_year
    ).select_related('party_line__product')
    
    rx_dict = defaultdict(lambda: defaultdict(float))
    for rx in rx_qs:
        m = rx.party_line.report.month
        price = float(rx.party_line.product.price) if getattr(rx.party_line.product, 'price', None) else 0.0
        rx_dict[rx.doctor_id][m] += (rx.mapped_billed_qty * price)

    # 🌟 MATRIX DATA FINALIZATION
    report_data = []
    gt_monthly = {m: {'inv_qty': 0, 'inv_val': 0.0, 'rx_val': 0.0} for m in months_range}
    gt_inv_qty, gt_inv_val, gt_rx_val = 0, 0.0, 0.0
    
    for doc_id, data in agg_data.items():
        monthly_list = []
        doc_tot_rx = 0.0
        
        for m in months_range:
            i_q = data['monthly'][m]['inv_qty']
            i_v = data['monthly'][m]['inv_val']
            r_v = rx_dict[doc_id][m]
            doc_tot_rx += r_v
            
            roi = (r_v / i_v * 100) if i_v > 0 else 0.0
            monthly_list.append({
                'inv_qty': i_q, 'inv_val': i_v, 'rx_val': r_v, 'roi': roi
            })
            
            # Clean addition without duplicate risks
            gt_monthly[m]['inv_qty'] += i_q
            gt_monthly[m]['inv_val'] += i_v
            gt_monthly[m]['rx_val'] += r_v 
                
        gt_inv_qty += data['tot_inv_qty']
        gt_inv_val += data['tot_inv_val']
        gt_rx_val += doc_tot_rx
        
        tot_roi = (doc_tot_rx / data['tot_inv_val'] * 100) if data['tot_inv_val'] > 0 else 0.0
        
        # HTML design ke liye multiple items ko merge karna (e.g., Pen + Diary)
        item_names = " + ".join(sorted([i[0] for i in data['items']]))
        item_types = set([i[1] for i in data['items']])
        final_item_type = 'HighValue' if 'HighValue' in item_types else 'Routine'

        report_data.append({
            'doc_name': data['doc_name'],
            'item_name': item_names,
            'item_type': final_item_type,
            'emp_name': data['emp_name'],
            'monthly_list': monthly_list,
            'tot_inv_qty': data['tot_inv_qty'],
            'tot_inv_val': data['tot_inv_val'],
            'tot_rx_val': doc_tot_rx,
            'tot_roi': tot_roi
        })
        
    report_data.sort(key=lambda x: x['doc_name'])
    gt_monthly_list = [gt_monthly[m] for m in months_range]
    overall_roi = (gt_rx_val / gt_inv_val * 100) if gt_inv_val > 0 else 0.0

    if request.GET.get('export') == 'excel':
        filename = f"Doctor_ROI_Matrix_{calendar.month_abbr[from_month]}_to_{calendar.month_abbr[to_month]}_{selected_year}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ROI Matrix"

        ws.append(['DOCTOR ROI MATRIX (MONTH-WISE)'])
        ws.append(['Period:', f"{calendar.month_name[from_month]} to {calendar.month_name[to_month]} {selected_year}"])
        ws.append(['Filter:', inv_type.replace('_', ' ').title()])
        ws.append([''])
        ws['A1'].font = Font(bold=True, size=14, color="107C41")

        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")

        headers1 = ['Doctor Name', 'Item Details', 'Given By']
        headers2 = ['', '', '']
        
        for m_num, m_name in months_headers:
            headers1.extend([f"{m_name}", "", "", ""])
            headers2.extend(['Inv Qty', 'Inv Val', 'Rx Val', 'ROI %'])
            
        headers1.extend(["GRAND TOTAL", "", "", ""])
        headers2.extend(['Inv Qty', 'Inv Val', 'Rx Val', 'ROI %'])

        ws.append(headers1)
        ws.append(headers2)

        for row in report_data:
            row_data = [f"Dr. {row['doc_name']}", row['item_name'], row['emp_name']]
            for m in row['monthly_list']:
                row_data.extend([m['inv_qty'], m['inv_val'], m['rx_val'], round(m['roi'], 1)])
            row_data.extend([row['tot_inv_qty'], row['tot_inv_val'], row['tot_rx_val'], round(row['tot_roi'], 1)])
            ws.append(row_data)

        gt_row = ['GRAND TOTAL', '', '']
        for m_gt in gt_monthly_list:
            m_roi = (m_gt['rx_val'] / m_gt['inv_val'] * 100) if m_gt['inv_val'] > 0 else 0.0
            gt_row.extend([m_gt['inv_qty'], m_gt['inv_val'], m_gt['rx_val'], round(m_roi, 1)])
        gt_row.extend([gt_inv_qty, gt_inv_val, gt_rx_val, round(overall_roi, 1)])
        ws.append(gt_row)
        
        for cell in ws[ws.max_row]: cell.font = Font(bold=True)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return render(request, 'doctor_roi_report.html', {
        'team_employees': team_employees,
        'selected_emp_id': int(selected_emp_id) if selected_emp_id else '',
        'inv_type': inv_type,
        'from_month': from_month,
        'to_month': to_month,
        'selected_year': selected_year,
        'months_choices': [(i, calendar.month_name[i]) for i in range(1, 13)],
        'months_headers': months_headers,
        'report_data': report_data,
        'gt_monthly_list': gt_monthly_list,
        'gt_inv_qty': gt_inv_qty,
        'gt_inv_val': gt_inv_val,
        'gt_rx_val': gt_rx_val,
        'overall_roi': overall_roi,
        'is_manager_view': is_manager_view,
    })

@employee_required
def dr_visit_history_view(request, employee):
    team_employees = get_dropdown_team(employee)
    default_emp_id = str(employee.id)
    if employee.designation != 'MR':
        first_sub = team_employees.exclude(id=employee.id).first()
        if first_sub: default_emp_id = str(first_sub.id)
            
    selected_emp_id = request.GET.get('employee_id', default_emp_id)
    selected_emp = get_object_or_404(Employee, id=selected_emp_id)
    
    today = timezone.localdate()
    from_month = int(request.GET.get('from_month', 1))
    to_month = int(request.GET.get('to_month', today.month))
    year = int(request.GET.get('year', today.year))
    
    if from_month > to_month: from_month, to_month = to_month, from_month
    months_range = list(range(from_month, to_month + 1))
    
    doctors = Doctor.objects.filter(allocated_to=selected_emp, status='Approved').order_by('name')
    visits = DCRVisit.objects.filter(daily_dcr__employee=selected_emp, daily_dcr__date__year=year, daily_dcr__date__month__gte=from_month, daily_dcr__date__month__lte=to_month, doctor__isnull=False).select_related('doctor', 'daily_dcr').order_by('daily_dcr__date')
    
    visit_dict = defaultdict(lambda: defaultdict(list))
    for v in visits: visit_dict[v.doctor_id][v.daily_dcr.date.month].append(v.daily_dcr.date.strftime("%d"))
        
    report_data = []
    for doc in doctors:
        doc_months_list = []; doc_months_dict = {}; total_visits = 0
        for m in months_range:
            dates = visit_dict[doc.id].get(m, [])
            dates_str = ", ".join(dates) if dates else "-"
            doc_months_dict[m] = dates_str
            doc_months_list.append({'name': calendar.month_name[m][:3], 'dates': dates_str})
            total_visits += len(dates)
            
        report_data.append({'doctor_name': doc.name, 'specialty': doc.get_specialty_display() if doc.specialty else 'N/A', 'category': doc.category, 'months_dict': doc_months_dict, 'months_list': doc_months_list, 'total_visits': total_visits})

    if request.GET.get('export') == 'xlsx':
        filename = f"Dr_Visit_Matrix_{selected_emp.name}_{year}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Visit Matrix"
        
        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        
        headers = ['Doctor Name', 'Specialty']
        for m in months_range: headers.append(calendar.month_name[m])
        headers.append('Total Visits')
        ws.append(headers)
        
        for col_num, cell in enumerate(ws[1], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20 if col_num <= 2 else 15

        for row in report_data:
            data_row = [f"Dr. {row['doctor_name']}", row['specialty']]
            for m in months_range: data_row.append(row['months_dict'][m])
            data_row.append(row['total_visits'])
            ws.append(data_row)
            for cell in ws[ws.max_row][2:]: cell.alignment = center_align

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return render(request, 'dr_visit_history.html', {
        'team_employees': team_employees, 'selected_emp_id': int(selected_emp_id), 'selected_emp': selected_emp,
        'from_month': from_month, 'to_month': to_month, 'year': year, 'months': [(i, calendar.month_name[i]) for i in range(1, 13)],
        'report_data': report_data
    })
