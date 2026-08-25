import io
import calendar
from collections import defaultdict
from datetime import date, datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.utils import timezone
from django.db.models import Sum, Count, Q

from SFA.models import (
    Employee, Doctor, Chemist, Product, DailyDCR, DCRVisit, DCRProductDetail,
    DayStart, MonthlyTargetMaster, TerritoryTarget, LeaveApplication, Holiday,
    StockistProductStatement, PartyWiseSaleLine
)
from .auth import get_dropdown_team, get_full_team_employees
from SFA.decorators import employee_required

# ==============================================================================
# 1. 📈 ANALYSIS HUB VIEW (Fully Speed Optimized)
# ==============================================================================
@employee_required
def analysis_hub_view(request, employee):
    team_employees = get_dropdown_team(employee, ordered=False)
    is_manager_view = employee.designation != 'MR'

    default_emp_id = str(employee.id)
    if is_manager_view:
        first_sub = team_employees.exclude(id=employee.id).first()
        if first_sub: default_emp_id = str(first_sub.id)

    selected_emp_id = request.GET.get('employee_id', default_emp_id)
    selected_emp = get_object_or_404(Employee, id=int(selected_emp_id))

    today = timezone.localdate()
    from_month = int(request.GET.get('from_month') or today.month)
    to_month = int(request.GET.get('to_month') or today.month)
    selected_year = int(request.GET.get('year') or today.year)

    if from_month > to_month: from_month, to_month = to_month, from_month

    months_range = list(range(from_month, to_month + 1))
    months_headers = [(m, calendar.month_name[m][:3]) for m in months_range]

    is_mr = selected_emp.designation == 'MR'
    total_doctors = Doctor.objects.filter(allocated_to=selected_emp, status='Approved').count() if is_mr else 0
    total_chemists = Chemist.objects.filter(allocated_to=selected_emp, status='Approved').count() if is_mr else 0

    metrics = {
        'field_days': {'label': '🏃‍♂️ Field Days', 'data': {m: 0 for m in months_range}, 'tot': 0},
        'joint_days': {'label': '🤝 Joint Days', 'data': {m: 0 for m in months_range}, 'tot': 0},
        'meeting_transit': {'label': '🏢 Meet/Transit', 'data': {m: 0 for m in months_range}, 'tot': 0},
        'leaves': {'label': '🏖️ Leaves', 'data': {m: 0 for m in months_range}, 'tot': 0},
        'holidays': {'label': '⛱️ Holidays', 'data': {m: 0 for m in months_range}, 'tot': 0},
        'pob': {'label': '💰 Tot POB', 'data': {m: 0 for m in months_range}, 'tot': 0.0, 'is_currency': True},
        'dr_avg': {'label': '👨‍⚕️ Dr Avg', 'data': {m: 0 for m in months_range}, 'tot': 0.0, 'is_avg': True},
        'chem_avg': {'label': '💊 Chem Avg', 'data': {m: 0 for m in months_range}, 'tot': 0.0, 'is_avg': True},
    }

    if is_mr:
        metrics['dr_in_list'] = {'label': '📋 Tot Docs in List', 'data': {m: 0 for m in months_range}, 'tot': total_doctors}
        metrics['dr_visited'] = {'label': '✅ Visited Docs', 'data': {m: 0 for m in months_range}, 'tot': 0}
        metrics['dr_cov'] = {'label': '🎯 Dr Cov %', 'data': {m: 0 for m in months_range}, 'tot': 0.0, 'is_pct': True}
        
        metrics['chem_in_list'] = {'label': '📋 Tot Chems in List', 'data': {m: 0 for m in months_range}, 'tot': total_chemists}
        metrics['chem_visited'] = {'label': '✅ Visited Chems', 'data': {m: 0 for m in months_range}, 'tot': 0}
        metrics['chem_cov'] = {'label': '🎯 Chem Cov %', 'data': {m: 0 for m in months_range}, 'tot': 0.0, 'is_pct': True}

    gt_dr_visits, gt_chem_visits, gt_field_days = 0, 0, 0
    gt_visited_docs, gt_visited_chems = set(), set()
    joint_workers = {}

    # =========================================================
    # 🚀 SPEED HACK: N+1 QUERIES KO BAHAR LAAYE (BULK FETCHING)
    # =========================================================
    start_date = date(selected_year, from_month, 1)
    end_date = date(selected_year, to_month, calendar.monthrange(selected_year, to_month)[1])

    # 1. Fetch All DayStarts in ONE query
    all_day_starts = DayStart.objects.filter(
        employee=selected_emp, date__gte=start_date, date__lte=end_date
    ).select_related('joint_worked_with')
    
    ds_by_month = defaultdict(list)
    for ds in all_day_starts:
        ds_by_month[ds.date.month].append(ds)

    # 2. Fetch All Leaves in ONE query
    all_leaves = LeaveApplication.objects.filter(
        employee=selected_emp, status='Approved', start_date__lte=end_date, end_date__gte=start_date
    )

    # 3. Fetch All DCRs in ONE query
    all_dcrs = DailyDCR.objects.filter(
        employee=selected_emp, date__gte=start_date, date__lte=end_date
    ).prefetch_related('visits__doctor', 'visits__chemist', 'visits__product_details__product')
    
    dcr_by_month = defaultdict(list)
    for dcr in all_dcrs:
        dcr_by_month[dcr.date.month].append(dcr)

    # =========================================================
    # 🚀 FAST IN-MEMORY LOOP (NO DATABASE HITS INSIDE)
    # =========================================================
    for m in months_range:
        num_days = calendar.monthrange(selected_year, m)[1]
        m_start = date(selected_year, m, 1)
        m_end = date(selected_year, m, num_days)
        
        f_days = 0; j_days = 0; m_t_days = 0; h_days = 0
        
        # In-memory evaluation
        for ds in ds_by_month[m]:
            if ds.work_type == 'Field Work':
                f_days += 1
                if ds.joint_worked_with:
                    j_days += 1
                    jw_name = ds.joint_worked_with.name
                    if jw_name not in joint_workers:
                        joint_workers[jw_name] = {'label': f'↳ Joint: {jw_name}', 'data': {mm: 0 for mm in months_range}, 'tot': 0, 'is_sub_row': True}
                    joint_workers[jw_name]['data'][m] += 1
                    joint_workers[jw_name]['tot'] += 1
            elif ds.work_type in ['Meeting', 'Transit']:
                m_t_days += 1
            elif ds.work_type == 'Holiday':
                h_days += 1
        
        # Fast Leave Calculation
        l_days = 0
        for l in all_leaves:
            start = max(l.start_date, m_start)
            end = min(l.end_date, m_end)
            if start <= end:
                l_days += (end - start).days + 1
            
        m_dr_visits, m_chem_visits, m_pob = 0, 0, 0.0
        m_visited_docs, m_visited_chems = set(), set()
        
        for dcr in dcr_by_month[m]:
            for v in dcr.visits.all():
                if v.doctor_id:
                    m_dr_visits += 1; m_visited_docs.add(v.doctor_id); gt_visited_docs.add(v.doctor_id)
                if v.chemist_id:
                    m_chem_visits += 1; m_visited_chems.add(v.chemist_id); gt_visited_chems.add(v.chemist_id)
                for pd in v.product_details.all():
                    price = float(pd.product.price) if getattr(pd.product, 'price', None) else 0.0
                    m_pob += pd.order_qty * price
                    
        gt_dr_visits += m_dr_visits; gt_chem_visits += m_chem_visits; gt_field_days += f_days
        
        metrics['field_days']['data'][m] = f_days; metrics['field_days']['tot'] += f_days
        metrics['joint_days']['data'][m] = j_days; metrics['joint_days']['tot'] += j_days
        metrics['meeting_transit']['data'][m] = m_t_days; metrics['meeting_transit']['tot'] += m_t_days
        metrics['leaves']['data'][m] = l_days; metrics['leaves']['tot'] += l_days
        metrics['holidays']['data'][m] = h_days; metrics['holidays']['tot'] += h_days
        metrics['pob']['data'][m] = m_pob; metrics['pob']['tot'] += m_pob
        metrics['dr_avg']['data'][m] = round(m_dr_visits / f_days, 1) if f_days > 0 else 0
        metrics['chem_avg']['data'][m] = round(m_chem_visits / f_days, 1) if f_days > 0 else 0
        
        if is_mr:
            metrics['dr_in_list']['data'][m] = total_doctors
            metrics['dr_visited']['data'][m] = len(m_visited_docs)
            metrics['dr_cov']['data'][m] = round((len(m_visited_docs) / total_doctors * 100), 1) if total_doctors > 0 else 0
            
            metrics['chem_in_list']['data'][m] = total_chemists
            metrics['chem_visited']['data'][m] = len(m_visited_chems)
            metrics['chem_cov']['data'][m] = round((len(m_visited_chems) / total_chemists * 100), 1) if total_chemists > 0 else 0

    metrics['dr_avg']['tot'] = round(gt_dr_visits / gt_field_days, 1) if gt_field_days > 0 else 0
    metrics['chem_avg']['tot'] = round(gt_chem_visits / gt_field_days, 1) if gt_field_days > 0 else 0
    
    if is_mr:
        metrics['dr_in_list']['tot'] = total_doctors
        metrics['dr_visited']['tot'] = len(gt_visited_docs)
        metrics['dr_cov']['tot'] = round((len(gt_visited_docs) / total_doctors * 100), 1) if total_doctors > 0 else 0
        
        metrics['chem_in_list']['tot'] = total_chemists
        metrics['chem_visited']['tot'] = len(gt_visited_chems)
        metrics['chem_cov']['tot'] = round((len(gt_visited_chems) / total_chemists * 100), 1) if total_chemists > 0 else 0

    report_data = []
    def add_metric(key):
        if key in metrics:
            item = metrics[key]
            item['monthly_list'] = [item['data'][m] for m in months_range]
            report_data.append(item)

    add_metric('field_days')
    add_metric('joint_days')
    
    for jw_name in sorted(joint_workers.keys()):
        jw_item = joint_workers[jw_name]
        jw_item['monthly_list'] = [jw_item['data'][m] for m in months_range]
        report_data.append(jw_item)
        
    add_metric('meeting_transit')
    add_metric('leaves')
    add_metric('holidays')
    add_metric('pob')
    
    if is_mr:
        add_metric('dr_in_list')
        add_metric('dr_visited')
        add_metric('dr_avg')
        add_metric('dr_cov')
        add_metric('chem_in_list')
        add_metric('chem_visited')
        add_metric('chem_avg')
        add_metric('chem_cov')
    else:
        add_metric('dr_avg')
        add_metric('chem_avg')

    if request.GET.get('export') == 'excel':
        filename = f"Analytical_Matrix_{selected_emp.name}_M{from_month}-M{to_month}_{selected_year}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Analytics Matrix"
        
        period_str = f"{calendar.month_name[from_month][:3]} to {calendar.month_name[to_month][:3]} {selected_year}"
        ws.append(['ANALYTICAL HUB - PERFORMANCE TREND MATRIX'])
        ws.append(['Employee:', selected_emp.name, 'Period:', period_str])
        ws.append([''])
        ws['A1'].font = Font(bold=True, size=14, color="107C41")
        ws['A2'].font = Font(bold=True); ws['C2'].font = Font(bold=True)
        
        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        
        headers = ['Key Performance Indicator (KPI)']
        for m_num, m_name in months_headers: headers.append(m_name)
        headers.append('Grand Total / Average')
        ws.append(headers)
        
        for col_num, cell in enumerate(ws[5], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 35 if col_num == 1 else 18
            
        for row in report_data:
            clean_label = row['label'].split(' ', 1)[1] if ' ' in row['label'] else row['label']
            row_data = [clean_label]
            for val in row['monthly_list']:
                if row.get('is_currency'): val = f"₹{round(val, 2)}"
                elif row.get('is_pct'): val = f"{val}%"
                row_data.append(val)
                
            tot = row['tot']
            if row.get('is_currency'): tot = f"₹{round(tot, 2)}"
            elif row.get('is_pct'): tot = f"{tot}%"
            row_data.append(tot)
            ws.append(row_data)
            
        for cell in ws[ws.max_row]: cell.font = Font(bold=True)
            
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return render(request, 'analysis_hub.html', {
        'team_employees': team_employees, 'selected_emp_id': int(selected_emp_id), 'selected_emp': selected_emp,
        'from_month': from_month, 'to_month': to_month, 'selected_year': selected_year,
        'months_choices': [(i, calendar.month_name[i]) for i in range(1, 13)], 'months_headers': months_headers,
        'report_data': report_data, 'is_manager_view': is_manager_view
    })

# ==============================================================================
# 2. 📊 SALES SUMMARY REPORT VIEW
# ==============================================================================
@employee_required
def sales_summary_report_view(request, employee):
    team_employees = get_dropdown_team(employee)
    default_emp_id = str(employee.id)
    if employee.designation != 'MR':
        first_sub = team_employees.exclude(id=employee.id).first()
        if first_sub:
            default_emp_id = str(first_sub.id)
            
    selected_emp_id = request.GET.get('employee_id', default_emp_id)
    selected_emp = get_object_or_404(Employee, id=int(selected_emp_id))
    
    today = timezone.now().date()
    from_month = int(request.GET.get('from_month') or today.month)
    to_month = int(request.GET.get('to_month') or today.month)
    selected_year = int(request.GET.get('year') or today.year)

    if from_month > to_month: from_month, to_month = to_month, from_month
    months_range = list(range(from_month, to_month + 1))
    months_headers = [(m, calendar.month_name[m][:3]) for m in months_range]

    target_team = get_full_team_employees(selected_emp)
    area_territory_ids = target_team.exclude(headquarter__isnull=True).values_list('headquarter_id', flat=True)

    span = to_month - from_month + 1
    curr_start_val = selected_year * 12 + from_month
    curr_end_val = selected_year * 12 + to_month

    prev_end_val = curr_start_val - 1
    prev_start_val = prev_end_val - span + 1
    
    ly_start_val = (selected_year - 1) * 12 + from_month
    ly_end_val = (selected_year - 1) * 12 + to_month

    years_needed = list(set([selected_year, selected_year - 1, (prev_start_val - 1) // 12]))

    team_territories = [emp.headquarter for emp in target_team if emp.headquarter]
    approved_territory_ids = MonthlyTargetMaster.objects.filter(
        territory__in=team_territories, 
        status='Approved', 
        year__in=years_needed
    ).values_list('territory_id', flat=True)

    all_targets = TerritoryTarget.objects.filter(territory_id__in=approved_territory_ids, year__in=years_needed).select_related('product')
    
    all_primary = StockistProductStatement.objects.filter(stockist__territory_id__in=area_territory_ids, year__in=years_needed).select_related('product')
    all_secondary = PartyWiseSaleLine.objects.filter(report__stockist__territory_id__in=area_territory_ids, report__year__in=years_needed).select_related('product')

    def calc_growth(curr, prev):
        if prev == 0: return 100.0 if curr > 0 else 0.0
        return ((curr - prev) / prev) * 100.0

    def calc_ach(val, tgt):
        if tgt == 0: return 100.0 if val > 0 else 0.0
        return (val / tgt) * 100.0

    products = Product.objects.filter(company=selected_emp.company).order_by('name')
    p_dict = {}
    
    gt = {
        'monthly': {m: {'t_qty':0, 't_val':0.0, 'p_qty':0, 'p_val':0.0, 's_qty':0, 's_val':0.0} for m in months_range},
        't_qty': 0, 't_val': 0.0, 'p_qty': 0, 'p_val': 0.0, 's_qty': 0, 's_val': 0.0,
        'prev_p_val': 0.0, 'prev_s_val': 0.0, 'ly_p_val': 0.0, 'ly_s_val': 0.0
    }

    for p in products:
        price = float(p.price) if getattr(p, 'price', None) else 0.0
        p_dict[p.id] = {
            'name': p.name, 'price': price,
            'monthly': {m: {'t_qty':0, 't_val':0.0, 'p_qty':0, 'p_val':0.0, 's_qty':0, 's_val':0.0} for m in months_range},
            'curr_t_qty': 0, 'curr_t_val': 0.0, 'curr_p_qty': 0, 'curr_p_val': 0.0, 'curr_s_qty': 0, 'curr_s_val': 0.0,
            'prev_p_val': 0.0, 'prev_s_val': 0.0, 'ly_p_val': 0.0, 'ly_s_val': 0.0
        }

    for t in all_targets:
        val = t.year * 12 + t.month
        if curr_start_val <= val <= curr_end_val and t.month in p_dict[t.product_id]['monthly']:
            t_val = t.target_qty * float(t.product.price)
            p_dict[t.product_id]['monthly'][t.month]['t_qty'] += t.target_qty; p_dict[t.product_id]['monthly'][t.month]['t_val'] += t_val
            p_dict[t.product_id]['curr_t_qty'] += t.target_qty; p_dict[t.product_id]['curr_t_val'] += t_val
            gt['monthly'][t.month]['t_qty'] += t.target_qty; gt['monthly'][t.month]['t_val'] += t_val
            gt['t_qty'] += t.target_qty; gt['t_val'] += t_val

    for p in all_primary:
        val = p.year * 12 + p.month
        p_val = p.received_qty * float(p.product.price)
        if curr_start_val <= val <= curr_end_val and p.month in p_dict[p.product_id]['monthly']:
            p_dict[p.product_id]['monthly'][p.month]['p_qty'] += p.received_qty; p_dict[p.product_id]['monthly'][p.month]['p_val'] += p_val
            p_dict[p.product_id]['curr_p_qty'] += p.received_qty; p_dict[p.product_id]['curr_p_val'] += p_val
            gt['monthly'][p.month]['p_qty'] += p.received_qty; gt['monthly'][p.month]['p_val'] += p_val
            gt['p_qty'] += p.received_qty; gt['p_val'] += p_val
        elif prev_start_val <= val <= prev_end_val:
            p_dict[p.product_id]['prev_p_val'] += p_val; gt['prev_p_val'] += p_val
        elif ly_start_val <= val <= ly_end_val:
            p_dict[p.product_id]['ly_p_val'] += p_val; gt['ly_p_val'] += p_val

    for s in all_secondary:
        val = s.report.year * 12 + s.report.month
        s_val = s.billed_qty * float(s.product.price)
        s_qty = s.billed_qty + s.free_qty
        if curr_start_val <= val <= curr_end_val and s.report.month in p_dict[s.product_id]['monthly']:
            p_dict[s.product_id]['monthly'][s.report.month]['s_qty'] += s_qty; p_dict[s.product_id]['monthly'][s.report.month]['s_val'] += s_val
            p_dict[s.product_id]['curr_s_qty'] += s_qty; p_dict[s.product_id]['curr_s_val'] += s_val
            gt['monthly'][s.report.month]['s_qty'] += s_qty; gt['monthly'][s.report.month]['s_val'] += s_val
            gt['s_qty'] += s_qty; gt['s_val'] += s_val
        elif prev_start_val <= val <= prev_end_val:
            p_dict[s.product_id]['prev_s_val'] += s_val; gt['prev_s_val'] += s_val
        elif ly_start_val <= val <= ly_end_val:
            p_dict[s.product_id]['ly_s_val'] += s_val; gt['ly_s_val'] += s_val

    report_data = []
    for pid, d in p_dict.items():
        if d['curr_t_val'] > 0 or d['curr_p_val'] > 0 or d['curr_s_val'] > 0:
            d['p_ach'] = calc_ach(d['curr_p_val'], d['curr_t_val'])
            d['s_ach'] = calc_ach(d['curr_s_val'], d['curr_t_val'])
            d['p_m2m'] = calc_growth(d['curr_p_val'], d['prev_p_val'])
            d['p_y2y'] = calc_growth(d['curr_p_val'], d['ly_p_val'])
            d['s_m2m'] = calc_growth(d['curr_s_val'], d['prev_s_val'])
            d['s_y2y'] = calc_growth(d['curr_s_val'], d['ly_s_val'])
            d['monthly_list'] = [d['monthly'][m] for m in months_range]
            report_data.append(d)

    gt['p_ach'] = calc_ach(gt['p_val'], gt['t_val']); gt['s_ach'] = calc_ach(gt['s_val'], gt['t_val'])
    gt['p_m2m'] = calc_growth(gt['p_val'], gt['prev_p_val']); gt['p_y2y'] = calc_growth(gt['p_val'], gt['ly_p_val'])
    gt['s_m2m'] = calc_growth(gt['s_val'], gt['prev_s_val']); gt['s_y2y'] = calc_growth(gt['s_val'], gt['ly_s_val'])
    gt['monthly_list'] = [gt['monthly'][m] for m in months_range]

    if request.GET.get('export') == 'excel':
        filename = f"Sales_Summary_{selected_emp.name}_M{from_month}-M{to_month}_{selected_year}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sales Summary Matrix"
        
        hq_name = selected_emp.headquarter.name if selected_emp.headquarter else "N/A"
        period_str = f"{calendar.month_name[from_month][:3]} to {calendar.month_name[to_month][:3]} {selected_year}"
        
        ws.append(['MASTER SALES SUMMARY REPORT (QTY & VAL MATRIX)'])
        ws.append(['Employee:', selected_emp.name, 'HQ:', hq_name])
        ws.append(['Period:', period_str])
        ws.append([''])
        
        ws['A1'].font = Font(bold=True, size=14, color="107C41")
        for cell in ws['A2:C3']: cell[0].font = Font(bold=True)
        
        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        
        headers1 = ['Product Name']
        headers2 = ['']
        for m in months_headers:
            headers1.extend([f"{m[1]} DATA", "", "", "", "", ""])
            headers2.extend(['Tgt Qty', 'Tgt (₹)', 'Pri Qty', 'Pri (₹)', 'Sec Qty', 'Sec (₹)'])
            
        headers1.extend(['OVERALL GRAND TOTAL', '', '', '', '', '', '', '', '', '', '', ''])
        headers2.extend(['Tot Tgt Qty', 'Tot Tgt (₹)', 'Tot Pri Qty', 'Tot Pri (₹)', 'Tot Sec Qty', 'Tot Sec (₹)', 'Pri Ach%', 'Pri M2M%', 'Pri Y2Y%', 'Sec Ach%', 'Sec M2M%', 'Sec Y2Y%'])
        
        ws.append(headers1)
        ws.append(headers2)
        
        for row_idx in [5, 6]:
            for col_num, cell in enumerate(ws[row_idx], 1):
                cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
                if row_idx == 6: ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 18 if col_num == 1 else 12
            
        for row in report_data:
            row_data = [row['name']]
            for m_data in row['monthly_list']:
                row_data.extend([m_data['t_qty'], round(m_data['t_val'], 2), m_data['p_qty'], round(m_data['p_val'], 2), m_data['s_qty'], round(m_data['s_val'], 2)])
            row_data.extend([
                row['curr_t_qty'], round(row['curr_t_val'], 2), row['curr_p_qty'], round(row['curr_p_val'], 2), row['curr_s_qty'], round(row['curr_s_val'], 2),
                f"{round(row['p_ach'], 1)}%", f"{round(row['p_m2m'], 1)}%", f"{round(row['p_y2y'], 1)}%", f"{round(row['s_ach'], 1)}%", f"{round(row['s_m2m'], 1)}%", f"{round(row['s_y2y'], 1)}%"
            ])
            ws.append(row_data)
            
        gt_row = ['OVERALL TOTAL']
        for m_data in gt['monthly_list']:
            gt_row.extend([m_data['t_qty'], round(m_data['t_val'], 2), m_data['p_qty'], round(m_data['p_val'], 2), m_data['s_qty'], round(m_data['s_val'], 2)])
            
        gt_row.extend([
            gt['t_qty'], round(gt['t_val'], 2), gt['p_qty'], round(gt['p_val'], 2), gt['s_qty'], round(gt['s_val'], 2),
            f"{round(gt['p_ach'], 1)}%", f"{round(gt['p_m2m'], 1)}%", f"{round(gt['p_y2y'], 1)}%", f"{round(gt['s_ach'], 1)}%", f"{round(gt['s_m2m'], 1)}%", f"{round(gt['s_y2y'], 1)}%"
        ])
        ws.append(gt_row)
        for cell in ws[ws.max_row]: cell.font = Font(bold=True)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return render(request, 'sales_summary_report.html', {
        'team_employees': team_employees, 'selected_emp_id': int(selected_emp_id), 'selected_employee_name': selected_emp.name,
        'from_month': from_month, 'to_month': to_month, 'selected_year': selected_year,
        'months_choices': [(i, calendar.month_name[i]) for i in range(1, 13)], 'months_headers': months_headers,
        'report_data': report_data, 'gt': gt, 'is_manager_view': employee.designation != 'MR'
    })


# ==============================================================================
# 3. 🎯 TARGET SETTING VIEW
# ==============================================================================
@employee_required
def target_setting_view(request, employee):
    selected_month = int(request.GET.get('month') or timezone.now().month)
    selected_year = int(request.GET.get('year') or timezone.now().year)
    
    products = Product.objects.filter(company=employee.company)
    
    master = None
    is_readonly = True
    
    if employee.headquarter:
        master, _ = MonthlyTargetMaster.objects.get_or_create(
            territory=employee.headquarter, 
            month=selected_month, 
            year=selected_year
        )
        is_readonly = master.status not in ['Draft', 'Rejected']

    if request.method == "POST" and not is_readonly and employee.headquarter:
        action = request.POST.get('action')
        
        if action in ['Save_Draft', 'Submit_Manager']:
            for p in products:
                t_qty = int(request.POST.get(f'target_{p.id}', 0) or 0)
                if t_qty > 0: 
                    TerritoryTarget.objects.update_or_create(
                        territory=employee.headquarter, 
                        product=p, 
                        month=selected_month, 
                        year=selected_year, 
                        defaults={'target_qty': t_qty}
                    )
                else: 
                    TerritoryTarget.objects.filter(
                        territory=employee.headquarter, 
                        product=p, 
                        month=selected_month, 
                        year=selected_year
                    ).delete()
        
        if action == 'Save_Draft':
            master.status = 'Draft'
            messages.success(request, "Target Draft saved!")
        elif action == 'Submit_Manager':
            master.status = 'Pending_Manager'
            master.approved_by_managers = []
            messages.success(request, "Target submitted for Manager approval! 🚀")
            
        master.save()
        return redirect(f"{request.path}?month={selected_month}&year={selected_year}")

    existing_targets = {}
    if employee.headquarter:
        existing_targets = {
            t.product_id: t.target_qty 
            for t in TerritoryTarget.objects.filter(territory=employee.headquarter, month=selected_month, year=selected_year)
        }
        
    targets = [{'product': p, 'target_qty': existing_targets.get(p.id, 0)} for p in products]

    return render(request, 'target_setting.html', {
        'selected_month': selected_month, 
        'selected_year': selected_year, 
        'targets': targets, 
        'master': master, 
        'is_readonly': is_readonly, 
        'months_choices': [(i, calendar.month_name[i]) for i in range(1, 13)], 
        'user_emp': employee
    })


# ==============================================================================
# 4. 📝 REVIEW TARGET VIEW
# ==============================================================================
@employee_required
def review_target_view(request, employee, target_id):
    master = get_object_or_404(MonthlyTargetMaster, id=target_id)
    
    target_emp = Employee.objects.filter(company=employee.company, headquarter=master.territory, is_active=True).first()
    products = Product.objects.filter(company=employee.company)

    if request.method == "POST":
        action = request.POST.get('action')
        if action in ['Approve', 'Reject', 'Save_Only']:
            for p in products:
                t_qty = int(request.POST.get(f'target_{p.id}', 0) or 0)
                if t_qty > 0: 
                    TerritoryTarget.objects.update_or_create(
                        territory=master.territory, 
                        product=p, 
                        month=master.month, 
                        year=master.year, 
                        defaults={'target_qty': t_qty}
                    )
                else: 
                    TerritoryTarget.objects.filter(
                        territory=master.territory, 
                        product=p, 
                        month=master.month, 
                        year=master.year
                    ).delete()

        if action == 'Reject':
            master.status = 'Rejected'
            master.manager_remark = request.POST.get('manager_remark', '')
            messages.error(request, "Target Rejected! ❌")
            master.save()
            return redirect('manager_approvals')

        elif action == 'Approve':
            master.manager_remark = request.POST.get('manager_remark', '')
            if employee.designation == 'Admin':
                master.status = 'Approved'
                messages.success(request, "Target Officially Approved by Admin! ✅")
            else:
                if employee.id not in master.approved_by_managers:
                    new_list = list(master.approved_by_managers)
                    new_list.append(employee.id)
                    master.approved_by_managers = new_list

                chain_managers = []
                creator_manager = target_emp.manager if target_emp else None
                while creator_manager is not None:
                    chain_managers.append(creator_manager.id)
                    creator_manager = creator_manager.manager

                if len(master.approved_by_managers) >= len(chain_managers):
                    master.status = 'Pending_Admin'
                    messages.success(request, "Approved by all managers! Pending Admin Approval.")
                else:
                    master.status = 'Pending_Manager'
                    messages.success(request, "Approved & Forwarded to next level.")
            master.save()
            return redirect('manager_approvals')

        elif action == 'Save_Only':
            messages.success(request, "Target modifications saved!")
            master.save()
            return redirect(request.path)

    existing_targets = {}
    if master.territory:
        existing_targets = {
            t.product_id: t.target_qty 
            for t in TerritoryTarget.objects.filter(territory=master.territory, month=master.month, year=master.year)
        }
    targets = [{'product': p, 'target_qty': existing_targets.get(p.id, 0)} for p in products]

    return render(request, 'review_target.html', {'targets': targets, 'master': master, 'target_emp': target_emp})


# ==============================================================================
# 5. 🏥 DOCTOR VISIT HISTORY
# ==============================================================================
@employee_required
def dr_visit_history_view(request, employee):
    team_employees = get_dropdown_team(employee)
    default_emp_id = str(employee.id)
    if employee.designation != 'MR':
        first_sub = team_employees.exclude(id=employee.id).first()
        if first_sub: default_emp_id = str(first_sub.id)
            
    selected_emp_id = request.GET.get('employee_id', default_emp_id)
    selected_emp = get_object_or_404(Employee, id=selected_emp_id)
    
    today = timezone.localdate()
    from_month = int(request.GET.get('from_month', 1))
    to_month = int(request.GET.get('to_month', today.month))
    year = int(request.GET.get('year', today.year))
    
    if from_month > to_month: from_month, to_month = to_month, from_month
    months_range = list(range(from_month, to_month + 1))
    
    doctors = Doctor.objects.filter(allocated_to=selected_emp, status='Approved').order_by('name')
    visits = DCRVisit.objects.filter(daily_dcr__employee=selected_emp, daily_dcr__date__year=year, daily_dcr__date__month__gte=from_month, daily_dcr__date__month__lte=to_month, doctor__isnull=False).select_related('doctor', 'daily_dcr').order_by('daily_dcr__date')
    
    visit_dict = defaultdict(lambda: defaultdict(list))
    for v in visits: visit_dict[v.doctor_id][v.daily_dcr.date.month].append(v.daily_dcr.date.strftime("%d"))
        
    report_data = []
    for doc in doctors:
        doc_months_list = []; doc_months_dict = {}; total_visits = 0
        for m in months_range:
            dates = visit_dict[doc.id].get(m, [])
            dates_str = ", ".join(dates) if dates else "-"
            doc_months_dict[m] = dates_str
            doc_months_list.append({'name': calendar.month_name[m][:3], 'dates': dates_str})
            total_visits += len(dates)
            
        report_data.append({'doctor_name': doc.name, 'specialty': doc.get_specialty_display() if doc.specialty else 'N/A', 'category': doc.category, 'months_dict': doc_months_dict, 'months_list': doc_months_list, 'total_visits': total_visits})

    if request.GET.get('export') == 'xlsx':
        filename = f"Dr_Visit_Matrix_{selected_emp.name}_{year}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Visit Matrix"
        
        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        
        headers = ['Doctor Name', 'Specialty']
        for m in months_range: headers.append(calendar.month_name[m])
        headers.append('Total Visits')
        ws.append(headers)
        
        for col_num, cell in enumerate(ws[1], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20 if col_num <= 2 else 15

        for row in report_data:
            data_row = [f"Dr. {row['doctor_name']}", row['specialty']]
            for m in months_range: data_row.append(row['months_dict'][m])
            data_row.append(row['total_visits'])
            ws.append(data_row)
            for cell in ws[ws.max_row][2:]: cell.alignment = center_align

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return render(request, 'dr_visit_history.html', {
        'team_employees': get_dropdown_team(employee), 'selected_emp_id': int(selected_emp_id), 'selected_emp': selected_emp,
        'from_month': from_month, 'to_month': to_month, 'year': year, 'months': [(i, calendar.month_name[i]) for i in range(1, 13)],
        'report_data': report_data
    })