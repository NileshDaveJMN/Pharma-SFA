"""
SFA/api/reports_free_claim.py
==============================
Free Claim Readonly Report API for Flutter.
Separated to keep backend modular and clean!
"""

import io
import calendar
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from django.http import HttpResponse
from django.utils import timezone

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from SFA.models import Employee, Stockist, FreeQtyClaimMaster
from SFA.services.team import get_dropdown_team
from .reports_helpers import _resolve_selected_employee, _employee_brief

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_free_claim_readonly(request):
    try:
        employee = request.user.employee
    except AttributeError:
        return Response({'error': 'Employee profile missing'}, status=400)

    selected_emp, team_employees = _resolve_selected_employee(request, employee)
    is_manager_view = employee.designation != 'MR'

    today = timezone.now().date()
    selected_month = int(request.GET.get('month') or today.month)
    selected_year = int(request.GET.get('year') or today.year)

    # Stockist fetch logic for the selected employee
    my_terr_ids = [selected_emp.headquarter_id] if selected_emp.headquarter_id else []
    available_stockists = Stockist.objects.filter(territory_id__in=my_terr_ids).order_by('name')

    selected_stockist_id = request.GET.get('stockist_id')
    if selected_stockist_id and not available_stockists.filter(id=selected_stockist_id).exists():
        selected_stockist_id = None
    if not selected_stockist_id and available_stockists.exists():
        selected_stockist_id = str(available_stockists.first().id)

    selected_stockist = available_stockists.filter(id=selected_stockist_id).first()

    master_data = None
    lines_data = []
    grand_total = 0.0

    if selected_stockist:
        master = FreeQtyClaimMaster.objects.filter(
            employee=selected_emp, stockist=selected_stockist, month=selected_month, year=selected_year
        ).first()

        if master:
            master_data = {
                'id': master.id,
                'status': master.get_status_display() if hasattr(master, 'get_status_display') else master.status,
                'status_raw': master.status
            }
            for line in master.claim_lines.select_related('product').order_by('product__name'):
                val = float(line.claim_value)
                lines_data.append({
                    'product_name': line.product.name,
                    'billed_qty': line.total_billed_qty,
                    'free_qty': line.total_free_qty,
                    'value': val
                })
                grand_total += val

    # EXCEL EXPORT LOGIC
    if request.GET.get('export') == 'excel':
        filename = f"Free_Claim_{selected_emp.name}_{selected_month}_{selected_year}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Free Claims"

        ws.append(['FREE SCHEME CLAIM REPORT'])
        ws.append(['Employee:', selected_emp.name, 'Stockist:', selected_stockist.name if selected_stockist else 'N/A'])
        ws.append(['Period:', f"{calendar.month_name[selected_month]} {selected_year}"])
        ws.append([''])

        ws['A1'].font = Font(bold=True, size=14, color="107C41")

        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")

        headers = ['Product Name', 'Total Billed', 'Total Free', 'Value (₹)']
        ws.append(headers)

        for col_num, cell in enumerate(ws[5], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20

        for line in lines_data:
            ws.append([line['product_name'], line['billed_qty'], line['free_qty'], line['value']])

        ws.append(['GRAND TOTAL', '', '', grand_total])
        for cell in ws[ws.max_row]: cell.font = Font(bold=True)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return Response({
        'team_employees': [_employee_brief(e) for e in team_employees] if is_manager_view else [],
        'is_manager_view': is_manager_view,
        'selected_emp_id': selected_emp.id,
        'month': selected_month,
        'year': selected_year,
        'stockists': [{'id': s.id, 'name': s.name} for s in available_stockists],
        'selected_stockist_id': int(selected_stockist_id) if selected_stockist_id else None,
        'months_choices': [(i, calendar.month_name[i]) for i in range(1, 13)],
        'master': master_data,
        'lines': lines_data,
        'grand_total': grand_total
    })
