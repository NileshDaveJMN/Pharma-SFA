"""
SFA/api/reports_gift_distribution.py
=====================================
Gift / Input Distribution Report API.
Separated to keep reports_core.py clean!
"""
import io
import calendar
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Sum

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from SFA.models import DoctorROILedger, GiftCampaignPlan
from SFA.services.team import get_dropdown_team

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_gift_distribution_report(request):
    try:
        employee = request.user.employee
    except AttributeError:
        return Response({'error': 'Employee profile missing'}, status=400)

    team_employees = get_dropdown_team(employee)
    is_manager_view = employee.designation != 'MR'

    today = timezone.now().date()
    from_month = int(request.GET.get('from_month') or today.month)
    to_month = int(request.GET.get('to_month') or today.month)
    selected_year = int(request.GET.get('year') or today.year)

    if from_month > to_month:
        from_month, to_month = to_month, from_month

    selected_emp_id = request.GET.get('employee_id', '')
    inv_type = request.GET.get('inv_type', 'any_gift')

    # 1. DISTRIBUTED GIFTS LOGIC
    if is_manager_view and selected_emp_id: 
        qs = DoctorROILedger.objects.filter(employee_id=selected_emp_id)
        plan_qs = GiftCampaignPlan.objects.filter(employee_id=selected_emp_id)
    elif is_manager_view: 
        qs = DoctorROILedger.objects.filter(employee__in=team_employees)
        plan_qs = GiftCampaignPlan.objects.filter(employee__in=team_employees)
    else: 
        qs = DoctorROILedger.objects.filter(employee=employee)
        plan_qs = GiftCampaignPlan.objects.filter(employee=employee)

    qs = qs.filter(
        date_given__month__gte=from_month,
        date_given__month__lte=to_month,
        date_given__year=selected_year
    ).select_related('doctor', 'employee', 'item').order_by('-date_given')

    if inv_type == 'high_value':
        qs = qs.filter(item__item_type='HighValue')
    elif inv_type == 'normal':
        qs = qs.filter(item__item_type='Routine')

    total_qty = qs.aggregate(t_qty=Sum('quantity'))['t_qty'] or 0
    total_val = qs.aggregate(t_val=Sum('total_value'))['t_val'] or 0

    distributed_list = []
    for log in qs:
        distributed_list.append({
            'date': log.date_given.strftime('%d %b, %y'),
            'doctor_name': log.doctor.name,
            'specialty': log.doctor.specialty or '-',
            'employee_name': log.employee.name if log.employee else '-',
            'item_name': log.item.name,
            'item_type': log.item.item_type,
            'qty': log.quantity,
            'value': float(log.total_value)
        })

    # 2. PENDING HV GIFTS LOGIC
    plan_qs = plan_qs.filter(
        month__gte=from_month,
        month__lte=to_month,
        year=selected_year,
        status='Approved'
    ).select_related('doctor', 'item', 'employee')
    
    delivered_combinations = set(
        qs.filter(item__item_type='HighValue').values_list('employee_id', 'doctor_id', 'item_id')
    )
    
    pending_hv_list = []
    for plan in plan_qs:
        if (plan.employee_id, plan.doctor_id, plan.item_id) not in delivered_combinations:
            pending_hv_list.append({
                'month': plan.month,
                'year': plan.year,
                'doctor_name': plan.doctor.name,
                'specialty': plan.doctor.specialty or '-',
                'employee_name': plan.employee.name if plan.employee else '-',
                'item_name': plan.item.name
            })

    if request.GET.get('export') == 'excel':
        filename = f"Gift_Distribution_{selected_year}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Gift Distribution"
        
        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        ws.append(['GIFT DISTRIBUTION LOGS'])
        ws.append(['Period:', f"{calendar.month_name[from_month]} to {calendar.month_name[to_month]} {selected_year}"])
        ws.append([''])
        
        headers = ['Date', 'Doctor Name', 'Specialty', 'Item Name', 'Category', 'Quantity', 'Total Value (₹)', 'Distributed By']
        ws.append(headers)
        for col_num in range(1, 9):
            ws.cell(row=4, column=col_num).fill = header_fill
            ws.cell(row=4, column=col_num).font = header_font
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20

        for log in distributed_list:
            ws.append([log['date'], f"Dr. {log['doctor_name']}", log['specialty'], log['item_name'], log['item_type'], log['qty'], log['value'], log['employee_name']])

        ws.append(['', '', '', '', 'GRAND TOTAL', total_qty, total_val, ''])
        for cell in ws[ws.max_row]: cell.font = Font(bold=True)
        
        if pending_hv_list:
            ws.append([''])
            ws.append(['PENDING HIGH-VALUE DELIVERIES'])
            ws.append(['Month/Year', 'Doctor Name', 'Specialty', 'Proposed Item', 'Status', 'Allocated MR'])
            r = ws.max_row
            for col_num in range(1, 7):
                ws.cell(row=r, column=col_num).fill = PatternFill(start_color="D9534F", end_color="D9534F", fill_type="solid")
                ws.cell(row=r, column=col_num).font = Font(color="FFFFFF", bold=True)
            for p in pending_hv_list:
                ws.append([f"{p['month']}/{p['year']}", f"Dr. {p['doctor_name']}", p['specialty'], p['item_name'], 'Pending Delivery', p['employee_name']])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return Response({
        'team_employees': [{'id': e.id, 'name': e.name} for e in team_employees] if is_manager_view else [],
        'is_manager_view': is_manager_view,
        'distributed_list': distributed_list,
        'pending_hv_list': pending_hv_list,
        'total_qty': total_qty,
        'total_val': total_val
    })
