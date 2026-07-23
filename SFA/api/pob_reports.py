"""
SFA/api/pob_reports.py
======================
POB & Samples Report API for Flutter.
"""
import io
import calendar
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from django.utils import timezone
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.http import HttpResponse

from SFA.models import Employee, DCRProductDetail
from SFA.services.team import get_dropdown_team, get_full_team_employees
from .reports_helpers import _resolve_selected_employee, _employee_brief

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_pob_report(request):
    try:
        employee = request.user.employee
    except AttributeError:
        return Response({'error': 'Employee profile missing'}, status=400)

    selected_emp, team_employees = _resolve_selected_employee(request, employee)
    
    today = timezone.now().date()
    from_month = int(request.GET.get('from_month') or today.month)
    to_month = int(request.GET.get('to_month') or today.month)
    selected_year = int(request.GET.get('year') or today.year)

    if from_month > to_month: 
        from_month, to_month = to_month, from_month
        
    months_range = list(range(from_month, to_month + 1))
    months_headers = [(m, calendar.month_name[m][:3]) for m in months_range]

    target_team = get_full_team_employees(selected_emp)
    emp_ids = target_team.values_list('id', flat=True)

    # Fetch DCR Details
    details = DCRProductDetail.objects.filter(
        visit__daily_dcr__employee_id__in=emp_ids,
        visit__daily_dcr__date__year=selected_year,
        visit__daily_dcr__date__month__in=months_range
    ).select_related('visit__daily_dcr__employee', 'product', 'visit__daily_dcr__employee__headquarter')

    data_dict = {}

    for d in details:
        eid = d.visit.daily_dcr.employee_id
        pid = d.product_id
        m = d.visit.daily_dcr.date.month

        if eid not in data_dict:
            emp_obj = d.visit.daily_dcr.employee
            data_dict[eid] = {
                'emp_name': emp_obj.name,
                'hq': emp_obj.headquarter.name if emp_obj.headquarter else 'N/A',
                'products': {}
            }

        if pid not in data_dict[eid]['products']:
            price = float(d.product.price) if getattr(d.product, 'price', None) else 0.0
            data_dict[eid]['products'][pid] = {
                'name': d.product.name,
                'price': price,
                'monthly': {month: {'samples': 0, 'orders': 0} for month in months_range},
                'total_samples': 0, 'total_orders': 0, 'total_val': 0.0
            }

        sq = d.sample_qty or 0
        oq = d.order_qty or 0
        val = oq * data_dict[eid]['products'][pid]['price']

        data_dict[eid]['products'][pid]['monthly'][m]['samples'] += sq
        data_dict[eid]['products'][pid]['monthly'][m]['orders'] += oq
        data_dict[eid]['products'][pid]['total_samples'] += sq
        data_dict[eid]['products'][pid]['total_orders'] += oq
        data_dict[eid]['products'][pid]['total_val'] += val

    pob_data = []
    gt_monthly = {m: {'samples': 0, 'orders': 0, 'val': 0.0} for m in months_range}
    gt_samples, gt_orders, gt_val = 0, 0, 0.0

    for eid, edata in data_dict.items():
        prod_list = []
        for pid, pdata in edata['products'].items():
            pdata['monthly_list'] = [pdata['monthly'][m] for m in months_range]
            prod_list.append(pdata)

            for m in months_range:
                gt_monthly[m]['samples'] += pdata['monthly'][m]['samples']
                gt_monthly[m]['orders'] += pdata['monthly'][m]['orders']
                gt_monthly[m]['val'] += pdata['monthly'][m]['orders'] * pdata['price']

            gt_samples += pdata['total_samples']
            gt_orders += pdata['total_orders']
            gt_val += pdata['total_val']

        edata['products'] = sorted(prod_list, key=lambda x: x['name'])
        pob_data.append(edata)

    pob_data = sorted(pob_data, key=lambda x: x['emp_name'])
    gt_monthly_list = [gt_monthly[m] for m in months_range]

    # EXCEL EXPORT
    if request.GET.get('export') == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "POB and Samples"
        
        ws.append(['POB & SAMPLES REPORT'])
        ws.append(['Period:', f"{calendar.month_name[from_month][:3]} to {calendar.month_name[to_month][:3]} {selected_year}"])
        ws.append([''])
        
        ws['A1'].font = Font(bold=True, size=14, color="107C41")
        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        
        h1 = ['Employee Name', 'HQ', 'Product Name']
        h2 = ['', '', '']
        for m in months_headers:
            h1.extend([f"{m[1]} DATA", ""])
            h2.extend(['Samples', 'Orders'])
        h1.extend(['GRAND TOTAL', '', ''])
        h2.extend(['Tot Samples', 'Tot Orders', 'Tot Val (₹)'])
        
        ws.append(h1)
        ws.append(h2)
        
        for row_idx in [4, 5]:
            for col_num, cell in enumerate(ws[row_idx], 1):
                cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
                if row_idx == 5: ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 15
        
        for emp in pob_data:
            for p in emp['products']:
                row = [emp['emp_name'], emp['hq'], p['name']]
                for m_data in p['monthly_list']:
                    row.extend([m_data['samples'], m_data['orders']])
                row.extend([p['total_samples'], p['total_orders'], round(p['total_val'], 2)])
                ws.append(row)
                
        # Grand Total
        gt_row = ['GRAND TOTAL', '', '']
        for m in gt_monthly_list:
            gt_row.extend([m['samples'], m['orders']])
        gt_row.extend([gt_samples, gt_orders, round(gt_val, 2)])
        ws.append(gt_row)
        for cell in ws[ws.max_row]: cell.font = Font(bold=True)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="POB_Report_{selected_year}.xlsx"'
        return response

    return Response({
        'team_employees': [_employee_brief(e) for e in team_employees] if employee.designation != 'MR' else [],
        'months_choices': [(i, calendar.month_name[i]) for i in range(1, 13)],
        'months_headers': months_headers,
        'pob_data': pob_data,
        'gt_monthly_list': gt_monthly_list,
        'gt_samples': gt_samples,
        'gt_orders': gt_orders,
        'gt_val': gt_val
    })
