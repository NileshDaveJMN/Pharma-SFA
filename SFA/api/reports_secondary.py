"""
SFA/api/reports_secondary.py
=============================
🌟 NAYA FILE: "Smart Stockist Statement" (Secondary Sale stock movement)
report, Flutter ke liye. Web ke `smart_secondary_report_view`
(SFA/views/sales.py) jaisi hi calculation hai — Opening / Primary /
Secondary / Closing qty+value, Stockist aur Product ke hisaab se.

Excel-export web-only feature hai, isliye yahan nahi banaya — Flutter
sirf JSON data dikhayega.

Wiring karne ke liye (2 chhoti si lines add karni hain):

1. SFA/api/reports.py (ya jahan bhi wildcard-shim hai) me ek line add karo:
       from .reports_secondary import * # noqa: F401,F403

2. SFA/api/urls.py me ek path add karo:
       path('reports/stockist-statement/', reports_api.api_stockist_statement,
            name='api_stockist_statement'),

Endpoint:
    GET /api/reports/stockist-statement/?employee_id=&month=&year=
"""

import io
import calendar
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from collections import defaultdict

from django.utils import timezone
from django.db.models import Sum
from django.http import HttpResponse

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from SFA.models import (
    Employee, Stockist, Product,
    StockistProductStatement, PartyWiseSaleLine,
)
from SFA.services.team import get_dropdown_team, get_full_team_employees


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_stockist_statement(request):
    """
    Smart Stockist Statement — Opening/Primary/Secondary/Closing stock
    movement, Stockist aur Product ke hisaab se.

    GET params:
        employee_id  (optional, Manager ke liye — kis employee/team ka data)
        month, year  (optional, default current month/year)
    """
    try:
        employee = request.user.employee
    except AttributeError:
        return Response({'error': 'Employee profile missing'}, status=400)

    team_employees = get_dropdown_team(employee, ordered=False)
    is_manager_view = employee.designation != 'MR'

    emp_id = request.GET.get('employee_id', str(employee.id))
    try:
        # 🌟 FIX: IDOR Prevention - Cross-company employee fetch rokna
        selected_emp = Employee.objects.get(id=emp_id, company=employee.company)
    except (Employee.DoesNotExist, ValueError):
        selected_emp = employee

    # 🌟 Security: sirf apni team ka employee select kar sakte hain
    if not team_employees.filter(id=selected_emp.id).exists():
        return Response({'error': 'Access denied'}, status=403)

    today = timezone.now().date()
    try:
        selected_month = int(request.GET.get('month') or today.month)
        selected_year = int(request.GET.get('year') or today.year)
    except ValueError:
        return Response({'error': 'Invalid month/year'}, status=400)

    # Manager ho to poori team ke territories combine honge, MR ho to sirf khud ka HQ
    selected_team_scope = get_full_team_employees(selected_emp)
    area_territory_ids = selected_team_scope.exclude(
        headquarter__isnull=True
    ).values_list('headquarter_id', flat=True)

    # 🌟 company= explicitly bhi laga diya (defense-in-depth, territory se already scoped hai)
    stockists = Stockist.objects.filter(
        territory_id__in=area_territory_ids, company=employee.company
    )
    stockist_dict = {st.id: st.name for st in stockists}

    products_dict = {
        p.id: {'name': p.name, 'price': float(p.price) if getattr(p, 'price', None) else 0.0}
        for p in Product.objects.filter(company=employee.company).order_by('name')
    }

    curr_val = selected_year * 12 + selected_month

    statements = StockistProductStatement.objects.filter(
        stockist_id__in=stockist_dict.keys()
    ).values('stockist_id', 'product_id', 'month', 'year').annotate(
        op=Sum('opening_qty'), rec=Sum('received_qty')
    )
    secondaries = PartyWiseSaleLine.objects.filter(
        report__stockist_id__in=stockist_dict.keys()
    ).values('report__stockist_id', 'product_id', 'report__month', 'report__year').annotate(
        tb=Sum('billed_qty'), tf=Sum('free_qty')
    )

    data_map = defaultdict(lambda: defaultdict(lambda: {'op': 0, 'rec': 0, 'tb': 0, 'tf': 0}))

    for st in statements:
        st_val = st['year'] * 12 + st['month']
        if st_val < curr_val:
            data_map[st['stockist_id']][st['product_id']]['op'] += (st['op'] or 0) + (st['rec'] or 0)
        elif st_val == curr_val:
            data_map[st['stockist_id']][st['product_id']]['op'] += (st['op'] or 0)
            data_map[st['stockist_id']][st['product_id']]['rec'] += (st['rec'] or 0)

    for sec in secondaries:
        sec_val = sec['report__year'] * 12 + sec['report__month']
        if sec_val < curr_val:
            data_map[sec['report__stockist_id']][sec['product_id']]['op'] -= (
                (sec['tb'] or 0) + (sec['tf'] or 0)
            )
        elif sec_val == curr_val:
            data_map[sec['report__stockist_id']][sec['product_id']]['tb'] += (sec['tb'] or 0)
            data_map[sec['report__stockist_id']][sec['product_id']]['tf'] += (sec['tf'] or 0)

    report_data = []
    gt = {'op_qty': 0, 'op_val': 0.0, 'pr_qty': 0, 'pr_val': 0.0,
          'sec_b': 0, 'sec_f': 0, 'sec_val': 0.0, 'cl_qty': 0, 'cl_val': 0.0}

    for s_id, s_name in stockist_dict.items():
        if s_id not in data_map:
            continue
        products_out = []

        for p_id, p_info in products_dict.items():
            if p_id not in data_map[s_id]:
                continue
            m = data_map[s_id][p_id]
            op, rec, tb, tf = m['op'], m['rec'], m['tb'], m['tf']
            cl = op + rec - (tb + tf)

            if op == 0 and rec == 0 and tb == 0 and tf == 0 and cl == 0:
                continue

            price = p_info['price']
            op_val, pr_val, sec_val, cl_val = op * price, rec * price, tb * price, cl * price

            products_out.append({
                'name': p_info['name'],
                'opening_qty': op, 'opening_val': round(op_val, 2),
                'primary_qty': rec, 'primary_val': round(pr_val, 2),
                'secondary_billed_qty': tb, 'secondary_free_qty': tf,
                'secondary_val': round(sec_val, 2),
                'closing_qty': cl, 'closing_val': round(cl_val, 2),
            })

            gt['op_qty'] += op; gt['op_val'] += op_val
            gt['pr_qty'] += rec; gt['pr_val'] += pr_val
            gt['sec_b'] += tb; gt['sec_f'] += tf; gt['sec_val'] += sec_val
            gt['cl_qty'] += cl; gt['cl_val'] += cl_val

        if products_out:
            report_data.append({'stockist_name': s_name, 'products': products_out})

    report_data.sort(key=lambda x: x['stockist_name'])
    
    # 🌟 EXCEL EXPORT LOGIC
    if request.GET.get('export') == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stockist Statement"
        
        ws.append(['SMART STOCKIST STATEMENT'])
        ws.append(['Employee:', selected_emp.name])
        period_str = f"{calendar.month_name[selected_month]} {selected_year}"
        ws.append(['Period:', period_str])
        ws.append([''])
        
        ws['A1'].font = Font(bold=True, size=14, color="107C41")
        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        
        headers = [
            'Stockist Name', 'Product Name', 
            'Opening Qty', 'Opening Val (₹)', 
            'Primary Qty', 'Primary Val (₹)',
            'Sec Billed Qty', 'Sec Free Qty', 'Sec Val (₹)',
            'Closing Qty', 'Closing Val (₹)'
        ]
        ws.append(headers)
        
        for col_num, cell in enumerate(ws[5], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 15
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

        for party in report_data:
            for prod in party['products']:
                ws.append([
                    party['stockist_name'], prod['name'],
                    prod['opening_qty'], round(prod['opening_val'], 2),
                    prod['primary_qty'], round(prod['primary_val'], 2),
                    prod['secondary_billed_qty'], prod['secondary_free_qty'], round(prod['secondary_val'], 2),
                    prod['closing_qty'], round(prod['closing_val'], 2)
                ])

        # Grand Total Row
        gt_row = ['GRAND TOTAL', '']
        gt_row.extend([
            gt['op_qty'], round(gt['op_val'], 2),
            gt['pr_qty'], round(gt['pr_val'], 2),
            gt['sec_b'], gt['sec_f'], round(gt['sec_val'], 2),
            gt['cl_qty'], round(gt['cl_val'], 2)
        ])
        ws.append(gt_row)
        for cell in ws[ws.max_row]: 
            cell.font = Font(bold=True)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Stockist_Statement_{selected_year}_{selected_month}.xlsx"'
        return response

    return Response({
        'employee': {'id': selected_emp.id, 'name': selected_emp.name},
        'is_manager_view': is_manager_view,
        'team_employees': (
            [{'id': e.id, 'name': e.name} for e in team_employees]
            if is_manager_view else []
        ),
        'month': selected_month,
        'year': selected_year,
        'report_data': report_data,
        'grand_total': {
            'opening_qty': gt['op_qty'], 'opening_val': round(gt['op_val'], 2),
            'primary_qty': gt['pr_qty'], 'primary_val': round(gt['pr_val'], 2),
            'secondary_billed_qty': gt['sec_b'], 'secondary_free_qty': gt['sec_f'],
            'secondary_val': round(gt['sec_val'], 2),
            'closing_qty': gt['cl_qty'], 'closing_val': round(gt['cl_val'], 2),
        },
    })
