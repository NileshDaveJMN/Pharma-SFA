"""
SFA/api/reports_ops.py
=======================
Network report, product master, doctor visit history, inventory,
free claims, tour plan report, expense report, holiday list.
(reports.py se split kiya gaya — 1000+ line limit ke wajah se)
"""

import calendar
from collections import defaultdict
from datetime import datetime

from django.utils import timezone
from django.db.models import Sum, Count
from django.shortcuts import get_object_or_404

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status

from SFA.models import (
    Employee, Doctor, Chemist, Product, DailyDCR, DCRVisit, DCRProductDetail,
    MonthlyTourProgram, MonthlyExpenseReport, Route, Holiday,
    MonthlyTargetMaster, LeaveApplication, LeaveBalance, DayStart, SystemSetting,
    FreeQtyClaimMaster, GiftCampaignPlan, ChemistEditRequest, DoctorEditRequest,
)
from SFA.services.team import (
    get_full_team_employees,
    get_team_territory_ids,
    get_team_requested_routes,
    get_dropdown_team,
)
from SFA.views.reports import _get_target_chain_starter


from .reports_helpers import _resolve_selected_employee, _employee_brief

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_network_report(request):
    """
    Approved Doctors/Chemists ki listing, route/specialty/category filters
    ke saath.

    Query params:
        employee_id, tab (doctor|chemist, default doctor),
        route, specialty, category (sab optional filters)

    Response:
    {
        "selected_employee": {...}, "team_employees": [...],
        "active_tab": "doctor",
        "routes": [{"id": 3, "name": "Dadar West"}],
        "specialty_choices": [["GP", "General Physician (GP)"], ...],
        "category_choices": [["A", "Category A"], ...],
        "doctors": [
            {"id": 1, "name": "Dr. Sharma", "specialty": "GP", "category": "A",
             "route": "Dadar West", "territory": "Dadar", "allocated_to": "Amit MR"}
        ],
        "chemists": [...]   // sirf jab tab=chemist
    }
    """
    try:
        employee = request.user.employee
    except AttributeError:
        return Response({'error': 'Employee profile missing'}, status=400)

    selected_emp, team_employees = _resolve_selected_employee(request, employee)

    sub_team = get_dropdown_team(selected_emp, ordered=False)
    my_terr_ids = get_team_territory_ids(sub_team)
    routes = get_team_requested_routes(sub_team, my_terr_ids)

    active_tab = request.GET.get('tab', 'doctor')
    route_id = request.GET.get('route', '')
    specialty = request.GET.get('specialty', '')
    category = request.GET.get('category', '')

    doctors_out, chemists_out = [], []

    if active_tab == 'chemist':
        chemists = Chemist.objects.filter(allocated_to__in=sub_team, status='Approved').select_related('allocated_to', 'route', 'territory')
        if route_id:
            chemists = chemists.filter(route_id=route_id)
        for c in chemists:
            terr_name = getattr(c.territory, 'name', None) if hasattr(c, 'territory') else None
            if not terr_name and c.route and getattr(c.route, 'territory', None):
                terr_name = c.route.territory.name
            chemists_out.append({
                'id': c.id, 'name': c.name,
                'route': c.route.name if c.route else None,
                'territory': terr_name,
                'allocated_to': c.allocated_to.name if c.allocated_to else None,
                'latitude': float(c.latitude) if c.latitude else None,   # 🌟 FIX: GPS status ke liye
                'longitude': float(c.longitude) if c.longitude else None,
            })
    else:
        doctors = Doctor.objects.filter(allocated_to__in=sub_team, status='Approved').select_related('allocated_to', 'route', 'territory')
        if route_id:
            doctors = doctors.filter(route_id=route_id)
        if specialty:
            doctors = doctors.filter(specialty=specialty)
        if category:
            doctors = doctors.filter(category=category)
        for d in doctors:
            terr_name = getattr(d.territory, 'name', None)
            if not terr_name and d.route and getattr(d.route, 'territory', None):
                terr_name = d.route.territory.name
            doctors_out.append({
                'id': d.id, 'name': d.name,
                'specialty': d.specialty, 'category': d.category,
                'route': d.route.name if d.route else None,
                'territory': terr_name,
                'allocated_to': d.allocated_to.name if d.allocated_to else None,
                'latitude': float(d.latitude) if d.latitude else None,   # 🌟 FIX: GPS status ke liye
                'longitude': float(d.longitude) if d.longitude else None,
            })
    if request.GET.get('export') == 'excel':
        import openpyxl
        import io
        from django.http import HttpResponse
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")

        if active_tab == 'doctor':
            filename = f"Doctor_Network_{selected_emp.name}.xlsx"
            ws.title = "Doctor List"
            ws.append(['MASTER DOCTOR DIRECTORY'])
            ws.append(['Employee:', selected_emp.name])
            ws.append([''])
            ws['A1'].font = Font(bold=True, size=14, color="107C41")

            headers = ['Doctor Name', 'Specialty', 'Category', 'Route / Patch', 'Territory / HQ', 'Allocated MR']
            ws.append(headers)
            for col_num, cell in enumerate(ws[4], 1):
                cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20

            for doc in doctors_out:
                ws.append([
                    f"Dr. {doc['name']}",
                    doc.get('specialty') or '-',
                    doc.get('category') or '-',
                    doc.get('route') or '-',
                    doc.get('territory') or '-',
                    doc.get('allocated_to') or '-'
                ])
        else:
            filename = f"Chemist_Network_{selected_emp.name}.xlsx"
            ws.title = "Chemist List"
            ws.append(['MASTER CHEMIST DIRECTORY'])
            ws.append(['Employee:', selected_emp.name])
            ws.append([''])
            ws['A1'].font = Font(bold=True, size=14, color="107C41")

            headers = ['Pharmacy / Chemist Name', 'Route / Patch', 'Territory / HQ', 'Allocated MR']
            ws.append(headers)
            for col_num, cell in enumerate(ws[4], 1):
                cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 25

            for chem in chemists_out:
                ws.append([
                    chem['name'],
                    chem.get('route') or '-',
                    chem.get('territory') or '-',
                    chem.get('allocated_to') or '-'
                ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


    # 🌟 FIX: web jaisa hi — Admin ne location re-capture allow ki hai ya nahi
    setting = SystemSetting.objects.filter(company=employee.company).first()
    allow_location_capture = setting.allow_location_capture if setting else True

    return Response({
        'selected_employee': _employee_brief(selected_emp),
        'team_employees': [_employee_brief(e) for e in team_employees] if employee.designation != 'MR' else [],
        'active_tab': active_tab,
        'routes': [{'id': r.id, 'name': r.name} for r in routes],
        'specialty_choices': list(Doctor.SPECIALTY_CHOICES),
        'category_choices': list(Doctor.CATEGORY_CHOICES),
        'allow_location_capture': allow_location_capture,  # 🌟 FIX
        'doctors': doctors_out,
        'chemists': chemists_out,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_route_playback(request, employee_id, date_str):
    """
    Ek din ke Day Start -> Visits -> Day End ka GPS waypoints trail —
    web ke route_playback_view jaisa hi (route_playback.html me jo
    map-plotting hoti hai, wahi data yahan se).

    🌟 Permission: khud ka din, ya apni team ke kisi member ka —
    (web me ye check nahi tha kyunki URL directly team-listing se hi
    aata tha, lekin API ko khud is se protect karna hoga)

    URL: GET api/reports/route-playback/<employee_id>/<date_str>/
         date_str format: YYYY-MM-DD

    Response:
    {
        "employee": {"id": 7, "name": "Devarshi joshi"},
        "date": "2026-07-06",
        "has_data": true,
        "waypoints": [
            {"lat": 23.03, "lng": 72.58, "title": "Day Start", "time": "09:15 AM", "type": "start"},
            {"lat": 23.04, "lng": 72.59, "title": "Dr. Sharma", "time": "10:30 AM", "type": "visit"},
            {"lat": 23.05, "lng": 72.60, "title": "Day End", "time": "06:45 PM", "type": "end"}
        ]
    }

    Error (403): { "error": "Ye record aapki team ka nahi hai." }
    Error (404): { "error": "Invalid date format." }
    """
    from SFA.models import DayEnd  # 🌟 local import — file-level import list disturb nahi karna

    try:
        employee = request.user.employee
    except AttributeError:
        return Response({'error': 'Employee profile missing'}, status=400)

    # 🌟 FIX: Company check ke sath fetch karo
    emp = get_object_or_404(Employee, id=employee_id, company=employee.company)

    # 🌟 Permission check: khud ka data, ya apni team ke kisi member ka
    allowed_ids = set(get_full_team_employees(employee).values_list('id', flat=True))
    if emp.id not in allowed_ids:
        return Response({'error': 'Ye record aapki team ka nahi hai.'}, status=403)

    try:
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return Response({'error': 'Invalid date format. Use YYYY-MM-DD.'}, status=400)

    day_start = DayStart.objects.filter(employee=emp, date=target_date).first()
    day_end = DayEnd.objects.filter(employee=emp, date=target_date).first()
    daily_dcr = DailyDCR.objects.filter(employee=emp, date=target_date).first()

    waypoints = []
    if day_start and day_start.latitude and day_start.longitude:
        waypoints.append({
            'lat': float(day_start.latitude), 'lng': float(day_start.longitude),
            'title': 'Day Start', 'time': day_start.started_at.strftime('%I:%M %p'), 'type': 'start',
        })
    if daily_dcr:
        for v in daily_dcr.visits.all().order_by('created_at'):
            if v.latitude and v.longitude:
                name = f"Dr. {v.doctor.name}" if v.doctor_id else (v.chemist.name if v.chemist_id else 'Visit')
                waypoints.append({
                    'lat': float(v.latitude), 'lng': float(v.longitude),
                    'title': name, 'time': v.created_at.strftime('%I:%M %p'), 'type': 'visit',
                })
    if day_end and day_end.latitude and day_end.longitude:
        waypoints.append({
            'lat': float(day_end.latitude), 'lng': float(day_end.longitude),
            'title': 'Day End', 'time': day_end.closed_at.strftime('%I:%M %p'), 'type': 'end',
        })

    return Response({
        'employee': _employee_brief(emp),
        'date': str(target_date),
        'has_data': len(waypoints) > 0,
        'waypoints': waypoints,
    })


# ==============================================================================
# 📦 5. PRODUCT MASTER
# ==============================================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_product_master(request):
    """
    Saari products ki master list (name, pack-size, price).

    Response:
    [
        {"id": 1, "name": "PPI", "pack_size": "10x10", "price": 109.0}
    ]
    """
    products = Product.objects.filter(company=employee.company).order_by('name')  # 🌟 FIX: company-scoped
    return Response([
        {
            'id': p.id, 'name': p.name,
            'pack_size': getattr(p, 'pack_size', None),
            'price': float(p.price) if getattr(p, 'price', None) else 0.0,
        } for p in products
    ])


# ==============================================================================
# 👨‍⚕️ 6. DOCTOR VISIT HISTORY (month-range matrix)
# ==============================================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_doctor_visit_history(request):
    """
    Har doctor ke liye, month-range ke andar konsi tareekhon pe visit hui —
    matrix format.

    Query params: employee_id, from_month, to_month, year (sab optional)

    Response:
    {
        "selected_employee": {...}, "team_employees": [...],
        "from_month": 1, "to_month": 6, "year": 2026,
        "months_headers": ["Jan", "Feb", ..., "Jun"],
        "doctors": [
            {
                "doctor_name": "Sharma", "specialty": "General Physician (GP)",
                "category": "A",
                "monthly": [{"month": "Jan", "dates": "5, 12, 20"}, ...],
                "total_visits": 14
            }
        ]
    }
    """
    try:
        employee = request.user.employee
    except AttributeError:
        return Response({'error': 'Employee profile missing'}, status=400)

    selected_emp, team_employees = _resolve_selected_employee(request, employee)

    today = datetime.today()
    from_month = int(request.GET.get('from_month') or 1)
    to_month = int(request.GET.get('to_month') or today.month)
    year = int(request.GET.get('year') or today.year)
    if from_month > to_month:
        from_month, to_month = to_month, from_month
    months_range = list(range(from_month, to_month + 1))

    doctors = Doctor.objects.filter(allocated_to=selected_emp, status='Approved').order_by('name')
    visits = DCRVisit.objects.filter(
        daily_dcr__employee=selected_emp, daily_dcr__date__year=year,
        daily_dcr__date__month__gte=from_month, daily_dcr__date__month__lte=to_month,
        doctor__isnull=False,
    ).select_related('doctor', 'daily_dcr').order_by('daily_dcr__date')

    visit_dict = defaultdict(lambda: defaultdict(list))
    for v in visits:
        visit_dict[v.doctor_id][v.daily_dcr.date.month].append(v.daily_dcr.date.strftime('%d'))

    doctors_out = []
    for doc in doctors:
        monthly = []
        total_visits = 0
        for m in months_range:
            dates = visit_dict[doc.id].get(m, [])
            monthly.append({'month': calendar.month_name[m][:3], 'dates': ", ".join(dates) if dates else None})
            total_visits += len(dates)

        doctors_out.append({
            'doctor_name': doc.name,
            'specialty': doc.get_specialty_display() if doc.specialty else None,
            'category': doc.category,
            'monthly': monthly,
            'total_visits': total_visits,
        })

    return Response({
        'selected_employee': _employee_brief(selected_emp),
        'team_employees': [_employee_brief(e) for e in team_employees] if employee.designation != 'MR' else [],
        'from_month': from_month, 'to_month': to_month, 'year': year,
        'months_headers': [calendar.month_name[m][:3] for m in months_range],
        'doctors': doctors_out,
    })


# 🌟 Naye models jo niche ke 2 functions (api_inventory, api_free_claims) use karte hain
from SFA.models import (
    MRInventory, PromoDispatch, FreeQtyClaimLine,
    Stockist, PartyWiseSaleLine,
)


# ====================================================================
# 7. INVENTORY & DISPATCH RECEIVE API (From reports.py -> mr_inventory_view)
# ====================================================================
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def api_inventory(request):
    employee = request.user.employee

    # 📥 GET: Current Stock aur In-Transit items
    if request.method == 'GET':
        stock_items = MRInventory.objects.filter(employee=employee, stock_qty__gt=0).select_related('item')
        current_stock = [{
            'id': s.id,
            'item_name': s.item.name,
            'item_type': s.item.item_type,
            'qty': s.stock_qty
        } for s in stock_items]

        dispatches = PromoDispatch.objects.filter(employee=employee, status='In-Transit').select_related('item')
        in_transit = [{
            'dispatch_id': d.id,
            'item_name': d.item.name,
            'dispatch_qty': d.quantity,
            'dispatch_date': d.dispatch_date.strftime('%Y-%m-%d') if d.dispatch_date else None
        } for d in dispatches]

        return Response({
            'success': True,
            'current_stock': current_stock,
            'in_transit': in_transit
        })

    # 📤 POST: In-Transit item ko 'Receive' karna
    if request.method == 'POST':
        dispatch_id = request.data.get('dispatch_id')
        if not dispatch_id:
            return Response({'success': False, 'error': 'Dispatch ID is required.'}, status=400)

        dispatch = get_object_or_404(PromoDispatch, id=dispatch_id, employee=employee, status='In-Transit')
        
        # Inventory mein add karo
        inv, created = MRInventory.objects.get_or_create(employee=employee, item=dispatch.item, defaults={'stock_qty': 0})
        inv.stock_qty += dispatch.quantity
        inv.save()

        # Status update karo
        dispatch.status = 'Received'
        dispatch.save()

        return Response({'success': True, 'message': f'{dispatch.item.name} successfully received into stock!'})


# ====================================================================
# 8. FREE QUANTITY CLAIMS API (From reports.py -> free_claim_view)
# ====================================================================
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def api_free_claims(request):
    employee = request.user.employee
    today = timezone.now().date()
    current_month = today.month
    current_year = today.year

    # 🌟 NAYA: Manager View setup
    team_employees = get_dropdown_team(employee, ordered=False)
    is_manager_view = employee.designation != 'MR'

    if request.method == 'GET':
        month = int(request.query_params.get('month', current_month))
        year = int(request.query_params.get('year', current_year))
        stockist_id = request.query_params.get('stockist_id')
        
        # 🌟 NAYA: Manager ne agar employee_id diya hai toh use lo, warna khud ko
        emp_id = request.query_params.get('employee_id', str(employee.id))
        try:
            selected_emp = Employee.objects.get(id=emp_id, company=employee.company)
        except Employee.DoesNotExist:
            selected_emp = employee
            
        # 🌟 NAYA: selected_emp ke hisaab se stockists laao
        my_terr_ids = [selected_emp.headquarter_id] if selected_emp.headquarter_id else []
        stockists = Stockist.objects.filter(territory_id__in=my_terr_ids, company=employee.company).order_by('name')
        
        res_data = {
            'success': True,
            'is_manager_view': is_manager_view, # 🌟 NAYA
            'team_employees': [{'id': e.id, 'name': e.name} for e in team_employees] if is_manager_view else [], # 🌟 NAYA
            'stockists': [{'id': s.id, 'name': s.name} for s in stockists],
            'master': None,
            'lines': [],
            'grand_total': 0.0
        }

        if stockist_id:
            # 🌟 NAYA: selected_emp use kiya employee ki jagah
            master = FreeQtyClaimMaster.objects.filter(employee=selected_emp, stockist_id=stockist_id, month=month, year=year).first()
            if master:
                res_data['master'] = {
                    'id': master.id,
                    'status': master.status,
                    'manager_remark': master.manager_remark or '',
                    'admin_remark': master.admin_remark or '',
                }
                
                lines_agg = FreeQtyClaimLine.objects.filter(master=master).values(
                    'product__name'
                ).annotate(
                    tot_b=Sum('total_billed_qty'),
                    tot_f=Sum('total_free_qty'),
                    tot_val=Sum('claim_value')
                ).order_by('product__name')

                grand_total = 0.0
                for line in lines_agg:
                    res_data['lines'].append({
                        'product_name': line['product__name'],
                        'total_billed_qty': line['tot_b'],
                        'total_free_qty': line['tot_f'],
                        'claim_value': float(line['tot_val'])
                    })
                    grand_total += float(line['tot_val'])
                res_data['grand_total'] = grand_total

        return Response(res_data)

    if request.method == 'POST':
        stockist_id = request.data.get('stockist_id')
        month = int(request.data.get('month', current_month))
        year = int(request.data.get('year', current_year))
        action = request.data.get('action')
        
        # 🌟 NAYA: POST mein bhi selected_emp ko handle karo
        emp_id = request.data.get('employee_id', employee.id)
        try:
            selected_emp = Employee.objects.get(id=emp_id, company=employee.company)
        except Employee.DoesNotExist:
            selected_emp = employee
            
        setting = SystemSetting.objects.filter(company=employee.company).first()
        deadline = setting.free_claim_deadline_day if setting and setting.free_claim_deadline_day else 4
        prev_month, prev_year = (12, today.year - 1) if today.month == 1 else (today.month - 1, today.year)
        is_immediate_prev_month = (month == prev_month and year == prev_year)
        is_locked = (today.day > deadline) if is_immediate_prev_month else True

        if is_locked and employee.designation not in ['Admin', 'System Administrator']:
            return Response({'success': False, 'error': f'Claim ki entry sirf {deadline} tareekh tak allow hoti hai.'}, status=403)

        if not stockist_id:
            return Response({'success': False, 'error': 'Stockist select karna zaroori hai.'}, status=400)

        stockist = get_object_or_404(Stockist, id=stockist_id, company=employee.company)
        master = FreeQtyClaimMaster.objects.filter(employee=selected_emp, stockist=stockist, month=month, year=year).first()

        if action == 'submit':
            if not master:
                return Response({'success': False, 'error': 'Pehle claim generate karein.'}, status=400)
            master.status = 'Pending_Manager'
            master.save()
            return Response({'success': True, 'message': 'Claim Manager ko submit kar diya gaya hai! 🚀'})
        
        elif action == 'generate':
            sales_agg = PartyWiseSaleLine.objects.filter(
                report__employee=selected_emp, report__stockist=stockist, report__month=month, report__year=year, free_qty__gt=0
            ).values('product_id').annotate(
                tot_billed=Sum('billed_qty'),
                tot_free=Sum('free_qty')
            )

            if not sales_agg.exists():
                return Response({'success': False, 'error': 'Is stockist ki secondary sale me koi free scheme entry nahi mili.'}, status=400)

            if master:
                if master.status not in ['Draft', 'Rejected']:
                    return Response({'success': False, 'error': 'Approved ya Pending claim ko sync nahi kar sakte.'}, status=400)
                FreeQtyClaimLine.objects.filter(master=master).delete()
            else:
                master = FreeQtyClaimMaster.objects.create(
                    employee=selected_emp, stockist=stockist, month=month, year=year, status='Draft'
                )

            for s in sales_agg:
                prod = Product.objects.get(id=s['product_id'])
                price = float(prod.price) if getattr(prod, 'price', None) else 0.0
                FreeQtyClaimLine.objects.create(
                    master=master, stockist=stockist, product=prod,
                    total_billed_qty=s['tot_billed'], total_free_qty=s['tot_free'], claim_value=(s['tot_free'] * price)
                )
            
            if master.status != 'Rejected':
                master.status = 'Draft'
            master.save()

            return Response({'success': True, 'message': 'Claim successfully Generate/Sync ho gaya! 🎉'})
# ==============================================================================
# 📅 9. TOUR PLAN (MTP) REPORT
# ==============================================================================
from SFA.models import MonthlyTourProgram, DailyTourPlan

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_tour_plan_report(request):
    try:
        employee = request.user.employee
    except AttributeError:
        return Response({'error': 'Employee profile missing'}, status=400)

    selected_emp, team_employees = _resolve_selected_employee(request, employee)

    today = timezone.now().date()
    month = int(request.GET.get('month') or today.month)
    year = int(request.GET.get('year') or today.year)

    mtp_master = MonthlyTourProgram.objects.filter(employee=selected_emp, month=month, year=year).first()

    days_out = []
    status_val = 'Not Created'
    
    if mtp_master:
        status_val = mtp_master.status
        # 🌟 FIX: route__territory ko bhi select_related me daal diya taaki fast load ho
        daily_mtps = DailyTourPlan.objects.filter(mtp=mtp_master).select_related('route', 'route__territory').order_by('date')
        
        for d in daily_mtps:
            days_out.append({
                'id': d.id,
                'date': str(d.date),
                'routes': d.route.name if d.route else '',
                # 🌟 NAYA: Territory ka naam fetch karke bhej rahe hain
                'territory': d.route.territory.name if d.route and getattr(d.route, 'territory', None) else '',
            })

    return Response({
        'selected_employee': _employee_brief(selected_emp),
        'team_employees': [_employee_brief(e) for e in team_employees] if employee.designation != 'MR' else [],
        'month': month, 'year': year,
        'mtp_status': status_val,
        'days': days_out
    })

# ==============================================================================
# 💸 10. EXPENSE REPORT (Monthly & Daily Breakdown)
# ==============================================================================
from SFA.models import MonthlyExpenseReport, DailyExpense, DayStart

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_expense_report(request):
    try:
        employee = request.user.employee
    except AttributeError:
        return Response({'error': 'Employee profile missing'}, status=400)

    # 🌟 Manager Team logic
    selected_emp, team_employees = _resolve_selected_employee(request, employee)

    today = timezone.now().date()
    month_str = request.GET.get('month', '')
    year = int(request.GET.get('year') or today.year)

    qs = MonthlyExpenseReport.objects.filter(employee=selected_emp, year=year)
    if month_str:  
        qs = qs.filter(month=int(month_str))

    qs = qs.order_by('-month')

    # 🌟 NAYA LOGIC: DayStart se Route aur Territory fetch karna
    day_starts = DayStart.objects.filter(
        employee=selected_emp, date__year=year
    ).prefetch_related('routes', 'territory')
    ds_dict = {ds.date: ds for ds in day_starts}

    expenses_out = []
    
    # Python List for Excel building
    excel_expenses = []

    for exp in qs:
        lines = DailyExpense.objects.filter(monthly_report=exp).order_by('date')
        
        grand_total = 0.0
        daily_lines_out = []
        
        for line in lines:
            ta = float(line.approved_ta if line.approved_ta is not None else line.ta_amount)
            da = float(line.approved_da if line.approved_da is not None else line.da_amount)
            misc = float(line.approved_misc if line.approved_misc is not None else line.misc_amount)
            total = ta + da + misc
            grand_total += total
            
            # 🌟 Route / HQ map karna
            line_date = line.date if hasattr(line.date, 'year') else line.date.date()
            ds = ds_dict.get(line_date)
            location_display = ""
            if ds:
                route_str = ", ".join([r.name for r in ds.routes.all()])
                hq_str = ds.territory.name if ds.territory else ""
                if route_str:
                    location_display = f"{route_str} ({hq_str})" if hq_str else route_str
                elif hq_str:
                    location_display = hq_str

            daily_lines_out.append({
                'date': str(line.date),
                'category': line.get_territory_category_display() if line.territory_category else 'HQ',
                'location_detail': location_display or '-',
                'distance_km': float(line.distance_km),
                'ta': ta,
                'da': da,
                'misc': misc,
                'total': total,
            })

        exp_data = {
            'id': exp.id,
            'month': exp.month,
            'year': exp.year,
            'status': exp.status,
            'is_modified': exp.is_modified,
            'manager_remark': exp.manager_remark,
            'grand_total': round(grand_total, 2),
            'daily_lines': daily_lines_out,
        }
        expenses_out.append(exp_data)
        excel_expenses.append(exp_data)

    # 📥 EXCEL EXPORT LOGIC
    if request.GET.get('export') == 'excel':
        import calendar
        month_name = calendar.month_name[int(month_str)] if month_str else "All Months"
        filename = f"Expense_Report_{selected_emp.name}_{month_name}_{year}.xlsx"
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Expense Report"

        ws.append(['MONTHLY EXPENSE REPORT'])
        ws.append(['Employee:', selected_emp.name, 'Period:', f"{month_name} {year}"])
        ws.append([''])
        
        ws['A1'].font = Font(bold=True, size=14, color="107C41")
        ws['A2'].font = Font(bold=True); ws['C2'].font = Font(bold=True)

        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")

        headers = ['Date', 'Category', 'Route / HQ', 'Distance (KM)', 'TA (Rs)', 'DA (Rs)', 'Misc (Rs)', 'Total (Rs)', 'Status']
        ws.append(headers)

        for col_num, cell in enumerate(ws[5], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 15
        ws.column_dimensions['C'].width = 30 # Route column thoda choda

        for exp in excel_expenses:
            for line in exp['daily_lines']:
                date_obj = timezone.datetime.strptime(line['date'], '%Y-%m-%d').date()
                ws.append([
                    date_obj.strftime('%d-%b-%Y'),
                    line['category'],
                    line['location_detail'],
                    line['distance_km'],
                    line['ta'], line['da'], line['misc'], line['total'],
                    exp['status']
                ])
                
        # Save to memory and return binary file
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return Response({
        'selected_employee': _employee_brief(selected_emp),
        'team_employees': [_employee_brief(e) for e in team_employees] if employee.designation != 'MR' else [],
        'selected_month': int(month_str) if month_str else None,
        'selected_year': year,
        'expenses': expenses_out
    })
    
# ==============================================================================
# 🏖️ 11. HOLIDAY LIST REPORT
# ==============================================================================
from SFA.models import Holiday
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import io
from django.http import HttpResponse

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_holiday_list(request):
    try:
        employee = request.user.employee
    except AttributeError:
        return Response({'error': 'Employee profile missing'}, status=400)

    # 🌟 Logic: Find the RBM to get region-specific holidays
    rbm_emp = None
    curr = employee
    while curr:
        if curr.designation == 'RBM':
            rbm_emp = curr
            break
        curr = curr.manager
        
    admin_ids = list(Employee.objects.filter(designation='Admin', company=employee.company).values_list('id', flat=True))  # 🌟 FIX: company-scoped
    holiday_creators = admin_ids.copy()
    if rbm_emp: 
        holiday_creators.append(rbm_emp.id)
        
    year = request.GET.get('year')
    
    holidays_qs = Holiday.objects.filter(proposed_by_id__in=holiday_creators, status='Approved').order_by('date')
    if year:
        holidays_qs = holidays_qs.filter(date__year=int(year))

    holidays_data = []
    for h in holidays_qs:
        holidays_data.append({
            'id': h.id,
            'name': h.name,
            'date': str(h.date),
            'day': h.date.strftime('%A')
        })

    region_name = rbm_emp.name if rbm_emp else "Consolidated"

    # 📥 EXCEL EXPORT LOGIC (.XLSX)
    if request.GET.get('export') == 'excel':
        filename = f"Holidays_{region_name}_{year or 'All'}.xlsx"
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Holidays"
        
        ws.append(['APPROVED HOLIDAYS LIST'])
        ws.append(['Region:', region_name])
        ws.append([''])
        
        ws['A1'].font = Font(bold=True, size=14, color="107C41")
        ws['A2'].font = Font(bold=True); ws['B2'].font = Font(bold=True)
        
        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        
        headers = ['Date', 'Day', 'Holiday Name']
        ws.append(headers)
        
        for col_num, cell in enumerate(ws[4], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 25
        
        for h in holidays_data:
            date_obj = timezone.datetime.strptime(h['date'], '%Y-%m-%d').date()
            ws.append([date_obj.strftime('%d-%b-%Y'), h['day'], h['name']])
            
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return Response({
        'region_name': region_name,
        'holidays': holidays_data
    })
