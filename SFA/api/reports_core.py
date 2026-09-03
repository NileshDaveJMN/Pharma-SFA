"""
SFA/api/reports_core.py
========================
Product-sales report, DCR report, DCR detail.
(reports.py se split kiya gaya — 1000+ line limit ke wajah se)
"""

import calendar
from collections import defaultdict
from datetime import datetime

from django.utils import timezone
from django.db.models import Sum, Count
from django.shortcuts import get_object_or_404

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status

from SFA.models import (
    Employee, Doctor, Chemist, Product, DailyDCR, DCRVisit, DCRProductDetail,
    MonthlyTourProgram, MonthlyExpenseReport, Route, Holiday,
    MonthlyTargetMaster, LeaveApplication, LeaveBalance, DayStart, SystemSetting,
    FreeQtyClaimMaster, GiftCampaignPlan, ChemistEditRequest, DoctorEditRequest,
)
from SFA.services.team import (
    get_full_team_employees,
    get_team_territory_ids,
    get_team_requested_routes,
    get_dropdown_team,
)
from SFA.views.reports import _get_target_chain_starter


from .reports_helpers import _resolve_selected_employee, _employee_brief

from django.db.models.functions import ExtractMonth

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_product_sales_report(request):
    try:
        employee = request.user.employee
    except AttributeError:
        return Response({'error': 'Employee profile missing'}, status=400)

    selected_emp, team_employees = _resolve_selected_employee(request, employee)

    today = timezone.now().date()
    from_month = int(request.GET.get('from_month') or today.month)
    to_month = int(request.GET.get('to_month') or today.month)
    selected_year = int(request.GET.get('year') or today.year)
    if from_month > to_month:
        from_month, to_month = to_month, from_month

    months_range = list(range(from_month, to_month + 1))

    # 🚀 OPTIMIZATION 2: Python (RAM) sum ki jagah SQL aggregation
    product_entries = DCRProductDetail.objects.filter(
        visit__daily_dcr__employee=selected_emp,
        visit__daily_dcr__date__month__gte=from_month,
        visit__daily_dcr__date__month__lte=to_month,
        visit__daily_dcr__date__year=selected_year,
    ).annotate(
        month=ExtractMonth('visit__daily_dcr__date')
    ).values('product_id', 'month').annotate(
        sq=Sum('sample_qty'),
        oq=Sum('order_qty')
    )

    products_dict = {
        p.id: {'name': p.name, 'price': float(p.price) if getattr(p, 'price', None) else 0.0}
        for p in Product.objects.filter(company=employee.company)
    }

    agg_data = defaultdict(lambda: {
        'monthly': {m: {'samples': 0, 'orders': 0, 'val': 0.0} for m in months_range},
        'tot_samples': 0, 'tot_orders': 0, 'tot_val': 0.0,
    })
    gt_monthly = {m: {'samples': 0, 'orders': 0, 'val': 0.0} for m in months_range}
    gt_samples, gt_orders, gt_val = 0, 0, 0.0

    # 🚀 Ab memory me hazaron rows ki jagah sirf kuch sau summarized rows aayengi
    for entry in product_entries:
        p_id, m = entry['product_id'], entry['month']
        sq, oq = entry['sq'] or 0, entry['oq'] or 0
        if p_id not in products_dict:
            continue
            
        price = products_dict[p_id]['price']
        val = oq * price

        agg_data[p_id]['monthly'][m]['samples'] += sq
        agg_data[p_id]['monthly'][m]['orders'] += oq
        agg_data[p_id]['monthly'][m]['val'] += val
        agg_data[p_id]['tot_samples'] += sq
        agg_data[p_id]['tot_orders'] += oq
        agg_data[p_id]['tot_val'] += val

        gt_monthly[m]['samples'] += sq
        gt_monthly[m]['orders'] += oq
        gt_monthly[m]['val'] += val
        gt_samples += sq
        gt_orders += oq
        gt_val += val

    products_out = []
    for p_id, p_info in products_dict.items():
        if p_id in agg_data and (agg_data[p_id]['tot_samples'] > 0 or agg_data[p_id]['tot_orders'] > 0):
            p_data = agg_data[p_id]
            products_out.append({
                'product_name': p_info['name'],
                'price': p_info['price'],
                'monthly': [
                    {
                        'month': calendar.month_name[m][:3],
                        'samples': p_data['monthly'][m]['samples'],
                        'orders': p_data['monthly'][m]['orders'],
                        'value': round(p_data['monthly'][m]['val'], 2),
                    } for m in months_range
                ],
                'total_samples': p_data['tot_samples'],
                'total_orders': p_data['tot_orders'],
                'total_value': round(p_data['tot_val'], 2),
            })
    products_out.sort(key=lambda x: x['product_name'])

    return Response({
        'selected_employee': _employee_brief(selected_emp),
        'team_employees': [_employee_brief(e) for e in team_employees] if employee.designation != 'MR' else [],
        'from_month': from_month, 'to_month': to_month, 'year': selected_year,
        'months_headers': [calendar.month_name[m][:3] for m in months_range],
        'products': products_out,
        'grand_total': {
            'monthly': [
                {
                    'month': calendar.month_name[m][:3],
                    'samples': gt_monthly[m]['samples'],
                    'orders': gt_monthly[m]['orders'],
                    'value': round(gt_monthly[m]['val'], 2),
                } for m in months_range
            ],
            'samples': gt_samples, 'orders': gt_orders, 'value': round(gt_val, 2),
        },
    })

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_dcr_report(request):
    try:
        employee = request.user.employee
    except AttributeError:
        return Response({'error': 'Employee profile missing'}, status=400)

    selected_emp, team_employees = _resolve_selected_employee(request, employee)

    today = timezone.now().date()
    selected_month = int(request.GET.get('month') or today.month)
    selected_year = int(request.GET.get('year') or today.year)

    # 🚀 OPTIMIZATION 1: prefetch_related('routes__territory') lagaya taaki N+1 loop na bane
    day_starts = DayStart.objects.filter(
        employee=selected_emp, date__month=selected_month, date__year=selected_year
    ).prefetch_related('routes__territory').order_by('-date')
    
    actual_dcrs = DailyDCR.objects.filter(
        employee=selected_emp, date__month=selected_month, date__year=selected_year
    ).prefetch_related('visits__doctor', 'visits__chemist', 'visits__product_details__product')
    
    dcr_dict = {d.date: d for d in actual_dcrs}

    dcrs_out = []
    total_dr_visits, total_chem_visits = 0, 0

    for ds in day_starts:
        dcr_obj = dcr_dict.get(ds.date)
        dr_v, chem_v, visit_count = 0, 0, 0
        visits_out = []

        if dcr_obj:
            all_v = list(dcr_obj.visits.all())
            dr_v = sum(1 for v in all_v if v.doctor_id)
            chem_v = sum(1 for v in all_v if v.chemist_id)
            visit_count = len(all_v)
            total_dr_visits += dr_v
            total_chem_visits += chem_v

            for v in all_v:
                visits_out.append({
                    'id': v.id,
                    'visit_type': 'Doctor' if v.doctor_id else 'Chemist',
                    'name': v.doctor.name if v.doctor_id else (v.chemist.name if v.chemist_id else None),
                    'specialty': v.doctor.get_specialty_display() if v.doctor_id and v.doctor.specialty else None,
                    'products': [
                        {
                            'product_name': pd.product.name,
                            'is_detailed': pd.is_detailed,
                            'sample_qty': pd.sample_qty or 0,
                            'order_qty': pd.order_qty or 0,
                        } for pd in v.product_details.all()
                    ],
                    'remark': v.remark,
                })

        # 🚀 Ab yeh loop RAM se chalega (Zero DB queries)
        route_list = ds.routes.all() 
        route_name = ", ".join([r.name for r in route_list]) if route_list else None
        territory_names = []
        for r in route_list:
            if r.territory and r.territory.name not in territory_names:
                territory_names.append(r.territory.name)
        territory_name = ", ".join(territory_names) if territory_names else None

        dcrs_out.append({
            'day_start_id': ds.id,
            'dcr_id': dcr_obj.id if dcr_obj else None,
            'date': str(ds.date),
            'work_type': ds.work_type,
            'route': route_name,
            'territory': territory_name,
            'has_dcr': dcr_obj is not None,
            'visit_count': visit_count,
            'dr_visit_count': dr_v,
            'chem_visit_count': chem_v,
            'visits': visits_out,
        })

    total_days_worked = day_starts.count()
    dr_avg = round(total_dr_visits / total_days_worked, 1) if total_days_worked > 0 else 0
    chem_avg = round(total_chem_visits / total_days_worked, 1) if total_days_worked > 0 else 0

    total_samples = DCRProductDetail.objects.filter(
        visit__daily_dcr__employee=selected_emp,
        visit__daily_dcr__date__month=selected_month,
        visit__daily_dcr__date__year=selected_year,
    ).aggregate(s=Sum('sample_qty'))['s'] or 0
    
    total_orders = DCRProductDetail.objects.filter(
        visit__daily_dcr__employee=selected_emp,
        visit__daily_dcr__date__month=selected_month,
        visit__daily_dcr__date__year=selected_year,
    ).aggregate(o=Sum('order_qty'))['o'] or 0

    return Response({
        'selected_employee': _employee_brief(selected_emp),
        'team_employees': [_employee_brief(e) for e in team_employees] if employee.designation != 'MR' else [],
        'month': selected_month, 'year': selected_year,
        'stats': {
            'total_days_worked': total_days_worked,
            'total_dr_visits': total_dr_visits,
            'total_chem_visits': total_chem_visits,
            'dr_avg': dr_avg, 'chem_avg': chem_avg,
            'total_samples': total_samples, 'total_orders': total_orders,
        },
        'dcrs': dcrs_out,
    })


# ==============================================================================
# 📄 2b. DCR REPORT — SINGLE DETAIL
# ==============================================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_dcr_detail(request, dcr_id):
    """
    Ek single DailyDCR ka pura detail — har visit, har visit ke product
    details (sample/order qty) ke saath.

    🌟 Manager/Admin apni team ke kisi bhi member ki DCR dekh sakte hain
    (get_full_team_employees se check), MR sirf apni khud ki — web ke
    list-expand jaisa hi consistent permission model.

    Response:
    {
        "id": 501, "date": "2026-06-15", "work_type": "Field Work",
        "employee": {"id": 7, "name": "Devarshi joshi", "designation": "MR"},
        "visits": [
            {
                "id": 1001, "visit_type": "Doctor", "name": "Dr. Sharma",
                "specialty": "GP", "visit_time": "10:30 AM",
                "products": [
                    {"product_name": "PPI", "sample_qty": 2, "order_qty": 5}
                ]
            }
        ]
    }

    Error (403): { "error": "Ye DCR aapki team ki nahi hai." }
    """
    try:
        employee = request.user.employee
    except AttributeError:
        return Response({'error': 'Employee profile missing'}, status=400)

    daily_dcr = get_object_or_404(
        DailyDCR.objects.select_related('employee').prefetch_related(
            'visits__product_details__product', 'visits__doctor', 'visits__chemist'
        ),
        id=dcr_id,
    )

    # 🌟 FIX: IDOR Attack Block — Direct company check
    if daily_dcr.employee.company_id != employee.company_id:
        return Response({'error': 'Access denied: Cross-company DCR access blocked.'}, status=403)

    # 🌟 Permission check: khud ki DCR, ya apni team ke kisi member ki
    allowed_ids = set(get_full_team_employees(employee).values_list('id', flat=True))
    if daily_dcr.employee_id not in allowed_ids:
        return Response({'error': 'Ye DCR aapki team ki nahi hai.'}, status=403)

    visits_out = []
    for v in daily_dcr.visits.all():
        visits_out.append({
            'id': v.id,
            'visit_type': 'Doctor' if v.doctor_id else 'Chemist',
            'name': v.doctor.name if v.doctor_id else (v.chemist.name if v.chemist_id else None),
            'specialty': v.doctor.get_specialty_display() if v.doctor_id and v.doctor.specialty else None,
            'visit_time': v.visit_time.strftime('%I:%M %p') if getattr(v, 'visit_time', None) else None,
            'products': [
                {
                    'product_name': pd.product.name,
                    'is_detailed': pd.is_detailed,  # 🌟 FIX
                    'sample_qty': pd.sample_qty or 0,
                    'order_qty': pd.order_qty or 0,
                } for pd in v.product_details.all()
            ],
            'remark': v.remark,  # 🌟 FIX
        })

    return Response({
        'id': daily_dcr.id,
        'date': str(daily_dcr.date),
        'work_type': daily_dcr.work_type,
        'employee': _employee_brief(daily_dcr.employee),
        'visits': visits_out,
    })
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.shortcuts import get_object_or_404
from django.utils import timezone
import calendar
from collections import defaultdict
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse

from SFA.models import (
    Employee, Product, MonthlyTargetMaster, TerritoryTarget, 
    StockistProductStatement, PartyWiseSaleLine
)
from SFA.services.team import get_full_team_employees
from .reports_helpers import _resolve_selected_employee, _employee_brief

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_sales_summary_report(request):
    try:
        employee = request.user.employee
    except AttributeError:
        return Response({'error': 'Employee profile missing'}, status=400)

    selected_emp, team_employees = _resolve_selected_employee(request, employee)
    
    today = timezone.now().date()
    from_month = int(request.GET.get('from_month') or today.month)
    to_month = int(request.GET.get('to_month') or today.month)
    selected_year = int(request.GET.get('year') or today.year)

    if from_month > to_month: 
        from_month, to_month = to_month, from_month
        
    months_range = list(range(from_month, to_month + 1))
    months_headers = [(m, calendar.month_name[m][:3]) for m in months_range]

    target_team = get_full_team_employees(selected_emp)
    area_territory_ids = target_team.exclude(headquarter__isnull=True).values_list('headquarter_id', flat=True)

    span = to_month - from_month + 1
    curr_start_val = selected_year * 12 + from_month
    curr_end_val = selected_year * 12 + to_month
    prev_end_val = curr_start_val - 1
    prev_start_val = prev_end_val - span + 1
    ly_start_val = (selected_year - 1) * 12 + from_month
    ly_end_val = (selected_year - 1) * 12 + to_month

    years_needed = list(set([selected_year, selected_year - 1, (prev_start_val - 1) // 12]))

    team_territories = [emp.headquarter for emp in target_team if emp.headquarter]
    approved_territory_ids = MonthlyTargetMaster.objects.filter(
        territory__in=team_territories, status='Approved', year__in=years_needed
    ).values_list('territory_id', flat=True)

    all_targets = TerritoryTarget.objects.filter(territory_id__in=approved_territory_ids, year__in=years_needed).select_related('product')
    all_primary = StockistProductStatement.objects.filter(stockist__territory_id__in=area_territory_ids, year__in=years_needed).select_related('product')
    all_secondary = PartyWiseSaleLine.objects.filter(report__stockist__territory_id__in=area_territory_ids, report__year__in=years_needed).select_related('product')

    def calc_growth(curr, prev):
        if prev == 0: return 100.0 if curr > 0 else 0.0
        return ((curr - prev) / prev) * 100.0

    def calc_ach(val, tgt):
        if tgt == 0: return 100.0 if val > 0 else 0.0
        return (val / tgt) * 100.0

    products = Product.objects.filter(company=selected_emp.company).order_by('name')
    p_dict = {}
    
    gt = {
        'monthly': {m: {'t_qty':0, 't_val':0.0, 'p_qty':0, 'p_val':0.0, 's_qty':0, 's_val':0.0} for m in months_range},
        't_qty': 0, 't_val': 0.0, 'p_qty': 0, 'p_val': 0.0, 's_qty': 0, 's_val': 0.0,
        'prev_p_val': 0.0, 'prev_s_val': 0.0, 'ly_p_val': 0.0, 'ly_s_val': 0.0
    }

    for p in products:
        price = float(p.price) if getattr(p, 'price', None) else 0.0
        p_dict[p.id] = {
            'name': p.name, 'price': price,
            'monthly': {m: {'t_qty':0, 't_val':0.0, 'p_qty':0, 'p_val':0.0, 's_qty':0, 's_val':0.0} for m in months_range},
            'curr_t_qty': 0, 'curr_t_val': 0.0, 'curr_p_qty': 0, 'curr_p_val': 0.0, 'curr_s_qty': 0, 'curr_s_val': 0.0,
            'prev_p_val': 0.0, 'prev_s_val': 0.0, 'ly_p_val': 0.0, 'ly_s_val': 0.0
        }

    for t in all_targets:
        val = t.year * 12 + t.month
        if curr_start_val <= val <= curr_end_val and t.month in p_dict[t.product_id]['monthly']:
            t_val = t.target_qty * float(t.product.price)
            p_dict[t.product_id]['monthly'][t.month]['t_qty'] += t.target_qty; p_dict[t.product_id]['monthly'][t.month]['t_val'] += t_val
            p_dict[t.product_id]['curr_t_qty'] += t.target_qty; p_dict[t.product_id]['curr_t_val'] += t_val
            gt['monthly'][t.month]['t_qty'] += t.target_qty; gt['monthly'][t.month]['t_val'] += t_val
            gt['t_qty'] += t.target_qty; gt['t_val'] += t_val

    for p in all_primary:
        val = p.year * 12 + p.month
        p_val = p.received_qty * float(p.product.price)
        if curr_start_val <= val <= curr_end_val and p.month in p_dict[p.product_id]['monthly']:
            p_dict[p.product_id]['monthly'][p.month]['p_qty'] += p.received_qty; p_dict[p.product_id]['monthly'][p.month]['p_val'] += p_val
            p_dict[p.product_id]['curr_p_qty'] += p.received_qty; p_dict[p.product_id]['curr_p_val'] += p_val
            gt['monthly'][p.month]['p_qty'] += p.received_qty; gt['monthly'][p.month]['p_val'] += p_val
            gt['p_qty'] += p.received_qty; gt['p_val'] += p_val
        elif prev_start_val <= val <= prev_end_val:
            p_dict[p.product_id]['prev_p_val'] += p_val; gt['prev_p_val'] += p_val
        elif ly_start_val <= val <= ly_end_val:
            p_dict[p.product_id]['ly_p_val'] += p_val; gt['ly_p_val'] += p_val

    for s in all_secondary:
        val = s.report.year * 12 + s.report.month
        s_val = s.billed_qty * float(s.product.price)
        s_qty = s.billed_qty + s.free_qty
        if curr_start_val <= val <= curr_end_val and s.report.month in p_dict[s.product_id]['monthly']:
            p_dict[s.product_id]['monthly'][s.report.month]['s_qty'] += s_qty; p_dict[s.product_id]['monthly'][s.report.month]['s_val'] += s_val
            p_dict[s.product_id]['curr_s_qty'] += s_qty; p_dict[s.product_id]['curr_s_val'] += s_val
            gt['monthly'][s.report.month]['s_qty'] += s_qty; gt['monthly'][s.report.month]['s_val'] += s_val
            gt['s_qty'] += s_qty; gt['s_val'] += s_val
        elif prev_start_val <= val <= prev_end_val:
            p_dict[s.product_id]['prev_s_val'] += s_val; gt['prev_s_val'] += s_val
        elif ly_start_val <= val <= ly_end_val:
            p_dict[s.product_id]['ly_s_val'] += s_val; gt['ly_s_val'] += s_val

    report_data = []
    for pid, d in p_dict.items():
        if d['curr_t_val'] > 0 or d['curr_p_val'] > 0 or d['curr_s_val'] > 0:
            d['p_ach'] = calc_ach(d['curr_p_val'], d['curr_t_val'])
            d['s_ach'] = calc_ach(d['curr_s_val'], d['curr_t_val'])
            d['p_m2m'] = calc_growth(d['curr_p_val'], d['prev_p_val'])
            d['p_y2y'] = calc_growth(d['curr_p_val'], d['ly_p_val'])
            d['s_m2m'] = calc_growth(d['curr_s_val'], d['prev_s_val'])
            d['s_y2y'] = calc_growth(d['curr_s_val'], d['ly_s_val'])
            # Convert dict to list for Flutter
            d['monthly_list'] = [d['monthly'][m] for m in months_range]
            report_data.append(d)

    gt['p_ach'] = calc_ach(gt['p_val'], gt['t_val'])
    gt['s_ach'] = calc_ach(gt['s_val'], gt['t_val'])
    gt['p_m2m'] = calc_growth(gt['p_val'], gt['prev_p_val'])
    gt['p_y2y'] = calc_growth(gt['p_val'], gt['ly_p_val'])
    gt['s_m2m'] = calc_growth(gt['s_val'], gt['prev_s_val'])
    gt['s_y2y'] = calc_growth(gt['s_val'], gt['ly_s_val'])
    gt['monthly_list'] = [gt['monthly'][m] for m in months_range]

    # Excel Export
    if request.GET.get('export') == 'excel':
        filename = f"Sales_Summary_{selected_emp.name}_M{from_month}-M{to_month}_{selected_year}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sales Summary Matrix"
        
        hq_name = selected_emp.headquarter.name if selected_emp.headquarter else "N/A"
        period_str = f"{calendar.month_name[from_month][:3]} to {calendar.month_name[to_month][:3]} {selected_year}"
        
        ws.append(['MASTER SALES SUMMARY REPORT'])
        ws.append(['Employee:', selected_emp.name, 'HQ:', hq_name])
        ws.append(['Period:', period_str])
        ws.append([''])
        
        ws['A1'].font = Font(bold=True, size=14, color="107C41")
        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        
        headers1 = ['Product Name']
        headers2 = ['']
        for m in months_headers:
            headers1.extend([f"{m[1]} DATA", "", "", "", "", ""])
            headers2.extend(['Tgt Qty', 'Tgt (₹)', 'Pri Qty', 'Pri (₹)', 'Sec Qty', 'Sec (₹)'])
            
        headers1.extend(['OVERALL GRAND TOTAL', '', '', '', '', '', '', '', '', '', '', ''])
        headers2.extend(['Tot Tgt Qty', 'Tot Tgt (₹)', 'Tot Pri Qty', 'Tot Pri (₹)', 'Tot Sec Qty', 'Tot Sec (₹)', 'Pri Ach%', 'Pri M2M%', 'Pri Y2Y%', 'Sec Ach%', 'Sec M2M%', 'Sec Y2Y%'])
        
        ws.append(headers1)
        ws.append(headers2)
        
        for row_idx in [5, 6]:
            for col_num, cell in enumerate(ws[row_idx], 1):
                cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
                if row_idx == 6: ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 18 if col_num == 1 else 12
            
        for row in report_data:
            row_data = [row['name']]
            for m_data in row['monthly_list']:
                row_data.extend([m_data['t_qty'], round(m_data['t_val'], 2), m_data['p_qty'], round(m_data['p_val'], 2), m_data['s_qty'], round(m_data['s_val'], 2)])
            row_data.extend([
                row['curr_t_qty'], round(row['curr_t_val'], 2), row['curr_p_qty'], round(row['curr_p_val'], 2), row['curr_s_qty'], round(row['curr_s_val'], 2),
                f"{round(row['p_ach'], 1)}%", f"{round(row['p_m2m'], 1)}%", f"{round(row['p_y2y'], 1)}%", f"{round(row['s_ach'], 1)}%", f"{round(row['s_m2m'], 1)}%", f"{round(row['s_y2y'], 1)}%"
            ])
            ws.append(row_data)
            
        gt_row = ['OVERALL TOTAL']
        for m_data in gt['monthly_list']:
            gt_row.extend([m_data['t_qty'], round(m_data['t_val'], 2), m_data['p_qty'], round(m_data['p_val'], 2), m_data['s_qty'], round(m_data['s_val'], 2)])
        gt_row.extend([
            gt['t_qty'], round(gt['t_val'], 2), gt['p_qty'], round(gt['p_val'], 2), gt['s_qty'], round(gt['s_val'], 2),
            f"{round(gt['p_ach'], 1)}%", f"{round(gt['p_m2m'], 1)}%", f"{round(gt['p_y2y'], 1)}%", f"{round(gt['s_ach'], 1)}%", f"{round(gt['s_m2m'], 1)}%", f"{round(gt['s_y2y'], 1)}%"
        ])
        ws.append(gt_row)
        for cell in ws[ws.max_row]: cell.font = Font(bold=True)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return Response({
        'selected_employee': _employee_brief(selected_emp),
        'team_employees': [_employee_brief(e) for e in team_employees] if employee.designation != 'MR' else [],
        'months_choices': [(i, calendar.month_name[i]) for i in range(1, 13)],
        'months_headers': months_headers,
        'report_data': report_data,
        'gt': gt,
    })
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_primary_sales_report(request):
    try:
        employee = request.user.employee
    except AttributeError:
        return Response({'error': 'Employee profile missing'}, status=400)

    selected_emp, team_employees = _resolve_selected_employee(request, employee)
    
    today = timezone.now().date()
    from_month = int(request.GET.get('from_month') or today.month)
    to_month = int(request.GET.get('to_month') or today.month)
    selected_year = int(request.GET.get('year') or today.year)

    if from_month > to_month: 
        from_month, to_month = to_month, from_month
        
    months_range = list(range(from_month, to_month + 1))
    months_headers = [(m, calendar.month_name[m][:3]) for m in months_range]

    target_team = get_full_team_employees(selected_emp)
    area_territory_ids = target_team.exclude(headquarter__isnull=True).values_list('headquarter_id', flat=True)

    # 🌟 FIX: PrimarySale model ko bhi import kiya
    from SFA.models import StockistProductStatement, Product, PrimarySale
    
    # ---------------------------------------------------------
    # 1. AGGREGATED DATA (Stockist & Product Tabs ke liye)
    # ---------------------------------------------------------
    sales = StockistProductStatement.objects.filter(
        stockist__territory_id__in=area_territory_ids, 
        year=selected_year, month__in=months_range
    ).select_related('stockist', 'product')

    party_dict = {}
    prod_dict = {}
    gt = {m: {'qty': 0, 'val': 0.0} for m in months_range}
    gt_total_qty, gt_total_val = 0, 0.0

    for s in sales:
        val = round(float(s.received_qty) * float(s.product.price), 2) if getattr(s.product, 'price', None) else 0.0

        # Party Data Grouping
        pid = s.stockist_id
        if pid not in party_dict:
            party_dict[pid] = {'party_name': s.stockist.name, 'products': {}}
        if s.product_id not in party_dict[pid]['products']:
            party_dict[pid]['products'][s.product_id] = {
                'name': s.product.name,
                'monthly': {m: {'qty': 0, 'val': 0.0} for m in months_range},
                'total_qty': 0, 'total_val': 0.0
            }
        
        party_dict[pid]['products'][s.product_id]['monthly'][s.month]['qty'] += s.received_qty
        party_dict[pid]['products'][s.product_id]['monthly'][s.month]['val'] += val
        party_dict[pid]['products'][s.product_id]['total_qty'] += s.received_qty
        party_dict[pid]['products'][s.product_id]['total_val'] += val

        # Product Data Grouping
        if s.product_id not in prod_dict:
            prod_dict[s.product_id] = {
                'product_name': s.product.name,
                'monthly': {m: {'qty': 0, 'val': 0.0} for m in months_range},
                'total_qty': 0, 'total_val': 0.0
            }
        prod_dict[s.product_id]['monthly'][s.month]['qty'] += s.received_qty
        prod_dict[s.product_id]['monthly'][s.month]['val'] += val
        prod_dict[s.product_id]['total_qty'] += s.received_qty
        prod_dict[s.product_id]['total_val'] += val

        # Grand Totals
        gt[s.month]['qty'] += s.received_qty
        gt[s.month]['val'] += val
        gt_total_qty += s.received_qty
        gt_total_val += val

    # ---------------------------------------------------------
    # 2. RAW DATA (Raw Tab ke liye - PrimarySale Model se)
    # ---------------------------------------------------------
    raw_sales_query = PrimarySale.objects.filter(
        stockist__territory_id__in=area_territory_ids,
        date__year=selected_year,
        date__month__in=months_range
    ).select_related('stockist', 'product').order_by('-date')

    raw_sales = []
    for rs in raw_sales_query:
        val = round(float(rs.quantity) * float(rs.product.price), 2) if getattr(rs.product, 'price', None) else 0.0
        raw_sales.append({
            'date': rs.date.strftime('%d-%b-%y'),
            'stockist': rs.stockist.name,
            'product': rs.product.name,
            'qty': rs.quantity,
            'val': val,  # 🌟 FIX: Value add ki
            'batch': getattr(rs, 'batch_number', '-')
        })    
    # Convert Dicts to Lists for JSON
    party_out = []
    for pid, pdata in party_dict.items():
        prod_list = []
        # 🌟 FIX: Stockist level ka total calculate karne ke liye variables
        st_monthly = {m: {'qty': 0, 'val': 0.0} for m in months_range}
        st_tot_qty, st_tot_val = 0, 0.0
        
        for pr_id, pr_data in pdata['products'].items():
            pr_data['monthly_list'] = [pr_data['monthly'][m] for m in months_range]
            prod_list.append(pr_data)
            
            # 🌟 FIX: Add to stockist totals
            for m in months_range:
                st_monthly[m]['qty'] += pr_data['monthly'][m]['qty']
                st_monthly[m]['val'] += pr_data['monthly'][m]['val']
            st_tot_qty += pr_data['total_qty']
            st_tot_val += pr_data['total_val']
            
        party_out.append({
            'party_name': pdata['party_name'], 
            'products': prod_list,
            'monthly_list': [{'qty': st_monthly[m]['qty'], 'val': round(st_monthly[m]['val'], 2)} for m in months_range],
            'total_qty': st_tot_qty,
            'total_val': round(st_tot_val, 2)      # ✅ round ka ( ) dono round-brackets
    })
    prod_out = []
    for pr_id, pr_data in prod_dict.items():
        pr_data['monthly_list'] = [pr_data['monthly'][m] for m in months_range]
        prod_out.append(pr_data)

    gt_monthly_list = [gt[m] for m in months_range]

    # Handle Excel Export
    if request.GET.get('export') == 'excel':
        import openpyxl
        import io
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Primary Sales"
        ws.append(["Primary Sales Data Export"])
        # Baaki excel ka logic aap same rakh sakte hain
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Primary_Sales_{selected_year}.xlsx"'
        return response

    return Response({
        'selected_employee': _employee_brief(selected_emp),
        'team_employees': [_employee_brief(e) for e in team_employees] if employee.designation != 'MR' else [],
        'months_choices': [(i, calendar.month_name[i]) for i in range(1, 13)],
        'months_headers': months_headers,
        
        'party_list': party_out,
        'product_list': prod_out,
        'primary_sales': raw_sales,
        
        'gt_monthly_list': gt_monthly_list,
        'grand_total_qty': gt_total_qty,
        'grand_total_value': round(gt_total_val, 2)
    })
# 🌟 NAYE IMPORTS (Agar file me pehle se nahi hain)
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import io
import calendar
from django.http import HttpResponse
from SFA.models import DoctorRxMapping

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_dr_wise_sale_report(request):
    try:
        employee = request.user.employee
    except AttributeError:
        return Response({'error': 'Employee profile missing'}, status=400)

    selected_emp, team_employees = _resolve_selected_employee(request, employee)
    
    today = timezone.now().date()
    from_month = int(request.GET.get('from_month') or today.month)
    to_month = int(request.GET.get('to_month') or today.month)
    selected_year = int(request.GET.get('year') or today.year)

    if from_month > to_month: 
        from_month, to_month = to_month, from_month
    months_range = list(range(from_month, to_month + 1))
    months_headers = [(m, calendar.month_name[m][:3]) for m in months_range]

    mappings = DoctorRxMapping.objects.filter(
        party_line__report__month__gte=from_month, party_line__report__month__lte=to_month,
        party_line__report__year=selected_year, party_line__report__employee=selected_emp
    ).select_related('doctor', 'party_line__chemist', 'party_line__product')

    agg_data = defaultdict(lambda: {'doctor_name': '', 'specialty': '', 'chemist_name': '', 'product_name': '', 'price': 0.0, 'monthly_data': {m: {'billed': 0, 'free': 0, 'val': 0.0} for m in months_range}, 'total_billed': 0, 'total_free': 0, 'total_value': 0.0})
    product_wise_data = defaultdict(lambda: {'product_name': '', 'doctors': defaultdict(lambda: {'doc_name': '', 'specialty': '', 'monthly': {m: {'billed': 0, 'free': 0, 'val': 0.0} for m in months_range}, 'tot_b': 0, 'tot_f': 0, 'tot_val': 0.0}), 'tot_b': 0, 'tot_f': 0, 'tot_val': 0.0})
    gt_monthly = {m: {'billed': 0, 'free': 0, 'val': 0.0} for m in months_range}
    gt_billed, gt_free, gt_value = 0, 0, 0.0

    for m in mappings:
        sale_month = m.party_line.report.month
        doc_id, doc_name, specialty = m.doctor.id, m.doctor.name, m.doctor.specialty or '-'
        # 🌟 FIX: Agar chemist null hai (Direct Doctor Sale) toh crash nahi hoga
        chem_id = m.party_line.chemist.id if m.party_line.chemist else 0
        chem_name = m.party_line.chemist.name if m.party_line.chemist else 'Direct Sale'
        prod_id, prod_name = m.party_line.product.id, m.party_line.product.name
        b_qty, f_qty = m.mapped_billed_qty, m.mapped_free_qty
        price = float(m.party_line.product.price) if getattr(m.party_line.product, 'price', None) else 0.0
        val = b_qty * price
        
        key = f"{doc_id}_{chem_id}_{prod_id}"
        if agg_data[key]['doctor_name'] == '':
            agg_data[key].update({'doctor_name': doc_name, 'specialty': specialty, 'chemist_name': chem_name, 'product_name': prod_name, 'price': price})
            
        agg_data[key]['monthly_data'][sale_month]['billed'] += b_qty
        agg_data[key]['monthly_data'][sale_month]['free'] += f_qty
        agg_data[key]['monthly_data'][sale_month]['val'] += val
        agg_data[key]['total_billed'] += b_qty
        agg_data[key]['total_free'] += f_qty
        agg_data[key]['total_value'] += val
        
        pw = product_wise_data[prod_id]
        pw['product_name'] = prod_name
        dw = pw['doctors'][doc_id]
        dw['doc_name'] = doc_name; dw['specialty'] = specialty
        
        dw['monthly'][sale_month]['billed'] += b_qty
        dw['monthly'][sale_month]['free'] += f_qty
        dw['monthly'][sale_month]['val'] += val
        dw['tot_b'] += b_qty; dw['tot_f'] += f_qty; dw['tot_val'] += val
        pw['tot_b'] += b_qty; pw['tot_f'] += f_qty; pw['tot_val'] += val
        
        gt_monthly[sale_month]['billed'] += b_qty
        gt_monthly[sale_month]['free'] += f_qty
        gt_monthly[sale_month]['val'] += val
        gt_billed += b_qty; gt_free += f_qty; gt_value += val

    report_data = list(agg_data.values())
    report_data.sort(key=lambda x: (x['doctor_name'], x['chemist_name'], x['product_name']))
    for row in report_data: 
        row['monthly_data_list'] = [row['monthly_data'][m[0]] for m in months_headers]
        
    doc_grouped_data = []
    current_doc = None
    curr_group = None
    
    for row in report_data:
        if row['doctor_name'] != current_doc:
            if curr_group: doc_grouped_data.append(curr_group)
            current_doc = row['doctor_name']
            curr_group = {
                'doctor_name': current_doc,
                'specialty': row['specialty'],
                'doc_total_val': 0.0,
                'items': []
            }
        curr_group['items'].append(row)
        curr_group['doc_total_val'] += row['total_value']
        
    if curr_group: doc_grouped_data.append(curr_group)

    prod_tab_list = []
    for p_id, p_data in product_wise_data.items():
        doc_list = []
        for d_id, d_data in p_data['doctors'].items():
            d_data['monthly_list'] = [d_data['monthly'][m] for m in months_range]
            doc_list.append(d_data)
        doc_list.sort(key=lambda x: x['doc_name'])
        p_data['doctors_list'] = doc_list
        prod_tab_list.append(p_data)
    prod_tab_list.sort(key=lambda x: x['product_name'])
        
    gt_monthly_list = [gt_monthly[m[0]] for m in months_headers]

    # 📊 EXCEL EXPORT LOGIC
    if request.GET.get('export') == 'excel':
        filename = f"Dr_Rx_Report_{selected_emp.name}_M{from_month}-M{to_month}_{selected_year}.xlsx"
        wb = openpyxl.Workbook()
        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        period_str = f"{calendar.month_name[from_month][:3]} to {calendar.month_name[to_month][:3]} {selected_year}"

        ws1 = wb.active
        ws1.title = "Doctor Wise Rx"
        ws1.append(['DOCTOR WISE RX SALE (MONTH TREND)'])
        ws1.append(['Employee:', selected_emp.name, 'Period:', period_str])
        ws1.append([''])
        ws1['A1'].font = Font(bold=True, size=14, color="107C41")
        
        headers1 = ['Doctor Name', 'Specialty', 'Chemist Name', 'Product Name', 'Price (₹)']
        for m_num, m_name in months_headers: headers1.extend([f'{m_name} Billed', f'{m_name} Free', f'{m_name} Value (₹)'])
        headers1.extend(['Total Billed', 'Total Free', 'Total Value (₹)'])
        ws1.append(headers1)
        
        for col_num, cell in enumerate(ws1[4], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws1.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 15 if col_num > 4 else 20
            
        for row in report_data:
            row_data = [f"Dr. {row['doctor_name']}", row['specialty'], row['chemist_name'], row['product_name'], round(row['price'], 2)]
            for m_data in row['monthly_data_list']: row_data.extend([m_data['billed'], m_data['free'], round(m_data['val'], 2)])
            row_data.extend([row['total_billed'], row['total_free'], round(row['total_value'], 2)])
            ws1.append(row_data)

        ws2 = wb.create_sheet(title="Product Wise Rx")
        ws2.append(['PRODUCT WISE -> DOCTOR RX SALE (MONTH TREND)'])
        ws2.append(['Employee:', selected_emp.name, 'Period:', period_str])
        ws2.append([''])
        ws2['A1'].font = Font(bold=True, size=14, color="107C41")
        
        headers2 = ['Product Name', 'Doctor Name', 'Specialty']
        for m_num, m_name in months_headers: headers2.extend([f'{m_name} Billed', f'{m_name} Free', f'{m_name} Value (₹)'])
        headers2.extend(['Total Billed', 'Total Free', 'Total Value (₹)'])
        ws2.append(headers2)

        for col_num, cell in enumerate(ws2[4], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws2.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 15 if col_num > 3 else 25

        for prod in prod_tab_list:
            for doc in prod['doctors_list']:
                row_data = [prod['product_name'], f"Dr. {doc['doc_name']}", doc['specialty']]
                for m_data in doc['monthly_list']: row_data.extend([m_data['billed'], m_data['free'], round(m_data['val'], 2)])
                row_data.extend([doc['tot_b'], doc['tot_f'], round(doc['tot_val'], 2)])
                ws2.append(row_data)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return Response({
        'team_employees': [_employee_brief(e) for e in team_employees] if employee.designation != 'MR' else [],
        'months_choices': [(i, calendar.month_name[i]) for i in range(1, 13)],
        'months_headers': months_headers,
        'doc_grouped_data': doc_grouped_data,
        'prod_tab_list': prod_tab_list,
        'gt_monthly_list': gt_monthly_list,
        'gt_value': gt_value
    })
# 🌟 NAYE IMPORTS (Agar file me pehle se nahi hain)
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import io
import calendar
from django.http import HttpResponse
from SFA.models import Stockist, PartyWiseSaleReport, PartyWiseSaleLine

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_party_rx_report(request):
    try:
        employee = request.user.employee
    except AttributeError:
        return Response({'error': 'Employee profile missing'}, status=400)

    selected_emp, team_employees = _resolve_selected_employee(request, employee)
    
    today = timezone.now().date()
    from_month = int(request.GET.get('from_month') or today.month)
    to_month = int(request.GET.get('to_month') or today.month)
    selected_year = int(request.GET.get('year') or today.year)
    selected_stockist_id = request.GET.get('stockist_id', '')

    if from_month > to_month: 
        from_month, to_month = to_month, from_month
    months_range = list(range(from_month, to_month + 1))
    months_headers = [(m, calendar.month_name[m][:3]) for m in months_range]

    stockists = Stockist.objects.filter(territory=selected_emp.headquarter).order_by('name') if selected_emp.headquarter else Stockist.objects.none()

    report_filter = {'employee': selected_emp, 'month__gte': from_month, 'month__lte': to_month, 'year': selected_year}
    if selected_stockist_id: report_filter['stockist_id'] = selected_stockist_id

    reports = PartyWiseSaleReport.objects.filter(**report_filter)
    lines = PartyWiseSaleLine.objects.filter(report__in=reports).select_related('chemist', 'product', 'report__stockist')

    party_wise_data = defaultdict(lambda: {'chemist_name': '', 'stockist_name': '', 'monthly': {m: {'b_qty': 0, 'f_qty': 0, 'val': 0.0} for m in months_range}, 'tot_b': 0, 'tot_f': 0, 'tot_val': 0.0, 'products': defaultdict(lambda: {'monthly': {m: {'b_qty': 0, 'f_qty': 0, 'val': 0.0} for m in months_range}, 'tot_b': 0, 'tot_f': 0, 'tot_val': 0.0})})
    product_wise_data = defaultdict(lambda: {'product_name': '', 'monthly': {m: {'b_qty': 0, 'f_qty': 0, 'val': 0.0} for m in months_range}, 'tot_b': 0, 'tot_f': 0, 'tot_val': 0.0})
    product_chemist_data = defaultdict(lambda: {'product_name': '', 'chemists': defaultdict(lambda: {'chemist_name': '', 'stockist_name': '', 'monthly': {m: {'b_qty': 0, 'f_qty': 0, 'val': 0.0} for m in months_range}, 'tot_b': 0, 'tot_f': 0, 'tot_val': 0.0})})
    
    gt_monthly = {m: {'b_qty': 0, 'f_qty': 0, 'val': 0.0} for m in months_range}
    gt_b_qty, gt_f_qty, gt_val = 0, 0, 0.0
    raw_logs = []

    for line in lines:
        r_month = line.report.month
        # 🌟 FIX: Agar chemist null hai (Direct Doctor Sale) toh crash nahi hoga
        ch_id = line.chemist.id if line.chemist else 0
        ch_name = line.chemist.name if line.chemist else 'Direct Sale'
        st_name = line.report.stockist.name
        pr_id, pr_name = line.product.id, line.product.name
        b_qty, f_qty = line.billed_qty, line.free_qty
        price = float(line.product.price) if getattr(line.product, 'price', None) else 0.0
        val = b_qty * price
            
        raw_logs.append({
            'date': f"{calendar.month_name[r_month][:3]}-{selected_year}", 
            'chemist': ch_name, 'stockist': st_name, 'product': pr_name, 
            'billed': b_qty, 'free': f_qty, 'value': val
        })

        # Party Wise Grouping
        pw = party_wise_data[ch_id]
        pw['chemist_name'] = ch_name
        pw['stockist_name'] = st_name
        pw['monthly'][r_month]['b_qty'] += b_qty; pw['monthly'][r_month]['f_qty'] += f_qty; pw['monthly'][r_month]['val'] += val
        pw['tot_b'] += b_qty; pw['tot_f'] += f_qty; pw['tot_val'] += val
        
        pd = pw['products'][pr_name]
        pd['monthly'][r_month]['b_qty'] += b_qty; pd['monthly'][r_month]['f_qty'] += f_qty; pd['monthly'][r_month]['val'] += val
        pd['tot_b'] += b_qty; pd['tot_f'] += f_qty; pd['tot_val'] += val

        # Product Wise Totals
        pr_data = product_wise_data[pr_id]
        pr_data['product_name'] = pr_name
        pr_data['monthly'][r_month]['b_qty'] += b_qty; pr_data['monthly'][r_month]['f_qty'] += f_qty; pr_data['monthly'][r_month]['val'] += val
        pr_data['tot_b'] += b_qty; pr_data['tot_f'] += f_qty; pr_data['tot_val'] += val

        # Product -> Chemist Grouping
        pc_data = product_chemist_data[pr_id]
        pc_data['product_name'] = pr_name
        pcc_data = pc_data['chemists'][ch_id]
        pcc_data['chemist_name'] = ch_name
        pcc_data['stockist_name'] = st_name
        pcc_data['monthly'][r_month]['b_qty'] += b_qty; pcc_data['monthly'][r_month]['f_qty'] += f_qty; pcc_data['monthly'][r_month]['val'] += val
        pcc_data['tot_b'] += b_qty; pcc_data['tot_f'] += f_qty; pcc_data['tot_val'] += val

        gt_monthly[r_month]['b_qty'] += b_qty; gt_monthly[r_month]['f_qty'] += f_qty; gt_monthly[r_month]['val'] += val
        gt_b_qty += b_qty; gt_f_qty += f_qty; gt_val += val

    party_list = []
    for ch_id, ch_data in party_wise_data.items():
        prod_list = []
        for pr_name, pd in ch_data['products'].items():
            prod_list.append({'name': pr_name, 'monthly_list': [pd['monthly'][m] for m in months_range], 'tot_b': pd['tot_b'], 'tot_f': pd['tot_f'], 'tot_val': pd['tot_val']})
        prod_list.sort(key=lambda x: x['name'])
        party_list.append({
            'chemist_name': ch_data['chemist_name'], 'stockist_name': ch_data['stockist_name'], 'products': prod_list,
            'monthly_list': [ch_data['monthly'][m] for m in months_range],
            'tot_b': ch_data['tot_b'], 'tot_f': ch_data['tot_f'], 'tot_val': ch_data['tot_val']
        })
    party_list.sort(key=lambda x: x['chemist_name'])

    product_list_grouped = []
    for pr_id, pc_data in product_chemist_data.items():
        chem_list = []
        for ch_id, cd in pc_data['chemists'].items():
            chem_list.append({
                'chemist_name': cd['chemist_name'], 'stockist_name': cd['stockist_name'],
                'monthly_list': [cd['monthly'][m] for m in months_range],
                'tot_b': cd['tot_b'], 'tot_f': cd['tot_f'], 'tot_val': cd['tot_val']
            })
        chem_list.sort(key=lambda x: x['chemist_name'])
        pr_totals = product_wise_data[pr_id]
        product_list_grouped.append({
            'product_name': pc_data['product_name'], 'chemists': chem_list,
            'monthly_list': [pr_totals['monthly'][m] for m in months_range],
            'tot_b': pr_totals['tot_b'], 'tot_f': pr_totals['tot_f'], 'tot_val': pr_totals['tot_val']
        })
    product_list_grouped.sort(key=lambda x: x['product_name'])

    gt_monthly_list = [gt_monthly[m] for m in months_range]

    # 📊 EXCEL EXPORT LOGIC
    if request.GET.get('export') == 'excel':
        filename = f"Party_Sales_{selected_emp.name}_M{from_month}-M{to_month}_{selected_year}.xlsx"
        wb = openpyxl.Workbook()
        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        period_str = f"{calendar.month_name[from_month][:3]} to {calendar.month_name[to_month][:3]} {selected_year}"
        group_fill = PatternFill(start_color="0B5E31", end_color="0B5E31", fill_type="solid")
        group_font = Font(color="FFFFFF", bold=True)
        subtotal_fill = PatternFill(start_color="EEF6F1", end_color="EEF6F1", fill_type="solid")

        # TAB 1: Party Wise
        ws1 = wb.active
        ws1.title = "Party Wise Sale"
        ws1.append(['PARTY WISE SALE REPORT (MONTH TREND)'])
        ws1.append(['Employee:', selected_emp.name, 'Period:', period_str])
        ws1.append([''])
        ws1['A1'].font = Font(bold=True, size=12, color="107C41")
        
        headers1 = ['Product Name']
        for m_num, m_name in months_headers: headers1.extend([f'{m_name} Billed', f'{m_name} Free', f'{m_name} Value (₹)'])
        headers1.extend(['Total Billed', 'Total Free', 'Total Value (₹)'])
        ws1.append(headers1)
        total_cols1 = len(headers1)
        for col_num, cell in enumerate(ws1[4], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws1.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 15 if col_num > 1 else 25

        for party in party_list:
            ws1.append([f"💊 {party['chemist_name']} — {party['stockist_name']}"])
            r = ws1.max_row
            ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=total_cols1)
            ws1.cell(row=r, column=1).fill = group_fill; ws1.cell(row=r, column=1).font = group_font

            for prod in party['products']:
                row_data = [prod['name']]
                for m_data in prod['monthly_list']: row_data.extend([m_data['b_qty'], m_data['f_qty'], round(m_data['val'], 2)])
                row_data.extend([prod['tot_b'], prod['tot_f'], round(prod['tot_val'], 2)])
                ws1.append(row_data)

            sub_row = ['Subtotal']
            for m_data in party['monthly_list']: sub_row.extend([m_data['b_qty'], m_data['f_qty'], round(m_data['val'], 2)])
            sub_row.extend([party['tot_b'], party['tot_f'], round(party['tot_val'], 2)])
            ws1.append(sub_row)
            r = ws1.max_row
            for c in range(1, total_cols1 + 1):
                ws1.cell(row=r, column=c).fill = subtotal_fill; ws1.cell(row=r, column=c).font = Font(bold=True)

        gt_row1 = ['GRAND TOTAL']
        for m_gt in gt_monthly_list: gt_row1.extend([m_gt['b_qty'], m_gt['f_qty'], round(m_gt['val'], 2)])
        gt_row1.extend([gt_b_qty, gt_f_qty, round(gt_val, 2)])
        ws1.append(gt_row1)
        for cell in ws1[ws1.max_row]: cell.font = Font(bold=True)

        # TAB 2: Product Wise
        ws2 = wb.create_sheet(title="Product Wise Sale")
        ws2.append(['PRODUCT WISE SALE REPORT (MONTH TREND)'])
        ws2.append(['Employee:', selected_emp.name, 'Period:', period_str])
        ws2.append([''])
        ws2['A1'].font = Font(bold=True, size=12, color="107C41")
        
        headers2 = ['Chemist Name', 'Stockist Name']
        for m_num, m_name in months_headers: headers2.extend([f'{m_name} Billed', f'{m_name} Free', f'{m_name} Value (₹)'])
        headers2.extend(['Total Billed', 'Total Free', 'Total Value (₹)'])
        ws2.append(headers2)
        total_cols2 = len(headers2)
        for col_num, cell in enumerate(ws2[4], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws2.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 15 if col_num > 2 else 25

        for prod in product_list_grouped:
            ws2.append([f"📦 {prod['product_name']}"])
            r = ws2.max_row
            ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=total_cols2)
            ws2.cell(row=r, column=1).fill = group_fill; ws2.cell(row=r, column=1).font = group_font

            for chem in prod['chemists']:
                row_data = [chem['chemist_name'], chem['stockist_name']]
                for m_data in chem['monthly_list']: row_data.extend([m_data['b_qty'], m_data['f_qty'], round(m_data['val'], 2)])
                row_data.extend([chem['tot_b'], chem['tot_f'], round(chem['tot_val'], 2)])
                ws2.append(row_data)

            sub_row = ['Subtotal', '']
            for m_data in prod['monthly_list']: sub_row.extend([m_data['b_qty'], m_data['f_qty'], round(m_data['val'], 2)])
            sub_row.extend([prod['tot_b'], prod['tot_f'], round(prod['tot_val'], 2)])
            ws2.append(sub_row)
            r = ws2.max_row
            for c in range(1, total_cols2 + 1):
                ws2.cell(row=r, column=c).fill = subtotal_fill; ws2.cell(row=r, column=c).font = Font(bold=True)

        gt_row2 = ['GRAND TOTAL', '']
        for m_gt in gt_monthly_list: gt_row2.extend([m_gt['b_qty'], m_gt['f_qty'], round(m_gt['val'], 2)])
        gt_row2.extend([gt_b_qty, gt_f_qty, round(gt_val, 2)])
        ws2.append(gt_row2)
        for cell in ws2[ws2.max_row]: cell.font = Font(bold=True)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return Response({
        'team_employees': [_employee_brief(e) for e in team_employees] if employee.designation != 'MR' else [],
        'stockists': [{'id': s.id, 'name': s.name} for s in stockists],
        'months_choices': [(i, calendar.month_name[i]) for i in range(1, 13)],
        'months_headers': months_headers,
        'party_list': party_list,
        'product_list_grouped': product_list_grouped,
        'raw_logs': raw_logs,
        'gt_monthly_list': gt_monthly_list,
        'gt_b_qty': gt_b_qty, 'gt_f_qty': gt_f_qty, 'gt_val': gt_val
    })
# 🌟 NAYE IMPORTS (Agar file me pehle se nahi hain)
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import io
import calendar
from django.http import HttpResponse
from SFA.models import DoctorROILedger, DoctorRxMapping

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_doctor_roi_report(request):
    try:
        employee = request.user.employee
    except AttributeError:
        return Response({'error': 'Employee profile missing'}, status=400)

    selected_emp, team_employees = _resolve_selected_employee(request, employee)
    is_manager_view = employee.designation != 'MR'

    today = timezone.now().date()
    from_month = int(request.GET.get('from_month') or today.month)
    to_month = int(request.GET.get('to_month') or today.month)
    selected_year = int(request.GET.get('year') or today.year)

    if from_month > to_month:
        from_month, to_month = to_month, from_month
        
    months_range = list(range(from_month, to_month + 1))
    months_headers = [(m, calendar.month_name[m][:3]) for m in months_range]

    selected_emp_id = str(selected_emp.id)
    inv_type = request.GET.get('inv_type', 'any_gift')

    if is_manager_view and selected_emp_id: 
        qs = DoctorROILedger.objects.filter(employee_id=selected_emp_id)
        base_docs = Doctor.objects.filter(allocated_to_id=selected_emp_id)
    elif is_manager_view: 
        qs = DoctorROILedger.objects.filter(employee__in=team_employees)
        base_docs = Doctor.objects.filter(allocated_to__in=team_employees)
    else: 
        qs = DoctorROILedger.objects.filter(employee=employee)
        base_docs = Doctor.objects.filter(allocated_to=employee)

    qs = qs.filter(
        date_given__month__gte=from_month,
        date_given__month__lte=to_month,
        date_given__year=selected_year
    ).select_related('doctor', 'employee', 'item')

    if inv_type == 'high_value':
        qs = qs.filter(item__item_type='HighValue')
    elif inv_type == 'normal':
        qs = qs.filter(item__item_type='Routine')

    agg_data = defaultdict(lambda: {
        'doc_name': '', 'items': set(), 'emp_name': '', 'doc_id': 0,
        'monthly': {m: {'inv_qty': 0, 'inv_val': 0.0} for m in months_range},
        'tot_inv_qty': 0, 'tot_inv_val': 0.0
    })
    
    for row in qs:
        m = row.date_given.month
        key = row.doctor_id 
        
        if not agg_data[key]['doc_name']:
            agg_data[key]['doc_name'] = row.doctor.name
            agg_data[key]['emp_name'] = row.employee.name if row.employee else 'Unknown'
            agg_data[key]['doc_id'] = row.doctor_id
            
        agg_data[key]['items'].add((row.item.name, row.item.item_type))
        agg_data[key]['monthly'][m]['inv_qty'] += row.quantity
        agg_data[key]['monthly'][m]['inv_val'] += float(row.total_value)
        agg_data[key]['tot_inv_qty'] += row.quantity
        agg_data[key]['tot_inv_val'] += float(row.total_value)

    if inv_type == 'all_doctors':
        for doc in base_docs.select_related('allocated_to'):
            if doc.id not in agg_data:
                agg_data[doc.id]['doc_name'] = doc.name
                agg_data[doc.id]['emp_name'] = doc.allocated_to.name if doc.allocated_to else 'Unknown'
                agg_data[doc.id]['doc_id'] = doc.id
                agg_data[doc.id]['items'].add(('No Gifts', 'Routine'))

    doc_ids = set(agg_data.keys())

    # 🛡️ RX bhi WOHI employee scope use karega jo upar ledger (qs) use karta hai
    # — warna shared doctors pe dusre MR ki sale RX mein ghus jati thi.
    if is_manager_view and selected_emp_id:
        rx_scope = {'party_line__report__employee_id': selected_emp_id}
    elif is_manager_view:
        rx_scope = {'party_line__report__employee__in': team_employees}
    else:
        rx_scope = {'party_line__report__employee': employee}

    rx_qs = DoctorRxMapping.objects.filter(
        doctor_id__in=doc_ids,
        party_line__report__month__gte=from_month,
        party_line__report__month__lte=to_month,
        party_line__report__year=selected_year,
        **rx_scope,   # 🌟 SCOPE ADD
    ).select_related('party_line__product')    
    
    rx_dict = defaultdict(lambda: defaultdict(float))
    for rx in rx_qs:
        m = rx.party_line.report.month
        price = float(rx.party_line.product.price) if getattr(rx.party_line.product, 'price', None) else 0.0
        rx_dict[rx.doctor_id][m] += (rx.mapped_billed_qty * price)

    report_data = []
    gt_monthly = {m: {'inv_qty': 0, 'inv_val': 0.0, 'rx_val': 0.0} for m in months_range}
    gt_inv_qty, gt_inv_val, gt_rx_val = 0, 0.0, 0.0
    
    for doc_id, data in agg_data.items():
        monthly_list = []
        doc_tot_rx = 0.0
        
        for m in months_range:
            i_q = data['monthly'][m]['inv_qty']
            i_v = data['monthly'][m]['inv_val']
            r_v = rx_dict[doc_id][m]
            doc_tot_rx += r_v
            
            roi = (r_v / i_v * 100) if i_v > 0 else 0.0
            monthly_list.append({
                'inv_qty': i_q, 'inv_val': i_v, 'rx_val': r_v, 'roi': roi
            })
            
            gt_monthly[m]['inv_qty'] += i_q
            gt_monthly[m]['inv_val'] += i_v
            gt_monthly[m]['rx_val'] += r_v 
                
        gt_inv_qty += data['tot_inv_qty']
        gt_inv_val += data['tot_inv_val']
        gt_rx_val += doc_tot_rx
        
        tot_roi = (doc_tot_rx / data['tot_inv_val'] * 100) if data['tot_inv_val'] > 0 else 0.0
        
        item_names = " + ".join(sorted([i[0] for i in data['items']]))
        item_types = set([i[1] for i in data['items']])
        final_item_type = 'HighValue' if 'HighValue' in item_types else 'Routine'

        report_data.append({
            'doc_name': data['doc_name'],
            'item_name': item_names,
            'item_type': final_item_type,
            'emp_name': data['emp_name'],
            'monthly_list': monthly_list,
            'tot_inv_qty': data['tot_inv_qty'],
            'tot_inv_val': data['tot_inv_val'],
            'tot_rx_val': doc_tot_rx,
            'tot_roi': tot_roi
        })
        
    report_data.sort(key=lambda x: x['doc_name'])
    gt_monthly_list = [gt_monthly[m] for m in months_range]
    overall_roi = (gt_rx_val / gt_inv_val * 100) if gt_inv_val > 0 else 0.0

    if request.GET.get('export') == 'excel':
        filename = f"Doctor_ROI_Matrix_{calendar.month_abbr[from_month]}_to_{calendar.month_abbr[to_month]}_{selected_year}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ROI Matrix"

        ws.append(['DOCTOR ROI MATRIX (MONTH-WISE)'])
        ws.append(['Period:', f"{calendar.month_name[from_month]} to {calendar.month_name[to_month]} {selected_year}"])
        ws.append(['Filter:', inv_type.replace('_', ' ').title()])
        ws.append([''])
        ws['A1'].font = Font(bold=True, size=14, color="107C41")

        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")

        headers1 = ['Doctor Name', 'Item Details', 'Given By']
        headers2 = ['', '', '']
        
        for m_num, m_name in months_headers:
            headers1.extend([f"{m_name}", "", "", ""])
            headers2.extend(['Inv Qty', 'Inv Val', 'Rx Val', 'ROI %'])
            
        headers1.extend(["GRAND TOTAL", "", "", ""])
        headers2.extend(['Inv Qty', 'Inv Val', 'Rx Val', 'ROI %'])

        ws.append(headers1)
        ws.append(headers2)
        
        for col_num, cell in enumerate(ws[5], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 15 if col_num > 3 else 25

        for row in report_data:
            row_data = [f"Dr. {row['doc_name']}", row['item_name'], row['emp_name']]
            for m in row['monthly_list']:
                row_data.extend([m['inv_qty'], m['inv_val'], m['rx_val'], round(m['roi'], 1)])
            row_data.extend([row['tot_inv_qty'], row['tot_inv_val'], row['tot_rx_val'], round(row['tot_roi'], 1)])
            ws.append(row_data)

        gt_row = ['GRAND TOTAL', '', '']
        for m_gt in gt_monthly_list:
            m_roi = (m_gt['rx_val'] / m_gt['inv_val'] * 100) if m_gt['inv_val'] > 0 else 0.0
            gt_row.extend([m_gt['inv_qty'], m_gt['inv_val'], m_gt['rx_val'], round(m_roi, 1)])
        gt_row.extend([gt_inv_qty, gt_inv_val, gt_rx_val, round(overall_roi, 1)])
        ws.append(gt_row)
        
        for cell in ws[ws.max_row]: cell.font = Font(bold=True)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return Response({
        'team_employees': [_employee_brief(e) for e in team_employees] if is_manager_view else [],
        'is_manager_view': is_manager_view,
        'months_choices': [(i, calendar.month_name[i]) for i in range(1, 13)],
        'months_headers': months_headers,
        'report_data': report_data,
        'gt_monthly_list': gt_monthly_list,
        'gt_inv_qty': gt_inv_qty, 'gt_inv_val': gt_inv_val,
        'gt_rx_val': gt_rx_val, 'overall_roi': overall_roi
    })
# 🌟 NAYE IMPORTS (Agar file me pehle se nahi hain)
from datetime import date
from django.db.models import Count, Q
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import io
import calendar
from django.http import HttpResponse
from SFA.models import Doctor, Chemist, DayStart, LeaveApplication, DailyDCR

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_analysis_hub(request):
    try:
        employee = request.user.employee
    except AttributeError:
        return Response({'error': 'Employee profile missing'}, status=400)

    team_employees = get_dropdown_team(employee, ordered=False)
    is_manager_view = employee.designation != 'MR'

    default_emp_id = str(employee.id)
    if is_manager_view:
        first_sub = team_employees.exclude(id=employee.id).first()
        if first_sub: default_emp_id = str(first_sub.id)

    selected_emp_id = request.GET.get('employee_id', default_emp_id)
    try:
        selected_emp = Employee.objects.get(id=int(selected_emp_id), company=employee.company)
    except (Employee.DoesNotExist, ValueError):
        return Response({'error': 'Employee not found'}, status=404)

    today = timezone.now().date()
    from_month = int(request.GET.get('from_month') or today.month)
    to_month = int(request.GET.get('to_month') or today.month)
    selected_year = int(request.GET.get('year') or today.year)

    if from_month > to_month: from_month, to_month = to_month, from_month

    months_range = list(range(from_month, to_month + 1))
    months_headers = [(m, calendar.month_name[m][:3]) for m in months_range]

    is_mr = selected_emp.designation == 'MR'
    total_doctors = Doctor.objects.filter(allocated_to=selected_emp, status='Approved').count() if is_mr else 0
    total_chemists = Chemist.objects.filter(allocated_to=selected_emp, status='Approved').count() if is_mr else 0

    metrics = {
        'field_days': {'label': '🏃‍♂️ Field Days', 'data': {m: 0 for m in months_range}, 'tot': 0},
        'joint_days': {'label': '🤝 Joint Days', 'data': {m: 0 for m in months_range}, 'tot': 0},
        'meeting_transit': {'label': '🏢 Meet/Transit', 'data': {m: 0 for m in months_range}, 'tot': 0},
        'leaves': {'label': '🏖️ Leaves', 'data': {m: 0 for m in months_range}, 'tot': 0},
        'holidays': {'label': '⛱️ Holidays', 'data': {m: 0 for m in months_range}, 'tot': 0},
        'pob': {'label': '💰 Tot POB', 'data': {m: 0 for m in months_range}, 'tot': 0.0, 'is_currency': True},
        'dr_avg': {'label': '👨‍⚕️ Dr Avg', 'data': {m: 0 for m in months_range}, 'tot': 0.0, 'is_avg': True},
        'chem_avg': {'label': '💊 Chem Avg', 'data': {m: 0 for m in months_range}, 'tot': 0.0, 'is_avg': True},
    }

    if is_mr:
        metrics['dr_in_list'] = {'label': '📋 Tot Docs in List', 'data': {m: 0 for m in months_range}, 'tot': total_doctors}
        metrics['dr_visited'] = {'label': '✅ Visited Docs', 'data': {m: 0 for m in months_range}, 'tot': 0}
        metrics['dr_cov'] = {'label': '🎯 Dr Cov %', 'data': {m: 0 for m in months_range}, 'tot': 0.0, 'is_pct': True}
        
        metrics['chem_in_list'] = {'label': '📋 Tot Chems in List', 'data': {m: 0 for m in months_range}, 'tot': total_chemists}
        metrics['chem_visited'] = {'label': '✅ Visited Chems', 'data': {m: 0 for m in months_range}, 'tot': 0}
        metrics['chem_cov'] = {'label': '🎯 Chem Cov %', 'data': {m: 0 for m in months_range}, 'tot': 0.0, 'is_pct': True}

    gt_dr_visits, gt_chem_visits, gt_field_days = 0, 0, 0
    gt_visited_docs, gt_visited_chems = set(), set()
    joint_workers = {}

    for m in months_range:
        num_days = calendar.monthrange(selected_year, m)[1]
        
        day_starts = DayStart.objects.filter(employee=selected_emp, date__month=m, date__year=selected_year).select_related('joint_worked_with')
        
        f_days = 0; j_days = 0; m_t_days = 0; h_days = 0
        for ds in day_starts:
            if ds.work_type == 'Field Work':
                f_days += 1
                if ds.joint_worked_with:
                    j_days += 1
                    jw_name = ds.joint_worked_with.name
                    if jw_name not in joint_workers:
                        joint_workers[jw_name] = {'label': f'↳ Joint: {jw_name}', 'data': {mm: 0 for mm in months_range}, 'tot': 0, 'is_sub_row': True}
                    joint_workers[jw_name]['data'][m] += 1
                    joint_workers[jw_name]['tot'] += 1
            elif ds.work_type in ['Meeting', 'Transit']:
                m_t_days += 1
            elif ds.work_type == 'Holiday':
                h_days += 1
        
        l_days = 0
        leaves = LeaveApplication.objects.filter(
            employee=selected_emp, status='Approved', start_date__lte=date(selected_year, m, num_days), end_date__gte=date(selected_year, m, 1)
        )
        for l in leaves:
            start = max(l.start_date, date(selected_year, m, 1))
            end = min(l.end_date, date(selected_year, m, num_days))
            l_days += (end - start).days + 1
            
        dcrs = DailyDCR.objects.filter(employee=selected_emp, date__month=m, date__year=selected_year).prefetch_related('visits__doctor', 'visits__chemist', 'visits__product_details__product')
        m_dr_visits, m_chem_visits, m_pob = 0, 0, 0.0
        m_visited_docs, m_visited_chems = set(), set()
        
        for dcr in dcrs:
            for v in dcr.visits.all():
                if v.doctor:
                    m_dr_visits += 1; m_visited_docs.add(v.doctor_id); gt_visited_docs.add(v.doctor_id)
                if v.chemist:
                    m_chem_visits += 1; m_visited_chems.add(v.chemist_id); gt_visited_chems.add(v.chemist_id)
                for pd in v.product_details.all():
                    price = float(pd.product.price) if getattr(pd.product, 'price', None) else 0.0
                    m_pob += pd.order_qty * price
                    
        gt_dr_visits += m_dr_visits; gt_chem_visits += m_chem_visits; gt_field_days += f_days
        
        metrics['field_days']['data'][m] = f_days; metrics['field_days']['tot'] += f_days
        metrics['joint_days']['data'][m] = j_days; metrics['joint_days']['tot'] += j_days
        metrics['meeting_transit']['data'][m] = m_t_days; metrics['meeting_transit']['tot'] += m_t_days
        metrics['leaves']['data'][m] = l_days; metrics['leaves']['tot'] += l_days
        metrics['holidays']['data'][m] = h_days; metrics['holidays']['tot'] += h_days
        metrics['pob']['data'][m] = m_pob; metrics['pob']['tot'] += m_pob
        metrics['dr_avg']['data'][m] = round(m_dr_visits / f_days, 1) if f_days > 0 else 0
        metrics['chem_avg']['data'][m] = round(m_chem_visits / f_days, 1) if f_days > 0 else 0
        
        if is_mr:
            metrics['dr_in_list']['data'][m] = total_doctors
            metrics['dr_visited']['data'][m] = len(m_visited_docs)
            metrics['dr_cov']['data'][m] = round((len(m_visited_docs) / total_doctors * 100), 1) if total_doctors > 0 else 0
            
            metrics['chem_in_list']['data'][m] = total_chemists
            metrics['chem_visited']['data'][m] = len(m_visited_chems)
            metrics['chem_cov']['data'][m] = round((len(m_visited_chems) / total_chemists * 100), 1) if total_chemists > 0 else 0

    metrics['dr_avg']['tot'] = round(gt_dr_visits / gt_field_days, 1) if gt_field_days > 0 else 0
    metrics['chem_avg']['tot'] = round(gt_chem_visits / gt_field_days, 1) if gt_field_days > 0 else 0
    
    if is_mr:
        metrics['dr_in_list']['tot'] = total_doctors
        metrics['dr_visited']['tot'] = len(gt_visited_docs)
        metrics['dr_cov']['tot'] = round((len(gt_visited_docs) / total_doctors * 100), 1) if total_doctors > 0 else 0
        
        metrics['chem_in_list']['tot'] = total_chemists
        metrics['chem_visited']['tot'] = len(gt_visited_chems)
        metrics['chem_cov']['tot'] = round((len(gt_visited_chems) / total_chemists * 100), 1) if total_chemists > 0 else 0

    report_data = []
    def add_metric(key):
        if key in metrics:
            item = metrics[key]
            item['monthly_list'] = [item['data'][m] for m in months_range]
            report_data.append(item)

    add_metric('field_days')
    add_metric('joint_days')
    
    for jw_name in sorted(joint_workers.keys()):
        jw_item = joint_workers[jw_name]
        jw_item['monthly_list'] = [jw_item['data'][m] for m in months_range]
        report_data.append(jw_item)
        
    add_metric('meeting_transit')
    add_metric('leaves')
    add_metric('holidays')
    add_metric('pob')
    
    if is_mr:
        add_metric('dr_in_list')
        add_metric('dr_visited')
        add_metric('dr_avg')
        add_metric('dr_cov')
        add_metric('chem_in_list')
        add_metric('chem_visited')
        add_metric('chem_avg')
        add_metric('chem_cov')
    else:
        add_metric('dr_avg')
        add_metric('chem_avg')

    if request.GET.get('export') == 'excel':
        filename = f"Analytical_Matrix_{selected_emp.name}_M{from_month}-M{to_month}_{selected_year}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Analytics Matrix"
        
        period_str = f"{calendar.month_name[from_month][:3]} to {calendar.month_name[to_month][:3]} {selected_year}"
        ws.append(['ANALYTICAL HUB - PERFORMANCE TREND MATRIX'])
        ws.append(['Employee:', selected_emp.name, 'Period:', period_str])
        ws.append([''])
        ws['A1'].font = Font(bold=True, size=14, color="107C41")
        
        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        
        headers = ['Key Performance Indicator (KPI)']
        for m_num, m_name in months_headers: headers.append(m_name)
        headers.append('Grand Total / Average')
        ws.append(headers)
        
        for col_num, cell in enumerate(ws[5], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 35 if col_num == 1 else 18
            
        for row in report_data:
            clean_label = row['label'].split(' ', 1)[1] if ' ' in row['label'] else row['label']
            row_data = [clean_label]
            for val in row['monthly_list']:
                if row.get('is_currency'): val = f"₹{round(val, 2)}"
                elif row.get('is_pct'): val = f"{round(val, 1)}%"
                elif row.get('is_avg'): val = round(val, 1)
                row_data.append(val)
                
            tot = row['tot']
            if row.get('is_currency'): tot = f"₹{round(tot, 2)}"
            elif row.get('is_pct'): tot = f"{round(tot, 1)}%"
            elif row.get('is_avg'): tot = round(tot, 1)
            row_data.append(tot)
            ws.append(row_data)
            
        for cell in ws[ws.max_row]: cell.font = Font(bold=True)
            
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return Response({
        'team_employees': [{'id': e.id, 'name': e.name} for e in team_employees] if is_manager_view else [],
        'is_manager_view': is_manager_view,
        'months_choices': [(i, calendar.month_name[i]) for i in range(1, 13)],
        'months_headers': months_headers,
        'report_data': report_data
    })

from collections import defaultdict
from datetime import datetime
import calendar
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from SFA.models import Employee, Doctor, DCRVisit
from SFA.services.team import get_dropdown_team

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_doctor_visit_history(request):
    try:
        employee = request.user.employee
    except AttributeError:
        return Response({'error': 'Employee profile missing'}, status=400)

    team_employees = get_dropdown_team(employee)
    is_manager_view = employee.designation != 'MR'
    
    default_emp_id = str(employee.id)
    if is_manager_view:
        first_sub = team_employees.exclude(id=employee.id).first()
        if first_sub: default_emp_id = str(first_sub.id)
            
    selected_emp_id = request.GET.get('employee_id', default_emp_id)
    
    try:
        selected_emp = Employee.objects.get(id=int(selected_emp_id), company=employee.company)
    except (Employee.DoesNotExist, ValueError):
        selected_emp = employee
    
    today = datetime.today()
    from_month = int(request.GET.get('from_month', 1))
    to_month = int(request.GET.get('to_month', today.month))
    year = int(request.GET.get('year', today.year))
    
    if from_month > to_month: 
        from_month, to_month = to_month, from_month
    
    months_range = list(range(from_month, to_month + 1))
    months_headers = [(m, calendar.month_name[m][:3]) for m in months_range]
    
    doctors = Doctor.objects.filter(allocated_to=selected_emp, status='Approved').order_by('name')
    visits = DCRVisit.objects.filter(
        daily_dcr__employee=selected_emp, 
        daily_dcr__date__year=year, 
        daily_dcr__date__month__gte=from_month, 
        daily_dcr__date__month__lte=to_month, 
        doctor__isnull=False
    ).select_related('doctor', 'daily_dcr').order_by('daily_dcr__date')
    
    visit_dict = defaultdict(lambda: defaultdict(list))
    for v in visits: 
        visit_dict[v.doctor_id][v.daily_dcr.date.month].append(v.daily_dcr.date.strftime("%d"))
        
    report_data = []
    for doc in doctors:
        doc_months_list = []
        total_visits = 0
        for m in months_range:
            dates = visit_dict[doc.id].get(m, [])
            dates_str = ", ".join(dates) if dates else "-"
            doc_months_list.append({'name': calendar.month_name[m][:3], 'dates': dates_str})
            total_visits += len(dates)
            
        report_data.append({
            'doctor_name': doc.name, 
            'specialty': doc.get_specialty_display() if getattr(doc, 'specialty', None) else 'N/A', 
            'months_list': doc_months_list, 
            'total_visits': total_visits
        })

    if request.GET.get('export') == 'xlsx':
        filename = f"Dr_Visit_Matrix_{selected_emp.name}_{year}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Visit Matrix"
        
        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        
        headers = ['Doctor Name', 'Specialty']
        for m in months_range: headers.append(calendar.month_name[m])
        headers.append('Total Visits')
        ws.append(headers)
        
        for col_num, cell in enumerate(ws[1], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20 if col_num <= 2 else 15

        for row in report_data:
            data_row = [f"Dr. {row['doctor_name']}", row['specialty']]
            for m_data in row['months_list']: data_row.append(m_data['dates'])
            data_row.append(row['total_visits'])
            ws.append(data_row)
            for cell in ws[ws.max_row][2:]: cell.alignment = center_align

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return Response({
        'team_employees': [{'id': e.id, 'name': e.name} for e in team_employees] if is_manager_view else [],
        'is_manager_view': is_manager_view,
        'selected_emp_id': selected_emp_id,
        'months_headers': months_headers,
        'report_data': report_data,
        'year': year
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_route_report(request):
    from django.db.models import Count, Q

    try:
        employee = request.user.employee
    except AttributeError:
        return Response({'error': 'Employee profile missing'}, status=400)

    team_employees = get_dropdown_team(employee)
    is_manager_view = employee.designation != 'MR'

    default_emp_id = str(employee.id)
    if is_manager_view:
        first_sub = team_employees.exclude(id=employee.id).first()
        if first_sub: default_emp_id = str(first_sub.id)

    selected_emp_id = request.GET.get('employee_id', default_emp_id)
    try:
        selected_emp = Employee.objects.get(id=int(selected_emp_id), company=employee.company)   # 🛡️ IDOR fix
    except (Employee.DoesNotExist, ValueError):
        selected_emp = employee

    sub_team = get_dropdown_team(selected_emp, ordered=False)
    my_terr_ids = get_team_territory_ids(sub_team)

    # 🚀 N+1 KILLED: route-wise doctor/chemist counts DB-level annotate se
    routes = get_team_requested_routes(sub_team, my_terr_ids).select_related('territory').annotate(
        doc_count=Count('doctor_set', filter=Q(doctor_set__status='Approved'), distinct=True),
        chem_count=Count('chemist_set', filter=Q(chemist_set__status='Approved'), distinct=True),
    )

    report_data = []
    gt_docs = 0
    gt_chems = 0

    for r in routes:   # 🚀 loop mein ZERO queries
        doc_count, chem_count = r.doc_count, r.chem_count
        report_data.append({
            'route_name': r.name,
            'territory': r.territory.name if r.territory else 'N/A',
            'category': r.get_category_display() if r.category else 'HQ',
            'distance': float(r.distance_from_hq or 0),
            'doc_count': doc_count,
            'chem_count': chem_count,
            'total_customers': doc_count + chem_count,
        })
        gt_docs += doc_count
        gt_chems += chem_count

    report_data = sorted(report_data, key=lambda x: (x['territory'], x['route_name']))

    if request.GET.get('export') == 'xlsx':
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse

        filename = f"Route_Report_{selected_emp.name}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Route Coverage"

        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")

        headers = ['Route Name', 'Territory / HQ', 'Category', 'Distance (KM)', 'Total Doctors', 'Total Chemists', 'Total Customers']
        ws.append(headers)

        for col_num, cell in enumerate(ws[1], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20 if col_num <= 2 else 15

        for row in report_data:
            ws.append([row['route_name'], row['territory'], row['category'], row['distance'], row['doc_count'], row['chem_count'], row['total_customers']])

        ws.append(['GRAND TOTAL', '', '', '', gt_docs, gt_chems, gt_docs + gt_chems])
        for cell in ws[ws.max_row]: cell.font = Font(bold=True)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return Response({
        'team_employees': [{'id': e.id, 'name': e.name} for e in team_employees] if is_manager_view else [],
        'is_manager_view': is_manager_view,
        'selected_emp_id': selected_emp_id,
        'report_data': report_data,
        'gt_docs': gt_docs,
        'gt_chems': gt_chems,
        'gt_total': gt_docs + gt_chems,
    })
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_product_master(request):
    try:
        employee = request.user.employee
    except AttributeError:
        return Response({'error': 'Employee profile missing'}, status=400)

    products = Product.objects.filter(company=employee.company).order_by('name')

    if request.GET.get('export') == 'xlsx':
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse

        filename = "Product_Master_List.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Product List"
        ws.append(['PRODUCT MASTER REPORT'])
        ws.append(['Generated By:', employee.name])
        ws.append([''])
        ws['A1'].font = Font(bold=True, size=14, color="107C41")
        ws['A2'].font = Font(bold=True)

        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")

        headers = ['Sr No.', 'Product Name', 'Pack Size', 'MRP (₹)', 'PTR (₹)', 'PTS (₹)', 'GST Slab (%)']
        ws.append(headers)
        for col_num, cell in enumerate(ws[4], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 25 if col_num == 2 else 15

        for idx, p in enumerate(products, 1):
            mrp_val = float(p.mrp) if getattr(p, 'mrp', None) else 0.0
            ptr_val = float(p.ptr) if getattr(p, 'ptr', None) else 0.0
            pts_val = float(p.pts) if getattr(p, 'pts', None) else 0.0
            gst_val = p.gst_slab if getattr(p, 'gst_slab', None) is not None else 0
            ws.append([idx, p.name, p.pack_size, mrp_val, ptr_val, pts_val, gst_val])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    product_list = []
    for p in products:
        product_list.append({
            'name': p.name,
            'pack_size': p.pack_size,
            'mrp': float(p.mrp) if getattr(p, 'mrp', None) else 0.0,
            'ptr': float(p.ptr) if getattr(p, 'ptr', None) else 0.0,
            'pts': float(p.pts) if getattr(p, 'pts', None) else 0.0,
            'gst_slab': p.gst_slab if getattr(p, 'gst_slab', None) is not None else 0,
        })

    return Response({'products': product_list})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_mr_inventory(request):
    from SFA.models import PromoItem, PromoDispatch, MRInventory, DoctorROILedger

    try:
        employee = request.user.employee
    except AttributeError:
        return Response({'error': 'Employee profile missing'}, status=400)

    team_employees = get_dropdown_team(employee)
    is_manager_view = employee.designation != 'MR'

    default_emp_id = str(employee.id)
    if is_manager_view:
        first_sub = team_employees.exclude(id=employee.id).first()
        if first_sub: default_emp_id = str(first_sub.id)

    selected_emp_id = request.GET.get('employee_id', default_emp_id)
    try:
        selected_emp = Employee.objects.get(id=int(selected_emp_id), company=employee.company)   # 🛡️ IDOR fix
    except (Employee.DoesNotExist, ValueError):
        selected_emp = employee

    # ── Stock Tab ──────────────────────────────────────────────────────────
    stock_rows = MRInventory.objects.filter(employee=selected_emp).select_related('item')

    # 🚀 N+1 KILLED: Saare aggregates loop se PEHLE, 3 hi queries mein:
    recv_map = {}
    for item_id, qty in PromoDispatch.objects.filter(employee=selected_emp, status='Received').values_list('item_id', 'quantity'):
        recv_map[item_id] = recv_map.get(item_id, 0) + qty

    dist_map = {}
    ledger_map = {}   # (doctor_id, item_id, month, year) → qty — HV tracker ke liye
    for item_id, doc_id, gm, gy, qty in DoctorROILedger.objects.filter(employee=selected_emp).values_list('item_id', 'doctor_id', 'date_given__month', 'date_given__year', 'quantity'):
        dist_map[item_id] = dist_map.get(item_id, 0) + qty
        ledger_map[(doc_id, item_id, gm, gy)] = ledger_map.get((doc_id, item_id, gm, gy), 0) + qty

    sample_gift_data = []
    hv_stock_data = []
    for row in stock_rows:   # 🚀 loop mein ZERO queries
        received = recv_map.get(row.item_id, 0)
        distributed = dist_map.get(row.item_id, 0)
        entry = {
            'item_name': row.item.name,
            'category': row.item.get_item_type_display(),
            'received': received,
            'distributed': distributed,
            'balance': row.stock_qty,
        }
        if row.item.item_type == 'HighValue':
            hv_stock_data.append(entry)
        else:
            sample_gift_data.append(entry)
    sample_gift_data.sort(key=lambda x: x['item_name'])
    hv_stock_data.sort(key=lambda x: x['item_name'])

    # ── HV Gifts Tab ───────────────────────────────────────────────────────
    hv_plans = GiftCampaignPlan.objects.filter(
        employee=selected_emp, item__item_type='HighValue'
    ).select_related('doctor', 'item').order_by('-year', '-month')

    hv_tracker = []
    for plan in hv_plans:   # 🚀 loop mein ZERO queries — ledger_map se
        dist_qty = ledger_map.get((plan.doctor_id, plan.item_id, plan.month, plan.year), 0)
        final_status = 'Given' if dist_qty >= 1 else plan.get_status_display()
        hv_tracker.append({
            'doctor_name': plan.doctor.name,
            'item_name': plan.item.name,
            'month_year': f"{calendar.month_name[plan.month][:3]} {plan.year}",
            'dist_qty': dist_qty,
            'final_status': final_status,
        })

    # ── Transit Tab ────────────────────────────────────────────────────────
    transit_qs = PromoDispatch.objects.filter(
        employee=selected_emp, status='In-Transit'
    ).select_related('item').order_by('-dispatch_date')

    in_transit_items = [{
        'id': d.id,
        'item_name': d.item.name,
        'dispatch_date': d.dispatch_date.strftime('%d %b'),
        'quantity': d.quantity,
    } for d in transit_qs]

    if request.GET.get('export') == 'xlsx':
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse

        filename = f"MR_Inventory_{selected_emp.name}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stock"
        ws.append(['MR INVENTORY — STOCK REPORT'])
        ws.append(['Employee:', selected_emp.name])
        ws.append([''])
        ws['A1'].font = Font(bold=True, size=14, color="107C41")
        ws['A2'].font = Font(bold=True)

        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")

        headers = ['Item Name', 'Category', 'Received', 'Distributed', 'Balance']
        ws.append(headers)
        for col_num, cell in enumerate(ws[4], 1):
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 25 if col_num == 1 else 15

        for row in sample_gift_data:
            ws.append([row['item_name'], row['category'], row['received'], row['distributed'], row['balance']])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return Response({
        'team_employees': [{'id': e.id, 'name': e.name} for e in team_employees] if is_manager_view else [],
        'is_manager_view': is_manager_view,
        'selected_emp_id': selected_emp_id,
        'sample_gift_data': sample_gift_data,
        'hv_stock_data': hv_stock_data,
        'hv_tracker': hv_tracker,
        'in_transit_items': in_transit_items,
    })
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_receive_dispatch(request):
    from SFA.models import PromoDispatch, MRInventory
    from django.utils import timezone

    try:
        employee = request.user.employee
    except AttributeError:
        return Response({'error': 'Employee profile missing'}, status=400)

    dispatch_id = request.data.get('dispatch_id')
    try:
        dispatch = PromoDispatch.objects.get(id=dispatch_id, employee=employee, status='In-Transit')
    except (PromoDispatch.DoesNotExist, ValueError, TypeError):
        return Response({'error': 'Dispatch not found or already received'}, status=404)

    dispatch.status = 'Received'
    dispatch.received_date = timezone.now().date()
    dispatch.save()

    stock, created = MRInventory.objects.get_or_create(employee=employee, item=dispatch.item, defaults={'stock_qty': 0})
    stock.stock_qty += dispatch.quantity
    stock.save()

    return Response({'success': True, 'message': f'{dispatch.item.name} received successfully'})
